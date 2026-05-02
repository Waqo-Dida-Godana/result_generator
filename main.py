# -*- coding: utf-8 -*-
"""
MOAS School Management System - Complete Merged App
Pure Tkinter • Modern UI • Full Features • Legacy DB Compatible
"""

import os
import random
import shutil
import sys
import json
import re
import uuid
import smtplib
import tempfile
import threading
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
from email.message import EmailMessage
from PIL import Image, ImageTk
from database import db
from extract_letterhead import extract_letterhead
from promotion import promotion_manager
from exam_analytics import exam_analytics
import matplotlib

matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import csv
import pandas as pd
from datetime import datetime
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, portrait, landscape
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image as RLImage,
    PageBreak,
    CondPageBreak,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.graphics.shapes import Drawing, Rect
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader
from fpdf import FPDF

from students_tab import StudentsTab

# ====================== DESIGN TOKENS ======================
# Lemon green + olive theme
LEMON_ACCENT = "#D7F171"
LEMON_SOFT = "#EEF7C7"
OLIVE_PRIMARY = "#6B764B"
OLIVE_DARK = "#55603A"
OLIVE_MID = "#889660"
CREAM_BG = "#F7F8EE"

SIDEBAR_BG = OLIVE_PRIMARY
SIDEBAR_ACTIVE = OLIVE_DARK
SIDEBAR_HOVER = OLIVE_MID
SIDEBAR_TEXT = LEMON_ACCENT
SIDEBAR_TEXT_ACT = "#ffffff"  # White

CONTENT_BG = CREAM_BG

# Card themes: (border_rgb_hex, inner_surface_hex) — soft tinted panels for variety
CARD_THEMES = {
    "cream": ("#bec9a0", "#fffef9"),
    "mint": ("#7fb89e", "#eef8f2"),
    "sky": ("#8fa0d4", "#f2f4ff"),
    "peach": ("#d4a088", "#fff3eb"),
    "lilac": ("#a894c9", "#f4effc"),
    "sand": ("#c9b87a", "#fff8e8"),
    "azure": ("#6dadc4", "#edf7fa"),
    "blossom": ("#c97fa0", "#fcedf4"),
}


def _card_colors(theme_key):
    return CARD_THEMES.get(str(theme_key), CARD_THEMES["cream"])


BORDER_CLR, CARD_BG = _card_colors("cream")

TEXT_PRIMARY = "#2f3521"
TEXT_SECONDARY = "#66704b"

BLUE = LEMON_ACCENT
GREEN = OLIVE_PRIMARY
ORANGE = "#C7E36A"
PURPLE = OLIVE_MID

GRADE_COLORS = {
    "EE": "#2ecc71",
    "ME": "#3498db",
    "AE": "#f39c12",
    "BE": "#e67e22",
    "IE": "#e74c3c",
}
DEFAULT_GRADE_LABELS = {
    "EE": "Exceeding Expectations",
    "ME": "Meeting Expectations",
    "AE": "Approaching Expectations",
    "BE": "Below Expectations",
    "IE": "Inadequate",
}
GRADE_LABELS = DEFAULT_GRADE_LABELS.copy()

SUBJECT_PALETTE = [
    "#2E7D32",
    "#1565C0",
    "#EF6C00",
    "#8E24AA",
    "#C62828",
    "#00897B",
    "#6D4C41",
    "#3949AB",
    "#7CB342",
    "#F9A825",
    "#5E35B1",
    "#D81B60",
    "#039BE5",
    "#43A047",
    "#FB8C00",
    "#546E7A",
    "#00ACC1",
    "#9E9D24",
    "#8D6E63",
    "#1E88E5",
]


def grade_base_code(grade_code):
    match = re.match(r"[A-Za-z]+", str(grade_code or "").upper())
    return match.group(0) if match else "IE"


def _hex_to_rgb(value):
    value = str(value or "").strip().lstrip("#")
    if len(value) != 6:
        return (0, 0, 0)
    return tuple(int(value[i : i + 2], 16) for i in (0, 2, 4))


def _rgb_to_hex(rgb):
    r, g, b = [max(0, min(255, int(v))) for v in rgb]
    return f"#{r:02x}{g:02x}{b:02x}"


def _mix_hex(color1, color2, ratio):
    ratio = max(0.0, min(1.0, float(ratio)))
    rgb1 = _hex_to_rgb(color1)
    rgb2 = _hex_to_rgb(color2)
    return _rgb_to_hex(tuple(rgb1[i] * (1 - ratio) + rgb2[i] * ratio for i in range(3)))


class ImportCancelledError(Exception):
    """Raised when a long-running import is cancelled by the user."""


FF = "Segoe UI"  # font family

# ====================== TOPBAR TOKENS ======================
TOPBAR_H = 50  # height in px
TOPBAR_RIGHT_BG = OLIVE_DARK
TOPBAR_BTN_BG = OLIVE_PRIMARY
TOPBAR_BTN_HOV = OLIVE_MID
TOPBAR_ICON_FG = "#ffffff"
TOPBAR_USER_BG = OLIVE_PRIMARY
TOPBAR_YR_BG = OLIVE_MID

# ====================== CBC LEVELS & SUBJECTS ======================
# Kenyan Competency Based Curriculum (CBC) Structure

# Education Levels
DEFAULT_LEVELS = [
    "Pre-Primary (PP1-PP2)",
    "Lower Primary (Grade 1-3)",
    "Upper Primary (Grade 4-6)",
    "Junior School (Grade 7-9)",
]
LEVELS = []
ALL_SCHOOL_LEVEL = "All School (All Levels)"

# ====================== RUNTIME-LOADED CONFIG FROM DATABASE =====================
# These globals are populated by refresh_dynamic_school_config() at startup
EXAM_TYPES = []
CBC_GRADE_LEVELS = {}
LEVEL_ORDER = []
SUBJECT_SHORT_NAMES = {}
SCHOOL_PROFILE = {}

# ====================== DEFAULT SEED DATA =====================
# Fallback defaults used only if database is empty; serve as seed on first run
DEFAULT_EXAM_TYPES = ["Opener", "Mid-Term", "End-Term"]

DEFAULT_CBC_GRADE_LEVELS = {
    "EE1": {"min": 90, "max": 100},
    "EE2": {"min": 75, "max": 89},
    "ME1": {"min": 60, "max": 74},
    "ME2": {"min": 50, "max": 59},
    "AE1": {"min": 35, "max": 49},
    "AE2": {"min": 25, "max": 34},
    "BE1": {"min": 12, "max": 24},
    "BE2": {"min": 0, "max": 11},
}

DEFAULT_LEVEL_ORDER = [
    "Pre-Primary (PP1-PP2)",
    "Lower Primary (Grade 1-3)",
    "Upper Primary (Grade 4-6)",
    "Junior School (Grade 7-9)",
]

LEGACY_DEFAULT_SUBJECT_SHORT_NAMES = {
    "Kiswahili / Kenyan Sign Language": "Kiswahili /\\nKSL",
    "Integrated Science": "Integrated\\nScience",
    "Health Education": "Health\\nEducation",
    "Pre-Technical Studies": "Pre-Technical\\nStudies",
    "Social Studies": "Social\\nStudies",
    "Religious Education (CRE/IRE/HRE)": "Religious\\nEducation",
    "Sports & Physical Education": "Sports &\\nPHE",
    "Life Skills Education": "Life Skills\\nEducation",
    "Visual Arts": "Visual\\nArts",
    "Performing Arts": "Performing\\nArts",
    "Foreign Languages (French, German, Arabic)": "Foreign\\nLanguages",
    "Science & Technology": "Science &\\nTechnology",
    "Christian Religious Education (CRE)": "CRE",
    "Islamic Religious Education (IRE)": "IRE",
    "Hindu Religious Education (HRE)": "HRE",
    "English": "English",
    "Kiswahili / KSL": "Kiswahili",
    "Mathematics": "Math",
    "Agriculture": "Agriculture",
    "Computer Science": "Computer\\nScience",
    "French": "French",
    "German": "German",
    "Arabic": "Arabic",
    "Kenyan Sign Language": "KSL",
}

DEFAULT_SUBJECT_SHORT_NAMES = {
    "Mathematics": "Math",
    "English": "English",
    "Kiswahili": "Kiswahili",
    "Integrated Science": "Integrated Science",
    "Agriculture": "Agriculture",
    "Social Studies": "Social Studies",
    "Christian Religious Education (CRE)": "CRE",
    "Creative Arts & Sports": "C/A",
    "Pre-Technical Studies": "Pre-Technical Studies",
    "French": "French",
    "Language": "Lang",
    "Environmental": "Envi",
}

DEFAULT_SCHOOL_PROFILE = {
    "school_name": "MT OLIVES ADVENTIST SCHOOL",
    "school_address": "Sajin Close, Along Ngong-Matasia Road, Next to Oryx Petrol Station, Ngong",
    "school_contact_line": "school@mountolivessda.org | +254 788 700073 | https://mountolivessda.org/",
    "school_motto": "In God We Excel",
    "school_location": "Nairobi, Kenya",
    "school_app_title": "MT OLIVES ADVENTIST SCHOOL, NGONG",
    "school_sidebar_title": "MT OLIVES",
    "school_sidebar_subtitle": "ADVENTIST SCHOOL",
}

# Dynamic database-backed helpers

def _load_grade_labels_from_db():
    labels = {}
    for row in db.get_grading_scales():
        code = str(row.get("grade_code") or "").strip().upper()
        name = str(row.get("grade_name") or "").strip()
        if code and name:
            labels.setdefault(code, name)
    for code, label in DEFAULT_GRADE_LABELS.items():
        labels.setdefault(code, label)
    return labels


def _load_exam_types_from_db():
    """Load exam types from database. Fall back to defaults if not configured."""
    exam_types_json = db.get_setting("exam_types", "")
    if exam_types_json:
        try:
            return json.loads(exam_types_json)
        except (json.JSONDecodeError, ValueError):
            pass
    return DEFAULT_EXAM_TYPES[:]


def _load_cbc_grade_levels_from_db():
    """Load CBC grade level ranges from database. Fall back to defaults if not configured."""
    cbc_levels_json = db.get_setting("cbc_grade_levels", "")
    if cbc_levels_json:
        try:
            return json.loads(cbc_levels_json)
        except (json.JSONDecodeError, ValueError):
            pass
    return DEFAULT_CBC_GRADE_LEVELS.copy()


def _load_level_order_from_db():
    """Load the order of levels from database. Fall back to defaults if not configured."""
    level_order_json = db.get_setting("level_order", "")
    if level_order_json:
        try:
            return json.loads(level_order_json)
        except (json.JSONDecodeError, ValueError):
            pass
    return DEFAULT_LEVEL_ORDER[:]


def _load_subject_short_names_from_db():
    """Load subject short/display names from database. Fall back to defaults if not configured."""
    subject_names_json = db.get_setting("subject_short_names", "")
    if subject_names_json:
        try:
            db_names = json.loads(subject_names_json)
            if db_names == LEGACY_DEFAULT_SUBJECT_SHORT_NAMES:
                db.set_setting(
                    "subject_short_names",
                    json.dumps(DEFAULT_SUBJECT_SHORT_NAMES),
                )
                return DEFAULT_SUBJECT_SHORT_NAMES.copy()
            if isinstance(db_names, dict) and db_names:
                return {
                    str(key).strip(): str(value).strip()
                    for key, value in db_names.items()
                    if str(key).strip()
                }
        except (json.JSONDecodeError, ValueError):
            pass
    return DEFAULT_SUBJECT_SHORT_NAMES.copy()


def _load_school_profile_from_db():
    profile = DEFAULT_SCHOOL_PROFILE.copy()
    for key, default_value in DEFAULT_SCHOOL_PROFILE.items():
        value = str(db.get_setting(key, "") or "").strip()
        if value:
            profile[key] = value
    return profile


def _seed_school_config_to_db():
    """Seed default school configuration to database if not already set.
    Called during app initialization to populate defaults.
    """
    # Seed exam types if not already configured
    if not db.get_setting("exam_types", ""):
        db.set_setting("exam_types", json.dumps(DEFAULT_EXAM_TYPES))
    
    # Seed CBC grade levels if not already configured
    if not db.get_setting("cbc_grade_levels", ""):
        db.set_setting("cbc_grade_levels", json.dumps(DEFAULT_CBC_GRADE_LEVELS))
    
    # Seed level order if not already configured
    if not db.get_setting("level_order", ""):
        db.set_setting("level_order", json.dumps(DEFAULT_LEVEL_ORDER))
    
    # Seed subject short names if not already configured
    if not db.get_setting("subject_short_names", ""):
        db.set_setting("subject_short_names", json.dumps(DEFAULT_SUBJECT_SHORT_NAMES))

    for key, default_value in DEFAULT_SCHOOL_PROFILE.items():
        if not db.get_setting(key, ""):
            db.set_setting(key, str(default_value))


def _sort_class_name(name):
    match = re.search(r"(\d+)", str(name or ""))
    return (int(match.group(1)) if match else 999, str(name or ""))


def _is_legacy_subject_level(level):
    normalized = str(level or "").strip()
    return normalized in {
        "Primary",
        "Pre-Primary",
        "Lower Primary",
        "Upper Primary",
        "Junior Secondary",
    }


def _canonicalize_subject_level(level, fallback=""):
    normalized = str(level or "").strip()
    if not normalized:
        return str(fallback or "").strip()
    if normalized in LEVELS or normalized == ALL_SUBJECT_LEVEL:
        return normalized

    lookup = {
        "pre-primary": "Pre-Primary (PP1-PP2)",
        "pre primary": "Pre-Primary (PP1-PP2)",
        "lower primary": "Lower Primary (Grade 1-3)",
        "upper primary": "Upper Primary (Grade 4-6)",
        "junior secondary": "Junior School (Grade 7-9)",
        "junior school": "Junior School (Grade 7-9)",
    }
    lowered = normalized.lower()
    if lowered == "primary":
        return str(fallback or "").strip() or "Lower Primary (Grade 1-3)"
    return lookup.get(lowered, normalized)


def _build_subject_catalog_from_db():
    rows = db.get_subjects_by_level()
    if not rows:
        return {}

    catalog = {}
    for row in rows:
        level = str(row.get("level") or "").strip()
        if not level or _is_legacy_subject_level(level):
            continue
        entry = (
            str(row.get("code") or "").strip(),
            str(row.get("name") or "").strip(),
            str(row.get("category") or "Core").strip(),
            bool(int(row.get("is_optional") or 0)),
        )
        level_entries = catalog.setdefault(level, [])
        level_entries.append(entry)

    normalized = {}
    for level, entries in catalog.items():
        core = [entry for entry in entries if not entry[3] and entry[2].strip().lower() != "optional"]
        optional = [entry for entry in entries if entry[3] or entry[2].strip().lower() == "optional"]
        normalized[level] = {"core": core, "optional": optional} if optional else core
    return normalized


def _build_subjects_by_level_from_db(subject_catalog):
    subjects_by_level = {}
    for level, entries in subject_catalog.items():
        if isinstance(entries, dict):
            subjects_by_level[level] = [
                name for _, name, _, _ in entries.get("core", [])
            ] + [name for _, name, _, _ in entries.get("optional", [])]
        else:
            subjects_by_level[level] = [name for _, name, _, _ in entries]
    return subjects_by_level


def _load_levels_from_db():
    levels = []
    for row in db.get_all_classes():
        level = str(row.get("level") or "").strip()
        if level and level not in levels:
            levels.append(level)
    for row in db.get_subjects_by_level():
        level = str(row.get("level") or "").strip()
        if level and not _is_legacy_subject_level(level) and level not in levels:
            levels.append(level)
    if not levels:
        return DEFAULT_LEVELS

    ordered = [level for level in LEVEL_ORDER if level in levels]
    ordered += [level for level in levels if level not in ordered]
    return ordered


def _load_classes_by_level_from_db():
    rows = db.get_all_classes()
    if not rows:
        return DEFAULT_CLASSES_BY_LEVEL

    classes_by_level = {}
    for row in rows:
        level = str(row.get("level") or "").strip()
        name = str(row.get("name") or "").strip()
        if not level or not name:
            continue
        classes_by_level.setdefault(level, []).append(name)

    for level, names in classes_by_level.items():
        classes_by_level[level] = sorted(set(names), key=_sort_class_name)
    return classes_by_level


def refresh_dynamic_school_config():
    global LEVELS, LEVEL_ORDER, SUBJECT_CATALOG, SUBJECTS_BY_LEVEL, CLASSES_BY_LEVEL, ALL_CLASSES, GRADE_LABELS
    global EXAM_TYPES, CBC_GRADE_LEVELS, SUBJECT_SHORT_NAMES, SCHOOL_PROFILE

    LEVEL_ORDER = _load_level_order_from_db()
    LEVELS = _load_levels_from_db()
    CLASSES_BY_LEVEL = _load_classes_by_level_from_db()
    ALL_CLASSES = [
        class_name
        for level_classes in CLASSES_BY_LEVEL.values()
        for class_name in level_classes
    ]
    SUBJECT_CATALOG = _build_subject_catalog_from_db()
    SUBJECTS_BY_LEVEL = _build_subjects_by_level_from_db(SUBJECT_CATALOG)
    GRADE_LABELS = _load_grade_labels_from_db()
    EXAM_TYPES = _load_exam_types_from_db()
    CBC_GRADE_LEVELS = _load_cbc_grade_levels_from_db()
    SUBJECT_SHORT_NAMES = _load_subject_short_names_from_db()
    SCHOOL_PROFILE = _load_school_profile_from_db()


# Subjects by Level
DEFAULT_SUBJECT_CATALOG = {
    "Pre-Primary (PP1-PP2)": [
        ("LANG", "Language Activities", "Core", False),
        ("MATH", "Mathematical Activities", "Core", False),
        ("CREA", "Creative Activities", "Core", False),
        ("ENVT", "Environmental Activities", "Core", False),
        ("RELG", "Religious Activities", "Core", False),
        ("PRIP", "Pastoral/ Religious Instruction Programme", "Core", False),
    ],
    "Lower Primary (Grade 1-3)": [
        ("LIT", "Literacy Activities", "Core", False),
        ("ENG", "English Activities", "Core", False),
        ("KIS", "Kiswahili Activities", "Core", False),
        ("MAT", "Mathematical Activities", "Core", False),
        ("ENV", "Environmental Activities", "Core", False),
        ("HNA", "Hygiene & Nutrition Activities", "Core", False),
        ("CRE", "Religious Education Activities", "Core", False),
        ("MCA", "Movement & Creative Activities", "Core", False),
    ],
    "Upper Primary (Grade 4-6)": [
        ("ENG", "English", "Core", False),
        ("KIS", "Kiswahili / KSL", "Core", False),
        ("MAT", "Mathematics", "Core", False),
        ("SCI", "Science & Technology", "Core", False),
        ("AGR", "Agriculture", "Core", False),
        ("SST", "Social Studies", "Core", False),
        ("CRE", "Christian Religious Education", "Core", False),
        ("IRE", "Islamic Religious Education", "Core", False),
        ("HRE", "Hindu Religious Education", "Core", False),
        ("ART", "Creative Arts", "Core", False),
        ("PHE", "Physical & Health Education", "Core", False),
        ("HSC", "Home Science", "Core", False),
    ],
    "Junior School (Grade 7-9)": {
        "core": [
            ("ENG", "English", "Core", False),
            ("KIS", "Kiswahili / KSL", "Core", False),
            ("MAT", "Mathematics", "Core", False),
            ("INT", "Integrated Science", "Core", False),
            ("SST", "Social Studies", "Core", False),
            ("CRE", "Christian Religious Education", "Core", False),
            ("IRE", "Islamic Religious Education", "Core", False),
            ("HRE", "Hindu Religious Education", "Core", False),
            ("AGR", "Agriculture", "Core", False),
            ("PTS", "Pre-Technical Studies", "Core", False),
            ("LSE", "Life Skills Education", "Core", False),
            ("HEA", "Health Education", "Core", False),
            ("SPE", "Sports & Physical Education", "Core", False),
            ("VIA", "Visual Arts", "Core", False),
            ("PFA", "Performing Arts", "Core", False),
        ],
        "optional": [
            ("CSC", "Computer Science", "Optional", True),
            ("FRE", "French", "Optional", True),
            ("GER", "German", "Optional", True),
            ("ARB", "Arabic", "Optional", True),
            ("KSL", "Kenyan Sign Language", "Optional", True),
        ],
    },
}

# Seed import-time globals with empty runtime state; refresh_dynamic_school_config()
# replaces these with database-backed values once module setup is complete.
SUBJECT_CATALOG = {}
SUBJECTS_BY_LEVEL = _build_subjects_by_level_from_db(SUBJECT_CATALOG)

# Classes by Level
DEFAULT_CLASSES_BY_LEVEL = {
    "Pre-Primary (PP1-PP2)": ["PP1", "PP2"],
    "Lower Primary (Grade 1-3)": ["Grade 1", "Grade 2", "Grade 3"],
    "Upper Primary (Grade 4-6)": ["Grade 4", "Grade 5", "Grade 6"],
    "Junior School (Grade 7-9)": ["Grade 7", "Grade 8", "Grade 9"],
}
CLASSES_BY_LEVEL = {}
ALL_CLASSES = []

# ====================== LEVEL DISPLAY CONFIG =======================
LEVEL_ORDER = [
    "Pre-Primary (PP1-PP2)",
    "Lower Primary (Grade 1-3)",
    "Upper Primary (Grade 4-6)",
    "Junior School (Grade 7-9)",
]

# Load any configured school structure from the database immediately on import.
# This keeps module-level globals in sync with persisted class/subject/grade data.
refresh_dynamic_school_config()

LEVEL_DISPLAY = {
    "Pre-Primary (PP1-PP2)": {
        "short": "Pre-Primary",
        "icon": "🌱",
        "hdr_bg": "#9d174d",
        "hdr_fg": "#ffffff",
        "row_bg": "#fce7f3",
        "btn_active_bg": "#be185d",
        "tag": "lv_pp",
        "hdr_tag": "lv_pp_hdr",
    },
    "Lower Primary (Grade 1-3)": {
        "short": "Lower",
        "icon": "📚",
        "hdr_bg": "#1e40af",
        "hdr_fg": "#ffffff",
        "row_bg": "#dbeafe",
        "btn_active_bg": "#2563eb",
        "tag": "lv_lower",
        "hdr_tag": "lv_lower_hdr",
    },
    "Upper Primary (Grade 4-6)": {
        "short": "Upper",
        "icon": "🎓",
        "hdr_bg": "#166534",
        "hdr_fg": "#ffffff",
        "row_bg": "#dcfce7",
        "btn_active_bg": "#16a34a",
        "tag": "lv_upper",
        "hdr_tag": "lv_upper_hdr",
    },
    "Junior School (Grade 7-9)": {
        "short": "Junior",
        "icon": "🏛",
        "hdr_bg": "#5b21b6",
        "hdr_fg": "#ffffff",
        "row_bg": "#ede9fe",
        "btn_active_bg": "#7c3aed",
        "tag": "lv_junior",
        "hdr_tag": "lv_junior_hdr",
    },
}

# Grading System by Level (CBC Competency Levels)
GRADING_BY_LEVEL = {
    "Pre-Primary (PP1-PP2)": {
        "levels": {
            "EE": {
                "label": "Exceeding Expectation",
                "description": "Learner performs above the required level",
                "points": 4,
            },
            "ME": {
                "label": "Meeting Expectation",
                "description": "Learner understands and performs well",
                "points": 3,
            },
            "AE": {
                "label": "Approaching Expectation",
                "description": "Learner is improving but needs support",
                "points": 2,
            },
            "BE": {
                "label": "Below Expectation",
                "description": "Learner needs more help",
                "points": 1,
            },
        }
    },
    "Lower Primary (Grade 1-3)": {
        "levels": {
            "EE": {
                "label": "Exceeding Expectation",
                "description": "Learner performs above the required level",
            },
            "ME": {
                "label": "Meeting Expectation",
                "description": "Learner understands and performs well",
            },
            "AE": {
                "label": "Approaching Expectation",
                "description": "Learner is improving but needs support",
            },
            "BE": {
                "label": "Below Expectation",
                "description": "Learner needs more help",
            },
        },
        "assessment_methods": "Class activities, Oral work, Practical activities, Teacher observation",
    },
    "Upper Primary (Grade 4-6)": {
        "levels": {
            "EE": {
                "label": "Exceeding Expectation",
                "description": "Learner performs above the required level",
            },
            "ME": {
                "label": "Meeting Expectation",
                "description": "Learner understands and performs well",
            },
            "AE": {
                "label": "Approaching Expectation",
                "description": "Learner is improving but needs support",
            },
            "BE": {
                "label": "Below Expectation",
                "description": "Learner needs more help",
            },
        },
        "assessment_components": "60% School Based Assessment (SBA) + 40% National Assessment (KNEC)",
    },
    "Junior School (Grade 7-9)": {
        "levels": {
            "EE": {
                "label": "Exceeding Expectation",
                "description": "Learner performs above the required level",
            },
            "ME": {
                "label": "Meeting Expectation",
                "description": "Learner understands and performs well",
            },
            "AE": {
                "label": "Approaching Expectation",
                "description": "Learner is improving but needs support",
            },
            "BE": {
                "label": "Below Expectation",
                "description": "Learner needs more help",
            },
        },
        "uses_percentage": True,
        "description": "Competency levels with percentage scores",
    },
}

# Legacy compatibility - Default to all-school view
SUBJECTS = []
CLASSES = list(ALL_CLASSES)

# ====================== CONSTANTS ==========================
TERMS = ["One", "Two", "Three"]
# EXAM_TYPES is now loaded dynamically from the database via refresh_dynamic_school_config()
DEFAULT_EXAM_TYPE = "End-Term"
EMAIL_SETTING_KEYS = [
    "smtp_host",
    "smtp_port",
    "smtp_username",
    "smtp_password",
    "smtp_sender_name",
    "smtp_use_tls",
]

CLASS_SUBJECTS_DONE_KEY = "class_subjects_done_map"
ALL_SUBJECT_LEVEL = "All Levels"
COLORS = {
    "primary": "#4CAF50",
    "sidebar": "#1b5e20",
    "card": "#ffffff",
    "text": "#333333",
    "text_sec": "#666666",
    "border": "#e0e0e0",
}


# ====================== CBC GRADE SUB-LEVELS ===============
def get_cbc_grade_sublevel(mark):
    """Return the CBC grade sub-level string (EE1, EE2, ME1, ME2, AE1, AE2, BE1, BE2).
    Uses database-backed CBC_GRADE_LEVELS if available, falls back to defaults.
    """
    try:
        mark = int(mark)
    except (TypeError, ValueError):
        return ""
    
    # Use the dynamic CBC_GRADE_LEVELS configuration
    levels = CBC_GRADE_LEVELS if CBC_GRADE_LEVELS else DEFAULT_CBC_GRADE_LEVELS
    
    for grade_code in ["EE1", "EE2", "ME1", "ME2", "AE1", "AE2", "BE1", "BE2"]:
        if grade_code in levels:
            level_range = levels[grade_code]
            if level_range["min"] <= mark <= level_range["max"]:
                return grade_code
    
    return ""


def get_grade_code(mark):
    """Return the CBC performance code for a numeric mark."""
    try:
        mark = float(mark)
    except (TypeError, ValueError):
        return "IE"
    if mark >= 80:
        return "EE"
    if mark >= 70:
        return "ME"
    if mark >= 60:
        return "AE"
    if mark >= 50:
        return "BE"
    return "IE"


def ensure_letterhead_assets():
    """Refresh extracted letterhead assets when the DOCX template changes."""
    docx_path = os.path.join("assets", "letterhead.docx")
    png_path = os.path.join("assets", "letterhead.png")
    footer_png_path = os.path.join("assets", "letterhead_footer.png")
    json_path = os.path.join("assets", "letterhead.json")
    if not os.path.exists(docx_path):
        return png_path if os.path.exists(png_path) else None

    try:
        needs_extract = (
            not os.path.exists(png_path)
            or not os.path.exists(footer_png_path)
            or not os.path.exists(json_path)
            or os.path.getmtime(docx_path) > os.path.getmtime(footer_png_path)
            or os.path.getmtime(docx_path) > os.path.getmtime(png_path)
            or os.path.getmtime(docx_path) > os.path.getmtime(json_path)
        )
        if needs_extract:
            extract_letterhead(docx_path=docx_path, output_dir="assets")
    except Exception as exc:
        print(f"Letterhead extraction failed: {exc}")

    return png_path if os.path.exists(png_path) else None


def get_letterhead_assets():
    """Return extracted header/footer image paths and text metadata."""
    header_path = ensure_letterhead_assets()
    footer_path = os.path.join("assets", "letterhead_footer.png")
    json_path = os.path.join("assets", "letterhead.json")
    data = {}

    try:
        if os.path.exists(json_path):
            with open(json_path, "r", encoding="utf-8") as handle:
                data = json.load(handle)
    except Exception as exc:
        print(f"Letterhead metadata read failed: {exc}")

    return {
        "header_path": header_path
        if header_path and os.path.exists(header_path)
        else None,
        "footer_path": footer_path if os.path.exists(footer_path) else None,
        "header_lines": [
            str(line).strip()
            for line in data.get("header_lines", [])
            if str(line).strip()
        ],
        "footer_lines": [
            str(line).strip()
            for line in data.get("footer_lines", [])
            if str(line).strip()
        ],
    }


def get_processed_letterhead_image(image_path, section="header"):
    """Return a cleaned PIL image for rendering letterhead assets."""
    if not image_path or not os.path.exists(image_path):
        return None

    img = Image.open(image_path).copy()
    if section == "header":
        # Trim the thin separator line embedded at the bottom of the header image.
        trim_bottom = max(6, int(img.height * 0.09))
        if img.height - trim_bottom > 20:
            img = img.crop((0, 0, img.width, img.height - trim_bottom))
    return img


def get_letterhead_print_lines():
    """Return text lines for printed report headers/footers."""
    profile = SCHOOL_PROFILE if SCHOOL_PROFILE else DEFAULT_SCHOOL_PROFILE
    default_lines = [
        profile.get("school_name", DEFAULT_SCHOOL_PROFILE["school_name"]),
        profile.get("school_address", DEFAULT_SCHOOL_PROFILE["school_address"]),
        profile.get(
            "school_contact_line", DEFAULT_SCHOOL_PROFILE["school_contact_line"]
        ),
        profile.get("school_motto", DEFAULT_SCHOOL_PROFILE["school_motto"]),
    ]

    assets = get_letterhead_assets()
    lines = assets["header_lines"] + assets["footer_lines"]
    if lines:
        return lines
    return default_lines


def get_school_profile():
    profile = SCHOOL_PROFILE if SCHOOL_PROFILE else DEFAULT_SCHOOL_PROFILE
    return profile.copy()


# ====================== HELPERS ============================
def _rr(canvas, x0, y0, x1, y1, r, fill, outline=None):
    """Draw a rounded rectangle on a Canvas."""
    oc = outline or fill
    canvas.create_arc(
        x0, y0, x0 + 2 * r, y0 + 2 * r, start=90, extent=90, fill=fill, outline=oc
    )
    canvas.create_arc(
        x1 - 2 * r, y0, x1, y0 + 2 * r, start=0, extent=90, fill=fill, outline=oc
    )
    canvas.create_arc(
        x0, y1 - 2 * r, x0 + 2 * r, y1, start=180, extent=90, fill=fill, outline=oc
    )
    canvas.create_arc(
        x1 - 2 * r, y1 - 2 * r, x1, y1, start=270, extent=90, fill=fill, outline=oc
    )
    canvas.create_rectangle(x0 + r, y0, x1 - r, y1, fill=fill, outline=fill)
    canvas.create_rectangle(x0, y0 + r, x1, y1 - r, fill=fill, outline=fill)


def rounded_badge(parent, icon_text, color, size=40):
    """Draw a coloured rounded-square icon badge on a Canvas."""
    c = tk.Canvas(parent, width=size, height=size, bg=CARD_BG, highlightthickness=0)
    r = 9
    x0, y0, x1, y1 = 2, 2, size - 2, size - 2
    # fill rounded rect via overlapping shapes
    c.create_arc(
        x0, y0, x0 + 2 * r, y0 + 2 * r, start=90, extent=90, fill=color, outline=color
    )
    c.create_arc(
        x1 - 2 * r, y0, x1, y0 + 2 * r, start=0, extent=90, fill=color, outline=color
    )
    c.create_arc(
        x0, y1 - 2 * r, x0 + 2 * r, y1, start=180, extent=90, fill=color, outline=color
    )
    c.create_arc(
        x1 - 2 * r, y1 - 2 * r, x1, y1, start=270, extent=90, fill=color, outline=color
    )
    c.create_rectangle(x0 + r, y0, x1 - r, y1, fill=color, outline=color)
    c.create_rectangle(x0, y0 + r, x1, y1 - r, fill=color, outline=color)
    c.create_text(
        size // 2, size // 2, text=icon_text, fill="white", font=(FF, size // 3, "bold")
    )
    return c


def make_card(parent, padx=20, pady=16, theme="cream", **grid_or_pack):
    """Return a bordered card frame; ``theme`` selects from ``CARD_THEMES``."""
    ob, ib = _card_colors(theme)
    outer = tk.Frame(parent, bg=ob)
    inner = tk.Frame(outer, bg=ib, padx=padx, pady=pady)
    inner.pack(fill="both", expand=True, padx=1, pady=1)
    return outer, inner


def scrollable_frame(parent, bg=CONTENT_BG):
    """Return (canvas, scrollbar, inner_frame) for a vertically scrollable area."""
    canvas = tk.Canvas(parent, bg=bg, highlightthickness=0)
    sb = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=sb.set)
    sb.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)
    inner = tk.Frame(canvas, bg=bg)
    win = canvas.create_window((0, 0), window=inner, anchor="nw")

    def _on_resize(e):
        canvas.itemconfig(win, width=e.width)

    canvas.bind("<Configure>", _on_resize)

    def _on_frame_resize(e):
        canvas.configure(scrollregion=canvas.bbox("all"))

    inner.bind("<Configure>", _on_frame_resize)

    _install_canvas_mousewheel(canvas)

    return canvas, sb, inner


def scrollable_frame_both(parent, bg=CONTENT_BG):
    """Return (canvas, v_scrollbar, inner_frame, h_scrollbar) for two-way scrolling."""
    outer = tk.Frame(parent, bg=bg)
    outer.pack(fill="both", expand=True)

    canvas = tk.Canvas(outer, bg=bg, highlightthickness=0)
    vbar = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
    hbar = ttk.Scrollbar(parent, orient="horizontal", command=canvas.xview)
    canvas.configure(yscrollcommand=vbar.set, xscrollcommand=hbar.set)

    vbar.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)
    hbar.pack(fill="x", side="bottom")

    inner = tk.Frame(canvas, bg=bg)
    win = canvas.create_window((0, 0), window=inner, anchor="nw")

    def _update_scrollregion(event=None):
        canvas.configure(scrollregion=canvas.bbox("all"))

    def _on_resize(e):
        min_width = inner.winfo_reqwidth()
        canvas.itemconfig(win, width=max(e.width, min_width))
        _update_scrollregion()

    canvas.bind("<Configure>", _on_resize)
    inner.bind("<Configure>", _update_scrollregion)

    _install_canvas_mousewheel(canvas, horizontal=True)

    return canvas, vbar, inner, hbar


def create_content_frame(parent, **kwargs):
    """Create a standardized content frame with CONTENT_BG background."""
    return tk.Frame(parent, bg=CONTENT_BG, **kwargs)


def create_card_frame(parent, **kwargs):
    """Create a standardized card frame with CARD_BG background."""
    return tk.Frame(parent, bg=CARD_BG, **kwargs)


def create_toolbar_button(parent, text, command, bg=BLUE, fg="white", **kwargs):
    """Create a standardized toolbar button."""
    return tk.Label(
        parent,
        text=text,
        bg=bg,
        fg=fg,
        font=(FF, 10, "bold"),
        padx=12,
        pady=6,
        cursor="hand2",
        **kwargs
    )


def create_section_header(parent, title, subtitle=""):
    """Create a standardized section header."""
    hdr = tk.Frame(parent, bg=CONTENT_BG)
    hdr.pack(fill="x", pady=(0, 16))
    
    top = tk.Frame(hdr, bg=CONTENT_BG)
    top.pack(fill="x")
    
    tk.Label(
        top,
        text=title,
        bg=CONTENT_BG,
        fg=TEXT_PRIMARY,
        font=(FF, 16, "bold"),
    ).pack(side="left")
    
    if subtitle:
        tk.Label(
            top,
            text=subtitle,
            bg=CONTENT_BG,
            fg=TEXT_SECONDARY,
            font=(FF, 10),
        ).pack(side="left", padx=(12, 0))
    
    tk.Frame(hdr, bg=BORDER_CLR, height=1).pack(fill="x", pady=(8, 0))
    
    return hdr


def _install_canvas_mousewheel(canvas, horizontal=False):
    """Bind mousewheel only while a canvas is active, and clean up on destroy."""

    def _scroll(units, axis="y"):
        try:
            if not canvas.winfo_exists():
                return "break"
            if axis == "x":
                canvas.xview_scroll(units, "units")
            else:
                canvas.yview_scroll(units, "units")
        except tk.TclError:
            return "break"
        return "break"

    def _on_mousewheel(event):
        delta = int(-1 * (getattr(event, "delta", 0) / 120))
        if delta:
            return _scroll(delta, "y")
        return "break"

    def _on_shift_mousewheel(event):
        delta = int(-1 * (getattr(event, "delta", 0) / 120))
        if delta:
            return _scroll(delta, "x")
        return "break"

    def _bind_global(_event=None):
        try:
            if canvas.winfo_exists():
                canvas.bind_all("<MouseWheel>", _on_mousewheel)
                if horizontal:
                    canvas.bind_all("<Shift-MouseWheel>", _on_shift_mousewheel)
        except tk.TclError:
            pass

    def _unbind_global(_event=None):
        try:
            canvas.unbind_all("<MouseWheel>")
            if horizontal:
                canvas.unbind_all("<Shift-MouseWheel>")
        except tk.TclError:
            pass

    canvas.bind("<Enter>", _bind_global)
    canvas.bind("<Leave>", _unbind_global)
    canvas.bind("<Destroy>", _unbind_global)


def short_subject_name(subject):
    """Compact subject labels for dense table headers.
    Uses database-backed SUBJECT_SHORT_NAMES if available, falls back to defaults.
    """
    subject = str(subject or "").strip()
    
    # Use the dynamic SUBJECT_SHORT_NAMES configuration
    short_map = SUBJECT_SHORT_NAMES if SUBJECT_SHORT_NAMES else DEFAULT_SUBJECT_SHORT_NAMES
    
    return short_map.get(subject, subject)


# ====================== TREEVIEW STYLE =====================
def setup_treeview_style():
    style = ttk.Style()
    style.theme_use("clam")
    style.configure(
        "App.Treeview",
        background=CARD_BG,
        foreground=TEXT_PRIMARY,
        rowheight=34,
        fieldbackground=CARD_BG,
        borderwidth=0,
        font=(FF, 10),
    )
    style.configure(
        "App.Treeview.Heading",
        background="#f8fafc",
        foreground=TEXT_SECONDARY,
        relief="flat",
        font=(FF, 10, "bold"),
        padding=8,
    )
    style.map(
        "App.Treeview",
        background=[("selected", "#eff6ff")],
        foreground=[("selected", TEXT_PRIMARY)],
    )
    style.configure(
        "App.Vertical.TScrollbar",
        background=BORDER_CLR,
        troughcolor="#f8fafc",
        arrowcolor=TEXT_SECONDARY,
        borderwidth=0,
    )
    style.configure(
        "App.TCombobox",
        fieldbackground=CARD_BG,
        background=CARD_BG,
        foreground=TEXT_PRIMARY,
        arrowcolor=TEXT_SECONDARY,
        bordercolor=BORDER_CLR,
        lightcolor=BORDER_CLR,
        darkcolor=BORDER_CLR,
        padding=6,
        font=(FF, 10),
    )
    style.configure(
        "App.TEntry",
        fieldbackground=CARD_BG,
        background=CARD_BG,
        foreground=TEXT_PRIMARY,
        bordercolor=BORDER_CLR,
        lightcolor=BORDER_CLR,
        darkcolor=BORDER_CLR,
        padding=8,
        font=(FF, 10),
    )


class AdvancedDataTable:
    """Reusable Treeview table with sorting, search, and pagination."""

    def __init__(
        self,
        parent,
        columns,
        page_size=20,
        search_label="Search",
        selectmode="browse",
        enable_select_all=False,
    ):
        self.parent = parent
        self.columns = columns
        self.page_size = max(1, int(page_size))
        self.sort_column = None
        self.sort_desc = False
        self.current_page = 0
        self.rows = []
        self.filtered_rows = []
        self.visible_rows = []
        self.row_index = {}
        self.search_var = tk.StringVar()
        self.enable_select_all = bool(enable_select_all)
        self.select_all_var = tk.BooleanVar(value=False)

        controls = tk.Frame(parent, bg=parent.cget("bg"))
        controls.pack(fill="x", padx=12, pady=(10, 8))

        tk.Label(
            controls,
            text=f"{search_label}:",
            bg=parent.cget("bg"),
            fg=TEXT_SECONDARY,
            font=(FF, 10, "bold"),
        ).pack(side="left", padx=(0, 6))
        self.search_entry = ttk.Entry(
            controls, textvariable=self.search_var, style="App.TEntry", width=28
        )
        self.search_entry.pack(side="left", ipady=4)
        self.search_var.trace_add(
            "write", lambda *_: self.apply_filters(reset_page=True)
        )
        self.search_entry.bind(
            "<Return>", lambda _e: self.apply_filters(reset_page=True)
        )

        self.search_btn = tk.Button(
            controls,
            text="🔍 Search",
            bg=OLIVE_PRIMARY,
            fg="white",
            activebackground=OLIVE_DARK,
            activeforeground="white",
            font=(FF, 9, "bold"),
            padx=10,
            pady=4,
            relief="flat",
            cursor="hand2",
            command=lambda: self.apply_filters(reset_page=True),
        )
        self.search_btn.pack(side="left", padx=(8, 0))

        if self.enable_select_all:
            self.select_all_cb = tk.Checkbutton(
                controls,
                text="Select all filtered",
                variable=self.select_all_var,
                bg=parent.cget("bg"),
                fg=TEXT_SECONDARY,
                activebackground=parent.cget("bg"),
                activeforeground=TEXT_PRIMARY,
                selectcolor=CARD_BG,
                font=(FF, 9, "bold"),
                command=self._on_select_all_toggle,
            )
            self.select_all_cb.pack(side="left", padx=(12, 0))

        self.status_label = tk.Label(
            controls,
            text="0 records",
            bg=parent.cget("bg"),
            fg=TEXT_SECONDARY,
            font=(FF, 9),
        )
        self.status_label.pack(side="right")

        tree_frame = tk.Frame(parent, bg=parent.cget("bg"))
        tree_frame.pack(fill="both", expand=True, padx=12, pady=(0, 8))

        col_keys = [c["key"] for c in columns]
        self.tree = ttk.Treeview(
            tree_frame,
            columns=col_keys,
            show="headings",
            style="App.Treeview",
            selectmode=selectmode,
        )
        for col in columns:
            key = col["key"]
            heading = col.get("title", key)
            width = col.get("width", 120)
            anchor = col.get("anchor", "w")
            self.tree.heading(
                key, text=heading, command=lambda k=key: self.toggle_sort(k)
            )
            self.tree.column(key, width=width, anchor=anchor)

        sb = ttk.Scrollbar(
            tree_frame,
            orient="vertical",
            command=self.tree.yview,
            style="App.Vertical.TScrollbar",
        )
        self.tree.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True)

        pager = tk.Frame(parent, bg=parent.cget("bg"))
        pager.pack(fill="x", padx=12, pady=(0, 10))

        self.prev_btn = tk.Button(
            pager,
            text="< Prev",
            bg="#eef2ff",
            fg=TEXT_PRIMARY,
            relief="flat",
            font=(FF, 9, "bold"),
            padx=12,
            pady=4,
            command=self.prev_page,
            cursor="hand2",
        )
        self.prev_btn.pack(side="left")

        self.next_btn = tk.Button(
            pager,
            text="Next >",
            bg="#eef2ff",
            fg=TEXT_PRIMARY,
            relief="flat",
            font=(FF, 9, "bold"),
            padx=12,
            pady=4,
            command=self.next_page,
            cursor="hand2",
        )
        self.next_btn.pack(side="left", padx=6)

        self.page_label = tk.Label(
            pager,
            text="Page 1/1",
            bg=parent.cget("bg"),
            fg=TEXT_SECONDARY,
            font=(FF, 9),
        )
        self.page_label.pack(side="left", padx=10)

        tk.Label(
            pager,
            text="Rows:",
            bg=parent.cget("bg"),
            fg=TEXT_SECONDARY,
            font=(FF, 9, "bold"),
        ).pack(side="right")
        self.page_size_var = tk.StringVar(value=str(self.page_size))
        self.page_size_cb = ttk.Combobox(
            pager,
            textvariable=self.page_size_var,
            values=["10", "20", "50", "100"],
            state="readonly",
            width=5,
            style="App.TCombobox",
        )
        self.page_size_cb.pack(side="right", padx=(0, 6))
        self.page_size_cb.bind("<<ComboboxSelected>>", self._on_page_size_changed)

    def set_rows(self, rows, reset_page=True):
        self.rows = list(rows or [])
        self.apply_filters(reset_page=reset_page)

    def get_selected(self):
        selected = self.tree.selection()
        if not selected:
            return None
        # Skip group header rows whose iid starts with "grp_"
        selected = [s for s in selected if not str(s).startswith("grp_")]
        if not selected:
            return None
        row = self.row_index.get(selected[0])
        if row:
            return row
        return None

    def get_selected_iids(self):
        if self.enable_select_all and self.select_all_var.get():
            return [
                row.get("iid")
                for row in self.filtered_rows
                if row.get("iid") and not str(row.get("iid", "")).startswith("grp_")
            ]
        return [iid for iid in self.tree.selection() if not str(iid).startswith("grp_")]

    def select_all_filtered(self):
        if not self.enable_select_all:
            visible_iids = [
                row.get("iid")
                for row in self.visible_rows
                if row.get("iid") and not str(row.get("iid", "")).startswith("grp_")
            ]
            if visible_iids:
                self.tree.selection_set(visible_iids)
            return
        self.select_all_var.set(True)
        self._on_select_all_toggle()

    def clear_selection(self):
        if self.enable_select_all:
            self.select_all_var.set(False)
        self.tree.selection_remove(self.tree.selection())

    def toggle_sort(self, key):
        if self.sort_column == key:
            self.sort_desc = not self.sort_desc
        else:
            self.sort_column = key
            self.sort_desc = False
        self.apply_filters(reset_page=True)

    def prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self._render_page()

    def next_page(self):
        total_pages = self._total_pages()
        if self.current_page < total_pages - 1:
            self.current_page += 1
            self._render_page()

    def apply_filters(self, reset_page=False):
        query = self.search_var.get().strip().lower()
        if query:
            filtered = []
            for row in self.rows:
                haystack = row.get("search", "")
                if not haystack:
                    haystack = " ".join(str(v) for v in row.get("values", ()))
                if query in haystack.lower():
                    filtered.append(row)
            self.filtered_rows = filtered
        else:
            self.filtered_rows = list(self.rows)

        if self.sort_column:
            self.filtered_rows.sort(
                key=lambda r: self._sort_value(r, self.sort_column),
                reverse=self.sort_desc,
            )

        if reset_page:
            self.current_page = 0
        else:
            self.current_page = min(self.current_page, max(0, self._total_pages() - 1))
        self._render_page()

    def _on_page_size_changed(self, _event=None):
        try:
            self.page_size = max(1, int(self.page_size_var.get()))
        except Exception:
            self.page_size = 20
        self.apply_filters(reset_page=True)

    def _on_select_all_toggle(self):
        if self.select_all_var.get():
            visible_iids = [
                row.get("iid")
                for row in self.visible_rows
                if row.get("iid") and not str(row.get("iid", "")).startswith("grp_")
            ]
            if visible_iids:
                self.tree.selection_set(visible_iids)
        else:
            self.tree.selection_remove(self.tree.selection())

    def _sort_value(self, row, key):
        value_map = row.get("value_map", {})
        value = value_map.get(key, "")
        if isinstance(value, (int, float)):
            return (0, value)
        text = str(value).strip()
        try:
            num = float(text)
            return (0, num)
        except Exception:
            return (1, text.lower())

    def _total_pages(self):
        if not self.filtered_rows:
            return 1
        return (len(self.filtered_rows) + self.page_size - 1) // self.page_size

    def _render_page(self):
        start = self.current_page * self.page_size
        end = start + self.page_size
        self.visible_rows = self.filtered_rows[start:end]

        self.row_index.clear()
        for item in self.tree.get_children():
            self.tree.delete(item)

        for idx, row in enumerate(self.visible_rows):
            row_iid = row.get("iid") or f"row_{start + idx}"
            self.tree.insert(
                "",
                "end",
                iid=row_iid,
                values=row.get("values", ()),
                tags=row.get("tags", ()),
            )
            self.row_index[row_iid] = row

        if self.enable_select_all and self.select_all_var.get():
            visible_iids = [
                row.get("iid") for row in self.visible_rows if row.get("iid")
            ]
            if visible_iids:
                self.tree.selection_set(visible_iids)

        total = len(self.filtered_rows)
        page_num = self.current_page + 1
        pages = self._total_pages()
        shown_from = 0 if total == 0 else start + 1
        shown_to = min(end, total)
        self.status_label.config(text=f"{shown_from}-{shown_to} of {total} records")
        self.page_label.config(text=f"Page {page_num}/{pages}")

        self.prev_btn.config(state="normal" if self.current_page > 0 else "disabled")
        self.next_btn.config(
            state="normal" if self.current_page < pages - 1 else "disabled"
        )


# ====================== MAIN APP ===========================
class SchoolReportApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(get_school_profile().get("school_app_title", "School Report"))
        self.root.geometry("1280x760")
        self.root.configure(bg=CONTENT_BG)
        self.root.minsize(960, 620)

        # UI constants for child pages
        self.CONTENT_BG = CONTENT_BG
        self.CARD_BG = CARD_BG
        self.TEXT_SECONDARY = TEXT_SECONDARY
        self.FF = FF
        self.BLUE = BLUE
        self.GREEN = GREEN
        self.ORANGE = ORANGE
        self.PURPLE = PURPLE

        # Expose AdvancedDataTable for students_tab.py
        self.AdvancedDataTable = AdvancedDataTable

        # CBC Level - Default to all-school view
        self.current_level = ALL_SCHOOL_LEVEL

        self.current_user = None
        self.user_role = "admin"  # Default role
        self.nav_frames: dict = {}
        self.active_nav: str = ""
        self.sidebar_collapsed = False
        self.sidebar_host = None
        self._topbar_clock_job = None
        self._topbar_clock_label = None
        self._topbar_bar = None
        self._topbar_clock_visible = True
        self._topbar_clock_generation = 0
        self._login_in_progress = False
        self._login_loader = None
        self._login_loader_after = None
        self._pending_login_notice = None
        self._pending_login_notice_after = None
        self._active_notice = None
        self._active_notice_after = None
        self._is_shutting_down = False
        self._report_email_settings_host = None
        self._report_email_settings_panel = None
        self._report_email_settings_status = None
        self._report_email_settings_entries = {}
        self._report_email_settings_tls_var = None
        self._failed_email_logs_host = None
        self._failed_email_logs_panel = None
        self._failed_email_logs_tree = None
        self._failed_email_logs_rows = []
        self._failed_email_logs_status = None
        self._failed_email_logs_title = None
        self.logo_img = self.load_logo()

        # Load logo image for use in dashboard
        self.logo_image = None
        try:
            # Get the directory where the script is located
            script_dir = os.path.dirname(os.path.abspath(__file__))
            logo_path = os.path.join(script_dir, "moas.jpg")
            img = Image.open(logo_path)
            # Resize logo to fit dashboard nicely
            img = img.resize((80, 80), Image.Resampling.LANCZOS)
            self.logo_image = ImageTk.PhotoImage(img)
        except Exception as e:
            print(f"Could not load logo image: {e}")

        try:
            self.root.protocol("WM_DELETE_WINDOW", self.shutdown)
        except Exception:
            pass

        # ── Window / taskbar icon (Windows-compatible .ico) ──────────────────
        try:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            jpg_path = os.path.join(script_dir, "moas.jpg")
            ico_path = os.path.join(script_dir, "moas.ico")
            # Generate .ico from jpg if not already present
            if not os.path.exists(ico_path):
                icon_img = Image.open(jpg_path).convert("RGBA")
                icon_img.save(
                    ico_path,
                    format="ICO",
                    sizes=[(16, 16), (32, 32), (48, 48), (64, 64)],
                )
            self.root.iconbitmap(ico_path)
        except Exception as e:
            print(f"Could not set window icon: {e}")

        self._ensure_default_class_catalog()
        self._ensure_default_grading_scales()
        refresh_dynamic_school_config()
        self.set_level(self.current_level)
        setup_treeview_style()
        self.show_login()

    # ------------------- utilities -------------------
    def _generate_short_label(self, text, kind="subject"):
        text = str(text or "").strip()
        if not text:
            return ""

        if kind == "class":
            match = re.search(r"grade\s*(\d+)", text, re.I)
            if match:
                return f"G{match.group(1)}"
            return "".join(word[0].upper() for word in text.split()[:3])

        if kind == "teacher":
            parts = [part for part in re.split(r"\s+", text) if part]
            if len(parts) >= 2:
                return "".join(part[0].upper() for part in parts[:2])
            return text[:4].upper()

        compact = short_subject_name(text).replace("\n", " / ").strip()
        if len(compact) <= 16:
            return compact
        words = [word for word in re.split(r"[^A-Za-z0-9]+", text) if word]
        if len(words) > 1:
            return "".join(word[0].upper() for word in words[:4])
        return text[:16]

    def _ensure_default_class_catalog(self):
        seed_key = "default_class_catalog_initialized"
        if db.get_setting(seed_key, "").strip():
            return

        if db.get_all_classes():
            db.set_setting(seed_key, "1")
            return

        for level, classes in DEFAULT_CLASSES_BY_LEVEL.items():
            for class_name in classes:
                if not db.get_class_by_name(class_name):
                    db.add_class(
                        class_name,
                        level,
                        abbreviation=self._generate_short_label(class_name, "class"),
                    )

        db.set_setting(seed_key, "1")

    def _ensure_default_subject_catalog(self):
        seed_key = "default_subject_catalog_initialized"
        if db.get_setting(seed_key, "").strip():
            return

        if db.get_subjects_by_level():
            db.set_setting(seed_key, "1")
            return

        for level in LEVELS:
            level_subjects = DEFAULT_SUBJECT_CATALOG.get(level, [])
            if isinstance(level_subjects, dict):
                catalog = list(level_subjects.get("core", [])) + list(
                    level_subjects.get("optional", [])
                )
            else:
                catalog = list(level_subjects)

            for code, name, category, is_optional in catalog:
                if not db.get_subject_by_name(name, level):
                    db.add_subject(
                        name,
                        level,
                        category,
                        is_optional=is_optional,
                        abbreviation=code,
                        code=code,
                    )

        db.set_setting(seed_key, "1")

    def _replace_subject_catalog_with_defaults(self):
        subject_rows = []
        for level in LEVELS:
            level_subjects = DEFAULT_SUBJECT_CATALOG.get(level, [])
            if isinstance(level_subjects, dict):
                catalog = list(level_subjects.get("core", [])) + list(
                    level_subjects.get("optional", [])
                )
            else:
                catalog = list(level_subjects)
            for code, name, category, is_optional in catalog:
                subject_rows.append(
                    {
                        "code": code,
                        "name": name,
                        "level": level,
                        "category": category,
                        "is_optional": is_optional,
                    }
                )
        db.replace_subject_catalog(subject_rows)
        refresh_dynamic_school_config()

    def _ensure_default_grading_scales(self):
        default_scale = [
            ("EE", "Exceeding Expectations", 80, 100, 1),
            ("ME", "Meeting Expectations", 70, 79, 2),
            ("AE", "Approaching Expectations", 60, 69, 3),
            ("BE", "Below Expectations", 50, 59, 4),
            ("IE", "Inadequate", 0, 49, 5),
        ]
        all_classes = []
        for classes in CLASSES_BY_LEVEL.values():
            for class_name in classes:
                if class_name not in all_classes:
                    all_classes.append(class_name)
        for row in db.get_all_classes():
            if row.get("name") and row["name"] not in all_classes:
                all_classes.append(row["name"])

        for class_name in all_classes:
            if db.get_grading_scales(class_name):
                continue
            for code, label, min_mark, max_mark, order in default_scale:
                db.add_grading_scale(class_name, min_mark, max_mark, code, label, order)

    def _get_class_meta(self, class_name):
        return db.get_class_by_name(class_name) or {}

    def _get_subject_meta(self, subject, class_name=""):
        level = self._get_level_for_class(class_name) if class_name else None
        meta = db.get_subject_by_name(subject, level)
        if meta:
            return meta
        if level:
            meta = db.get_subject_by_name(subject, ALL_SUBJECT_LEVEL)
            if meta:
                return meta
        meta = db.get_subject_by_name(subject) or {}
        if meta and not _is_legacy_subject_level(meta.get("level", "")):
            return meta
        return {}

    def _get_teacher_label(self, teacher):
        if isinstance(teacher, dict):
            abbreviation = str(teacher.get("abbreviation", "") or "").strip()
            return (
                abbreviation
                or teacher.get("full_name")
                or teacher.get("username")
                or ""
            )
        return self._generate_short_label(str(teacher or ""), "teacher")

    def _get_class_label(self, class_name):
        meta = self._get_class_meta(class_name)
        abbreviation = str(meta.get("abbreviation", "") or "").strip()
        return abbreviation or class_name

    def _format_class_stream_label(self, class_name, stream_name=""):
        class_label = self._get_class_label(class_name)
        stream_text = str(stream_name or "").strip()
        return f"{class_label} [{stream_text}]" if stream_text else class_label

    def _get_assignment_stream_options(self, class_name, include_whole_class=True):
        values = []
        if include_whole_class:
            values.append("Whole Class")
        values.extend(self._get_stream_names_for_class(class_name))
        return values

    def _get_subject_label(self, subject, class_name="", multiline=False):
        meta = self._get_subject_meta(subject, class_name)
        label = str(meta.get("abbreviation", "") or "").strip()
        if not label:
            label = self._generate_short_label(subject, "subject")
        return label.replace(" / ", "\n") if multiline else label

    def _get_class_grading_scale(self, class_name):
        if class_name in ("", ALL_SCHOOL_LEVEL, "All"):
            return []
        scales = db.get_grading_scales(class_name)
        if scales:
            return scales
        return []

    def _get_grade_code_for_class(self, mark, class_name=""):
        try:
            mark = float(mark)
        except (TypeError, ValueError):
            return "IE"
        for scale in self._get_class_grading_scale(class_name):
            if (
                float(scale.get("min_mark", 0))
                <= mark
                <= float(scale.get("max_mark", 0))
            ):
                return scale.get("grade_code", "IE")
        return get_grade_code(mark)

    def _get_grade_name_for_class(self, grade_code, class_name=""):
        for scale in self._get_class_grading_scale(class_name):
            if scale.get("grade_code") == grade_code:
                return scale.get("grade_name") or GRADE_LABELS.get(
                    grade_base_code(grade_code), grade_code
                )
        return GRADE_LABELS.get(grade_base_code(grade_code), grade_code)

    def _get_grade_color(self, grade_code):
        return GRADE_COLORS.get(
            str(grade_code or "").upper(),
            GRADE_COLORS.get(grade_base_code(grade_code), GREEN),
        )

    def _get_subject_color(self, subject, class_name=""):
        key = self._normalize_key(
            self._get_subject_label(subject, class_name) or subject
        )
        if not key:
            return SUBJECT_PALETTE[0]
        index = sum(ord(ch) for ch in key) % len(SUBJECT_PALETTE)
        return SUBJECT_PALETTE[index]

    def _get_subject_colors(self, subject, class_name=""):
        base = self._get_subject_color(subject, class_name)
        return {
            "base": base,
            "soft": _mix_hex(base, "#ffffff", 0.82),
            "mid": _mix_hex(base, "#ffffff", 0.65),
            "border": _mix_hex(base, "#223022", 0.18),
            "text": "#ffffff",
            "dark_text": _mix_hex(base, "#102018", 0.35),
        }

    def get_current_subjects(self):
        """Get subjects for the current user based on role"""
        # Check if user is a teacher with assigned subjects
        if self.current_user and hasattr(self, "user_role"):
            role = getattr(self, "user_role", "admin")
            teacher_id = self.current_user.get("id")

            # If teacher, return only assigned subjects
            if role == "teacher" and teacher_id:
                assignments = db.get_teacher_subjects(teacher_id)
                if assignments:
                    return [a["subject"] for a in assignments]

        # Default to CBC level subjects
        return self._get_subjects_for_level(self.current_level)

    def get_teacher_assigned_classes(self):
        """Get classes assigned to the current teacher"""
        if self.current_user and hasattr(self, "user_role"):
            role = getattr(self, "user_role", "admin")
            teacher_id = self.current_user.get("id")

            if role == "teacher" and teacher_id:
                # Get unique classes from subject assignments
                assignments = db.get_teacher_subjects(teacher_id)
                classes = list(set([a["class_name"] for a in assignments]))
                if classes:
                    return classes
            elif role == "class_teacher" and teacher_id:
                return db.get_teacher_classes(teacher_id)

        return self.get_current_classes()

    def get_current_classes(self):
        """Get classes for the current CBC level"""
        db_classes = [
            row.get("name", "") for row in db.get_all_classes() if row.get("name")
        ]
        if self.current_level == ALL_SCHOOL_LEVEL:
            return db_classes

        level_classes = [
            row.get("name", "")
            for row in db.get_all_classes()
            if row.get("name") and row.get("level") == self.current_level
        ]
        if level_classes:
            return level_classes

        if self.current_level == ALL_SCHOOL_LEVEL:
            return []
        return []

    def get_current_grading(self):
        """Get grading system for the current CBC level"""
        if self.current_level == ALL_SCHOOL_LEVEL:
            return GRADING_BY_LEVEL["Junior School (Grade 7-9)"]
        return GRADING_BY_LEVEL.get(
            self.current_level, GRADING_BY_LEVEL["Junior School (Grade 7-9)"]
        )

    def _get_subjects_for_level(self, level):
        """Flatten level subject config into a clean ordered list."""
        if level in LEVELS:
            level_rows = db.get_subjects_by_level(level)
            global_rows = db.get_subjects_by_level(ALL_SUBJECT_LEVEL)
            if level_rows or global_rows:
                ordered_names = []
                seen = set()

                catalog_subjects = SUBJECT_CATALOG.get(level, [])
                if isinstance(catalog_subjects, dict):
                    catalog_subjects = list(catalog_subjects.get("core", [])) + list(
                        catalog_subjects.get("optional", [])
                    )

                for _code, name, _category, _is_optional in catalog_subjects:
                    if name not in seen and (
                        any(row.get("name") == name for row in level_rows)
                        or any(row.get("name") == name for row in global_rows)
                    ):
                        ordered_names.append(name)
                        seen.add(name)

                for row in list(level_rows) + list(global_rows):
                    name = row.get("name", "")
                    if name and name not in seen:
                        ordered_names.append(name)
                        seen.add(name)

                return ordered_names

        if level == ALL_SCHOOL_LEVEL:
            all_rows = db.get_subjects_by_level()
            if all_rows:
                subjects = []
                seen = set()
                for level_name in LEVELS:
                    for subject in self._get_subjects_for_level(level_name):
                        if subject not in seen:
                            subjects.append(subject)
                            seen.add(subject)
                for row in all_rows:
                    name = row.get("name", "")
                    if (
                        name
                        and not _is_legacy_subject_level(row.get("level", ""))
                        and name not in seen
                    ):
                        subjects.append(name)
                        seen.add(name)
                return subjects
            return []
        level_subjects = SUBJECTS_BY_LEVEL.get(level, SUBJECTS)
        if isinstance(level_subjects, dict):
            return list(level_subjects.get("core", []))
        return list(level_subjects)

    def _get_subjects_for_selected_class(
        self, class_name, term="One", exam_type=DEFAULT_EXAM_TYPE
    ):
        """Return the right subject list for a concrete class selection."""
        if class_name and class_name not in ("All", ALL_SCHOOL_LEVEL):
            return self._get_subjects_for_class(class_name, term, exam_type)
        return self.get_current_subjects()

    def _match_subject_from_candidates(self, raw_subject, candidates, class_name=""):
        """Resolve a subject name against a candidate list using aliases and abbreviations."""
        raw = str(raw_subject or "").strip()
        if not raw:
            return ""

        cleaned_candidates = []
        for subject in candidates or []:
            subject_name = str(subject or "").strip()
            if subject_name and subject_name not in cleaned_candidates:
                cleaned_candidates.append(subject_name)
        if not cleaned_candidates:
            return ""

        lookup = {}
        for subject in cleaned_candidates:
            lookup.setdefault(self._normalize_text(subject), subject)
            lookup.setdefault(self._normalize_key(subject), subject)

            meta = self._get_subject_meta(subject, class_name)
            abbreviation = str(meta.get("abbreviation", "") or "").strip()
            if abbreviation:
                lookup.setdefault(self._normalize_text(abbreviation), subject)
                lookup.setdefault(self._normalize_key(abbreviation), subject)

            label = self._get_subject_label(subject, class_name)
            if label:
                lookup.setdefault(self._normalize_text(label), subject)
                lookup.setdefault(self._normalize_key(label), subject)

        raw_norm = self._normalize_text(raw)
        raw_key = self._normalize_key(raw)
        if raw_norm in lookup:
            return lookup[raw_norm]
        if raw_key in lookup:
            return lookup[raw_key]

        alias_map = {
            "eng": ["English", "English Activities", "English Language Activities"],
            "english": ["English", "English Activities", "English Language Activities"],
            "math": ["Mathematics", "Mathematical Activities"],
            "maths": ["Mathematics", "Mathematical Activities"],
            "mat": ["Mathematics", "Mathematical Activities"],
            "kis": [
                "Kiswahili / KSL",
                "Kiswahili / Kenyan Sign Language",
                "Kiswahili Activities",
                "Kiswahili Language Activities",
            ],
            "kiswahili": [
                "Kiswahili / KSL",
                "Kiswahili / Kenyan Sign Language",
                "Kiswahili Activities",
                "Kiswahili Language Activities",
            ],
            "intsci": ["Integrated Science"],
            "integratedscience": ["Integrated Science"],
            "sci": ["Science & Technology", "Integrated Science"],
            "science": ["Science & Technology", "Integrated Science"],
            "agri": ["Agriculture"],
            "sst": ["Social Studies"],
            "socialstudies": ["Social Studies"],
            "cre": [
                "Christian Religious Education (CRE)",
                "Christian Religious Education",
                "Religious Education (CRE/IRE/HRE)",
            ],
            "christianreligiouseducation": [
                "Christian Religious Education (CRE)",
                "Christian Religious Education",
                "Religious Education (CRE/IRE/HRE)",
            ],
            "christianreligiouseducationcre": [
                "Christian Religious Education (CRE)",
                "Christian Religious Education",
            ],
            "pretech": ["Pre-Technical Studies"],
            "pts": ["Pre-Technical Studies"],
            "french": ["French", "Foreign Languages (French, German, Arabic)"],
            "foreignlanguages": [
                "Foreign Languages (French, German, Arabic)",
                "French",
                "German",
                "Arabic",
            ],
            "via": ["Visual Arts"],
            "visualarts": ["Visual Arts"],
        }

        for alias in alias_map.get(raw_key, []):
            alias_norm = self._normalize_text(alias)
            alias_key = self._normalize_key(alias)
            if alias_norm in lookup:
                return lookup[alias_norm]
            if alias_key in lookup:
                return lookup[alias_key]

        for subject in cleaned_candidates:
            subject_key = self._normalize_key(subject)
            if raw_key and (raw_key in subject_key or subject_key in raw_key):
                return subject

        return ""

    def _determine_class_level(self, class_name):
        """Determine the appropriate level for a class name during import."""
        class_name_lower = str(class_name or "").lower()

        # Check against known level mappings
        for level, classes in CLASSES_BY_LEVEL.items():
            for known_class in classes:
                if (
                    known_class.lower() in class_name_lower
                    or class_name_lower in known_class.lower()
                ):
                    return level

        # Try to infer from class name patterns
        if any(
            word in class_name_lower
            for word in ["pp1", "pp2", "baby", "nursery", "pre"]
        ):
            return "Pre-Primary (PP1-PP2)"
        elif any(word in class_name_lower for word in ["grade", "class", "std"]):
            # Extract grade number
            import re

            match = re.search(r"(\d+)", class_name)
            if match:
                grade_num = int(match.group(1))
                if grade_num <= 3:
                    return "Lower Primary (Grade 1-3)"
                elif grade_num <= 6:
                    return "Upper Primary (Grade 4-6)"
                return "Junior School (Grade 7-9)"

        return LEVELS[0] if LEVELS else DEFAULT_LEVELS[0]

    def _get_level_for_class(self, class_name):
        """Get the school level for the given class name."""
        if not class_name:
            return self.current_level
        if class_name in ("All", ALL_SCHOOL_LEVEL):
            return ALL_SCHOOL_LEVEL

        class_name = str(class_name).strip()
        class_info = db.get_class_by_name(class_name)
        if class_info and class_info.get("level"):
            return class_info.get("level")

        # Support stream-qualified class names, e.g. Class A :: Stream
        if "::" in class_name:
            base_class = class_name.split("::", 1)[0].strip()
            class_info = db.get_class_by_name(base_class)
            if class_info and class_info.get("level"):
                return class_info.get("level")

        return self._determine_class_level(class_name)

    def _get_subjects_for_class(
        self,
        class_name,
        term="One",
        exam_type=DEFAULT_EXAM_TYPE,
        for_reporting=False,
        academic_year=None,
    ):
        """Return ordered class subject list, honoring class-specific 'subjects done' settings."""
        academic_year = str(academic_year or datetime.now().year)
        subjects = self._get_subject_pool_for_class(class_name)

        mapping = self._get_class_subjects_done_map()
        configured = mapping.get(class_name, [])
        if configured:
            ordered = []
            for subject in configured:
                matched = self._match_subject_from_candidates(
                    subject, subjects, class_name
                )
                chosen = matched or str(subject or "").strip()
                if chosen and chosen not in ordered:
                    ordered.append(chosen)

            if for_reporting:
                done_subjects = self._get_done_subjects_from_marks(
                    class_name, term, exam_type, academic_year
                )
                if done_subjects:
                    filtered = []
                    for subject in ordered:
                        matched = self._match_subject_from_candidates(
                            subject, done_subjects, class_name
                        )
                        if matched and matched not in filtered:
                            filtered.append(matched)
                    for subject in done_subjects:
                        if subject not in filtered:
                            filtered.append(subject)
                    return filtered

            return ordered

        if for_reporting:
            # Automatic report mode: show only subjects already entered for this class/term/exam.
            done_subjects = self._get_done_subjects_from_marks(
                class_name, term, exam_type, academic_year
            )
            return done_subjects

        return subjects

    def _get_subjects_for_scope(
        self,
        class_name,
        term="One",
        exam_type=DEFAULT_EXAM_TYPE,
        results=None,
        academic_year=None,
    ):
        """Return ordered subject columns for a report scope."""
        academic_year = str(academic_year or datetime.now().year)
        if class_name != "All":
            return self._get_subjects_for_class(
                class_name, term, exam_type, for_reporting=True, academic_year=academic_year
            )

        subjects = []
        subject_set = set()
        rows = results or []
        if not rows:
            rows = self._get_ranked_results(class_name, term, exam_type, academic_year)
        for result in rows:
            for subject in result.get("subjects", []):
                if subject not in subject_set:
                    subjects.append(subject)
                    subject_set.add(subject)
        return subjects

    def _get_subject_mode_badge(
        self,
        class_name,
        term="One",
        exam_type=DEFAULT_EXAM_TYPE,
        academic_year=None,
    ):
        """Return (label, bg, fg) for report subject mode badge."""
        academic_year = str(academic_year or datetime.now().year)
        mapping = self._get_class_subjects_done_map()
        if class_name and class_name != "All":
            if mapping.get(class_name):
                return ("Mode: Manual Subjects", "#14532d", "white")
            auto_subjects = self._get_done_subjects_from_marks(
                class_name, term, exam_type, academic_year
            )
            if auto_subjects:
                return ("Mode: Automatic (From Marks)", "#1d4ed8", "white")
            return ("Mode: Automatic (No Marks Yet)", "#92400e", "white")

        classes = self.get_current_classes()
        manual = 0
        auto = 0
        for cls in classes:
            if mapping.get(cls):
                manual += 1
            else:
                auto += 1
        if manual and auto:
            return (f"Mode: Mixed ({manual} Manual / {auto} Auto)", "#6b21a8", "white")
        if manual and not auto:
            return ("Mode: Manual Subjects (All Classes)", "#14532d", "white")
        return ("Mode: Automatic (All Classes)", "#1d4ed8", "white")

    def _get_ranked_results(
        self,
        class_name,
        term="One",
        exam_type=DEFAULT_EXAM_TYPE,
        academic_year=None,
    ):
        """Build whole-class ranked results with subject marks and summary fields."""
        academic_year = str(academic_year or datetime.now().year)
        if class_name == "All":
            allowed_classes = set(self.get_current_classes())
            students = [
                student
                for student in db.get_all_students()
                if student.get("class") in allowed_classes
            ]
        else:
            students = db.get_students_by_class(class_name)

        class_subjects = {}
        results = []

        for student in students:
            student_class = student.get("class", "")
            if student_class not in class_subjects:
                class_subjects[student_class] = self._get_subjects_for_class(
                    student_class,
                    term,
                    exam_type,
                    for_reporting=True,
                    academic_year=academic_year,
                )

            subjects = class_subjects.get(student_class, [])
            marks = db.get_student_marks(student["id"], term, exam_type, academic_year)
            total = sum(int(marks.get(subject, 0) or 0) for subject in subjects)
            average = round(total / len(subjects), 1) if subjects else 0
            grade = self._get_grade_code_for_class(average, student_class)
            class_level = self._get_level_for_class(student_class)

            results.append(
                {
                    "student": student,
                    "marks": marks,
                    "academic_year": academic_year,
                    "subjects": subjects,
                    "exam_type": exam_type,
                    "subject_count": len(subjects),
                    "possible_total": len(subjects) * 100,
                    "total": total,
                    "average": average,
                    "grade": grade,
                    "level": self._get_grade_name_for_class(grade, student_class),
                    "class_level": class_level,
                }
            )

        results.sort(
            key=lambda row: (
                -row["total"],
                -row["average"],
                row["student"]["name"].lower(),
            )
        )

        last_key = None
        current_position = 0
        for index, row in enumerate(results, start=1):
            rank_key = (row["total"], row["average"])
            if rank_key != last_key:
                current_position = index
                last_key = rank_key
            row["position"] = current_position

        return results

    def _get_stream_names_for_class(self, class_name):
        class_row = db.get_class_by_name(class_name or "")
        if not class_row:
            return []
        streams = db.get_streams_for_class(class_row["id"])
        return [
            stream.get("name", "").strip()
            for stream in streams
            if stream.get("name", "").strip()
        ]

    def _get_selected_report_stream(self):
        if not hasattr(self, "rc_stream_cb"):
            return ""
        stream = (self.rc_stream_cb.get() or "").strip()
        return "" if stream == "All Streams" else stream

    def _rerank_results(self, results):
        rows = list(results or [])
        rows.sort(
            key=lambda row: (
                -row.get("total", 0),
                -row.get("average", 0),
                str(row.get("student", {}).get("name", "")).lower(),
            )
        )

        last_key = None
        current_position = 0
        for index, row in enumerate(rows, start=1):
            rank_key = (row.get("total", 0), row.get("average", 0))
            if rank_key != last_key:
                current_position = index
                last_key = rank_key
            row["position"] = current_position
        return rows

    def _get_selected_results_stream(self):
        if not hasattr(self, "rep_stream_cb"):
            return ""
        stream = (self.rep_stream_cb.get() or "").strip()
        return "" if stream == "All Streams" else stream

    def _refresh_results_streams(self, reload_results=True):
        if not hasattr(self, "rep_stream_cb"):
            return
        current = (self.rep_stream_cb.get() or "").strip()
        selected_class = self.rep_cls_cb.get() if hasattr(self, "rep_cls_cb") else ""
        if selected_class in ("All", ALL_SCHOOL_LEVEL, ""):
            values = ["All Streams"]
        else:
            values = ["All Streams"] + self._get_stream_names_for_class(selected_class)
        self.rep_stream_cb["values"] = values
        if current and current in values:
            self.rep_stream_cb.set(current)
        else:
            self.rep_stream_cb.set(values[0] if values else "All Streams")
        self._update_results_page_header()
        if reload_results:
            self.load_reports()

    def _get_selected_marks_stream(self):
        if hasattr(self, "marks_stream_cb"):
            stream = (self.marks_stream_cb.get() or "").strip()
            return "" if stream == "All Streams" else stream
        return getattr(self, "_selected_marks_stream", "")

    def _refresh_marks_streams(self, reload_results=True):
        if not hasattr(self, "marks_stream_cb"):
            return
        current = (self.marks_stream_cb.get() or "").strip()
        if not current:
            current = getattr(self, "_selected_marks_stream", "")

        selected_class = (
            self.marks_class_cb.get() if hasattr(self, "marks_class_cb") else ""
        )
        values = ["All Streams"] + self._get_stream_names_for_class(selected_class)
        self.marks_stream_cb["values"] = values
        if current and current in values:
            self.marks_stream_cb.set(current)
        else:
            self.marks_stream_cb.set(values[0] if values else "All Streams")

        self._selected_marks_stream = self._get_selected_marks_stream()
        if reload_results:
            self._load_marks_table()

    def _get_selected_chart_stream(self):
        if not hasattr(self, "ch_stream_cb"):
            return ""
        stream = (self.ch_stream_cb.get() or "").strip()
        return "" if stream == "All Streams" else stream

    def _refresh_chart_streams(self, reload_results=True):
        if not hasattr(self, "ch_stream_cb"):
            return
        current = (self.ch_stream_cb.get() or "").strip()
        selected_class = self.ch_cls_cb.get() if hasattr(self, "ch_cls_cb") else ""
        if selected_class in ("All", ""):
            values = ["All Streams"]
        else:
            values = ["All Streams"] + self._get_stream_names_for_class(selected_class)
        self.ch_stream_cb["values"] = values
        if current and current in values:
            self.ch_stream_cb.set(current)
        else:
            self.ch_stream_cb.set(values[0] if values else "All Streams")
        if reload_results:
            self.load_charts()

    def _refresh_exam_analytics_streams(self):
        if not hasattr(self, "analytics_stream_cb"):
            return
        current = (self.analytics_stream_var.get() or "").strip()
        selected_class = (
            self.analytics_class_var.get()
            if hasattr(self, "analytics_class_var")
            else ""
        )
        if selected_class in ("All Classes", ""):
            values = ["All Streams"]
        else:
            values = ["All Streams"] + self._get_stream_names_for_class(selected_class)
        self.analytics_stream_cb["values"] = values
        if current and current in values:
            self.analytics_stream_cb.set(current)
        else:
            self.analytics_stream_cb.set(values[0] if values else "All Streams")

    def _update_marks_page_header(self):
        if not hasattr(self, "page_sub_lbl") or not hasattr(self, "marks_class_cb"):
            return
        cls = self.marks_class_cb.get()
        stream = self._get_selected_marks_stream()
        subtitle = f"Enter Marks - {cls}" if cls else "Enter Marks"
        if stream:
            subtitle += f" (Stream {stream})"
        self.page_sub_lbl.config(text=subtitle)

    def _get_year_options(self):
        try:
            years = db.get_academic_years()
        except Exception:
            years = []
        current_year = str(datetime.now().year)
        if current_year not in years:
            years.insert(0, current_year)
        for offset in range(1, 6):
            year = str(datetime.now().year - offset)
            if year not in years:
                years.append(year)
        return years

    def _get_results_page_results(self, class_name, term, exam_type, academic_year=None):
        academic_year = str(academic_year or datetime.now().year)
        results = self._get_ranked_results(class_name, term, exam_type, academic_year)
        selected_stream = self._get_selected_results_stream()
        if selected_stream and class_name not in ("All", ALL_SCHOOL_LEVEL):
            results = [
                result
                for result in results
                if result.get("student", {}).get("stream", "").strip()
                == selected_stream
            ]
            results = self._rerank_results(results)
        return results

    def _update_results_page_header(self):
        if not hasattr(self, "page_title_lbl") or not hasattr(self, "page_sub_lbl"):
            return
        cls = self.rep_cls_cb.get() if hasattr(self, "rep_cls_cb") else ""
        stream = self._get_selected_results_stream()
        year = self.rep_year_cb.get() if hasattr(self, "rep_year_cb") else ""
        class_text = "All Classes" if cls == "All" else cls
        title = f"Results - {class_text}"
        if stream:
            title = f"{title} - {stream}"
        self.page_title_lbl.config(text=title)
        if stream:
            subtitle = f"View student performance and rankings for {class_text} stream {stream}"
        elif cls == "All":
            subtitle = "View student performance and rankings across all classes"
        else:
            subtitle = f"View student performance and rankings for {class_text}"
        if year:
            subtitle = f"{subtitle} in {year}"
        self.page_sub_lbl.config(text=subtitle)

    def _get_report_card_results(self):
        academic_year = (
            self.rc_year_cb.get()
            if hasattr(self, "rc_year_cb")
            else str(datetime.now().year)
        )
        results = self._get_ranked_results(
            self.rc_cls_cb.get(),
            self.rc_term_cb.get(),
            self.rc_exam_cb.get() or DEFAULT_EXAM_TYPE,
            academic_year,
        )
        selected_stream = self._get_selected_report_stream()
        if selected_stream:
            results = [
                result
                for result in results
                if result.get("student", {}).get("stream", "").strip()
                == selected_stream
            ]
        return results

    def _refresh_report_card_streams(self, reload_results=True):
        if not hasattr(self, "rc_stream_cb"):
            return
        current = (self.rc_stream_cb.get() or "").strip()
        values = ["All Streams"] + self._get_stream_names_for_class(
            self.rc_cls_cb.get()
        )
        self.rc_stream_cb["values"] = values
        if current and current in values:
            self.rc_stream_cb.set(current)
        else:
            self.rc_stream_cb.set(values[0] if values else "All Streams")
        if reload_results:
            self._load_rc()

    def _normalize_text(self, value):
        text = str(value or "").strip().lower()
        text = re.sub(r"\s+", " ", text)
        return text

    def _normalize_key(self, value):
        text = self._normalize_text(value)
        return re.sub(r"[^a-z0-9]+", "", text)

    def _get_known_class_names(self):
        class_names = []
        for classes in CLASSES_BY_LEVEL.values():
            for class_name in classes:
                if class_name not in class_names:
                    class_names.append(class_name)
        for row in db.get_all_classes():
            class_name = str(row.get("name", "") or "").strip()
            if class_name and class_name not in class_names:
                class_names.append(class_name)
        return class_names

    def _match_known_class_name(self, value):
        raw = str(value or "").strip()
        if not raw:
            return ""

        raw_norm = self._normalize_text(raw)
        raw_key = self._normalize_key(raw)
        known_classes = self._get_known_class_names()

        for class_name in known_classes:
            if self._normalize_text(class_name) == raw_norm:
                return class_name

        for class_name in known_classes:
            if self._normalize_key(class_name) == raw_key:
                return class_name

        pp_match = re.search(r"\b(?:pp|p|preprimary)\s*([12])\b", raw_key)
        if pp_match:
            candidate_key = f"pp{pp_match.group(1)}"
            for class_name in known_classes:
                if self._normalize_key(class_name) == candidate_key:
                    return class_name
            return f"PP{pp_match.group(1)}"

        grade_match = re.search(r"\b(?:grade|g)\s*(\d+)\b", raw_norm)
        if grade_match:
            candidate = f"Grade {int(grade_match.group(1))}"
            for class_name in known_classes:
                if self._normalize_key(class_name) == self._normalize_key(candidate):
                    return class_name
            return candidate

        bare_grade_match = re.fullmatch(r"(\d+)", raw_key)
        if bare_grade_match:
            candidate = f"Grade {int(bare_grade_match.group(1))}"
            for class_name in known_classes:
                if self._normalize_key(class_name) == self._normalize_key(candidate):
                    return class_name

        return ""

    def _get_known_stream_names(self, class_name=""):
        class_row = db.get_class_by_name(class_name or "")
        stream_names = []

        if class_row:
            for stream in db.get_streams_for_class(class_row["id"]):
                stream_name = str(stream.get("name", "") or "").strip()
                if stream_name and stream_name not in stream_names:
                    stream_names.append(stream_name)

        for student in db.get_students_by_class(class_name or ""):
            stream_name = str(student.get("stream", "") or "").strip()
            if stream_name and stream_name not in stream_names:
                stream_names.append(stream_name)

        return stream_names

    def _match_known_stream_name(self, value, class_name=""):
        raw = str(value or "").strip()
        if not raw:
            return ""

        raw_norm = self._normalize_text(raw)
        raw_key = self._normalize_key(raw)
        streams = self._get_known_stream_names(class_name)

        for stream_name in streams:
            if self._normalize_text(stream_name) == raw_norm:
                return stream_name

        for stream_name in streams:
            stream_key = self._normalize_key(stream_name)
            if raw_key and (
                raw_key == stream_key or raw_key in stream_key or stream_key in raw_key
            ):
                return stream_name

        if len(raw) <= 3 and raw.isalpha():
            return raw.upper()
        return raw.title()

    def _extract_class_from_text(self, text):
        text_norm = self._normalize_text(text)
        if not text_norm:
            return ""

        direct_match = self._match_known_class_name(text)
        if direct_match:
            return direct_match

        known_classes = self._get_known_class_names()
        for class_name in known_classes:
            class_norm = self._normalize_text(class_name)
            if text_norm == class_norm or class_norm in text_norm:
                return class_name

        for pattern in (
            r"\b(?:pp|p|preprimary)\s*([12])\b",
            r"grade\s+[a-z]+\s*\((\d+)\)",
            r"grade\s*(\d+)",
            r"\bg\s*(\d+)\b",
        ):
            match = re.search(pattern, text_norm)
            if not match:
                continue
            token = match.group(1)
            if pattern.startswith(r"\b(?:pp|p|preprimary)"):
                candidate = self._match_known_class_name(f"PP{token}")
                if candidate:
                    return candidate
                return f"PP{token}"

            candidate = self._match_known_class_name(f"Grade {int(token)}")
            if candidate:
                return candidate
            return f"Grade {int(token)}"
        return ""

    def _extract_stream_from_text(self, text, class_name=""):
        text_norm = self._normalize_text(text)
        if not text_norm:
            return ""

        ignored_tokens = {"overall", "summary", "report", "analysis"}

        if class_name:
            class_norm = self._normalize_text(class_name)
            if class_norm and text_norm.startswith(class_norm):
                remainder = re.sub(
                    r"^[\s\-_/:\[\]\(\)]+",
                    "",
                    text_norm[len(class_norm) :],
                ).strip()
                if remainder:
                    token_match = re.match(r"([a-z0-9]+)", remainder)
                    if token_match:
                        raw_stream = token_match.group(1)
                        if raw_stream not in ignored_tokens:
                            return (
                                self._match_known_stream_name(raw_stream, class_name)
                                or raw_stream.title()
                            )

        match = re.search(
            r"(?:grade\s*\d+|grade\s+[a-z]+\s*\(\d+\)|g\s*\d+|pp\s*[12]|p\s*[12]|preprimary\s*[12])\s*[-_/ ]*([a-z0-9]+)\b",
            text_norm,
        )
        if match:
            raw_stream = match.group(1)
            if raw_stream not in ignored_tokens:
                return (
                    self._match_known_stream_name(raw_stream, class_name)
                    or raw_stream.title()
                )
        return ""

    def _get_sheet_context(self, sheet_name, worksheet=None):
        header_text = ""
        if worksheet is not None:
            header_parts = []
            for row in worksheet.iter_rows(
                min_row=1, max_row=min(6, worksheet.max_row), values_only=True
            ):
                for value in row:
                    text = str(value or "").strip()
                    if text:
                        header_parts.append(text)
            header_text = " ".join(header_parts)

        combined_text = " ".join(
            part for part in [str(sheet_name or ""), header_text] if part
        ).strip()
        combined_norm = self._normalize_text(combined_text)
        is_summary = any(
            flag in combined_norm
            for flag in ("overall report", "overall", "summary", "analysis")
        )
        class_name = self._extract_class_from_text(combined_text)
        stream_name = self._extract_stream_from_text(sheet_name, class_name)

        return {
            "class_name": class_name,
            "stream_name": stream_name,
            "is_summary": is_summary,
        }

    def _sheet_title_to_class_name(self, title, worksheet=None):
        return self._get_sheet_context(title, worksheet).get("class_name") or None

    def _safe_excel_sheet_name(self, title, fallback="Sheet"):
        clean = re.sub(r"[:\\/?*\[\]]+", " ", str(title or "").strip())
        clean = re.sub(r"\s+", " ", clean).strip()
        return (clean or fallback)[:31]

    def _generate_admission_no(self, class_name, student_name):
        resolved_class = (
            self._match_known_class_name(class_name) or str(class_name or "").strip()
        )
        return db.get_next_class_admission_no(resolved_class)

    def _ensure_import_class_setup(self, class_name, stream_name=""):
        resolved_class = (
            self._match_known_class_name(class_name) or str(class_name or "").strip()
        )
        resolved_stream = (
            self._match_known_stream_name(stream_name, resolved_class)
            or str(stream_name or "").strip()
        )
        if not resolved_class:
            return "", resolved_stream

        class_row = db.get_class_by_name(resolved_class)
        level = self._get_level_for_class(resolved_class) or self.current_level
        abbreviation = self._generate_short_label(resolved_class, "class")

        if not class_row:
            db.add_class(
                resolved_class,
                level,
                resolved_stream or None,
                abbreviation,
            )
            class_row = db.get_class_by_name(resolved_class)
        elif not str(class_row.get("abbreviation", "") or "").strip():
            db.update_class(
                class_row["id"],
                resolved_class,
                class_row.get("level", "") or level,
                class_row.get("stream"),
                abbreviation,
            )
            class_row = db.get_class_by_name(resolved_class) or class_row

        if class_row and resolved_stream:
            existing_streams = db.get_streams_for_class(class_row["id"])
            stream_keys = {
                self._normalize_key(stream.get("name", ""))
                for stream in existing_streams
                if stream.get("name")
            }
            if self._normalize_key(resolved_stream) not in stream_keys:
                db.add_stream(resolved_stream, class_row["id"])

        return resolved_class, resolved_stream

    def _ensure_import_subject_record(self, raw_subject, class_name=""):
        raw = str(raw_subject or "").strip()
        if not raw:
            return ""

        level = self._get_level_for_class(class_name)
        candidates = []
        for subject in self._get_subject_pool_for_class(class_name):
            if subject and subject not in candidates:
                candidates.append(subject)
        for subject in self._get_catalog_subject_names_for_level(level):
            if subject and subject not in candidates:
                candidates.append(subject)
        for row in db.get_subjects_by_level():
            subject = str(row.get("name", "") or "").strip()
            if subject and subject not in candidates:
                candidates.append(subject)

        subject_name = self._match_subject_from_candidates(raw, candidates, class_name)

        if not subject_name:
            raw_key = self._normalize_key(raw)
            for row in db.get_subjects_by_level():
                for alias in (
                    row.get("name", ""),
                    row.get("abbreviation", ""),
                    row.get("code", ""),
                ):
                    if alias and self._normalize_key(alias) == raw_key:
                        subject_name = str(row.get("name", "") or "").strip()
                        break
                if subject_name:
                    break

        subject_name = subject_name or raw
        existing_meta = self._get_subject_meta(subject_name, class_name)
        if existing_meta and existing_meta.get("name"):
            return str(existing_meta.get("name") or "").strip()

        subject_code = re.sub(
            r"[^A-Z0-9]+",
            "",
            self._generate_short_label(subject_name, "subject").upper(),
        )[:12]
        target_level = level or ALL_SUBJECT_LEVEL
        db.add_subject(
            subject_name,
            target_level,
            "Core",
            False,
            subject_code,
            subject_code,
        )
        created_meta = self._get_subject_meta(subject_name, class_name)
        if created_meta and created_meta.get("name"):
            return str(created_meta.get("name") or "").strip()
        return subject_name

    def _merge_import_subjects_into_class_map(self, class_name, subjects):
        cls = str(class_name or "").strip()
        cleaned_subjects = [
            str(subject or "").strip()
            for subject in subjects
            if str(subject or "").strip()
        ]
        if not cls or not cleaned_subjects:
            return

        mapping = self._get_class_subjects_done_map()
        current = list(mapping.get(cls, []))
        changed = False
        for subject in cleaned_subjects:
            matched = self._match_subject_from_candidates(subject, current, cls)
            subject_name = matched or subject
            if subject_name not in current:
                current.append(subject_name)
                changed = True

        if changed:
            mapping[cls] = current
            self._save_class_subjects_done_map(mapping)

    def _map_sheet_subject(self, raw_subject, class_name=""):
        raw = str(raw_subject or "").strip()
        if not raw:
            return ""

        classes = (
            self._get_subjects_for_class(class_name, TERMS[0]) if class_name else []
        )
        class_lookup = {self._normalize_key(subject): subject for subject in classes}
        raw_key = self._normalize_key(raw)
        if raw_key in class_lookup:
            return class_lookup[raw_key]

        alias_map = {
            "eng": ["English", "English Activities", "English Language Activities"],
            "english": ["English", "English Activities", "English Language Activities"],
            "lang": [
                "Language Activities",
                "Literacy Activities",
                "English Language Activities",
                "English Activities",
                "English",
            ],
            "math": ["Mathematical Activities", "Mathematics"],
            "maths": ["Mathematical Activities", "Mathematics"],
            "mat": ["Mathematical Activities", "Mathematics"],
            "mathematicalactivities": ["Mathematical Activities", "Mathematics"],
            "kis": [
                "Kiswahili Activities",
                "Kiswahili / KSL",
                "Kiswahili / Kenyan Sign Language",
                "Kiswahili Language Activities",
                "Kenyan Sign Language",
            ],
            "kiswahili": [
                "Kiswahili Activities",
                "Kiswahili / KSL",
                "Kiswahili / Kenyan Sign Language",
                "Kiswahili Language Activities",
                "Kenyan Sign Language",
            ],
            "englishlanguageactivities": [
                "English Language Activities",
                "English Activities",
                "English",
            ],
            "kiswahililanguageactivities": [
                "Kiswahili Language Activities",
                "Kiswahili Activities",
                "Kiswahili / KSL",
            ],
            "intsci": ["Integrated Science"],
            "integratedscience": ["Integrated Science"],
            "sci": ["Science & Technology", "Integrated Science"],
            "science": ["Science & Technology", "Integrated Science"],
            "scitech": ["Science & Technology"],
            "scienceandtechnology": ["Science & Technology"],
            "environ": ["Environmental Activities"],
            "env": ["Environmental Activities"],
            "envi": ["Environmental Activities"],
            "creative": [
                "Creative Activities",
                "Movement & Creative Activities",
                "Creative Arts",
                "Visual Arts",
                "Performing Arts",
            ],
            "creativearts": [
                "Creative Arts",
                "Creative Activities",
                "Movement & Creative Activities",
            ],
            "carts": ["Creative Arts", "Creative Activities"],
            "ca": ["Creative Arts", "Visual Arts", "Performing Arts"],
            "casports": ["Sports & Physical Education"],
            "agri": ["Agriculture"],
            "agrinut": ["Agriculture"],
            "sst": ["Social Studies"],
            "socialstudies": ["Social Studies"],
            "cre": [
                "Religious Activities",
                "Religious Education Activities",
                "Christian Religious Education",
                "Christian Religious Education (CRE)",
            ],
            "pretech": ["Pre-Technical Studies"],
            "pretechnicalstudies": ["Pre-Technical Studies"],
            "french": ["French", "Foreign Languages (French, German, Arabic)"],
        }

        mapped_candidates = alias_map.get(raw_key, [raw])
        for mapped in mapped_candidates:
            mapped_key = self._normalize_key(mapped)
            if mapped_key in class_lookup:
                return class_lookup[mapped_key]

        for mapped in mapped_candidates:
            mapped_key = self._normalize_key(mapped)
            for subject in classes:
                subject_key = self._normalize_key(subject)
                if mapped_key and (
                    mapped_key == subject_key
                    or mapped_key in subject_key
                    or subject_key in mapped_key
                ):
                    return subject

        for subject in classes:
            subject_key = self._normalize_key(subject)
            if raw_key and (raw_key in subject_key or subject_key in raw_key):
                return subject

        return ""

    def _find_assessment_header_row(self, worksheet, class_name=""):
        name_aliases = {
            "learner",
            "learner name",
            "name",
            "student",
            "student name",
            "pupil",
            "pupil name",
            "full name",
            "full_name",
        }
        id_aliases = {
            "no",
            "no.",
            "adm",
            "adm no",
            "admission",
            "admission no",
            "admission_no",
            "reg no",
            "reg_no",
            "regno",
        }
        excluded = {
            "no",
            "no.",
            "learner",
            "learner name",
            "name",
            "student",
            "student name",
            "pupil",
            "pupil name",
            "full name",
            "full_name",
            "total",
            "total ",
            "avg",
            "average",
            "psn",
            "position",
            "level",
        }

        best_match = None
        max_scan_rows = min(20, worksheet.max_row)
        for row_idx in range(1, max_scan_rows + 1):
            raw_values = list(
                next(
                    worksheet.iter_rows(
                        min_row=row_idx, max_row=row_idx, values_only=True
                    )
                )
            )
            values = [self._normalize_text(cell) for cell in raw_values]

            name_col_idx = next(
                (idx + 1 for idx, value in enumerate(values) if value in name_aliases),
                None,
            )
            if not name_col_idx:
                continue

            has_id_col = any(value in id_aliases for value in values)
            subject_hits = 0
            for col_idx, cell in enumerate(raw_values, start=1):
                if col_idx == name_col_idx:
                    continue
                label = self._normalize_text(cell)
                if not label or label in excluded:
                    continue
                mapped = (
                    self._map_sheet_subject(cell, class_name)
                    if class_name
                    else str(cell or "").strip()
                )
                if mapped:
                    subject_hits += 1

            score = subject_hits + (2 if has_id_col else 0)
            if subject_hits >= 2 or (subject_hits >= 1 and has_id_col):
                if best_match is None or score > best_match["score"]:
                    best_match = {
                        "row": row_idx,
                        "name_col": name_col_idx,
                        "score": score,
                    }

        return best_match

    def _parse_assessment_sheet(self, worksheet):
        sheet_context = self._get_sheet_context(worksheet.title, worksheet)
        if sheet_context.get("is_summary"):
            return None

        class_name = sheet_context.get("class_name") or self._sheet_title_to_class_name(
            worksheet.title, worksheet
        )
        if not class_name:
            return None

        header_info = self._find_assessment_header_row(worksheet, class_name)
        if not header_info:
            return None
        header_row = header_info["row"]
        learner_col_idx = header_info["name_col"]
        stream_name = sheet_context.get("stream_name", "")

        excluded = {
            "no",
            "no.",
            "adm",
            "adm no",
            "admission",
            "admission no",
            "admission_no",
            "reg no",
            "reg_no",
            "regno",
            "learner",
            "learner name",
            "name",
            "student",
            "student name",
            "pupil",
            "pupil name",
            "full name",
            "full_name",
            "total",
            "total ",
            "avg",
            "average",
            "psn",
            "position",
            "level",
        }
        header_values = [
            cell
            for cell in next(
                worksheet.iter_rows(
                    min_row=header_row, max_row=header_row, values_only=True
                )
            )
        ]

        subject_columns = []
        for col_idx, cell in enumerate(header_values, start=1):
            label = self._normalize_text(cell)
            if col_idx == learner_col_idx or not label or label in excluded:
                continue
            mapped_subject = self._map_sheet_subject(cell, class_name)
            if mapped_subject:
                subject_columns.append((col_idx, mapped_subject))

        if not subject_columns:
            return None

        students = []
        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            learner = (
                str(row[learner_col_idx - 1] or "").strip()
                if len(row) >= learner_col_idx
                else ""
            )
            if not learner or self._is_summary_student_name(learner):
                continue

            marks = {}
            for col_idx, subject in subject_columns:
                raw = row[col_idx - 1] if len(row) >= col_idx else None
                if raw in (None, ""):
                    continue
                try:
                    marks[subject] = min(100, max(0, int(float(raw))))
                except (TypeError, ValueError):
                    continue

            if marks:
                students.append({"name": learner, "marks": marks})

        if not students:
            return None

        return {
            "sheet_name": worksheet.title,
            "class_name": class_name,
            "stream_name": stream_name,
            "students": students,
        }

    def _open_progress_dialog(self, title, message, details="", allow_cancel=False):
        dialog = tk.Toplevel(self.root)
        dialog.title(title)
        width, height = 720, 340
        dialog.configure(bg=CONTENT_BG)
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        dialog._cancel_requested = False

        self.root.update_idletasks()
        x = self.root.winfo_rootx() + max(0, (self.root.winfo_width() - width) // 2)
        y = self.root.winfo_rooty() + max(0, (self.root.winfo_height() - height) // 2)
        dialog.geometry(f"{width}x{height}+{x}+{y}")
        dialog.minsize(width, height)

        pr_bo, pr_bi = _card_colors("azure")
        outer = tk.Frame(dialog, bg=pr_bo)
        outer.pack(fill="both", expand=True, padx=18, pady=18)
        card = tk.Frame(outer, bg=pr_bi, padx=28, pady=26)
        card.pack(fill="both", expand=True, padx=1, pady=1)

        tk.Label(
            card, text=title, bg=pr_bi, fg=TEXT_PRIMARY, font=(FF, 16, "bold")
        ).pack(anchor="w")
        status_label = tk.Label(
            card,
            text=message,
            bg=pr_bi,
            fg=TEXT_SECONDARY,
            font=(FF, 12),
            wraplength=620,
            justify="left",
        )
        status_label.pack(anchor="w", fill="x", pady=(10, 12))

        detail_panel = tk.Frame(card, bg="#F5F8FD", padx=16, pady=14)
        detail_panel.pack(fill="x", pady=(0, 12))
        tk.Label(
            detail_panel,
            text="Live import details",
            bg="#F5F8FD",
            fg=TEXT_PRIMARY,
            font=(FF, 10, "bold"),
        ).pack(anchor="w", pady=(0, 4))

        detail_label = tk.Label(
            detail_panel,
            text=details,
            bg="#F5F8FD",
            fg=TEXT_SECONDARY,
            font=(FF, 10),
            wraplength=588,
            justify="left",
        )
        detail_label.pack(anchor="w", fill="x")

        progress_header = tk.Frame(card, bg=pr_bi)
        progress_header.pack(fill="x", pady=(12, 6))
        tk.Label(
            progress_header,
            text="Progress",
            bg=pr_bi,
            fg=TEXT_SECONDARY,
            font=(FF, 10, "bold"),
        ).pack(side="left")
        percent_label = tk.Label(
            progress_header, text="0%", bg=pr_bi, fg=GREEN, font=(FF, 13, "bold")
        )
        percent_label.pack(side="right")

        progress_shell = tk.Frame(card, bg="#D6E3C0", padx=1, pady=1)
        progress_shell.pack(fill="x", pady=(0, 4))

        progress = tk.Canvas(
            progress_shell,
            height=28,
            bg="#F4F7EC",
            highlightthickness=0,
            bd=0,
        )
        progress.pack(fill="x")
        progress.update_idletasks()
        fill_id = progress.create_rectangle(0, 0, 0, 28, fill=ORANGE, outline="")
        progress._fill_id = fill_id
        dialog._progress_detail_label = detail_label
        dialog._progress_status_label = status_label

        if allow_cancel:
            action_row = tk.Frame(card, bg=pr_bi)
            action_row.pack(fill="x", pady=(12, 0))
            tk.Label(
                action_row,
                text="You can cancel this import at any time.",
                bg=pr_bi,
                fg=TEXT_SECONDARY,
                font=(FF, 9),
            ).pack(side="left")
            cancel_button = tk.Button(
                action_row,
                text="Cancel Import",
                bg="#E85D5D",
                fg="white",
                font=(FF, 10, "bold"),
                relief="flat",
                padx=14,
                pady=6,
                command=lambda: self._request_progress_cancel(dialog),
            )
            cancel_button.pack(side="right")
            dialog._progress_cancel_button = cancel_button
            dialog.protocol(
                "WM_DELETE_WINDOW", lambda: self._request_progress_cancel(dialog)
            )
        else:
            dialog.protocol("WM_DELETE_WINDOW", lambda: None)

        dialog.update_idletasks()
        return dialog, status_label, percent_label, progress

    def _update_progress_dialog(
        self,
        dialog,
        status_label,
        percent_label,
        progress,
        current,
        total,
        message,
        details=None,
    ):
        total = max(1, int(total or 1))
        current = min(total, max(0, int(current)))
        percent = int((current / total) * 100)
        status_label.config(text=message)
        detail_label = getattr(dialog, "_progress_detail_label", None)
        if detail_label is not None and details is not None:
            detail_label.config(text=details)
        percent_label.config(text=f"{percent}%")
        if isinstance(progress, tk.Canvas):
            progress.update_idletasks()
            width = max(1, progress.winfo_width())
            height = max(1, progress.winfo_height())
            fill_width = int(width * (percent / 100))
            if percent < 35:
                fill_color = ORANGE
            elif percent < 70:
                fill_color = "#6D97C9"
            else:
                fill_color = GREEN
            progress.coords(
                getattr(progress, "_fill_id", None), 0, 0, fill_width, height
            )
            progress.itemconfig(getattr(progress, "_fill_id", None), fill=fill_color)
        else:
            progress["value"] = percent
        try:
            dialog.update()
        except tk.TclError:
            pass

    def _request_progress_cancel(self, dialog):
        if dialog is None or not dialog.winfo_exists():
            return
        if getattr(dialog, "_cancel_requested", False):
            return
        if not messagebox.askyesno(
            "Cancel Import",
            "Stop this import now?\n\nAny rows already imported before cancellation will remain saved.",
            parent=dialog,
        ):
            return
        dialog._cancel_requested = True
        cancel_button = getattr(dialog, "_progress_cancel_button", None)
        if cancel_button is not None and cancel_button.winfo_exists():
            cancel_button.config(state="disabled", text="Cancelling...")
        status_label = getattr(dialog, "_progress_status_label", None)
        if status_label is not None and status_label.winfo_exists():
            status_label.config(text="Cancelling import. Please wait...")
        try:
            dialog.update()
        except tk.TclError:
            pass

    def _progress_cancel_requested(self, dialog):
        return bool(dialog and getattr(dialog, "_cancel_requested", False))

    def _cancel_login_loader_timer(self):
        if self._login_loader_after is None:
            return
        try:
            self.root.after_cancel(self._login_loader_after)
        except Exception:
            pass
        self._login_loader_after = None

    def _schedule_login_loader(self):
        self._cancel_login_loader_timer()
        self._login_loader_after = self.root.after(350, self._show_login_loader)

    def _show_login_loader(self):
        self._login_loader_after = None
        if not self._login_in_progress or self._login_loader is not None:
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("Signing In")
        dialog.configure(bg=CONTENT_BG)
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        dialog.protocol("WM_DELETE_WINDOW", lambda: None)

        width, height = 380, 170
        self.root.update_idletasks()
        x = self.root.winfo_rootx() + max(0, (self.root.winfo_width() - width) // 2)
        y = self.root.winfo_rooty() + max(0, (self.root.winfo_height() - height) // 2)
        dialog.geometry(f"{width}x{height}+{x}+{y}")

        outer_bg, inner_bg = _card_colors("mint")
        outer = tk.Frame(dialog, bg=outer_bg)
        outer.place(relx=0.5, rely=0.5, anchor="center")
        card = tk.Frame(outer, bg=inner_bg, padx=22, pady=20)
        card.pack(padx=1, pady=1)

        tk.Label(
            card,
            text="Signing you in...",
            bg=inner_bg,
            fg=TEXT_PRIMARY,
            font=(FF, 13, "bold"),
        ).pack(anchor="w")
        tk.Label(
            card,
            text="Please wait while we verify your account details.",
            bg=inner_bg,
            fg=TEXT_SECONDARY,
            font=(FF, 10),
            wraplength=320,
            justify="left",
        ).pack(anchor="w", pady=(8, 14))

        progress = ttk.Progressbar(
            card, orient="horizontal", length=320, mode="indeterminate"
        )
        progress.pack(fill="x")
        progress.start(12)

        dialog.update_idletasks()
        self._login_loader = (dialog, progress)

    def _close_login_loader(self):
        self._cancel_login_loader_timer()
        if self._login_loader is None:
            return

        dialog, progress = self._login_loader
        self._login_loader = None
        try:
            progress.stop()
        except Exception:
            pass
        try:
            dialog.grab_release()
        except Exception:
            pass
        try:
            dialog.destroy()
        except Exception:
            pass

    def _format_user_display_name(self, user):
        raw_name = ""
        for key in ("name", "full_name", "username", "email"):
            value = str(user.get(key, "") or "").strip()
            if value:
                raw_name = value
                break

        if "@" in raw_name:
            raw_name = raw_name.split("@", 1)[0]

        cleaned = raw_name.replace(".", " ").replace("_", " ").strip()
        return cleaned.title() if cleaned else "User"

    def _cancel_pending_login_notice_timer(self):
        if self._pending_login_notice_after is None:
            return
        try:
            self.root.after_cancel(self._pending_login_notice_after)
        except Exception:
            pass
        self._pending_login_notice_after = None

    def _cancel_active_notice_timer(self):
        if self._active_notice_after is None:
            return
        try:
            self.root.after_cancel(self._active_notice_after)
        except Exception:
            pass
        self._active_notice_after = None

    def _cancel_topbar_clock_timer(self):
        if self._topbar_clock_job is None:
            return
        try:
            self.root.after_cancel(self._topbar_clock_job)
        except Exception:
            pass
        self._topbar_clock_job = None

    def _show_pending_login_notice(self):
        self._pending_login_notice_after = None
        if not self._pending_login_notice:
            return
        title, message = self._pending_login_notice
        self._pending_login_notice = None
        self._show_notice(title, message, kind="success")

    def _dismiss_active_notice(self):
        if self._active_notice is None:
            self._cancel_active_notice_timer()
            return
        self._dismiss_notice(self._active_notice)

    def _dismiss_notice(self, notice):
        if self._active_notice is notice:
            self._cancel_active_notice_timer()
            self._active_notice = None
        try:
            notice.destroy()
        except Exception:
            pass

    def _show_delete_result_notice(
        self, item_label, deleted_count, failed_count=0, duration_ms=3600
    ):
        label = str(item_label or "item").strip()
        if failed_count and deleted_count:
            self._show_notice(
                f"{label.title()} Deletion Completed",
                f"Deleted {deleted_count} {label}(s), but {failed_count} could not be removed.",
                kind="info",
                duration_ms=duration_ms,
            )
            return
        if failed_count:
            self._show_notice(
                f"{label.title()} Delete Failed",
                f"We could not delete {failed_count} {label}(s). Please try again.",
                kind="error",
                duration_ms=duration_ms,
            )
            return
        self._show_notice(
            f"{label.title()} Deleted",
            f"Deleted {deleted_count} {label}(s) successfully.",
            kind="success",
            duration_ms=duration_ms,
        )

    def _confirm_delete_action(
        self, item_label, count=1, scope="selected", details=None, parent=None
    ):
        label = str(item_label or "item").strip()
        count = max(1, int(count or 1))
        scope_key = str(scope or "selected").strip().lower()

        if scope_key == "all":
            prompt = f"Delete ALL {count} {label}(s)?"
        elif count == 1:
            prompt = f"Delete this {label}?"
        else:
            prompt = f"Delete {count} selected {label}(s)?"

        lines = [prompt]
        extra = str(details or "").strip()
        if extra:
            lines.extend(["", extra])
        lines.extend(["", "This action cannot be undone."])
        return messagebox.askyesno(
            "Confirm Delete",
            "\n".join(lines),
            parent=parent or self.root,
        )

    def _show_notice(self, title, message, kind="success", duration_ms=3200):
        if self._active_notice is not None:
            self._dismiss_notice(self._active_notice)

        # ── Kind config ─────────────────────────────────────────────
        KIND_CFG = {
            "success": {
                "strip": "#22c55e",
                "icon": "✓",
                "title_fg": "#15803d",
                "theme": "mint",
            },
            "error": {
                "strip": "#ef4444",
                "icon": "✕",
                "title_fg": "#b91c1c",
                "theme": "blossom",
            },
            "info": {
                "strip": "#3b82f6",
                "icon": "ℹ",
                "title_fg": "#1d4ed8",
                "theme": "azure",
            },
            "warning": {
                "strip": "#f59e0b",
                "icon": "⚠",
                "title_fg": "#b45309",
                "theme": "sand",
            },
        }
        cfg = KIND_CFG.get(kind, KIND_CFG["info"])
        border_bg, card_bg = _card_colors(cfg["theme"])

        # ── Window ───────────────────────────────────────────────────
        notice = tk.Toplevel(self.root)
        notice.overrideredirect(True)
        notice.configure(bg=border_bg)
        try:
            notice.attributes("-topmost", True)
        except Exception:
            pass

        outer = tk.Frame(notice, bg=border_bg)
        outer.pack()
        card = tk.Frame(outer, bg=card_bg)
        card.pack(padx=1, pady=1)

        # Coloured left accent strip
        tk.Frame(card, bg=cfg["strip"], width=5).pack(side="left", fill="y")

        body = tk.Frame(card, bg=card_bg, padx=14, pady=11)
        body.pack(side="left", fill="both", expand=True)

        # Title row: icon • title text • × close
        title_row = tk.Frame(body, bg=card_bg)
        title_row.pack(fill="x")

        tk.Label(
            title_row,
            text=cfg["icon"],
            bg=card_bg,
            fg=cfg["strip"],
            font=(FF, 12, "bold"),
        ).pack(side="left", padx=(0, 6))

        tk.Label(
            title_row,
            text=title,
            bg=card_bg,
            fg=cfg["title_fg"],
            font=(FF, 11, "bold"),
            anchor="w",
        ).pack(side="left", fill="x", expand=True)

        close_lbl = tk.Label(
            title_row,
            text="×",
            bg=card_bg,
            fg=TEXT_SECONDARY,
            font=(FF, 14, "bold"),
            cursor="hand2",
            padx=4,
        )
        close_lbl.pack(side="right")
        close_lbl.bind("<Button-1>", lambda _e: self._dismiss_notice(notice))

        # Optional message
        msg_text = str(message or "").strip()
        if msg_text:
            tk.Label(
                body,
                text=msg_text,
                bg=card_bg,
                fg=TEXT_SECONDARY,
                font=(FF, 9),
                justify="left",
                wraplength=250,
                anchor="w",
            ).pack(fill="x", pady=(4, 6))

        # Shrinking progress bar
        prog_bg = tk.Frame(body, bg="#dde1e7", height=3)
        prog_bg.pack(fill="x", pady=(6, 0))
        prog_bg.pack_propagate(False)
        prog_bar = tk.Frame(prog_bg, bg=cfg["strip"])
        prog_bar.place(relx=0, rely=0, relwidth=1.0, relheight=1.0)

        # ── Position: bottom-right of main window ───────────────────
        self.root.update_idletasks()
        notice.update_idletasks()
        nw = max(300, notice.winfo_reqwidth())
        nh = notice.winfo_reqheight()
        rx = self.root.winfo_rootx()
        ry = self.root.winfo_rooty()
        rw = self.root.winfo_width()
        rh = self.root.winfo_height()
        x = rx + rw - nw - 20
        y = ry + rh - nh - 20
        notice.geometry(f"{nw}x{nh}+{max(10, x)}+{max(10, y)}")

        self._active_notice = notice
        self._cancel_active_notice_timer()

        # ── Animated progress bar ────────────────────────────────────
        total_ms = max(1200, int(duration_ms or 3200))
        steps = 60
        step_ms = max(16, total_ms // steps)

        def _shrink(step=0):
            try:
                if notice is not self._active_notice or not notice.winfo_exists():
                    return
                frac = 1.0 - step / steps
                prog_bar.place(relwidth=max(0.0, frac))
                if step < steps:
                    self._active_notice_after = self.root.after(
                        step_ms, lambda: _shrink(step + 1)
                    )
                else:
                    self._dismiss_active_notice()
            except Exception:
                pass

        _shrink()

    def _complete_login_attempt(self, user, error=None):
        self._login_in_progress = False
        self._close_login_loader()

        if error is not None:
            messagebox.showerror(
                "Login Failed",
                f"Something went wrong while signing you in.\n\n{error}",
            )
            return

        if user:
            self.current_user = user
            self.user_role = user.get("role", "admin")
            display_name = self._format_user_display_name(user)
            self._pending_login_notice = (
                "Login Successful",
                f"Welcome back, {display_name}. Your account is ready.",
            )
            self.show_main()
            return

        pwd_entry = self._auth_entries.get("password")
        if isinstance(pwd_entry, tk.Entry):
            pwd_entry.focus_set()
            pwd_entry.selection_range(0, "end")

        self._show_notice(
            "Login Failed",
            "We could not sign you in with those details.\n\n"
            "Please check your email and password, then try again.\n\n"
            "Default: admin / admin123",
            kind="error",
            duration_ms=4200,
        )

    def set_level(self, level):
        """Set the current CBC level and update subjects/classes"""
        if level in LEVELS or level == ALL_SCHOOL_LEVEL:
            self.current_level = level
            # Update legacy compatibility variables
            global SUBJECTS, CLASSES
            SUBJECTS = self._get_subjects_for_level(level)
            CLASSES = self.get_current_classes()
            return True
        return False

    def load_logo(self):
        try:
            img = Image.open("moas.jpg").resize((60, 60))
            return ImageTk.PhotoImage(img)
        except:
            return None

    def _clear(self):
        self._cancel_topbar_clock_timer()
        self._topbar_clock_generation += 1
        self._topbar_clock_label = None
        self._topbar_bar = None
        self._topbar_clock_visible = True
        self._login_in_progress = False
        self._close_login_loader()
        self._cancel_pending_login_notice_timer()
        self._pending_login_notice = None
        if self._active_notice is not None:
            self._dismiss_notice(self._active_notice)
        for w in self.root.winfo_children():
            w.destroy()

    def clear_frame(self):
        for w in self.content_frame.winfo_children():
            w.destroy()

    def clear_root(self):
        self._clear()

    def shutdown(self):
        if self._is_shutting_down:
            return
        self._is_shutting_down = True
        try:
            self._clear()
        except Exception:
            pass
        try:
            if self.root.winfo_exists():
                self.root.quit()
        except Exception:
            pass
        try:
            if self.root.winfo_exists():
                self.root.destroy()
        except Exception:
            pass

    def _get_active_db_display(self):
        db_path = str(getattr(db, "db_name", "") or "").strip()
        if not db_path:
            return "DB: unknown"
        return f"DB: {os.path.basename(db_path)}"

    def _page_header(self, title: str, subtitle: str):
        hdr = tk.Frame(self.content_frame, bg=CONTENT_BG)
        hdr.pack(fill="x", pady=(0, 22))
        top = tk.Frame(hdr, bg=CONTENT_BG)
        top.pack(fill="x")
        self.page_title_lbl = tk.Label(
            top, text=title, bg=CONTENT_BG, fg=TEXT_PRIMARY, font=(FF, 26, "bold")
        )
        self.page_title_lbl.pack(side="left", anchor="w")
        tk.Label(
            top,
            text=self._get_active_db_display(),
            bg=CONTENT_BG,
            fg=TEXT_SECONDARY,
            font=(FF, 9, "bold"),
        ).pack(side="right", anchor="e")
        self.page_sub_lbl = tk.Label(
            hdr, text=subtitle, bg=CONTENT_BG, fg=TEXT_SECONDARY, font=(FF, 11)
        )
        self.page_sub_lbl.pack(anchor="w")

    def _toolbar_btn(self, parent, text, command, bg=BLUE, fg="white"):
        b = tk.Label(
            parent,
            text=text,
            bg=bg,
            fg=fg,
            font=(FF, 10, "bold"),
            padx=14,
            pady=7,
            cursor="hand2",
        )
        b.bind("<Button-1>", lambda e: command())
        hover_bg = (
            OLIVE_MID
            if bg in (BLUE, ORANGE, PURPLE, GREEN, OLIVE_PRIMARY, LEMON_ACCENT)
            else OLIVE_DARK
        )
        b.bind("<Enter>", lambda e: b.config(bg=hover_bg))
        b.bind("<Leave>", lambda e: b.config(bg=bg))
        return b

    # ─────────────────────── Auth screen ───────────────────────────────────
    _AUTH_BG = CREAM_BG
    _CARD_W = 400
    _BLUE = OLIVE_PRIMARY
    _DARK_BLUE = OLIVE_DARK
    _FIELD_BG = LEMON_SOFT
    _TEXT = TEXT_PRIMARY
    _GRAY_T = TEXT_SECONDARY

    def show_login(self):
        self._clear()
        self._auth_tab = "login"
        self._auth_outer = tk.Frame(self.root, bg=self._AUTH_BG)
        self._auth_outer.pack(fill="both", expand=True)
        self._build_auth_card()

    def _build_auth_card(self):
        for w in self._auth_outer.winfo_children():
            w.destroy()

        tab = self._auth_tab
        CW = self._CARD_W
        CARD_H = 492 if tab == "login" else 565
        PAD = 6

        # ── Canvas: shadow + rounded white card ─────────────────────────────
        canvas = tk.Canvas(
            self._auth_outer,
            width=CW + PAD * 2,
            height=CARD_H + PAD * 2,
            bg=self._AUTH_BG,
            highlightthickness=0,
        )
        canvas.place(relx=0.5, rely=0.5, anchor="center")

        # shadow
        _rr(canvas, PAD + 3, PAD + 5, CW + PAD + 3, CARD_H + PAD + 5, 16, "#c8cfac")
        # card
        _rr(canvas, PAD, PAD, CW + PAD, CARD_H + PAD, 16, CARD_BG)

        # content frame embedded in canvas
        frame = tk.Frame(canvas, bg=CARD_BG)
        canvas.create_window(
            CW // 2 + PAD, PAD + 1, window=frame, anchor="n", width=CW - 40
        )
        self._auth_frame = frame

        # ── Logo ────────────────────────────────────────────────────────────
        if self.logo_img:
            tk.Label(frame, image=self.logo_img, bg=CARD_BG).pack(pady=20)
        else:
            logo = tk.Canvas(
                frame, width=56, height=56, bg=CARD_BG, highlightthickness=0
            )
            logo.pack(pady=(24, 12))
            _rr(logo, 0, 0, 56, 56, 12, self._BLUE)
            logo.create_text(28, 28, text="\U0001f393", font=(FF, 22), fill="white")

        # ── Title ───────────────────────────────────────────────────────────
        tk.Label(
            frame, text="MOAS", bg=CARD_BG, fg=self._TEXT, font=(FF, 16, "bold")
        ).pack(pady=(5, 3))
        tk.Label(
            frame,
            text="School Management System",
            bg=CARD_BG,
            fg=self._GRAY_T,
            font=(FF, 11),
        ).pack(pady=(0, 18))

        # ── Tab switcher ────────────────────────────────────────────────────
        TH = 42
        tab_c = tk.Canvas(frame, height=TH, bg=CARD_BG, highlightthickness=0)
        tab_c.pack(fill="x", pady=(0, 14))

        def draw_tabs(w):
            tab_c.delete("all")
            _rr(tab_c, 0, 0, w, TH, 10, self._FIELD_BG)
            half = w // 2
            # active indicator
            if tab == "login":
                _rr(tab_c, 4, 4, half - 2, TH - 4, 7, CARD_BG)
            else:
                _rr(tab_c, half + 2, 4, w - 4, TH - 4, 7, CARD_BG)
            # Login label
            lo = tk.Label(
                tab_c,
                text="Login",
                bg=CARD_BG if tab == "login" else self._FIELD_BG,
                fg=self._TEXT if tab == "login" else self._GRAY_T,
                font=(FF, 11, "bold" if tab == "login" else "normal"),
                cursor="hand2",
            )
            tab_c.create_window(half // 2, TH // 2, window=lo, width=half - 10)
            lo.bind("<Button-1>", lambda e: self._switch_tab("login"))
            # Sign Up label
            su = tk.Label(
                tab_c,
                text="Sign Up",
                bg=CARD_BG if tab == "signup" else self._FIELD_BG,
                fg=self._TEXT if tab == "signup" else self._GRAY_T,
                font=(FF, 11, "bold" if tab == "signup" else "normal"),
                cursor="hand2",
            )
            tab_c.create_window(half + half // 2, TH // 2, window=su, width=half - 10)
            su.bind("<Button-1>", lambda e: self._switch_tab("signup"))

        tab_c.bind("<Configure>", lambda e: draw_tabs(e.width))

        # ── Form fields ─────────────────────────────────────────────────────
        self._auth_entries = {}

        def mk_field(label, key, show=""):
            tk.Label(
                frame,
                text=label,
                bg=CARD_BG,
                fg=self._TEXT,
                font=(FF, 11, "bold"),
                anchor="w",
            ).pack(fill="x", pady=(10, 4))
            wrap = tk.Frame(frame, bg=self._FIELD_BG, padx=14)
            wrap.pack(fill="x")
            e = tk.Entry(
                wrap,
                bg=self._FIELD_BG,
                fg=self._TEXT,
                relief="flat",
                font=(FF, 12),
                show=show,
                bd=0,
                highlightthickness=0,
                insertbackground=self._TEXT,
            )
            e.pack(fill="x", ipady=12)
            e.bind(
                "<Return>",
                lambda ev: self._do_login() if tab == "login" else self._do_register(),
            )
            self._auth_entries[key] = e
            return e

        if tab == "signup":
            mk_field("Full Name", "name")
        first = mk_field("Email", "email")
        mk_field("Password", "password", show="\u25cf")
        first.focus_set()

        # ── Action button ────────────────────────────────────────────────────
        btn_text = "Sign In" if tab == "login" else "Create Account"
        btn_cmd = self._do_login if tab == "login" else self._do_register

        btn_c = tk.Canvas(
            frame, height=46, bg=CARD_BG, highlightthickness=0, cursor="hand2"
        )
        btn_c.pack(fill="x", pady=(20, 0))

        def draw_btn(color=None):
            color = color or self._BLUE
            w = btn_c.winfo_width() or (CW - 40)
            btn_c.delete("all")
            _rr(btn_c, 0, 0, w, 46, 8, color)
            btn_c.create_text(
                w // 2, 23, text=btn_text, fill="white", font=(FF, 12, "bold")
            )

        btn_c.bind("<Configure>", lambda e: draw_btn())
        btn_c.bind("<Button-1>", lambda e: btn_cmd())
        btn_c.bind("<Enter>", lambda e: draw_btn(OLIVE_MID))
        btn_c.bind("<Leave>", lambda e: draw_btn(self._BLUE))

        # ── Forgot password / bottom spacer ─────────────────────────────────
        if tab == "login":
            fp = tk.Label(
                frame,
                text="Forgot password?",
                bg=CARD_BG,
                fg=self._BLUE,
                font=(FF, 10),
                cursor="hand2",
            )
            fp.pack(pady=(12, 24))
            fp.bind(
                "<Button-1>",
                lambda e: messagebox.showinfo(
                    "Forgot Password",
                    "Please contact your administrator to reset your password.",
                ),
            )

            # Demo Mode button
            tk.Button(
                frame,
                text="Demo Mode (Skip Login)",
                command=self.show_main,
                bg=LEMON_SOFT,
                fg=TEXT_PRIMARY,
                width=20,
                pady=5,
            ).pack(pady=5)
        else:
            tk.Frame(frame, bg=CARD_BG, height=24).pack()

    def _switch_tab(self, tab):
        if self._login_in_progress:
            return
        if self._auth_tab != tab:
            self._auth_tab = tab
            self._build_auth_card()

    def _do_login(self):
        if self._login_in_progress:
            return

        email = self._auth_entries.get("email", tk.Entry()).get().strip()
        pwd = self._auth_entries.get("password", tk.Entry()).get().strip()
        if not email or not pwd:
            messagebox.showerror("Error", "Please fill in all fields")
            return

        self._login_in_progress = True
        self._schedule_login_loader()

        def worker():
            try:
                user = db.authenticate(email, pwd)
                self.root.after(0, self._complete_login_attempt, user)
            except Exception as exc:
                self.root.after(0, self._complete_login_attempt, None, exc)

        threading.Thread(target=worker, daemon=True).start()

    def _do_register(self):
        name = self._auth_entries.get("name", tk.Entry()).get().strip()
        email = self._auth_entries.get("email", tk.Entry()).get().strip()
        pwd = self._auth_entries.get("password", tk.Entry()).get().strip()
        if not all([name, email, pwd]):
            messagebox.showerror("Error", "All fields are required")
            return
        if len(pwd) < 6:
            messagebox.showerror("Error", "Password must be at least 6 characters")
            return
        if db.register_user(name, email, pwd):
            messagebox.showinfo("Success", "Account created! Please sign in.")
            self._switch_tab("login")
        else:
            messagebox.showerror("Error", "This email is already registered.")

    def do_login(self):
        """Legacy – kept for compatibility"""
        email = self.email_entry.get()
        pwd = self.pwd_entry.get()
        if db.authenticate(email, pwd):
            self.show_main()
        else:
            messagebox.showerror(
                "Login Failed", "Invalid credentials. Try admin/admin123"
            )

    # ------------------- main layout -------------------
    def show_main(self):
        self._clear()

        wrapper = tk.Frame(self.root, bg=CONTENT_BG)
        wrapper.pack(fill="both", expand=True)

        # ── Top navbar ──────────────────────────────────────
        self._build_topbar(wrapper)

        # ── Bottom: sidebar + content ────────────────────────
        bottom = tk.Frame(wrapper, bg=CONTENT_BG)
        bottom.pack(fill="both", expand=True)

        self.sidebar_host = tk.Frame(bottom, bg=SIDEBAR_BG)
        self.sidebar_host.pack(side="left", fill="y")
        self._build_sidebar(self.sidebar_host)

        # 1-px divider between sidebar and content
        tk.Frame(bottom, bg=BORDER_CLR, width=1).pack(side="left", fill="y")

        # content scroller
        content_wrapper = tk.Frame(bottom, bg=CONTENT_BG)
        content_wrapper.pack(side="left", fill="both", expand=True)

        self.content = tk.Frame(content_wrapper, bg=CONTENT_BG)
        self.content.pack(fill="both", expand=True, padx=28, pady=24)
        self.content_frame = self.content

        self.show_dashboard()
        if self._pending_login_notice:
            self._cancel_pending_login_notice_timer()
            self._pending_login_notice_after = self.root.after(
                250, self._show_pending_login_notice
            )

    # ─────────────────────── Top Navbar ────────────────────────────────────
    def _draw_topbar_icon(self, c, icon_type, bg):
        """Draw crisp vector-style icons on a Canvas widget."""
        c.delete("all")
        c.config(bg=bg)
        FG = TOPBAR_ICON_FG
        if icon_type == "logout":
            # Door frame (left rectangle)
            c.create_rectangle(2, 2, 13, 20, outline=FG, width=1.5)
            # Door knob
            c.create_oval(9, 10, 11.5, 12.5, fill=FG, outline="")
            # Arrow shaft pointing right
            c.create_line(13, 11, 22, 11, fill=FG, width=2)
            # Arrow head
            c.create_line(18, 7, 22, 11, fill=FG, width=2)
            c.create_line(18, 15, 22, 11, fill=FG, width=2)
        elif icon_type == "about":
            # Circle outline
            c.create_oval(1, 1, 21, 21, outline=FG, width=1.5)
            # Question mark glyph
            c.create_text(11, 12, text="?", fill=FG, font=(FF, 11, "bold"))
        elif icon_type == "menu":
            c.create_line(4, 6, 18, 6, fill=FG, width=2, capstyle="round")
            c.create_line(4, 11, 18, 11, fill=FG, width=2, capstyle="round")
            c.create_line(4, 16, 18, 16, fill=FG, width=2, capstyle="round")

    def _topbar_btn(self, parent, icon_type, label, bg, command=None, side="right"):
        """Icon-over-text button with crisp Canvas-drawn icons."""
        hover = TOPBAR_BTN_HOV

        f = tk.Frame(parent, bg=bg, padx=16, cursor="hand2")
        f.pack(side=side, fill="y")

        # Canvas icon (22×22 drawing area)
        ic = tk.Canvas(f, width=22, height=22, bg=bg, highlightthickness=0)
        ic.pack(pady=(10, 2))
        self._draw_topbar_icon(ic, icon_type, bg)

        txt = tk.Label(f, text=label, bg=bg, fg=TOPBAR_ICON_FG, font=(FF, 8))
        txt.pack(pady=(0, 10))

        all_w = [f, ic, txt]
        if command:
            for w in all_w:
                w.bind("<Button-1>", lambda e: command())

        def _on_enter(e):
            for w in [f, txt]:
                w.config(bg=hover)
            ic.config(bg=hover)
            self._draw_topbar_icon(ic, icon_type, hover)

        def _on_leave(e):
            for w in [f, txt]:
                w.config(bg=bg)
            ic.config(bg=bg)
            self._draw_topbar_icon(ic, icon_type, bg)

        for w in all_w:
            w.bind("<Enter>", _on_enter)
            w.bind("<Leave>", _on_leave)
        return f

    def _topbar_compact_btn(
        self, parent, icon_type, label, bg, command=None, side="right"
    ):
        """Compact topbar button with icon and label on one line."""
        hover = TOPBAR_BTN_HOV

        f = tk.Frame(parent, bg=bg, padx=10, cursor="hand2")
        f.pack(side=side, fill="y")

        inner = tk.Frame(f, bg=bg)
        inner.pack(expand=True, pady=10)

        ic = tk.Canvas(inner, width=22, height=22, bg=bg, highlightthickness=0)
        ic.pack(side="left")
        self._draw_topbar_icon(ic, icon_type, bg)

        txt = tk.Label(
            inner, text=label, bg=bg, fg=TOPBAR_ICON_FG, font=(FF, 8, "bold")
        )
        txt.pack(side="left", padx=(6, 0))

        all_w = [f, inner, ic, txt]
        if command:
            for w in all_w:
                w.bind("<Button-1>", lambda e: command())

        def _on_enter(e):
            for w in [f, inner, txt]:
                w.config(bg=hover)
            ic.config(bg=hover)
            self._draw_topbar_icon(ic, icon_type, hover)

        def _on_leave(e):
            for w in [f, inner, txt]:
                w.config(bg=bg)
            ic.config(bg=bg)
            self._draw_topbar_icon(ic, icon_type, bg)

        for w in all_w:
            w.bind("<Enter>", _on_enter)
            w.bind("<Leave>", _on_leave)
        return f

    def _build_topbar(self, parent):
        """Build the Gestio-RPS-style top navigation bar."""
        now = datetime.now()
        acad_year = f"{now.year}/{now.year + 1}"
        uname = (
            self.current_user.get("username", "Admin") if self.current_user else "Admin"
        )
        role = uname.title()

        # ── Outer bar – right-area bg fills the whole bar ─────────────────
        bar = tk.Frame(parent, bg=TOPBAR_RIGHT_BG, height=TOPBAR_H)
        bar.pack(fill="x", side="top")
        bar.pack_propagate(False)
        self._topbar_bar = bar
        bar.bind("<Configure>", self._update_topbar_responsive)
        self._topbar_compact_btn(
            bar,
            "menu",
            "Menu",
            TOPBAR_BTN_BG,
            command=self._toggle_sidebar,
            side="left",
        )

        # ── Right buttons – packed right→left so visual order = left→right ─
        # About
        self._topbar_compact_btn(
            bar,
            "about",
            "About",
            TOPBAR_BTN_BG,
            command=lambda: messagebox.showinfo(
                "About", "School MIS v1.0\nMT Olives Adventist School"
            ),
        )
        # Log Out  (License removed)
        self._topbar_compact_btn(
            bar, "logout", "Log Out", TOPBAR_BTN_BG, command=self.logout
        )

        # ── User section (purple) ──────────────────────────────────────────
        usr_f = tk.Frame(bar, bg=TOPBAR_USER_BG, padx=10, cursor="hand2")
        usr_f.pack(side="right", fill="y")

        # Crisp Canvas person icon
        av = tk.Canvas(
            usr_f, width=28, height=28, bg=TOPBAR_USER_BG, highlightthickness=0
        )
        av.pack(side="left", pady=11)
        av.create_oval(7, 1, 21, 15, fill="white", outline="")  # head
        av.create_arc(
            1, 13, 27, 31, start=0, extent=180, fill="white", outline=""
        )  # shoulders

        tk.Label(
            usr_f,
            text=f"  admin | {role}",
            bg=TOPBAR_USER_BG,
            fg="white",
            font=(FF, 9, "bold"),
        ).pack(side="left", pady=14)

        # ── Academic year badge (green) ───────────────────────────────────
        yr_f = tk.Frame(bar, bg=TOPBAR_YR_BG, padx=10, cursor="hand2")
        yr_f.pack(side="right", fill="y", padx=(0, 2))

        # Canvas checkmark
        ck = tk.Canvas(yr_f, width=22, height=18, bg=TOPBAR_YR_BG, highlightthickness=0)
        ck.pack(side="left", pady=16)
        ck.create_line(
            2, 9, 8, 15, fill="white", width=2.5, capstyle="round", joinstyle="round"
        )
        ck.create_line(
            8, 15, 20, 4, fill="white", width=2.5, capstyle="round", joinstyle="round"
        )

        tk.Label(
            yr_f, text=acad_year, bg=TOPBAR_YR_BG, fg="white", font=(FF, 7, "bold")
        ).pack(side="left", padx=(4, 0), pady=16)

        # ── CBC Level Selector ────────────────────────────────────────────────
        level_frame = tk.Frame(bar, bg=OLIVE_PRIMARY, padx=8, cursor="hand2")
        level_frame.pack(side="right", fill="y", padx=(8, 2))

        tk.Label(
            level_frame, text="CBC:", bg=OLIVE_PRIMARY, fg="white", font=(FF, 8, "bold")
        ).pack(side="left", pady=14)

        # Level dropdown
        self.level_var = tk.StringVar(value=self.current_level)
        level_cb = ttk.Combobox(
            level_frame,
            textvariable=self.level_var,
            values=[ALL_SCHOOL_LEVEL] + LEVELS,
            state="readonly",
            width=18,
        )
        level_cb.pack(side="left", padx=(4, 0), pady=11)
        level_cb.bind("<<ComboboxSelected>>", lambda e: self._on_level_change())

        # ── Live clock ────────────────────────────────────────────────────
        dt_lbl = tk.Label(bar, bg=TOPBAR_RIGHT_BG, fg="#c8e6c9", font=(FF, 9), padx=10)
        dt_lbl.pack(side="right", fill="y", pady=14)
        self._topbar_clock_label = dt_lbl
        self._topbar_clock_visible = True
        self._update_topbar_responsive()
        self._topbar_clock_generation += 1
        self._tick_topbar_clock(self._topbar_clock_generation)

    def _tick_topbar_clock(self, generation=None):
        try:
            if generation is not None and generation != self._topbar_clock_generation:
                self._topbar_clock_job = None
                return
            if (
                not self._topbar_clock_label
                or not self._topbar_clock_label.winfo_exists()
            ):
                self._topbar_clock_job = None
                return
            self._topbar_clock_label.config(
                text=datetime.now().strftime("%a, %d %b %Y  %H:%M:%S")
            )
            self._topbar_clock_job = self.root.after(
                1000, lambda gen=generation: self._tick_topbar_clock(gen)
            )
        except Exception:
            self._topbar_clock_job = None

    def _update_topbar_responsive(self, event=None):
        """Hide lower-priority topbar elements when the window gets narrow."""
        if not self._topbar_bar or not self._topbar_bar.winfo_exists():
            return
        if not self._topbar_clock_label or not self._topbar_clock_label.winfo_exists():
            return

        width = event.width if event is not None else self._topbar_bar.winfo_width()
        should_show_clock = width >= 1500

        if should_show_clock and not self._topbar_clock_visible:
            self._topbar_clock_label.pack(side="right", fill="y", pady=14)
            self._topbar_clock_visible = True
        elif not should_show_clock and self._topbar_clock_visible:
            self._topbar_clock_label.pack_forget()
            self._topbar_clock_visible = False

    def _on_level_change(self):
        """Handle CBC level change - update subjects and classes"""
        new_level = self.level_var.get()
        if new_level and new_level != self.current_level:
            self.set_level(new_level)
            # Refresh current view if logged in
            if self.active_nav:
                # Navigate to dashboard to refresh
                self.show_dashboard()
            messagebox.showinfo(
                "CBC Level Changed",
                f"Switched to: {new_level}\n\nClasses available:\n"
                + "\n".join(f"• {c}" for c in self.get_current_classes()[:6])
                + ("..." if len(self.get_current_classes()) > 6 else "")
                + f"\n\nSubjects updated to:\n"
                + "\n".join(f"• {s}" for s in self.get_current_subjects()[:5])
                + ("..." if len(self.get_current_subjects()) > 5 else ""),
            )

    def _toggle_sidebar(self):
        """Collapse or expand the sidebar."""
        self.sidebar_collapsed = not self.sidebar_collapsed
        self._rebuild_sidebar()

    def _rebuild_sidebar(self):
        """Re-render the sidebar in its current collapsed/expanded state."""
        if not self.sidebar_host or not self.sidebar_host.winfo_exists():
            return
        for child in self.sidebar_host.winfo_children():
            child.destroy()
        self._build_sidebar(self.sidebar_host)
        if self.active_nav:
            self._set_nav(self.active_nav)

    def _build_sidebar(self, parent):
        sidebar_width = 72 if self.sidebar_collapsed else 220
        sb = tk.Frame(parent, bg=SIDEBAR_BG, width=sidebar_width)
        sb.pack(side="left", fill="y")
        sb.pack_propagate(False)

        # --- logo ---
        logo_row = tk.Frame(sb, bg=SIDEBAR_BG)
        logo_row.pack(fill="x", padx=8 if self.sidebar_collapsed else 14, pady=(8, 8))

        if self.logo_img:
            tk.Label(logo_row, image=self.logo_img, bg=SIDEBAR_BG).pack(side="left")
        else:
            icon_c = tk.Canvas(
                logo_row, width=40, height=40, bg=SIDEBAR_BG, highlightthickness=0
            )
            icon_c.pack(side="left")
            icon_c.create_oval(
                0, 0, 40, 40, fill=SIDEBAR_ACTIVE, outline=SIDEBAR_ACTIVE
            )
            icon_c.create_text(20, 20, text="\U0001f393", font=(FF, 17))

        if not self.sidebar_collapsed:
            title_box = tk.Frame(logo_row, bg=SIDEBAR_BG)
            title_box.pack(side="left", padx=8)
            tk.Label(
                title_box,
                text=get_school_profile().get(
                    "school_sidebar_title",
                    DEFAULT_SCHOOL_PROFILE["school_sidebar_title"],
                ),
                bg=SIDEBAR_BG,
                fg="white",
                font=(FF, 10, "bold"),
            ).pack(anchor="w")
            tk.Label(
                title_box,
                text=get_school_profile().get(
                    "school_sidebar_subtitle",
                    DEFAULT_SCHOOL_PROFILE["school_sidebar_subtitle"],
                ),
                bg=SIDEBAR_BG,
                fg=SIDEBAR_TEXT,
                font=(FF, 7),
            ).pack(anchor="w")

        tk.Frame(sb, bg="#2E7D32", height=1).pack(
            fill="x", padx=8 if self.sidebar_collapsed else 14, pady=8
        )

        # --- nav items ---
        nav_cfg = self._get_role_based_nav()
        nav_wrap = tk.Frame(sb, bg=SIDEBAR_BG)
        nav_wrap.pack(
            fill="both",
            expand=True,
            padx=(4, 2) if self.sidebar_collapsed else (8, 4),
            pady=(0, 4),
        )

        nav_canvas = tk.Canvas(nav_wrap, bg=SIDEBAR_BG, highlightthickness=0, bd=0)
        nav_scroll = ttk.Scrollbar(
            nav_wrap, orient="vertical", command=nav_canvas.yview
        )
        nav_box = tk.Frame(nav_canvas, bg=SIDEBAR_BG)

        nav_box.bind(
            "<Configure>",
            lambda e: nav_canvas.configure(scrollregion=nav_canvas.bbox("all")),
        )

        nav_window = nav_canvas.create_window((0, 0), window=nav_box, anchor="nw")

        def _resize_sidebar_nav(event):
            nav_canvas.itemconfigure(nav_window, width=event.width)

        def _scroll_sidebar_nav(event):
            nav_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        nav_canvas.bind("<Configure>", _resize_sidebar_nav)
        nav_canvas.bind("<MouseWheel>", _scroll_sidebar_nav)
        nav_canvas.bind("<Enter>", lambda e: nav_canvas.focus_set())
        nav_box.bind("<MouseWheel>", _scroll_sidebar_nav)

        nav_canvas.configure(yscrollcommand=nav_scroll.set, yscrollincrement=20)
        nav_canvas.pack(side="left", fill="both", expand=True)
        nav_scroll.pack(side="right", fill="y")
        self.nav_frames = {}

        for item in nav_cfg:
            if item[0] == "section":
                self._nav_section(nav_box, item[1])
            else:
                icon, label, cmd = item
                self._nav_item(nav_box, icon, label, cmd)

        # --- bottom ---
        bot = tk.Frame(sb, bg=SIDEBAR_BG)
        bot.pack(
            side="bottom", fill="x", padx=8 if self.sidebar_collapsed else 14, pady=12
        )
        tk.Frame(bot, bg="#2E7D32", height=1).pack(fill="x", pady=(0, 8))

        uname = (
            self.current_user.get("username", "admin") if self.current_user else "admin"
        )
        email = uname if "@" in uname else uname + "@school.ac"
        if not self.sidebar_collapsed:
            tk.Label(
                bot,
                text=email,
                bg=SIDEBAR_BG,
                fg=SIDEBAR_TEXT,
                font=(FF, 9),
                anchor="w",
            ).pack(fill="x", pady=(0, 6))
            tk.Label(
                bot,
                text=self._get_active_db_display(),
                bg=SIDEBAR_BG,
                fg="#d8ef9c",
                font=(FF, 8, "bold"),
                anchor="w",
            ).pack(fill="x", pady=(0, 6))

        so = tk.Frame(bot, bg=SIDEBAR_BG, cursor="hand2")
        so.pack(fill="x")
        sign_out_text = "\u2192" if self.sidebar_collapsed else "\u2192  Sign Out"
        lbl_so = tk.Label(
            so,
            text=sign_out_text,
            bg=SIDEBAR_BG,
            fg=SIDEBAR_TEXT,
            font=(FF, 10),
            anchor="center" if self.sidebar_collapsed else "w",
            padx=4,
            pady=5,
        )
        lbl_so.pack(fill="x")

        for w in (so, lbl_so):
            w.bind("<Button-1>", lambda e: self.logout())
            w.bind(
                "<Enter>",
                lambda e: [
                    so.config(bg=SIDEBAR_HOVER),
                    lbl_so.config(bg=SIDEBAR_HOVER),
                ],
            )
            w.bind(
                "<Leave>",
                lambda e: [so.config(bg=SIDEBAR_BG), lbl_so.config(bg=SIDEBAR_BG)],
            )

    def _get_role_based_nav(self):
        """Get navigation items based on user role."""
        role = getattr(self, "user_role", "admin")
        if role == "admin":
            return [
                ("section", "Main", None),
                ("🏠", "Dashboard", self.show_dashboard),
                ("section", "Academics", None),
                ("🏫", "Classes", self.show_settings_classes),
                ("📚", "Subjects", self.show_settings_subjects),
                ("👩‍🏫", "Teachers", self.show_settings_teachers),
                ("section", "Students", None),
                ("👥", "Students", self.show_students),
                ("📝", "Enter Marks", self.show_marks_entry),
                ("📊", "Results", self.show_reports),
                ("📄", "Report Cards", self.show_report_cards),
                ("🎓", "Promotions", self.show_promotions),
                ("section", "Analytics", None),
                ("📈", "Charts", self.show_charts),
                ("📊", "Exam Analytics", self.show_exam_analytics),
                ("📏", "Grading Scale", self.show_settings_grading),
                ("section", "System", None),
                ("⚙️", "Settings", self.show_settings),
            ]

        if role == "teacher":
            return [
                ("section", "Main", None),
                ("🏠", "Dashboard", self.show_dashboard),
                ("section", "Students", None),
                ("📝", "Enter Marks", self.show_marks_entry),
            ]

        if role == "class_teacher":
            return [
                ("section", "Main", None),
                ("🏠", "Dashboard", self.show_dashboard),
                ("section", "Students", None),
                ("👥", "My Students", self.show_class_students),
                ("📝", "Enter Marks", self.show_marks_entry),
                ("💬", "Add Comments", self.show_add_comments),
                ("📄", "Report Cards", self.show_report_cards),
            ]

        return [
            ("section", "Main", None),
            ("🏠", "Dashboard", self.show_dashboard),
        ]

    def _nav_section(self, parent, title: str):
        if self.sidebar_collapsed:
            tk.Frame(parent, bg="#2E7D32", height=1).pack(
                fill="x", padx=10, pady=(4, 3)
            )
            return
        tk.Label(
            parent,
            text=title.upper(),
            bg=SIDEBAR_BG,
            fg="#c8d6a2",
            font=(FF, 8, "bold"),
            anchor="w",
            padx=10,
            pady=4,
        ).pack(fill="x", pady=(4, 1))

    def _nav_item(self, parent, icon: str, label: str, cmd):
        frame = tk.Frame(parent, bg=SIDEBAR_BG, cursor="hand2")
        frame.pack(fill="x", pady=1)

        row = tk.Frame(
            frame, bg=SIDEBAR_BG, padx=6 if self.sidebar_collapsed else 10, pady=4
        )
        row.pack(fill="x")

        ico = tk.Label(
            row,
            text=icon,
            bg=SIDEBAR_BG,
            fg=SIDEBAR_TEXT,
            font=(FF, 12),
            width=2,
            anchor="center",
        )
        ico.pack(side="left")
        txt = tk.Label(
            row,
            text="" if self.sidebar_collapsed else f"  {label}",
            bg=SIDEBAR_BG,
            fg=SIDEBAR_TEXT,
            font=(FF, 10),
            anchor="w",
        )
        if not self.sidebar_collapsed:
            txt.pack(side="left", fill="x")

        widgets = [frame, row, ico, txt]
        self.nav_frames[label] = widgets

        def activate(e=None):
            self._set_nav(label)
            cmd()

        def hover_on(e=None):
            if self.active_nav != label:
                for w in widgets:
                    w.config(bg=SIDEBAR_HOVER)
                ico.config(fg="white")
                txt.config(fg="white")

        def hover_off(e=None):
            if self.active_nav != label:
                for w in widgets:
                    w.config(bg=SIDEBAR_BG)
                ico.config(fg=SIDEBAR_TEXT)
                txt.config(fg=SIDEBAR_TEXT)

        for w in widgets:
            w.bind("<Button-1>", activate)
            w.bind("<Enter>", hover_on)
            w.bind("<Leave>", hover_off)

    def _set_nav(self, label: str):
        # deactivate old
        if self.active_nav and self.active_nav in self.nav_frames:
            for w in self.nav_frames[self.active_nav]:
                w.config(bg=SIDEBAR_BG)
            self.nav_frames[self.active_nav][2].config(fg=SIDEBAR_TEXT)  # icon
            self.nav_frames[self.active_nav][3].config(fg=SIDEBAR_TEXT)  # text
        # activate new
        self.active_nav = label
        if label in self.nav_frames:
            for w in self.nav_frames[label]:
                w.config(bg=SIDEBAR_ACTIVE)
            self.nav_frames[label][2].config(fg="white")
            self.nav_frames[label][3].config(fg="white")

    def _all_tree_items(self, tree, parent=""):
        items = []
        for iid in tree.get_children(parent):
            items.append(iid)
            items.extend(self._all_tree_items(tree, iid))
        return items

    def _select_all_tree_rows(self, tree):
        items = self._all_tree_items(tree)
        if items:
            tree.selection_set(items)

    def _clear_tree_selection(self, tree):
        selected = tree.selection()
        if selected:
            tree.selection_remove(selected)

    def logout(self):
        self.current_user = None
        self.user_role = "admin"
        self.show_login()

    def change_password(self):
        old = simpledialog.askstring(
            "Change Password", "Current password:", show="*", parent=self.root
        )
        if not old:
            return
        new = simpledialog.askstring(
            "Change Password", "New password:", show="*", parent=self.root
        )
        if not new:
            return
        if new != simpledialog.askstring(
            "Change Password", "Confirm new password:", show="*", parent=self.root
        ):
            messagebox.showerror("Error", "Passwords do not match")
            return
        if db.change_password(self.current_user["username"], old, new):
            messagebox.showinfo("Success", "Password changed successfully")
        else:
            messagebox.showerror("Error", "Current password is incorrect")

    # ==================== CLASSES ====================
    def show_classes(self):
        """Show classes inside the main content area with exam history and results."""
        self.clear_frame()
        self._set_nav("Classes")
        self._page_header(
            "Classes",
            "Browse stored classes, exam sessions, and detailed class results",
        )

        classes_data = db.get_all_classes_exam_history()
        classes_map = {row.get("class_name", ""): row for row in classes_data}

        controls = tk.Frame(self.content_frame, bg=CONTENT_BG)
        controls.pack(fill="x", pady=(0, 12))

        tk.Label(
            controls, text="Class:", bg=CONTENT_BG, fg=TEXT_SECONDARY, font=(FF, 10)
        ).pack(side="left", padx=(10, 4))
        class_names = [row.get("class_name", "") for row in classes_data]
        class_var = tk.StringVar(value=class_names[0] if class_names else "")
        class_cb = ttk.Combobox(
            controls,
            textvariable=class_var,
            values=class_names,
            state="readonly",
            style="App.TCombobox",
            width=18,
        )
        class_cb.pack(side="left", ipady=4)

        tk.Label(
            controls, text="Stream:", bg=CONTENT_BG, fg=TEXT_SECONDARY, font=(FF, 10)
        ).pack(side="left", padx=(16, 4))
        stream_var = tk.StringVar(value="All Streams")
        stream_cb = ttk.Combobox(
            controls,
            textvariable=stream_var,
            values=["All Streams"],
            state="readonly",
            style="App.TCombobox",
            width=14,
        )
        stream_cb.pack(side="left", ipady=4)

        tk.Label(
            controls, text="Search:", bg=CONTENT_BG, fg=TEXT_SECONDARY, font=(FF, 10)
        ).pack(side="left", padx=(16, 4))
        search_var = tk.StringVar()
        search_entry = ttk.Entry(
            controls, textvariable=search_var, style="App.TEntry", width=26
        )
        search_entry.pack(side="left", ipady=4)
        tk.Button(
            controls,
            text="🔍 Search",
            bg=OLIVE_PRIMARY,
            fg="white",
            activebackground=OLIVE_DARK,
            activeforeground="white",
            font=(FF, 9, "bold"),
            padx=10,
            pady=4,
            relief="flat",
            cursor="hand2",
            command=lambda: apply_search(),
        ).pack(side="left", padx=(8, 0))

        tk.Label(
            controls, text="Exam Type:", bg=CONTENT_BG, fg=TEXT_SECONDARY, font=(FF, 10)
        ).pack(side="left", padx=(16, 4))
        exam_filter_var = tk.StringVar(value="All")
        exam_filter_cb = ttk.Combobox(
            controls,
            textvariable=exam_filter_var,
            values=["All"]
            + self._get_ordered_exam_type_options(canonicalize_output=True),
            state="readonly",
            style="App.TCombobox",
            width=12,
        )
        exam_filter_cb.pack(side="left", ipady=4)

        actions_top = tk.Frame(controls, bg=CONTENT_BG)
        actions_top.pack(side="right")

        body = tk.Frame(self.content_frame, bg=CONTENT_BG)
        body.pack(fill="both", expand=True)
        body.columnconfigure(0, weight=4)
        body.columnconfigure(1, weight=5)
        body.rowconfigure(0, weight=1)

        l_bo, l_bi = _card_colors("mint")
        left_outer = tk.Frame(body, bg=l_bo)
        left_outer.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        left_card = tk.Frame(left_outer, bg=l_bi)
        left_card.pack(fill="both", expand=True, padx=1, pady=1)

        tk.Label(
            left_card,
            text="Stored Classes",
            bg=l_bi,
            fg=TEXT_PRIMARY,
            font=(FF, 12, "bold"),
        ).pack(anchor="w", padx=14, pady=(12, 6))

        class_tree_frame = tk.Frame(left_card, bg=l_bi)
        class_tree_frame.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        class_cols = ("class", "level", "stream", "students", "exams", "avg")
        class_tree = ttk.Treeview(
            class_tree_frame, columns=class_cols, show="headings", style="App.Treeview"
        )
        class_tree.heading("class", text="Class")
        class_tree.heading("level", text="Level")
        class_tree.heading("stream", text="Stream")
        class_tree.heading("students", text="Students")
        class_tree.heading("exams", text="Exams")
        class_tree.heading("avg", text="Latest Avg")
        class_tree.column("class", width=120, anchor="w")
        class_tree.column("level", width=150, anchor="w")
        class_tree.column("stream", width=85, anchor="center")
        class_tree.column("students", width=70, anchor="center")
        class_tree.column("exams", width=70, anchor="center")
        class_tree.column("avg", width=85, anchor="center")
        class_scroll = ttk.Scrollbar(
            class_tree_frame,
            orient="vertical",
            command=class_tree.yview,
            style="App.Vertical.TScrollbar",
        )
        class_tree.configure(yscrollcommand=class_scroll.set)
        class_tree.pack(side="left", fill="both", expand=True)
        class_scroll.pack(side="right", fill="y")

        rt_bo, rt_bi = _card_colors("sky")
        right_outer = tk.Frame(body, bg=rt_bo)
        right_outer.grid(row=0, column=1, sticky="nsew")
        right_card = tk.Frame(right_outer, bg=rt_bi)
        right_card.pack(fill="both", expand=True, padx=1, pady=1)

        summary = tk.Frame(right_card, bg=rt_bi, padx=16, pady=14)
        summary.pack(fill="x")
        class_title = tk.Label(
            summary,
            text="Select a class",
            bg=rt_bi,
            fg=TEXT_PRIMARY,
            font=(FF, 14, "bold"),
        )
        class_title.pack(anchor="w")
        class_meta = tk.Label(
            summary,
            text="Choose a class from the list or use the class selector above.",
            bg=rt_bi,
            fg=TEXT_SECONDARY,
            font=(FF, 10),
        )
        class_meta.pack(anchor="w", pady=(4, 10))

        stats_row = tk.Frame(summary, bg=rt_bi)
        stats_row.pack(fill="x")
        stat_labels = {}
        for key, title, theme in [
            ("students", "Students", "mint"),
            ("exams", "Exam Sessions", "sky"),
            ("avg", "Latest Average", "sand"),
        ]:
            _, bg = _card_colors(theme)
            tile = tk.Frame(stats_row, bg=bg, padx=12, pady=10)
            tile.pack(side="left", padx=(0, 10))
            tk.Label(
                tile, text=title, bg=bg, fg=TEXT_SECONDARY, font=(FF, 9, "bold")
            ).pack(anchor="w")
            stat_labels[key] = tk.Label(
                tile, text="-", bg=bg, fg=TEXT_PRIMARY, font=(FF, 14, "bold")
            )
            stat_labels[key].pack(anchor="w", pady=(4, 0))

        history_wrap = tk.Frame(right_card, bg=rt_bi, padx=16, pady=10)
        history_wrap.pack(fill="both", expand=True)
        tk.Label(
            history_wrap,
            text="Exam History",
            bg=rt_bi,
            fg=TEXT_PRIMARY,
            font=(FF, 11, "bold"),
        ).pack(anchor="w", pady=(0, 8))

        history_frame = tk.Frame(history_wrap, bg=rt_bi)
        history_frame.pack(fill="both", expand=True)
        exam_cols = ("term", "exam_type", "students", "subjects", "average")
        exam_tree = ttk.Treeview(
            history_frame, columns=exam_cols, show="headings", style="App.Treeview"
        )
        exam_tree.heading("term", text="Term")
        exam_tree.heading("exam_type", text="Exam")
        exam_tree.heading("students", text="Students")
        exam_tree.heading("subjects", text="Subjects")
        exam_tree.heading("average", text="Class Avg")
        exam_tree.column("term", width=80, anchor="center")
        exam_tree.column("exam_type", width=120, anchor="center")
        exam_tree.column("students", width=80, anchor="center")
        exam_tree.column("subjects", width=80, anchor="center")
        exam_tree.column("average", width=90, anchor="center")
        exam_scroll = ttk.Scrollbar(
            history_frame,
            orient="vertical",
            command=exam_tree.yview,
            style="App.Vertical.TScrollbar",
        )
        exam_tree.configure(yscrollcommand=exam_scroll.set)
        exam_tree.pack(side="left", fill="both", expand=True)
        exam_scroll.pack(side="right", fill="y")

        def format_avg(value):
            return "-" if value is None else f"{value}%"

        sort_state = {"classes": {}, "history": {}}

        def _coerce_sort_value(raw_value):
            value = str(raw_value or "").strip()
            if value in ("", "-", "No exam history"):
                return (0, "")
            if value.endswith("%"):
                try:
                    return (1, float(value[:-1]))
                except ValueError:
                    pass
            try:
                return (1, float(value))
            except ValueError:
                return (2, value.lower())

        def _sort_tree(tree, column, group_key):
            descending = sort_state[group_key].get(column, False)
            rows = [(tree.set(item, column), item) for item in tree.get_children("")]
            rows.sort(key=lambda row: _coerce_sort_value(row[0]), reverse=descending)
            for index, (_, item_id) in enumerate(rows):
                tree.move(item_id, "", index)
            sort_state[group_key][column] = not descending

        def get_stream_options(class_info):
            options = []
            class_name = class_info.get("class_name", "")
            class_row = class_info.get("class") or {}

            inline_stream = str(
                class_info.get("stream", "") or class_row.get("stream", "")
            ).strip()
            if inline_stream and inline_stream != "-":
                options.append(inline_stream)

            class_id = class_row.get("id")
            if class_id:
                for row in db.get_streams_for_class(class_id):
                    name = str(row.get("name", "")).strip()
                    if name and name not in options:
                        options.append(name)

            try:
                conn = db.get_connection()
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT DISTINCT stream FROM students WHERE class = ? AND stream IS NOT NULL AND TRIM(stream) != '' ORDER BY stream",
                    (class_name,),
                )
                for row in cursor.fetchall():
                    name = str(row["stream"]).strip()
                    if name and name not in options:
                        options.append(name)
            except Exception:
                pass
            finally:
                try:
                    conn.close()
                except Exception:
                    pass

            return ["All Streams"] + options

        def clear_history_panel(
            message="Choose a class from the list or use the class selector above.",
        ):
            class_title.config(text="Select a class")
            class_meta.config(text=message)
            for key in stat_labels:
                stat_labels[key].config(text="-")
            for item in exam_tree.get_children():
                exam_tree.delete(item)

        def get_selected_class_info():
            selected = class_tree.selection()
            if not selected:
                return None
            return classes_map.get(selected[0])

        def load_class_history(class_name):
            class_info = classes_map.get(class_name)
            if not class_info:
                clear_history_panel()
                return
            class_var.set(class_name)
            stream_options = get_stream_options(class_info)
            stream_cb["values"] = stream_options
            selected_stream = (stream_var.get() or "").strip()
            if selected_stream not in stream_options:
                selected_stream = "All Streams"
                stream_var.set(selected_stream)

            selected_stream = stream_var.get().strip()
            stream_text = (
                f" | Stream: {selected_stream}"
                if selected_stream != "All Streams"
                else ""
            )
            class_title.config(text=class_info.get("class_name", ""))
            class_meta.config(text=f"{class_info.get('level', '')}{stream_text}")

            for item in exam_tree.get_children():
                exam_tree.delete(item)

            all_exams = class_info.get("exams", [])
            selected_exam_type = exam_filter_var.get().strip()
            allowed_student_ids = None
            selected_stream = stream_var.get().strip()
            if selected_stream and selected_stream != "All Streams":
                stream_students = [
                    s
                    for s in db.get_students_by_class_and_stream(
                        class_name, selected_stream
                    )
                    if not self._is_summary_student_name(s.get("name"))
                ]
                allowed_student_ids = {
                    s.get("id") for s in stream_students if s.get("id")
                }
                stat_labels["students"].config(text=str(len(stream_students)))
            else:
                stat_labels["students"].config(
                    text=str(class_info.get("student_count", 0))
                )

            exams = []
            for index, exam in enumerate(all_exams):
                exam_type = self._canonical_exam_type(exam.get("exam_type", ""))
                if selected_exam_type != "All" and exam_type != selected_exam_type:
                    continue
                exam_copy = dict(exam)
                exam_copy["_index"] = index
                exam_copy["exam_type"] = exam_type
                exams.append(exam_copy)

            visible_exam_count = 0
            latest_avg = None
            if not exams:
                stat_labels["exams"].config(text="0")
                stat_labels["avg"].config(text="-")
                if selected_exam_type == "All":
                    exam_tree.insert(
                        "", "end", values=("-", "No exam history", "-", "-", "-")
                    )
                else:
                    exam_tree.insert(
                        "",
                        "end",
                        values=("-", f"No {selected_exam_type} history", "-", "-", "-"),
                    )
                return

            for exam in exams:
                term = exam.get("term", "")
                exam_type = exam.get("exam_type", "")
                results = db.get_class_exam_details(class_name, term, exam_type)
                if allowed_student_ids is not None:
                    results = [
                        row
                        for row in results
                        if row.get("student_id") in allowed_student_ids
                    ]
                    if not results:
                        continue
                visible_exam_count += 1
                subject_count = len(results[0].get("marks", {})) if results else 0
                avg = (
                    round(
                        sum(row.get("average", 0) for row in results) / len(results), 1
                    )
                    if results
                    else None
                )
                if latest_avg is None:
                    latest_avg = avg
                exam_tree.insert(
                    "",
                    "end",
                    iid=f"{class_name}::{exam.get('_index', 0)}",
                    values=(
                        term,
                        exam_type,
                        len(results),
                        subject_count,
                        format_avg(avg),
                    ),
                )
            stat_labels["exams"].config(text=str(visible_exam_count))
            stat_labels["avg"].config(text=format_avg(latest_avg))
            if visible_exam_count == 0:
                exam_tree.insert(
                    "",
                    "end",
                    values=("-", "No exam history for selected stream", "-", "-", "-"),
                )

        def select_class_in_tree(class_name):
            if not class_name or class_name not in class_tree.get_children():
                return
            class_tree.selection_set(class_name)
            class_tree.focus(class_name)
            class_tree.see(class_name)
            load_class_history(class_name)

        def populate_classes(filtered_rows=None):
            rows = filtered_rows if filtered_rows is not None else classes_data
            visible_names = [row.get("class_name", "") for row in rows]
            class_cb["values"] = visible_names
            for item in class_tree.get_children():
                class_tree.delete(item)
            for cls_info in rows:
                class_tree.insert(
                    "",
                    "end",
                    iid=cls_info["class_name"],
                    values=(
                        cls_info.get("class_name", ""),
                        cls_info.get("level", ""),
                        cls_info.get("stream", "") or "-",
                        cls_info.get("student_count", 0),
                        cls_info.get("exam_count", 0),
                        format_avg(cls_info.get("latest_avg")),
                    ),
                )
            if rows:
                current = class_var.get().strip()
                target = current if current in visible_names else rows[0]["class_name"]
                select_class_in_tree(target)
            else:
                class_var.set("")
                clear_history_panel("No classes match your search.")

        def open_selected_exam(event=None):
            class_info = get_selected_class_info()
            if not class_info:
                return
            exam_sel = exam_tree.selection()
            selected_stream = (stream_var.get() or "").strip()
            stream_filter = "" if selected_stream == "All Streams" else selected_stream
            if not exam_sel:
                if class_info.get("exams"):
                    payload = dict(class_info)
                    if stream_filter:
                        payload["stream_filter"] = stream_filter
                    self._show_class_exam_details(payload)
                return
            exam_id = exam_sel[0]
            if "::" not in exam_id:
                return
            exam_index = int(exam_id.split("::", 1)[1])
            exams = class_info.get("exams", [])
            if 0 <= exam_index < len(exams):
                payload = dict(class_info)
                payload["exams"] = [exams[exam_index]]
                if stream_filter:
                    payload["stream_filter"] = stream_filter
                self._show_class_exam_details(payload)

        def apply_search(*_args):
            term = search_var.get().strip().lower()
            if not term:
                populate_classes()
                return
            filtered = []
            for cls_info in classes_data:
                haystack = " ".join(
                    [
                        cls_info.get("class_name", ""),
                        cls_info.get("level", ""),
                        cls_info.get("stream", ""),
                    ]
                ).lower()
                if term in haystack:
                    filtered.append(cls_info)
            populate_classes(filtered)

        def refresh_page():
            nonlocal classes_data
            classes_data = db.get_all_classes_exam_history()
            classes_map.clear()
            classes_map.update({row.get("class_name", ""): row for row in classes_data})
            class_names = [row.get("class_name", "") for row in classes_data]
            class_cb["values"] = class_names
            populate_classes()

        search_var.trace_add("write", apply_search)
        search_entry.bind("<Return>", lambda _e: apply_search())
        class_tree.heading(
            "class",
            text="Class",
            command=lambda: _sort_tree(class_tree, "class", "classes"),
        )
        class_tree.heading(
            "level",
            text="Level",
            command=lambda: _sort_tree(class_tree, "level", "classes"),
        )
        class_tree.heading(
            "stream",
            text="Stream",
            command=lambda: _sort_tree(class_tree, "stream", "classes"),
        )
        class_tree.heading(
            "students",
            text="Students",
            command=lambda: _sort_tree(class_tree, "students", "classes"),
        )
        class_tree.heading(
            "exams",
            text="Exams",
            command=lambda: _sort_tree(class_tree, "exams", "classes"),
        )
        class_tree.heading(
            "avg",
            text="Latest Avg",
            command=lambda: _sort_tree(class_tree, "avg", "classes"),
        )

        exam_tree.heading(
            "term",
            text="Term",
            command=lambda: _sort_tree(exam_tree, "term", "history"),
        )
        exam_tree.heading(
            "exam_type",
            text="Exam",
            command=lambda: _sort_tree(exam_tree, "exam_type", "history"),
        )
        exam_tree.heading(
            "students",
            text="Students",
            command=lambda: _sort_tree(exam_tree, "students", "history"),
        )
        exam_tree.heading(
            "subjects",
            text="Subjects",
            command=lambda: _sort_tree(exam_tree, "subjects", "history"),
        )
        exam_tree.heading(
            "average",
            text="Class Avg",
            command=lambda: _sort_tree(exam_tree, "average", "history"),
        )

        class_cb.bind(
            "<<ComboboxSelected>>",
            lambda e: select_class_in_tree(class_var.get().strip()),
        )
        stream_cb.bind(
            "<<ComboboxSelected>>",
            lambda e: (
                load_class_history(class_tree.selection()[0])
                if class_tree.selection()
                else clear_history_panel()
            ),
        )
        exam_filter_cb.bind(
            "<<ComboboxSelected>>",
            lambda e: (
                load_class_history(class_tree.selection()[0])
                if class_tree.selection()
                else clear_history_panel()
            ),
        )
        class_tree.bind(
            "<<TreeviewSelect>>",
            lambda e: (
                load_class_history(class_tree.selection()[0])
                if class_tree.selection()
                else clear_history_panel()
            ),
        )
        exam_tree.bind("<Double-1>", open_selected_exam)

        tk.Button(
            actions_top,
            text="View Detailed Results",
            bg=OLIVE_PRIMARY,
            fg="white",
            font=(FF, 10, "bold"),
            padx=16,
            pady=8,
            command=open_selected_exam,
        ).pack(side="left", padx=(0, 8))
        tk.Button(
            actions_top,
            text="Refresh",
            bg=BLUE,
            fg="white",
            font=(FF, 10, "bold"),
            padx=16,
            pady=8,
            command=refresh_page,
        ).pack(side="left")

        populate_classes()

    def _show_class_exam_details(self, class_info):
        """Show detailed exam results for a specific class."""
        class_name = class_info.get("class_name", "")
        exams = class_info.get("exams", [])
        stream_filter = (class_info.get("stream_filter") or "").strip()

        if not exams:
            messagebox.showinfo(
                "No Exams", f"No exam results available for {class_name}"
            )
            return

        # Create dialog for selecting exam
        dialog = tk.Toplevel(self.root)
        initial_exam = exams[0] if exams else {}
        initial_heading = self._format_results_heading(
            class_name,
            initial_exam.get("term", "One"),
            initial_exam.get("exam_type", DEFAULT_EXAM_TYPE),
            stream_name=stream_filter,
        )
        dialog.title(initial_heading)
        dialog.geometry("1200x700")
        dialog.transient(self.root)
        dialog.grab_set()

        # Header
        header = tk.Frame(dialog, bg=OLIVE_PRIMARY, padx=20, pady=15)
        header.pack(fill="x")
        title_label = tk.Label(
            header,
            text=initial_heading,
            bg=OLIVE_PRIMARY,
            fg="white",
            font=(FF, 14, "bold"),
        )
        title_label.pack(side="left")
        if stream_filter:
            tk.Label(
                header,
                text=f"Stream: {stream_filter}",
                bg=OLIVE_PRIMARY,
                fg="white",
                font=(FF, 11, "bold"),
            ).pack(side="right")

        # Exam selector
        selector_frame = tk.Frame(dialog, bg=CONTENT_BG, padx=20, pady=10)
        selector_frame.pack(fill="x")

        tk.Label(
            selector_frame,
            text="Select Exam:",
            bg=CONTENT_BG,
            fg=TEXT_PRIMARY,
            font=(FF, 11),
        ).pack(side="left", padx=(0, 10))

        exam_options = [
            f"Term {e.get('term', '')} - {e.get('exam_type', '')}" for e in exams
        ]
        exam_var = tk.StringVar(value=exam_options[0] if exam_options else "")

        exam_combo = ttk.Combobox(
            selector_frame,
            textvariable=exam_var,
            values=exam_options,
            state="readonly",
            font=(FF, 10),
            width=25,
        )
        exam_combo.pack(side="left", padx=(0, 10))

        # Results table frame
        table_frame = tk.Frame(dialog, bg=CONTENT_BG, padx=20, pady=10)
        table_frame.pack(fill="both", expand=True)

        # Create treeview - will be configured dynamically
        tree = ttk.Treeview(table_frame, show="headings", height=20)

        # Scrollbar
        tree_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=tree_scroll.set)

        tree.pack(side="left", fill="both", expand=True)
        tree_scroll.pack(side="right", fill="y")

        def configure_tree(subjects):
            """Configure treeview columns based on available subjects."""
            # Clear existing columns
            tree.delete(*tree.get_children())
            for col in tree["columns"]:
                tree.column(col, width=0)
                tree.heading(col, text="")

            # Define columns
            cols = (
                ["position", "name", "admission"]
                + subjects
                + ["total", "average", "grade"]
            )
            tree["columns"] = cols

            # Configure column headings and widths
            tree.heading("position", text="#")
            tree.column("position", width=40, anchor="center")

            tree.heading("name", text="Student Name")
            tree.column("name", width=150, anchor="w")

            tree.heading("admission", text="Adm No")
            tree.column("admission", width=90, anchor="center")

            # Dynamic subject columns
            for subj in subjects:
                # Shorten long subject names for display
                short_name = subj[:12] if len(subj) > 12 else subj
                tree.heading(subj, text=short_name)
                tree.column(subj, width=55, anchor="center")

            tree.heading("total", text="Total")
            tree.column("total", width=60, anchor="center")

            tree.heading("average", text="Avg")
            tree.column("average", width=60, anchor="center")

            tree.heading("grade", text="Grade")
            tree.column("grade", width=50, anchor="center")

        def load_results():
            # Clear existing
            for item in tree.get_children():
                tree.delete(item)

            # Get selected exam
            selected_idx = exam_combo.current()
            if selected_idx < 0 or selected_idx >= len(exams):
                return

            exam = exams[selected_idx]
            term = exam.get("term", "One")
            exam_type = exam.get("exam_type", "End-Term")
            heading_text = self._format_results_heading(
                class_name, term, exam_type, stream_name=stream_filter
            )
            dialog.title(heading_text)
            title_label.config(text=heading_text)

            # Get results
            results = db.get_class_exam_details(class_name, term, exam_type)
            if stream_filter:
                stream_students = [
                    s
                    for s in db.get_students_by_class_and_stream(
                        class_name, stream_filter
                    )
                    if not self._is_summary_student_name(s.get("name"))
                ]
                allowed_ids = {s.get("id") for s in stream_students if s.get("id")}
                results = [
                    row for row in results if row.get("student_id") in allowed_ids
                ]
                results.sort(key=lambda row: row.get("total", 0), reverse=True)
                for idx, row in enumerate(results, start=1):
                    row["position"] = idx

            # Get subjects from first result if available
            subjects = []
            if results:
                subjects = list(results[0].get("marks", {}).keys())

            # Configure tree with subjects
            configure_tree(subjects)

            # Insert into tree
            for r in results:
                marks = r.get("marks", {})
                values = [
                    r.get("position", ""),
                    r.get("student_name", ""),
                    r.get("admission_no", ""),
                ]

                # Add marks for each subject
                for subj in subjects:
                    values.append(marks.get(subj, "-"))

                values.extend(
                    [r.get("total", ""), r.get("average", ""), r.get("grade", "")]
                )
                tree.insert("", "end", values=values)

        # Load button
        def on_exam_select(event=None):
            load_results()

        exam_combo.bind("<<ComboboxSelected>>", on_exam_select)

        tk.Button(
            selector_frame,
            text="Load Results",
            bg=OLIVE_PRIMARY,
            fg="white",
            font=(FF, 10),
            padx=15,
            pady=5,
            command=load_results,
        ).pack(side="left")

        # Close button
        close_frame = tk.Frame(dialog, bg=CONTENT_BG, pady=10)
        close_frame.pack(fill="x")
        tk.Button(
            close_frame,
            text="Close",
            bg=LEMON_SOFT,
            fg=TEXT_PRIMARY,
            font=(FF, 11),
            padx=20,
            pady=8,
            command=dialog.destroy,
        ).pack()

        # Load initial results
        load_results()

    # ==================== SETTINGS ====================
    def show_settings(
        self, initial_tab="classes", nav_label="Settings", show_tabs=True
    ):
        """Show settings/admin page for managing classes, subjects, and teachers."""
        self.clear_frame()
        self._set_nav(nav_label)
        page_title = "Settings" if show_tabs else nav_label
        self._page_header(page_title, "Manage school configuration")

        tabs_bar = None
        if show_tabs:
            # Custom RGB tab bar
            tabs_outer = tk.Frame(self.content_frame, bg="#d1d5db", padx=1, pady=1)
            tabs_outer.pack(fill="x", padx=20, pady=(10, 0))
            tabs_bar = tk.Frame(tabs_outer, bg="#e5e7eb")
            tabs_bar.pack(fill="x", padx=1, pady=1)

        content_outer = tk.Frame(self.content_frame, bg="#d1d5db", padx=1, pady=1)
        top_pad = (0, 10) if show_tabs else (10, 10)
        content_outer.pack(fill="both", expand=True, padx=20, pady=top_pad)
        content_stack = tk.Frame(content_outer, bg=CONTENT_BG)
        content_stack.pack(fill="both", expand=True, padx=1, pady=1)

        tab_defs = [
            ("classes", "🏫 Classes", "#2563eb", self._build_classes_tab),
            ("subjects", "📚 Subjects", "#0f766e", self._build_subjects_tab),
            ("teachers", "👩‍🏫 Teachers", "#7c3aed", self._build_teachers_settings_tab),
            ("grading", "📏 Grading Scale", "#ea580c", self._build_grading_scale_tab),
        ]

        tab_buttons = {}
        tab_frames = {}
        neutral_bg = "#475569"

        for key, label, _color, builder in tab_defs:
            frame = tk.Frame(content_stack, bg=CONTENT_BG)
            frame.place(relx=0, rely=0, relwidth=1, relheight=1)
            builder(frame)
            tab_frames[key] = frame

        def activate_tab(active_key):
            for key, _label, color, _builder in tab_defs:
                if key == active_key:
                    tab_frames[key].lift()
                    if show_tabs:
                        btn = tab_buttons[key]
                        btn.configure(bg=color, fg="white", relief="sunken", bd=2)
                else:
                    if show_tabs:
                        btn = tab_buttons[key]
                        btn.configure(bg=neutral_bg, fg="white", relief="raised", bd=1)

        if show_tabs:
            for key, label, color, _builder in tab_defs:
                btn = tk.Button(
                    tabs_bar,
                    text=f"  {label}  ",
                    bg=neutral_bg,
                    fg="white",
                    activebackground=color,
                    activeforeground="white",
                    font=(FF, 10, "bold"),
                    padx=10,
                    pady=8,
                    relief="raised",
                    bd=1,
                    cursor="hand2",
                    command=lambda k=key: activate_tab(k),
                )
                btn.pack(side="left", padx=2, pady=2)
                tab_buttons[key] = btn

        start_tab = initial_tab if initial_tab in tab_frames else "classes"
        activate_tab(start_tab)

    def show_settings_classes(self):
        self.show_settings(initial_tab="classes", nav_label="Classes", show_tabs=False)

    def show_settings_streams(self):
        self.show_settings(initial_tab="classes", nav_label="Classes", show_tabs=False)

    def show_settings_subjects(self):
        self.show_settings(
            initial_tab="subjects", nav_label="Subjects", show_tabs=False
        )

    def show_settings_teachers(self):
        self.show_settings(
            initial_tab="teachers", nav_label="Teachers", show_tabs=False
        )

    def show_settings_grading(self):
        self.show_settings(
            initial_tab="grading", nav_label="Grading Scale", show_tabs=False
        )

    def show_settings_assignments(self):
        self.show_settings(
            initial_tab="teachers", nav_label="Teachers", show_tabs=False
        )

    def _build_classes_tab(self, parent, initial_scope="Both"):
        toolbar = tk.Frame(parent, bg=CONTENT_BG)
        toolbar.pack(fill="x", pady=10)

        tk.Button(
            toolbar,
            text="+ Add Class",
            bg=GREEN,
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: self._open_class_dialog(on_save=refresh_tree),
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="+ Add Stream",
            bg="#0891b2",
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: self._open_stream_dialog(on_save=refresh_tree),
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="📥 Template",
            bg=ORANGE,
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=self.download_classes_import_template,
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="📥 Import Excel",
            bg=BLUE,
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: self.import_classes_excel(on_complete=refresh_tree),
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Edit Selected",
            bg=BLUE,
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: edit_selected(),
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Delete Selected",
            bg="#e74c3c",
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: delete_selected(),
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Class Subjects Done",
            bg="#0f766e",
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=self._open_class_subjects_done_dialog,
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Export CSV",
            bg="#1d4ed8",
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: self._export_unified_classes_streams(tree),
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="🗑️ Delete ALL Classes",
            bg="#c41e3a",
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=self._delete_all_classes_batch,
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Refresh",
            bg="#666",
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: refresh_tree(),
        ).pack(side="left", padx=5)

        filters = tk.Frame(parent, bg=CONTENT_BG)
        filters.pack(fill="x", padx=10, pady=(0, 4))
        tk.Label(
            filters, text="View:", bg=CONTENT_BG, fg=TEXT_SECONDARY, font=(FF, 10)
        ).pack(side="left", padx=(0, 4))
        scope_var = tk.StringVar(
            value=initial_scope
            if initial_scope in ("Both", "Classes Only", "Streams Only")
            else "Both"
        )
        scope_cb = ttk.Combobox(
            filters,
            textvariable=scope_var,
            values=["Both", "Classes Only", "Streams Only"],
            state="readonly",
            style="App.TCombobox",
            width=14,
        )
        scope_cb.pack(side="left", ipady=3)
        tk.Label(
            filters, text="Search:", bg=CONTENT_BG, fg=TEXT_SECONDARY, font=(FF, 10)
        ).pack(side="left", padx=(14, 4))
        search_var = tk.StringVar()
        search_entry = ttk.Entry(
            filters, textvariable=search_var, style="App.TEntry", width=30
        )
        search_entry.pack(side="left", ipady=3)
        tk.Button(
            filters,
            text="🔍 Search",
            bg=OLIVE_PRIMARY,
            fg="white",
            activebackground=OLIVE_DARK,
            activeforeground="white",
            font=(FF, 9, "bold"),
            padx=10,
            pady=4,
            relief="flat",
            cursor="hand2",
            command=lambda: filtered(search_var.get(), scope_var.get()),
        ).pack(side="left", padx=(8, 0))

        list_frame = tk.Frame(parent, bg=CARD_BG, relief="flat", bd=1)
        list_frame.pack(fill="both", expand=True, padx=10, pady=5)

        cols = ("type", "label", "abbr", "level", "parent", "students", "subjects")
        tree = ttk.Treeview(
            list_frame,
            columns=cols,
            show="tree headings",
            style="App.Treeview",
            selectmode="extended",
        )
        tree.heading("#0", text="Item")
        tree.heading("type", text="Type")
        tree.heading("label", text="Name")
        tree.heading("abbr", text="Short Label")
        tree.heading("level", text="Level")
        tree.heading("parent", text="Parent Class")
        tree.heading("students", text="Students")
        tree.heading("subjects", text="Subjects")

        tree.column("#0", width=200, anchor="w")
        tree.column("type", width=95, anchor="center")
        tree.column("label", width=170, anchor="w")
        tree.column("abbr", width=105, anchor="center")
        tree.column("level", width=210, anchor="w")
        tree.column("parent", width=150, anchor="w")
        tree.column("students", width=80, anchor="center")
        tree.column("subjects", width=80, anchor="center")

        select_all_var = tk.BooleanVar(value=False)
        select_row = tk.Frame(list_frame, bg=CARD_BG)
        select_row.pack(fill="x", padx=10, pady=(8, 0))
        tk.Checkbutton(
            select_row,
            text="Select all classes/streams",
            variable=select_all_var,
            bg=CARD_BG,
            fg=TEXT_SECONDARY,
            activebackground=CARD_BG,
            activeforeground=TEXT_PRIMARY,
            selectcolor=CARD_BG,
            font=(FF, 9, "bold"),
            command=lambda: (
                self._select_all_tree_rows(tree)
                if select_all_var.get()
                else self._clear_tree_selection(tree)
            ),
        ).pack(side="left")

        yscroll = ttk.Scrollbar(
            list_frame,
            orient="vertical",
            command=tree.yview,
            style="App.Vertical.TScrollbar",
        )
        xscroll = ttk.Scrollbar(
            list_frame,
            orient="horizontal",
            command=tree.xview,
            style="App.Horizontal.TScrollbar",
        )
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)

        tree.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=10)
        yscroll.pack(side="right", fill="y", pady=(10, 0))
        xscroll.pack(side="bottom", fill="x", padx=10, pady=(0, 10))

        sort_state = {}
        class_cache = {}
        stream_cache = {}

        def _is_summary(student_row):
            return self._is_summary_student_name(student_row.get("name"))

        def _subject_count_for_class(class_name):
            try:
                return len(
                    self._get_subjects_for_class(
                        class_name, "One", DEFAULT_EXAM_TYPE, for_reporting=True
                    )
                )
            except Exception:
                return 0

        def _stream_subject_count(class_name, stream_name):
            subjects = set()
            students = [
                s
                for s in db.get_students_by_class_and_stream(class_name, stream_name)
                if not _is_summary(s)
            ]
            for student in students:
                sid = student.get("id")
                if not sid:
                    continue
                try:
                    conn = db.get_connection()
                    cursor = conn.cursor()
                    cursor.execute(
                        "SELECT DISTINCT subject FROM marks WHERE student_id = ? AND subject IS NOT NULL AND TRIM(subject) != ''",
                        (sid,),
                    )
                    for row in cursor.fetchall():
                        subjects.add(str(row["subject"]).strip())
                except Exception:
                    pass
                finally:
                    try:
                        conn.close()
                    except Exception:
                        pass
            return len(subjects)

        def _coerce_sort_value(value):
            s = str(value or "").strip()
            if s in ("", "-", "Class", "Stream"):
                return (0, "")
            try:
                return (1, float(s))
            except ValueError:
                return (2, s.lower())

        def _flatten_rows():
            rows = []
            for cid in tree.get_children(""):
                rows.append(cid)
                for sid in tree.get_children(cid):
                    rows.append(sid)
            return rows

        def sort_by(column):
            reverse = sort_state.get(column, False)
            items = _flatten_rows()
            keyed = [(tree.set(iid, column), iid, tree.parent(iid)) for iid in items]
            keyed.sort(
                key=lambda row: (_coerce_sort_value(row[0]), row[2] != ""),
                reverse=reverse,
            )
            # Rebuild top-level and child ordering while keeping hierarchy
            top = [iid for _v, iid, parent in keyed if parent == ""]
            for idx, iid in enumerate(top):
                tree.move(iid, "", idx)
                children = [sid for _v, sid, p in keyed if p == iid]
                for cidx, sid in enumerate(children):
                    tree.move(sid, iid, cidx)
            sort_state[column] = not reverse

        def filtered(term, scope):
            q = term.strip().lower()
            for item in tree.get_children():
                tree.delete(item)

            for class_id, cls in class_cache.items():
                class_name = cls.get("name", "")
                class_level = cls.get("level", "")
                class_abbr = cls.get("abbreviation", "") or self._generate_short_label(
                    class_name, "class"
                )
                class_students = [
                    s
                    for s in db.get_students_by_class_and_stream(class_name, "")
                    if not _is_summary(s)
                ]
                class_streams = stream_cache.get(class_id, [])
                stream_count_total = len(class_streams)
                class_subjects = _subject_count_for_class(class_name)

                class_haystack = " ".join([class_name, class_level, class_abbr]).lower()
                class_match = (not q) or (q in class_haystack)
                include_class_row = scope in ("Both", "Classes Only") and class_match
                include_stream_children = scope in ("Both", "Streams Only")

                parent_iid = f"class::{class_id}"
                if include_class_row:
                    class_label = f"{class_name} ({stream_count_total} streams)"
                    tree.insert(
                        "",
                        "end",
                        iid=parent_iid,
                        text=f"🏫 {class_label}",
                        values=(
                            "Class",
                            class_name,
                            class_abbr,
                            class_level,
                            "-",
                            len(class_students),
                            class_subjects,
                        ),
                    )

                stream_rows = []
                for stream in class_streams:
                    stream_name = stream.get("name", "")
                    stream_students = [
                        s
                        for s in db.get_students_by_class_and_stream(
                            class_name, stream_name
                        )
                        if not _is_summary(s)
                    ]
                    stream_subjects = _stream_subject_count(class_name, stream_name)
                    stream_haystack = " ".join(
                        [stream_name, class_name, class_level]
                    ).lower()
                    if q and q not in stream_haystack:
                        continue
                    stream_rows.append((stream, len(stream_students), stream_subjects))

                if include_stream_children and stream_rows:
                    attach_parent = parent_iid if include_class_row else ""
                    for stream, student_count, subject_count in stream_rows:
                        sid = stream.get("id", "")
                        tree.insert(
                            attach_parent,
                            "end",
                            iid=f"stream::{sid}",
                            text=f"🧩 {stream.get('name', '')}",
                            values=(
                                "Stream",
                                stream.get("name", ""),
                                "-",
                                class_level,
                                class_name,
                                student_count,
                                subject_count,
                            ),
                        )

        def refresh_tree():
            nonlocal class_cache, stream_cache
            classes = db.get_all_classes()
            class_cache = {row.get("id", ""): row for row in classes}
            stream_cache = {}
            for class_id in class_cache:
                stream_cache[class_id] = db.get_streams_for_class(class_id)
            filtered(search_var.get(), scope_var.get())

        def selected_payload():
            sel = tree.selection()
            if not sel:
                return None, None
            iid = sel[0]
            if iid.startswith("class::"):
                class_id = iid.split("::", 1)[1]
                return "class", class_cache.get(class_id)
            if iid.startswith("stream::"):
                stream_id = iid.split("::", 1)[1]
                for class_id, rows in stream_cache.items():
                    found = next(
                        (row for row in rows if row.get("id") == stream_id), None
                    )
                    if found:
                        payload = dict(found)
                        payload["class_id"] = class_id
                        payload["class_name"] = class_cache.get(class_id, {}).get(
                            "name", ""
                        )
                        return "stream", payload
            return None, None

        def edit_selected():
            selected = tree.selection()
            if len(selected) != 1:
                messagebox.showwarning(
                    "Select One", "Please select exactly one class/stream to edit"
                )
                return
            kind, payload = selected_payload()
            if not kind:
                messagebox.showwarning(
                    "Select", "Please select a class or stream to edit"
                )
                return
            if kind == "class":
                self._open_class_dialog(payload, on_save=refresh_tree)
                return
            self._open_stream_dialog(payload, on_save=refresh_tree)

        def delete_selected():
            selected = list(tree.selection())
            if not selected:
                messagebox.showwarning(
                    "Select", "Please select one or more classes/streams to delete"
                )
                return

            class_ids = set()
            stream_ids = set()
            for iid in selected:
                if iid.startswith("class::"):
                    class_ids.add(iid.split("::", 1)[1])
                elif iid.startswith("stream::"):
                    stream_ids.add(iid.split("::", 1)[1])

            # Skip streams whose parent class is already selected (class delete cascades stream delete)
            if class_ids and stream_ids:
                covered_streams = set()
                for class_id in class_ids:
                    for stream in stream_cache.get(class_id, []):
                        sid = stream.get("id")
                        if sid:
                            covered_streams.add(sid)
                stream_ids = {sid for sid in stream_ids if sid not in covered_streams}

            if not class_ids and not stream_ids:
                messagebox.showwarning("Select", "No valid classes/streams selected.")
                return

            total = len(class_ids) + len(stream_ids)
            if not self._confirm_delete_action(
                "class/stream",
                total,
                scope="selected",
                details="Selected classes will also remove their streams, students, marks, and assignments.",
            ):
                return

            errors = []
            # Delete classes first (cascades streams)
            for class_id in class_ids:
                if not db.delete_class(class_id):
                    errors.append(f"Class {class_id[:8]}")
            for stream_id in stream_ids:
                if not db.delete_stream(stream_id):
                    errors.append(f"Stream {stream_id[:8]}")

            self._show_delete_result_notice(
                "class/stream", total - len(errors), len(errors), duration_ms=4200
            )
            refresh_tree()

        for col, text in [
            ("type", "Type"),
            ("label", "Name"),
            ("abbr", "Short Label"),
            ("level", "Level"),
            ("parent", "Parent Class"),
            ("students", "Students"),
            ("subjects", "Subjects"),
        ]:
            tree.heading(col, text=text, command=lambda c=col: sort_by(c))

        tree.bind("<Double-1>", lambda _e: edit_selected())
        scope_cb.bind(
            "<<ComboboxSelected>>",
            lambda _e: filtered(search_var.get(), scope_var.get()),
        )
        search_entry.bind(
            "<Return>", lambda _e: filtered(search_var.get(), scope_var.get())
        )
        search_var.trace_add(
            "write", lambda *_: filtered(search_var.get(), scope_var.get())
        )
        refresh_tree()

    def _load_classes(self, tree):
        for item in tree.get_children():
            tree.delete(item)

        classes = db.get_all_classes()
        for cls in classes:
            tree.insert(
                "",
                "end",
                iid=cls.get("id", ""),
                values=(
                    cls.get("id", "")[:8],
                    cls.get("name", ""),
                    cls.get("abbreviation", "")
                    or self._generate_short_label(cls.get("name", ""), "class"),
                    cls.get("level", ""),
                    cls.get("stream", "") or "-",
                ),
            )

    def _open_class_dialog(self, class_row=None, on_save=None):
        dialog = tk.Toplevel(self.root)
        dialog.title("Edit Class" if class_row else "Add Class")
        dialog.geometry("430x390")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)

        tk.Label(dialog, text="Class Name:", font=(FF, 11)).pack(pady=(20, 5))
        name_entry = tk.Entry(dialog, font=(FF, 11))
        name_entry.pack(fill="x", padx=20)

        tk.Label(dialog, text="Short Label / Abbreviation:", font=(FF, 11)).pack(
            pady=(15, 5)
        )
        abbr_entry = tk.Entry(dialog, font=(FF, 11))
        abbr_entry.pack(fill="x", padx=20)

        tk.Label(dialog, text="Level:", font=(FF, 11)).pack(pady=(15, 5))
        level_var = tk.StringVar()
        level_cb = ttk.Combobox(
            dialog,
            textvariable=level_var,
            values=LEVELS,
            state="readonly",
            font=(FF, 10),
        )
        level_cb.pack(fill="x", padx=20)

        tk.Label(dialog, text="Stream (optional):", font=(FF, 11)).pack(pady=(15, 5))
        stream_entry = tk.Entry(dialog, font=(FF, 11))
        stream_entry.pack(fill="x", padx=20)

        if class_row:
            name_entry.insert(0, class_row.get("name", ""))
            abbr_entry.insert(0, class_row.get("abbreviation", ""))
            level_var.set(class_row.get("level", ""))
            stream_entry.insert(0, class_row.get("stream", "") or "")
        else:
            level_var.set(
                self.current_level if self.current_level in LEVELS else LEVELS[0]
            )

        def save():
            name = name_entry.get().strip()
            abbreviation = abbr_entry.get().strip()
            level = level_var.get().strip()
            stream = stream_entry.get().strip()

            if not name or not level:
                messagebox.showerror("Error", "Class name and level are required")
                return

            if class_row:
                success, msg = db.update_class(
                    class_row["id"], name, level, stream or None, abbreviation
                )
            else:
                success, msg = db.add_class(name, level, stream or None, abbreviation)
            if success:
                refresh_dynamic_school_config()
                dialog.destroy()
                if callable(on_save):
                    on_save()
                else:
                    self.show_settings()
            else:
                messagebox.showerror("Error", msg)

        tk.Button(
            dialog,
            text="Save",
            bg=GREEN,
            fg="white",
            font=(FF, 11),
            padx=20,
            pady=8,
            command=save,
        ).pack(pady=(12, 20))

    def _edit_class_dialog(self, tree):
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("Select", "Please select a class to edit")
            return
        class_id = selected[0]
        class_row = next(
            (row for row in db.get_all_classes() if row.get("id") == class_id), None
        )
        if not class_row:
            messagebox.showerror("Error", "Could not load the selected class")
            return
        self._open_class_dialog(class_row)

    def _delete_class(self, tree):
        selected = tree.selection()
        if not selected:
            self._show_notice(
                "Select Class", "Please select a class to delete.", kind="info"
            )
            return

        if not self._confirm_delete_action(
            "class",
            1,
            details="This will also remove the class streams, students, marks, and assignments.",
        ):
            return

        class_id = selected[0]

        if db.delete_class(class_id):
            refresh_dynamic_school_config()
            self._show_delete_result_notice("class", 1, 0)
            self._load_classes(tree)
        else:
            self._show_delete_result_notice("class", 0, 1)

    def _build_streams_tab(self, parent):
        """Unified classes + streams view focused on streams."""
        self._build_classes_tab(parent, initial_scope="Streams Only")

    def _open_stream_dialog(self, stream_row=None, on_save=None):
        """Dialog to add or edit a stream."""
        classes = db.get_all_classes()
        if not classes:
            messagebox.showwarning("No Classes", "Please add classes first")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("Edit Stream" if stream_row else "Add Stream")
        dialog.geometry("420x290")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)

        tk.Label(dialog, text="Stream Name:", font=(FF, 11)).pack(pady=(20, 5))
        name_entry = tk.Entry(dialog, font=(FF, 11))
        name_entry.pack(fill="x", padx=20)

        tk.Label(dialog, text="Class:", font=(FF, 11)).pack(pady=(15, 5))
        class_var = tk.StringVar()
        class_cb = ttk.Combobox(
            dialog,
            textvariable=class_var,
            values=[c["name"] for c in classes],
            state="readonly",
            font=(FF, 10),
        )
        class_cb.pack(fill="x", padx=20)

        if stream_row:
            name_entry.insert(0, stream_row.get("name", ""))
            class_name = stream_row.get("class_name", "")
            if not class_name and stream_row.get("class_id"):
                class_name = next(
                    (
                        c.get("name", "")
                        for c in classes
                        if c.get("id") == stream_row.get("class_id")
                    ),
                    "",
                )
            class_var.set(class_name if class_name else classes[0]["name"])
        else:
            class_var.set(classes[0]["name"])

        def save():
            name = name_entry.get().strip()
            class_name = class_var.get()

            if not name or not class_name:
                messagebox.showerror("Error", "All fields are required")
                return

            # Get class ID
            class_id = next((c["id"] for c in classes if c["name"] == class_name), None)
            if not class_id:
                messagebox.showerror("Error", "Invalid class")
                return

            if stream_row:
                success, msg = db.update_stream(
                    stream_row.get("id", ""), name, class_id
                )
            else:
                success, msg = db.add_stream(name, class_id)
            if success:
                messagebox.showinfo("Success", msg)
                dialog.destroy()
                if callable(on_save):
                    on_save()
                else:
                    self.show_settings()
            else:
                messagebox.showerror("Error", msg)

        btn_row = tk.Frame(dialog, bg=dialog.cget("bg"))
        btn_row.pack(fill="x", pady=(16, 20), padx=20)
        tk.Button(
            btn_row,
            text="Cancel",
            bg=LEMON_SOFT,
            fg=TEXT_PRIMARY,
            font=(FF, 10, "bold"),
            padx=18,
            pady=8,
            command=dialog.destroy,
        ).pack(side="left", padx=(0, 8))
        tk.Button(
            btn_row,
            text="Save",
            bg=GREEN,
            fg="white",
            font=(FF, 11),
            padx=20,
            pady=8,
            command=save,
        ).pack(side="left")

    def _add_stream_dialog(self):
        """Backward-compatible add stream entry point."""
        self._open_stream_dialog()

    def _export_unified_classes_streams(self, tree):
        """Export unified classes/streams table to CSV."""
        path = filedialog.asksaveasfilename(
            title="Export Classes & Streams",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            initialfile="classes_streams.csv",
        )
        if not path:
            return
        try:
            with open(path, "w", newline="", encoding="utf-8") as fh:
                writer = csv.writer(fh)
                writer.writerow(
                    [
                        "Type",
                        "Name",
                        "Short Label",
                        "Level",
                        "Parent Class",
                        "Students",
                        "Subjects",
                    ]
                )
                for parent_id in tree.get_children(""):
                    parent_vals = tree.item(parent_id, "values")
                    writer.writerow(list(parent_vals))
                    for child_id in tree.get_children(parent_id):
                        child_vals = tree.item(child_id, "values")
                        writer.writerow(list(child_vals))
            messagebox.showinfo(
                "Export Complete", f"Classes & Streams exported to:\n{path}"
            )
        except Exception as exc:
            messagebox.showerror("Export Failed", f"Could not export CSV.\n\n{exc}")

    def download_classes_import_template(self):
        file_path = filedialog.asksaveasfilename(
            title="Save Classes Template",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="classes_import_template.xlsx",
        )
        if not file_path:
            return

        try:
            rows = db.get_all_classes()
            template_rows = []
            for row in rows:
                template_rows.append(
                    {
                        "class_name": row.get("name", ""),
                        "level": row.get("level", ""),
                        "stream": row.get("stream", "") or "",
                        "abbreviation": row.get("abbreviation", "") or "",
                    }
                )
            if not template_rows:
                template_rows = [
                    {
                        "class_name": "Grade 1",
                        "level": LEVELS[1] if len(LEVELS) > 1 else LEVELS[0],
                        "stream": "",
                        "abbreviation": "G1",
                    }
                ]

            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                pd.DataFrame(
                    {
                        "Instructions": [
                            "Use one row per class.",
                            "stream is optional. If provided, the importer will add it under the class.",
                            "If a class already exists, its level/abbreviation can be updated.",
                            "Accepted levels are the same ones used in the app settings.",
                        ]
                    }
                ).to_excel(writer, sheet_name="Instructions", index=False)
                pd.DataFrame(template_rows).to_excel(
                    writer, sheet_name="Classes", index=False
                )

            messagebox.showinfo(
                "Template Ready", f"Classes template saved to:\n{file_path}"
            )
        except Exception as exc:
            messagebox.showerror(
                "Template Error", f"Could not create classes template.\n\n{exc}"
            )

    def import_classes_excel(self, on_complete=None):
        file_path = filedialog.askopenfilename(
            title="Import Classes Workbook",
            filetypes=[("Excel files", "*.xlsx *.xls")],
        )
        if not file_path:
            return

        progress_dialog = None
        try:
            workbook = pd.read_excel(file_path, sheet_name=None)
            if not isinstance(workbook, dict):
                workbook = {"Classes": workbook}

            aliases = {
                "class_name": {"class_name", "class", "name", "grade"},
                "level": {"level", "school_level"},
                "stream": {"stream", "class_stream"},
                "abbreviation": {"abbreviation", "abbr", "short_label", "short label"},
            }

            def clean(value):
                text = str(value or "").strip()
                return "" if text.lower() == "nan" else text

            prepared_rows = []
            for sheet_name, raw_df in workbook.items():
                if raw_df is None or raw_df.empty:
                    continue
                df = raw_df.copy()
                df.columns = [self._normalize_text(col) for col in df.columns]

                def find_col(alias_key):
                    return next(
                        (col for col in df.columns if col in aliases[alias_key]), None
                    )

                name_col = find_col("class_name")
                if not name_col:
                    continue
                level_col = find_col("level")
                stream_col = find_col("stream")
                abbr_col = find_col("abbreviation")

                for idx, (_, row) in enumerate(df.iterrows(), start=2):
                    class_name = clean(row.get(name_col, ""))
                    level = clean(row.get(level_col, "")) if level_col else ""
                    stream_name = clean(row.get(stream_col, "")) if stream_col else ""
                    abbreviation = clean(row.get(abbr_col, "")) if abbr_col else ""
                    if not any([class_name, level, stream_name, abbreviation]):
                        continue
                    prepared_rows.append(
                        {
                            "sheet": str(sheet_name),
                            "row": idx,
                            "class_name": class_name,
                            "level": level,
                            "stream_name": stream_name,
                            "abbreviation": abbreviation,
                        }
                    )

            if not prepared_rows:
                messagebox.showwarning(
                    "No Data",
                    "No class rows were found in this workbook.\n\nExpected a column like class_name or class.",
                )
                return

            progress_dialog, status_label, percent_label, progress = (
                self._open_progress_dialog(
                    "Importing Classes",
                    "Preparing class import...",
                    allow_cancel=True,
                )
            )

            def ensure_not_cancelled():
                if self._progress_cancel_requested(progress_dialog):
                    raise ImportCancelledError(
                        "Class import cancelled.\n\nAny classes or streams imported before cancellation were kept."
                    )

            def build_class_progress_details(
                current_sheet="",
                processed_rows=0,
                total_rows_count=0,
                classes_added=0,
                classes_updated=0,
                streams_added=0,
                skipped_rows=0,
            ):
                detail_lines = []
                if current_sheet:
                    detail_lines.append(f"Sheet: {current_sheet}")
                detail_lines.append(
                    f"Rows processed: {processed_rows}/{total_rows_count}"
                )
                detail_lines.append(
                    f"Classes added: {classes_added}   Updated: {classes_updated}"
                )
                detail_lines.append(
                    f"Streams added: {streams_added}   Skipped: {skipped_rows}"
                )
                return "\n".join(detail_lines)

            classes_added = 0
            classes_updated = 0
            streams_added = 0
            skipped_rows = 0
            runtime_skipped = []
            total_rows = len(prepared_rows)

            for index, item in enumerate(prepared_rows, start=1):
                self._update_progress_dialog(
                    progress_dialog,
                    status_label,
                    percent_label,
                    progress,
                    index - 1,
                    total_rows,
                    f"Importing class row {item['row']} from {item['sheet']}...",
                    build_class_progress_details(
                        current_sheet=item["sheet"],
                        processed_rows=index - 1,
                        total_rows_count=total_rows,
                        classes_added=classes_added,
                        classes_updated=classes_updated,
                        streams_added=streams_added,
                        skipped_rows=skipped_rows,
                    ),
                )
                ensure_not_cancelled()

                class_name = item["class_name"]
                if not class_name:
                    skipped_rows += 1
                    runtime_skipped.append(
                        f"{item['sheet']} row {item['row']}: class_name is required"
                    )
                    continue

                existing = db.get_class_by_name(class_name)
                resolved_level = (
                    item["level"]
                    or (existing.get("level", "") if existing else "")
                    or self._determine_class_level(class_name)
                    or (LEVELS[0] if LEVELS else "")
                )
                abbreviation = item["abbreviation"] or (
                    existing.get("abbreviation", "") if existing else ""
                )
                stream_name = item["stream_name"]

                if existing:
                    needs_update = resolved_level != existing.get("level", "") or (
                        abbreviation or ""
                    ) != (existing.get("abbreviation", "") or "")
                    if needs_update:
                        success, msg = db.update_class(
                            existing["id"],
                            class_name,
                            resolved_level,
                            existing.get("stream", ""),
                            abbreviation,
                        )
                        if success:
                            classes_updated += 1
                        else:
                            skipped_rows += 1
                            runtime_skipped.append(
                                f"{item['sheet']} row {item['row']}: {msg}"
                            )
                            continue
                    class_row = existing
                else:
                    success, msg = db.add_class(
                        class_name, resolved_level, None, abbreviation
                    )
                    if not success:
                        skipped_rows += 1
                        runtime_skipped.append(
                            f"{item['sheet']} row {item['row']}: {msg}"
                        )
                        continue
                    classes_added += 1
                    class_row = db.get_class_by_name(class_name)

                if stream_name and class_row:
                    existing_streams = {
                        self._normalize_key(stream.get("name", ""))
                        for stream in db.get_streams_for_class(class_row["id"])
                    }
                    if self._normalize_key(stream_name) not in existing_streams:
                        success, msg = db.add_stream(stream_name, class_row["id"])
                        if success:
                            streams_added += 1
                        else:
                            skipped_rows += 1
                            runtime_skipped.append(
                                f"{item['sheet']} row {item['row']}: {msg}"
                            )

            self._update_progress_dialog(
                progress_dialog,
                status_label,
                percent_label,
                progress,
                total_rows,
                total_rows,
                "Refreshing classes view...",
                build_class_progress_details(
                    processed_rows=total_rows,
                    total_rows_count=total_rows,
                    classes_added=classes_added,
                    classes_updated=classes_updated,
                    streams_added=streams_added,
                    skipped_rows=skipped_rows,
                ),
            )
            ensure_not_cancelled()
            refresh_dynamic_school_config()

            try:
                progress_dialog.destroy()
            except Exception:
                pass

            if callable(on_complete):
                on_complete()
            else:
                self.show_settings_classes()

            msg = (
                "Classes workbook imported successfully.\n\n"
                f"Classes added: {classes_added}\n"
                f"Classes updated: {classes_updated}\n"
                f"Streams added: {streams_added}"
            )
            if runtime_skipped:
                preview = "\n".join(runtime_skipped[:12])
                extra = (
                    ""
                    if len(runtime_skipped) <= 12
                    else f"\n...and {len(runtime_skipped) - 12} more"
                )
                msg += f"\n\nSkipped rows: {len(runtime_skipped)}\n{preview}{extra}"
            messagebox.showinfo("Import Complete", msg)
        except ImportCancelledError as exc:
            if progress_dialog is not None:
                try:
                    progress_dialog.destroy()
                except Exception:
                    pass
            messagebox.showinfo("Import Cancelled", str(exc))
        except Exception as exc:
            if progress_dialog is not None:
                try:
                    progress_dialog.destroy()
                except Exception:
                    pass
            messagebox.showerror(
                "Import Error", f"Failed to import classes workbook:\n{exc}"
            )

    def _build_subjects_tab(self, parent):
        toolbar = tk.Frame(parent, bg=CONTENT_BG)
        toolbar.pack(fill="x", pady=10)

        tk.Button(
            toolbar,
            text="+ Add Subject",
            bg=GREEN,
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: self._open_subject_dialog(),
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="📥 Template",
            bg=ORANGE,
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=self.download_subjects_import_template,
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="📥 Import Excel",
            bg=BLUE,
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: self.import_subjects_excel(
                on_complete=lambda: self._load_subjects(tree, search_var.get())
            ),
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Reset To Default List",
            bg=PURPLE,
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: self._reset_subject_catalog(tree),
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Edit Selected",
            bg=BLUE,
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: self._edit_subject_dialog(tree),
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Delete Selected",
            bg="#e74c3c",
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: self._delete_subject(tree),
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Class Subjects Done",
            bg="#0f766e",
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=self._open_class_subjects_done_dialog,
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Export Subjects",
            bg="#1d4ed8",
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: self._export_subjects_table(tree),
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Print Subjects",
            bg="#7c3aed",
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: self._print_subjects_table(tree),
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="🗑️ Delete ALL Subjects",
            bg="#c41e3a",
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=self._delete_all_subjects_batch,
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Refresh",
            bg="#666",
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: self._load_subjects(tree),
        ).pack(side="left", padx=5)

        search_row = tk.Frame(parent, bg=CONTENT_BG)
        search_row.pack(fill="x", padx=10, pady=(0, 4))
        tk.Label(
            search_row, text="Search:", bg=CONTENT_BG, fg=TEXT_SECONDARY, font=(FF, 10)
        ).pack(side="left", padx=(0, 6))
        search_var = tk.StringVar()
        search_entry = ttk.Entry(
            search_row, textvariable=search_var, style="App.TEntry", width=30
        )
        search_entry.pack(side="left", ipady=3)
        tk.Button(
            search_row,
            text="🔍 Search",
            bg=OLIVE_PRIMARY,
            fg="white",
            activebackground=OLIVE_DARK,
            activeforeground="white",
            font=(FF, 9, "bold"),
            padx=10,
            pady=4,
            relief="flat",
            cursor="hand2",
            command=lambda: self._load_subjects(tree, search_var.get()),
        ).pack(side="left", padx=(8, 0))

        list_frame = tk.Frame(parent, bg=CARD_BG, relief="flat", bd=1)
        list_frame.pack(fill="both", expand=True, padx=10, pady=5)

        cols = ("id", "name", "abbr", "level", "category", "optional", "classes_done")
        tree = ttk.Treeview(
            list_frame,
            columns=cols,
            show="headings",
            style="App.Treeview",
            selectmode="extended",
        )
        tree.heading("id", text="Code")
        tree.heading("name", text="Subject Name")
        tree.heading("abbr", text="Short Label")
        tree.heading("level", text="Level")
        tree.heading("category", text="Category")
        tree.heading("optional", text="Optional")
        tree.heading("classes_done", text="Classes Taking This Subject")

        tree.column("id", width=70, anchor="center")
        tree.column("name", width=220)
        tree.column("abbr", width=120, anchor="center")
        tree.column("level", width=180)
        tree.column("category", width=120, anchor="center")
        tree.column("optional", width=90, anchor="center")
        tree.column("classes_done", width=250)

        select_all_var = tk.BooleanVar(value=False)
        select_row = tk.Frame(list_frame, bg=CARD_BG)
        select_row.pack(fill="x", padx=10, pady=(8, 0))
        tk.Checkbutton(
            select_row,
            text="Select all subjects",
            variable=select_all_var,
            bg=CARD_BG,
            fg=TEXT_SECONDARY,
            activebackground=CARD_BG,
            activeforeground=TEXT_PRIMARY,
            selectcolor=CARD_BG,
            font=(FF, 9, "bold"),
            command=lambda: (
                self._select_all_tree_rows(tree)
                if select_all_var.get()
                else self._clear_tree_selection(tree)
            ),
        ).pack(side="left")

        tree.pack(fill="both", expand=True, padx=10, pady=10)
        tree.bind("<Double-1>", lambda e: self._edit_subject_dialog(tree))
        self._subjects_col_labels = {
            "id": "Code",
            "name": "Subject Name",
            "abbr": "Short Label",
            "level": "Level",
            "category": "Category",
            "optional": "Optional",
            "classes_done": "Classes Taking This Subject",
        }
        self._subjects_sort_state = {"column": None, "reverse": False}
        self._configure_subjects_sorting(tree)
        search_entry.bind(
            "<Return>", lambda _e: self._load_subjects(tree, search_var.get())
        )
        search_var.trace_add(
            "write", lambda *_: self._load_subjects(tree, search_var.get())
        )
        self._load_subjects(tree, search_var.get())

    def _configure_subjects_sorting(self, tree):
        for col in tree["columns"]:
            label = self._subjects_col_labels.get(col, col.title())
            tree.heading(
                col, text=label, command=lambda c=col: self._sort_subjects_tree(tree, c)
            )

    def _subject_sort_value(self, col, value):
        text = str(value or "").strip()
        if col == "optional":
            return 1 if text.lower() in ("yes", "true", "1") else 0
        return text.lower()

    def _sort_subjects_tree(self, tree, col):
        state = getattr(
            self, "_subjects_sort_state", {"column": None, "reverse": False}
        )
        reverse = state.get("column") == col and not state.get("reverse", False)
        rows = []
        for iid in tree.get_children(""):
            values = tree.item(iid).get("values", ())
            rows.append((iid, values))

        try:
            idx = list(tree["columns"]).index(col)
        except ValueError:
            return

        rows.sort(
            key=lambda item: self._subject_sort_value(
                col, item[1][idx] if idx < len(item[1]) else ""
            ),
            reverse=reverse,
        )
        for position, (iid, _vals) in enumerate(rows):
            tree.move(iid, "", position)

        self._subjects_sort_state = {"column": col, "reverse": reverse}
        for c in tree["columns"]:
            base = self._subjects_col_labels.get(c, c.title())
            if c == col:
                base = f"{base} {'▼' if reverse else '▲'}"
            tree.heading(
                c, text=base, command=lambda key=c: self._sort_subjects_tree(tree, key)
            )

    def _load_subjects(self, tree, search_query=""):
        for item in tree.get_children():
            tree.delete(item)

        # Get class subjects done mapping
        class_subjects_map = self._get_class_subjects_done_map()
        query = str(search_query or "").strip().lower()

        for subj in db.get_subjects_by_level():
            if _is_legacy_subject_level(subj.get("level", "")):
                continue
            subject_name = subj.get("name", "")

            # Find which classes take this subject
            classes_taking = []
            for class_name, subjects in class_subjects_map.items():
                if subject_name in subjects:
                    classes_taking.append(class_name)

            # If no explicit mapping, check if subject is in the level's default subjects
            if not classes_taking:
                subject_level = subj.get("level", "")
                if subject_level in CLASSES_BY_LEVEL:
                    # Check if this subject is in the default subjects for this level
                    level_subjects = self._get_subjects_for_level(subject_level)
                    if subject_name in level_subjects:
                        classes_taking = CLASSES_BY_LEVEL[subject_level]

            classes_text = (
                ", ".join(classes_taking) if classes_taking else "Not assigned"
            )

            if query:
                haystack = " ".join(
                    [
                        str(subj.get("code", "") or subj.get("abbreviation", "")),
                        str(subj.get("name", "")),
                        str(subj.get("abbreviation", "") or subj.get("code", "")),
                        str(subj.get("level", "")),
                        str(subj.get("category", "")),
                        classes_text,
                    ]
                ).lower()
                if query not in haystack:
                    continue

            tree.insert(
                "",
                "end",
                iid=subj.get("id", ""),
                values=(
                    subj.get("code", "") or subj.get("abbreviation", ""),
                    subj.get("name", ""),
                    subj.get("abbreviation", "")
                    or subj.get("code", "")
                    or self._generate_short_label(subj.get("name", ""), "subject"),
                    subj.get("level", ""),
                    subj.get("category", ""),
                    "Yes" if subj.get("is_optional") else "No",
                    classes_text,
                ),
            )

    def _export_subjects_table(self, tree):
        file_path = filedialog.asksaveasfilename(
            title="Export Subjects",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            initialfile="subjects_list.csv",
        )
        if not file_path:
            return

        headers = [
            self._subjects_col_labels.get(col, col.title()) for col in tree["columns"]
        ]
        try:
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                for iid in tree.get_children(""):
                    writer.writerow(tree.item(iid).get("values", ()))
            messagebox.showinfo(
                "Export Complete", f"Subjects exported to:\n{file_path}"
            )
        except Exception as exc:
            messagebox.showerror(
                "Export Failed", f"Could not export subjects.\n\n{exc}"
            )

    def download_subjects_import_template(self):
        file_path = filedialog.asksaveasfilename(
            title="Save Subjects Template",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="subjects_import_template.xlsx",
        )
        if not file_path:
            return

        try:
            rows = db.get_subjects_by_level()
            template_rows = []
            for row in rows:
                if _is_legacy_subject_level(row.get("level", "")):
                    continue
                template_rows.append(
                    {
                        "code": row.get("code", "")
                        or row.get("abbreviation", "")
                        or "",
                        "name": row.get("name", ""),
                        "abbreviation": row.get("abbreviation", "") or "",
                        "level": row.get("level", ""),
                        "category": row.get("category", "") or "Core",
                        "optional": "Yes" if row.get("is_optional") else "No",
                    }
                )
            if not template_rows:
                template_rows = [
                    {
                        "code": "MAT",
                        "name": "Mathematics",
                        "abbreviation": "MAT",
                        "level": LEVELS[1] if len(LEVELS) > 1 else LEVELS[0],
                        "category": "Core",
                        "optional": "No",
                    }
                ]

            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                pd.DataFrame(
                    {
                        "Instructions": [
                            "Use one row per subject.",
                            "code and name are recommended. If abbreviation is blank, code will be reused.",
                            "optional accepts values like Yes, No, True, False, 1, or 0.",
                            "If a subject with the same name and level already exists, the importer updates it.",
                        ]
                    }
                ).to_excel(writer, sheet_name="Instructions", index=False)
                pd.DataFrame(template_rows).to_excel(
                    writer, sheet_name="Subjects", index=False
                )

            messagebox.showinfo(
                "Template Ready", f"Subjects template saved to:\n{file_path}"
            )
        except Exception as exc:
            messagebox.showerror(
                "Template Error", f"Could not create subjects template.\n\n{exc}"
            )

    def import_subjects_excel(self, on_complete=None):
        file_path = filedialog.askopenfilename(
            title="Import Subjects Workbook",
            filetypes=[("Excel files", "*.xlsx *.xls")],
        )
        if not file_path:
            return

        progress_dialog = None
        try:
            workbook = pd.read_excel(file_path, sheet_name=None)
            if not isinstance(workbook, dict):
                workbook = {"Subjects": workbook}

            aliases = {
                "code": {"code", "subject_code", "subject code"},
                "name": {"name", "subject", "subject_name", "subject name"},
                "abbreviation": {"abbreviation", "abbr", "short_label", "short label"},
                "level": {"level", "subject_level"},
                "category": {"category", "type"},
                "optional": {"optional", "is_optional", "is optional"},
            }

            def clean(value):
                text = str(value or "").strip()
                return "" if text.lower() == "nan" else text

            def parse_bool(value):
                normalized = clean(value).lower()
                return normalized in {"1", "true", "yes", "y", "optional"}

            prepared_rows = []
            for sheet_name, raw_df in workbook.items():
                if raw_df is None or raw_df.empty:
                    continue
                df = raw_df.copy()
                df.columns = [self._normalize_text(col) for col in df.columns]

                def find_col(alias_key):
                    return next(
                        (col for col in df.columns if col in aliases[alias_key]), None
                    )

                name_col = find_col("name")
                if not name_col:
                    continue
                code_col = find_col("code")
                abbr_col = find_col("abbreviation")
                level_col = find_col("level")
                category_col = find_col("category")
                optional_col = find_col("optional")

                for idx, (_, row) in enumerate(df.iterrows(), start=2):
                    name = clean(row.get(name_col, ""))
                    code = clean(row.get(code_col, "")) if code_col else ""
                    abbreviation = clean(row.get(abbr_col, "")) if abbr_col else ""
                    level = clean(row.get(level_col, "")) if level_col else ""
                    category = clean(row.get(category_col, "")) if category_col else ""
                    optional = (
                        parse_bool(row.get(optional_col, "")) if optional_col else False
                    )
                    if not any([name, code, abbreviation, level, category]):
                        continue
                    prepared_rows.append(
                        {
                            "sheet": str(sheet_name),
                            "row": idx,
                            "name": name,
                            "code": code.upper(),
                            "abbreviation": abbreviation.upper(),
                            "level": level,
                            "category": category,
                            "is_optional": optional,
                        }
                    )

            if not prepared_rows:
                messagebox.showwarning(
                    "No Data",
                    "No subject rows were found in this workbook.\n\nExpected a column like name or subject_name.",
                )
                return

            progress_dialog, status_label, percent_label, progress = (
                self._open_progress_dialog(
                    "Importing Subjects",
                    "Preparing subject import...",
                    allow_cancel=True,
                )
            )

            def ensure_not_cancelled():
                if self._progress_cancel_requested(progress_dialog):
                    raise ImportCancelledError(
                        "Subject import cancelled.\n\nAny subjects imported before cancellation were kept."
                    )

            def build_subject_progress_details(
                current_sheet="",
                processed_rows=0,
                total_rows_count=0,
                added=0,
                updated=0,
                skipped=0,
            ):
                detail_lines = []
                if current_sheet:
                    detail_lines.append(f"Sheet: {current_sheet}")
                detail_lines.append(
                    f"Rows processed: {processed_rows}/{total_rows_count}"
                )
                detail_lines.append(f"Subjects added: {added}   Updated: {updated}")
                detail_lines.append(f"Skipped: {skipped}")
                return "\n".join(detail_lines)

            added = 0
            updated = 0
            skipped = 0
            runtime_skipped = []
            total_rows = len(prepared_rows)

            for index, item in enumerate(prepared_rows, start=1):
                self._update_progress_dialog(
                    progress_dialog,
                    status_label,
                    percent_label,
                    progress,
                    index - 1,
                    total_rows,
                    f"Importing subject row {item['row']} from {item['sheet']}...",
                    build_subject_progress_details(
                        current_sheet=item["sheet"],
                        processed_rows=index - 1,
                        total_rows_count=total_rows,
                        added=added,
                        updated=updated,
                        skipped=skipped,
                    ),
                )
                ensure_not_cancelled()

                name = item["name"]
                if not name:
                    skipped += 1
                    runtime_skipped.append(
                        f"{item['sheet']} row {item['row']}: subject name is required"
                    )
                    continue

                level = item["level"] or (
                    self.current_level if self.current_level in LEVELS else LEVELS[0]
                )
                level = _canonicalize_subject_level(
                    level,
                    self.current_level if self.current_level in LEVELS else LEVELS[0],
                )
                category = item["category"] or "Core"
                code = (
                    item["code"]
                    or item["abbreviation"]
                    or self._generate_short_label(name, "subject")
                )
                abbreviation = item["abbreviation"] or code
                existing = db.get_subject_by_name(name, level)

                if existing:
                    success, msg = db.update_subject(
                        existing["id"],
                        name,
                        level,
                        category,
                        item["is_optional"],
                        abbreviation,
                        code,
                    )
                    if success:
                        updated += 1
                    else:
                        skipped += 1
                        runtime_skipped.append(
                            f"{item['sheet']} row {item['row']}: {msg}"
                        )
                else:
                    success, msg = db.add_subject(
                        name,
                        level,
                        category,
                        item["is_optional"],
                        abbreviation,
                        code,
                    )
                    if success:
                        added += 1
                    else:
                        skipped += 1
                        runtime_skipped.append(
                            f"{item['sheet']} row {item['row']}: {msg}"
                        )

            self._update_progress_dialog(
                progress_dialog,
                status_label,
                percent_label,
                progress,
                total_rows,
                total_rows,
                "Refreshing subjects view...",
                build_subject_progress_details(
                    processed_rows=total_rows,
                    total_rows_count=total_rows,
                    added=added,
                    updated=updated,
                    skipped=skipped,
                ),
            )
            ensure_not_cancelled()

            try:
                progress_dialog.destroy()
            except Exception:
                pass

            if callable(on_complete):
                on_complete()
            else:
                self.show_settings_subjects()

            msg = (
                "Subjects workbook imported successfully.\n\n"
                f"Subjects added: {added}\n"
                f"Subjects updated: {updated}"
            )
            if runtime_skipped:
                preview = "\n".join(runtime_skipped[:12])
                extra = (
                    ""
                    if len(runtime_skipped) <= 12
                    else f"\n...and {len(runtime_skipped) - 12} more"
                )
                msg += f"\n\nSkipped rows: {len(runtime_skipped)}\n{preview}{extra}"
            messagebox.showinfo("Import Complete", msg)
        except ImportCancelledError as exc:
            if progress_dialog is not None:
                try:
                    progress_dialog.destroy()
                except Exception:
                    pass
            messagebox.showinfo("Import Cancelled", str(exc))
        except Exception as exc:
            if progress_dialog is not None:
                try:
                    progress_dialog.destroy()
                except Exception:
                    pass
            messagebox.showerror(
                "Import Error", f"Failed to import subjects workbook:\n{exc}"
            )

    def _print_subjects_table(self, tree):
        headers = [
            self._subjects_col_labels.get(col, col.title()) for col in tree["columns"]
        ]
        rows = [tree.item(iid).get("values", ()) for iid in tree.get_children("")]
        if not rows:
            messagebox.showwarning("No Data", "No subject rows available to print.")
            return

        col_widths = [len(h) for h in headers]
        for row in rows:
            for i, value in enumerate(row):
                col_widths[i] = max(col_widths[i], len(str(value)))

        def fmt_row(row_values):
            return " | ".join(
                str(v).ljust(col_widths[i]) for i, v in enumerate(row_values)
            )

        divider = "-+-".join("-" * w for w in col_widths)
        lines = [
            f"{get_school_profile().get('school_name', DEFAULT_SCHOOL_PROFILE['school_name'])} - SUBJECTS LIST",
            "",
            fmt_row(headers),
            divider,
        ]
        for row in rows:
            lines.append(fmt_row(row))

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        print_path = os.path.join(
            tempfile.gettempdir(), f"subjects_print_{timestamp}.txt"
        )
        try:
            with open(print_path, "w", encoding="utf-8") as f:
                f.write("\n".join(lines))
            try:
                os.startfile(print_path, "print")
                messagebox.showinfo("Print Sent", "Subjects list sent to printer.")
            except Exception:
                messagebox.showinfo(
                    "Print File Ready", f"Print file created:\n{print_path}"
                )
        except Exception as exc:
            messagebox.showerror(
                "Print Failed", f"Could not prepare print file.\n\n{exc}"
            )

    def _get_class_subjects_done_map(self):
        raw = db.get_setting(CLASS_SUBJECTS_DONE_KEY, "{}")
        try:
            data = json.loads(raw or "{}")
        except Exception:
            data = {}
        cleaned = {}
        if isinstance(data, dict):
            for class_name, subjects in data.items():
                class_key = str(class_name or "").strip()
                if not class_key:
                    continue
                if isinstance(subjects, list):
                    cleaned[class_key] = [
                        str(s).strip() for s in subjects if str(s).strip()
                    ]
        return cleaned

    def _save_class_subjects_done_map(self, mapping):
        payload = json.dumps(mapping or {}, ensure_ascii=True)
        db.set_setting(CLASS_SUBJECTS_DONE_KEY, payload)

    def _get_subject_pool_for_class(self, class_name):
        level = self._get_level_for_class(class_name)
        subjects = self._get_subjects_for_level(level)
        level_subjects = [
            row
            for row in db.get_subjects_by_level(level)
            if not _is_legacy_subject_level(row.get("level", ""))
        ]
        global_subjects = [
            row
            for row in db.get_subjects_by_level(ALL_SUBJECT_LEVEL)
            if not _is_legacy_subject_level(row.get("level", ""))
        ]
        combined_subjects = list(level_subjects or [])
        for row in global_subjects or []:
            name = row.get("name", "")
            if name and not any(
                existing.get("name") == name for existing in combined_subjects
            ):
                combined_subjects.append(row)

        if combined_subjects:
            custom_names = [
                row.get("name", "") for row in combined_subjects if row.get("name")
            ]
            ordered = [subject for subject in subjects if subject in custom_names]
            for subject in custom_names:
                if subject not in ordered:
                    ordered.append(subject)
            subjects = ordered
        return subjects

    def _get_done_subjects_from_marks(
        self, class_name, term="One", exam_type=DEFAULT_EXAM_TYPE, academic_year=None
    ):
        academic_year = str(academic_year or datetime.now().year)
        present = set()
        for student in db.get_students_by_class(class_name):
            for subject in db.get_student_marks(
                student["id"], term, exam_type, academic_year
            ).keys():
                present.add(subject)

        if not present:
            return []

        ordered = []
        remaining = sorted(present)
        for subject in self._get_subject_pool_for_class(class_name):
            matched = self._match_subject_from_candidates(
                subject, remaining, class_name
            )
            if matched and matched not in ordered:
                ordered.append(matched)
        return ordered

    def _open_class_subjects_done_dialog(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("Class Subjects Done")
        dialog.geometry("860x620")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)

        tk.Label(
            dialog,
            text="Class-specific subjects done for reports and report cards",
            font=(FF, 11, "bold"),
        ).pack(anchor="w", padx=20, pady=(16, 4))
        tk.Label(
            dialog,
            text="Only selected subjects will appear in Results/Report Cards for this class.",
            font=(FF, 9),
            fg=TEXT_SECONDARY,
        ).pack(anchor="w", padx=20, pady=(0, 10))

        legend = tk.Frame(dialog, bg=dialog.cget("bg"))
        legend.pack(fill="x", padx=20, pady=(0, 10))

        def legend_chip(parent, color, text, fg="white"):
            chip = tk.Frame(parent, bg=color, padx=10, pady=4)
            chip.pack(side="left", padx=(0, 8))
            tk.Label(chip, text=text, bg=color, fg=fg, font=(FF, 9, "bold")).pack()

        legend_chip(legend, "#16a34a", "Add Subject")
        legend_chip(legend, "#f59e0b", "Edit Subject", fg="#1f2937")
        legend_chip(legend, "#dc2626", "Remove Subject")
        legend_chip(legend, "#2563eb", "Auto From Marks")
        legend_chip(legend, "#7c3aed", "Automatic Mode")

        classes = [row.get("name") for row in db.get_all_classes() if row.get("name")]
        if not classes:
            classes = self.get_current_classes()

        top = tk.Frame(dialog, bg=dialog.cget("bg"))
        top.pack(fill="x", padx=20, pady=(0, 8))

        tk.Label(top, text="Class:", font=(FF, 10, "bold")).pack(side="left")
        class_var = tk.StringVar(value=classes[0] if classes else "")
        class_cb = ttk.Combobox(
            top, textvariable=class_var, values=classes, state="readonly", width=24
        )
        class_cb.pack(side="left", padx=(8, 14))

        tk.Label(top, text="Term:", font=(FF, 10, "bold")).pack(side="left")
        term_var = tk.StringVar(value=TERMS[0])
        term_cb = ttk.Combobox(
            top, textvariable=term_var, values=TERMS, state="readonly", width=8
        )
        term_cb.pack(side="left", padx=(8, 8))

        tk.Label(top, text="Exam:", font=(FF, 10, "bold")).pack(side="left")
        exam_var = tk.StringVar(value=DEFAULT_EXAM_TYPE)
        exam_cb = ttk.Combobox(
            top, textvariable=exam_var, values=EXAM_TYPES, state="readonly", width=12
        )
        exam_cb.pack(side="left", padx=(8, 0))

        list_wrap = tk.Frame(dialog, bg=CARD_BG, relief="flat", bd=1)
        list_wrap.pack(fill="both", expand=True, padx=20, pady=(4, 8))

        left_pane = tk.Frame(list_wrap, bg=CARD_BG)
        left_pane.pack(side="left", fill="both", expand=True, padx=(10, 6), pady=10)
        tk.Label(
            left_pane,
            text="Available Subjects",
            bg=CARD_BG,
            fg=TEXT_PRIMARY,
            font=(FF, 10, "bold"),
        ).pack(anchor="w", pady=(0, 6))
        available_list = tk.Listbox(
            left_pane, selectmode="extended", font=(FF, 10), activestyle="none"
        )
        available_list.pack(side="left", fill="both", expand=True)
        available_sb = ttk.Scrollbar(
            left_pane,
            orient="vertical",
            command=available_list.yview,
            style="App.Vertical.TScrollbar",
        )
        available_sb.pack(side="right", fill="y")
        available_list.configure(yscrollcommand=available_sb.set)

        middle_pane = tk.Frame(list_wrap, bg=CARD_BG)
        middle_pane.pack(side="left", fill="y", padx=8, pady=10)

        right_pane = tk.Frame(list_wrap, bg=CARD_BG)
        right_pane.pack(side="left", fill="both", expand=True, padx=(6, 10), pady=10)
        tk.Label(
            right_pane,
            text="Subjects Done (Used in Reports)",
            bg=CARD_BG,
            fg=TEXT_PRIMARY,
            font=(FF, 10, "bold"),
        ).pack(anchor="w", pady=(0, 6))
        selected_list = tk.Listbox(
            right_pane, selectmode="extended", font=(FF, 10), activestyle="none"
        )
        selected_list.pack(side="left", fill="both", expand=True)
        selected_sb = ttk.Scrollbar(
            right_pane,
            orient="vertical",
            command=selected_list.yview,
            style="App.Vertical.TScrollbar",
        )
        selected_sb.pack(side="right", fill="y")
        selected_list.configure(yscrollcommand=selected_sb.set)

        all_subjects = []
        selected_subjects = []

        def subject_line(subject, cls_name):
            label = self._get_subject_label(subject, cls_name).replace("\n", " / ")
            return f"{label}  -  {subject}"

        def render_lists():
            cls_name = class_var.get().strip()
            available_list.delete(0, tk.END)
            selected_list.delete(0, tk.END)
            for subject in all_subjects:
                if subject not in selected_subjects:
                    available_list.insert(tk.END, subject_line(subject, cls_name))
            for subject in selected_subjects:
                selected_list.insert(tk.END, subject_line(subject, cls_name))

        def reload_subject_list(use_auto=False):
            nonlocal all_subjects, selected_subjects
            cls = class_var.get().strip()
            if not cls:
                return
            pool = self._get_subject_pool_for_class(cls)
            marks_subjects = self._get_done_subjects_from_marks(
                cls, term_var.get(), exam_var.get()
            )
            mapping = self._get_class_subjects_done_map()
            configured = mapping.get(cls, [])
            all_subjects = list(pool)
            for subject in marks_subjects + configured:
                if subject not in all_subjects:
                    all_subjects.append(subject)

            selected = list(configured)
            if use_auto:
                selected = list(marks_subjects)
            elif not configured:
                selected = list(marks_subjects or pool)

            selected_subjects = [
                subject for subject in selected if subject in all_subjects
            ] + [subject for subject in selected if subject not in all_subjects]
            render_lists()

        class_cb.bind("<<ComboboxSelected>>", lambda _e: reload_subject_list())
        term_cb.bind("<<ComboboxSelected>>", lambda _e: reload_subject_list())
        exam_cb.bind("<<ComboboxSelected>>", lambda _e: reload_subject_list())

        def button(master, text, bg, fg="white", cmd=None):
            tk.Button(
                master,
                text=text,
                bg=bg,
                fg=fg,
                activebackground=bg,
                activeforeground=fg,
                font=(FF, 10, "bold"),
                padx=10,
                pady=8,
                relief="flat",
                cursor="hand2",
                command=cmd,
            ).pack(fill="x", pady=5)

        def add_subjects():
            chosen = [available_list.get(i) for i in available_list.curselection()]
            if not chosen:
                messagebox.showwarning(
                    "Select Subject",
                    "Select one or more available subjects to add.",
                    parent=dialog,
                )
                return
            cls = class_var.get().strip()
            for line in chosen:
                raw = line.split(" - ", 1)[1] if " - " in line else line
                raw = raw.strip()
                if raw and raw not in selected_subjects:
                    selected_subjects.append(raw)
                    if raw not in all_subjects:
                        all_subjects.append(raw)
            render_lists()

        def remove_subjects():
            chosen_idx = list(selected_list.curselection())
            if not chosen_idx:
                messagebox.showwarning(
                    "Select Subject",
                    "Select one or more subjects done to remove.",
                    parent=dialog,
                )
                return
            chosen_values = [selected_list.get(i) for i in chosen_idx]
            remove_raw = []
            for line in chosen_values:
                raw = line.split(" - ", 1)[1] if " - " in line else line
                remove_raw.append(raw.strip())
            selected_subjects[:] = [s for s in selected_subjects if s not in remove_raw]
            render_lists()

        def edit_subject():
            chosen_idx = list(selected_list.curselection())
            if len(chosen_idx) != 1:
                messagebox.showwarning(
                    "Select One",
                    "Select exactly one subject done to edit.",
                    parent=dialog,
                )
                return
            current_line = selected_list.get(chosen_idx[0])
            current_raw = (
                current_line.split(" - ", 1)[1].strip()
                if " - " in current_line
                else current_line.strip()
            )
            new_raw = simpledialog.askstring(
                "Edit Subject",
                "Update subject name (exact subject key used in reports):",
                initialvalue=current_raw,
                parent=dialog,
            )
            if new_raw is None:
                return
            new_raw = new_raw.strip()
            if not new_raw:
                messagebox.showwarning(
                    "Invalid Value", "Subject name cannot be empty.", parent=dialog
                )
                return
            selected_subjects[chosen_idx[0]] = new_raw
            if new_raw not in all_subjects:
                all_subjects.append(new_raw)
            render_lists()

        button(middle_pane, "Add  >>", "#16a34a", "white", add_subjects)
        button(middle_pane, "Edit", "#f59e0b", "#1f2937", edit_subject)
        button(middle_pane, "<<  Remove", "#dc2626", "white", remove_subjects)

        actions = tk.Frame(dialog, bg=dialog.cget("bg"))
        actions.pack(fill="x", padx=20, pady=(0, 12))

        def auto_pick_from_marks():
            reload_subject_list(use_auto=True)

        def clear_class_override():
            cls = class_var.get().strip()
            if not cls:
                return
            mapping = self._get_class_subjects_done_map()
            if cls in mapping:
                mapping.pop(cls, None)
                self._save_class_subjects_done_map(mapping)
            messagebox.showinfo(
                "Updated",
                f"{cls} now uses automatic subjects (from entered marks).",
                parent=dialog,
            )
            reload_subject_list()

        def save_selection():
            cls = class_var.get().strip()
            if not cls:
                messagebox.showwarning(
                    "Missing Class", "Please select a class.", parent=dialog
                )
                return
            mapping = self._get_class_subjects_done_map()
            mapping[cls] = list(selected_subjects)
            self._save_class_subjects_done_map(mapping)
            messagebox.showinfo(
                "Saved",
                f"Subjects done for {cls} saved ({len(selected_subjects)}).",
                parent=dialog,
            )

        tk.Button(
            actions,
            text="Auto From Marks",
            bg="#2563eb",
            fg="white",
            activebackground="#1d4ed8",
            activeforeground="white",
            font=(FF, 10, "bold"),
            padx=12,
            pady=8,
            command=auto_pick_from_marks,
        ).pack(side="left", padx=(0, 8))
        tk.Button(
            actions,
            text="Use Automatic Mode",
            bg="#7c3aed",
            fg="white",
            activebackground="#6d28d9",
            activeforeground="white",
            font=(FF, 10, "bold"),
            padx=12,
            pady=8,
            command=clear_class_override,
        ).pack(side="left", padx=(0, 8))
        tk.Button(
            actions,
            text="Save Subjects Done",
            bg="#0f766e",
            fg="white",
            activebackground="#0e675f",
            activeforeground="white",
            font=(FF, 10, "bold"),
            padx=12,
            pady=8,
            command=save_selection,
        ).pack(side="left")

        tk.Button(
            dialog,
            text="Close",
            bg="#cbd5e1",
            fg="#0f172a",
            activebackground="#94a3b8",
            activeforeground="#0f172a",
            font=(FF, 10, "bold"),
            padx=22,
            pady=8,
            command=dialog.destroy,
        ).pack(pady=(0, 16))

        reload_subject_list()

    def _open_subject_dialog(self, subject_row=None):
        dialog = tk.Toplevel(self.root)
        dialog.title("Edit Subject" if subject_row else "Add Subject")
        dialog.geometry("430x500")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)

        tk.Label(dialog, text="Subject Code:", font=(FF, 11)).pack(pady=(20, 5))
        code_entry = tk.Entry(dialog, font=(FF, 11))
        code_entry.pack(fill="x", padx=20)

        tk.Label(dialog, text="Subject Name:", font=(FF, 11)).pack(pady=(20, 5))
        name_entry = tk.Entry(dialog, font=(FF, 11))
        name_entry.pack(fill="x", padx=20)

        tk.Label(dialog, text="Short Label / Abbreviation:", font=(FF, 11)).pack(
            pady=(15, 5)
        )
        abbr_entry = tk.Entry(dialog, font=(FF, 11))
        abbr_entry.pack(fill="x", padx=20)

        tk.Label(dialog, text="Level:", font=(FF, 11)).pack(pady=(15, 5))
        level_var = tk.StringVar()
        level_options = [ALL_SUBJECT_LEVEL] + list(LEVELS)
        level_cb = ttk.Combobox(
            dialog,
            textvariable=level_var,
            values=level_options,
            state="readonly",
            font=(FF, 10),
        )
        level_cb.pack(fill="x", padx=20)

        tk.Label(dialog, text="Category:", font=(FF, 11)).pack(pady=(15, 5))
        category_entry = tk.Entry(dialog, font=(FF, 11))
        category_entry.pack(fill="x", padx=20)

        optional_var = tk.BooleanVar(value=False)
        tk.Checkbutton(
            dialog,
            text="Optional Subject",
            variable=optional_var,
            font=(FF, 11),
            bg=dialog.cget("bg"),
        ).pack(pady=(12, 16))

        if subject_row:
            code_entry.insert(
                0, subject_row.get("code", "") or subject_row.get("abbreviation", "")
            )
            name_entry.insert(0, subject_row.get("name", ""))
            abbr_entry.insert(0, subject_row.get("abbreviation", ""))
            level_var.set(subject_row.get("level", ""))
            category_entry.insert(0, subject_row.get("category", ""))
            optional_var.set(bool(subject_row.get("is_optional")))
        else:
            level_var.set(
                self.current_level if self.current_level in LEVELS else LEVELS[0]
            )
            category_entry.insert(0, "Core")

        def save():
            code = code_entry.get().strip().upper()
            name = name_entry.get().strip()
            abbreviation = code or abbr_entry.get().strip().upper()
            level = _canonicalize_subject_level(
                level_var.get().strip(),
                self.current_level if self.current_level in LEVELS else LEVELS[0],
            )
            category = category_entry.get().strip()

            if not code or not name or not level or not category:
                messagebox.showerror(
                    "Error", "Subject code, name, level, and category are required"
                )
                return

            abbr_entry.delete(0, tk.END)
            abbr_entry.insert(0, abbreviation)
            if subject_row:
                success, msg = db.update_subject(
                    subject_row["id"],
                    name,
                    level,
                    category,
                    optional_var.get(),
                    abbreviation,
                    code,
                )
            else:
                success, msg = db.add_subject(
                    name, level, category, optional_var.get(), abbreviation, code
                )
            if success:
                refresh_dynamic_school_config()
                dialog.destroy()
                self.show_settings()
            else:
                messagebox.showerror("Error", msg)

        tk.Button(
            dialog,
            text="Save",
            bg=GREEN,
            fg="white",
            font=(FF, 11),
            padx=20,
            pady=8,
            command=save,
        ).pack(pady=(4, 20))

    def _edit_subject_dialog(self, tree):
        selected = tree.selection()
        if len(selected) != 1:
            messagebox.showwarning(
                "Select One", "Please select exactly one subject to edit"
            )
            return
        subject_id = selected[0]
        subject_row = next(
            (row for row in db.get_subjects_by_level() if row.get("id") == subject_id),
            None,
        )
        if not subject_row:
            messagebox.showerror("Error", "Could not load the selected subject")
            return
        self._open_subject_dialog(subject_row)

    def _delete_subject(self, tree):
        selected = list(tree.selection())
        if not selected:
            self._show_notice(
                "Select Subjects",
                "Please select one or more subjects to delete.",
                kind="info",
            )
            return
        if not self._confirm_delete_action(
            "subject",
            len(selected),
            scope="selected",
            details="This will also remove marks and teacher assignments linked to those subjects.",
        ):
            return
        failures = 0
        for subject_id in selected:
            if not db.delete_subject(subject_id):
                failures += 1
        refresh_dynamic_school_config()
        self._load_subjects(tree)
        self._show_delete_result_notice(
            "subject", len(selected) - failures, failures, duration_ms=4200
        )

    def _reset_subject_catalog(self, tree):
        if not messagebox.askyesno(
            "Confirm Reset",
            "Delete all current subjects and replace them with the new default subject list?",
        ):
            return
        self._replace_subject_catalog_with_defaults()
        refresh_dynamic_school_config()
        self._load_subjects(tree)
        messagebox.showinfo(
            "Subjects Updated",
            "All subject records were replaced with the new coded subject catalog.",
        )

    def _build_teachers_settings_tab(self, parent):
        toolbar = tk.Frame(parent, bg=CONTENT_BG)
        toolbar.pack(fill="x", pady=10)

        tk.Button(
            toolbar,
            text="+ Add Teacher",
            bg=GREEN,
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: self._open_teacher_dialog(),
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Edit Selected",
            bg=BLUE,
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: self._edit_teacher_dialog(tree),
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Delete Selected",
            bg="#e74c3c",
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: self._delete_teacher(tree, reload_callback=refresh_all),
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Template",
            bg=ORANGE,
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=self.download_teacher_import_template,
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Export Excel",
            bg=GREEN,
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=self.export_teachers_excel,
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Import Excel",
            bg=BLUE,
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: self.import_teachers_excel(on_complete=refresh_all),
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Assign Subject",
            bg="#1d4ed8",
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: self._assign_subject_dialog(on_saved=refresh_all),
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Assign Grade Facilitator",
            bg=PURPLE,
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: self._assign_class_teacher_dialog(on_saved=refresh_all),
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Refresh",
            bg="#666",
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: refresh_all(),
        ).pack(side="left", padx=5)

        search_row = tk.Frame(parent, bg=CONTENT_BG)
        search_row.pack(fill="x", padx=10, pady=(0, 4))
        tk.Label(
            search_row, text="Search:", bg=CONTENT_BG, fg=TEXT_SECONDARY, font=(FF, 10)
        ).pack(side="left", padx=(0, 6))
        search_var = tk.StringVar()
        search_entry = ttk.Entry(
            search_row, textvariable=search_var, style="App.TEntry", width=30
        )
        search_entry.pack(side="left", ipady=3)
        tk.Button(
            search_row,
            text="🔍 Search",
            bg=OLIVE_PRIMARY,
            fg="white",
            activebackground=OLIVE_DARK,
            activeforeground="white",
            font=(FF, 9, "bold"),
            padx=10,
            pady=4,
            relief="flat",
            cursor="hand2",
            command=lambda: refresh_all(),
        ).pack(side="left", padx=(8, 0))

        list_frame = tk.Frame(parent, bg=CARD_BG, relief="flat", bd=1)
        list_frame.pack(fill="both", expand=True, padx=10, pady=(5, 8))

        cols = ("id", "abbr", "name", "username", "role")
        tree = ttk.Treeview(
            list_frame,
            columns=cols,
            show="headings",
            style="App.Treeview",
            selectmode="extended",
        )
        tree.heading("id", text="ID")
        tree.heading("abbr", text="Short Label")
        tree.heading("name", text="Full Name")
        tree.heading("username", text="Username")
        tree.heading("role", text="Role")

        tree.column("id", width=70, anchor="center")
        tree.column("abbr", width=100, anchor="center")
        tree.column("name", width=220)
        tree.column("username", width=150)
        tree.column("role", width=140, anchor="center")

        select_all_var = tk.BooleanVar(value=False)
        select_row = tk.Frame(list_frame, bg=CARD_BG)
        select_row.pack(fill="x", padx=10, pady=(8, 0))
        tk.Checkbutton(
            select_row,
            text="Select all teachers",
            variable=select_all_var,
            bg=CARD_BG,
            fg=TEXT_SECONDARY,
            activebackground=CARD_BG,
            activeforeground=TEXT_PRIMARY,
            selectcolor=CARD_BG,
            font=(FF, 9, "bold"),
            command=lambda: (
                self._select_all_tree_rows(tree)
                if select_all_var.get()
                else self._clear_tree_selection(tree)
            ),
        ).pack(side="left")

        tree.pack(fill="both", expand=True, padx=10, pady=10)
        tree.bind("<Double-1>", lambda e: self._edit_teacher_dialog(tree))

        assignments_wrap = tk.Frame(parent, bg=CONTENT_BG)
        assignments_wrap.pack(fill="both", expand=True, padx=10, pady=(0, 8))
        assignments_wrap.columnconfigure(0, weight=1)
        assignments_wrap.columnconfigure(1, weight=1)
        assignments_wrap.rowconfigure(0, weight=1)

        subject_card = tk.Frame(assignments_wrap, bg=CARD_BG, relief="flat", bd=1)
        subject_card.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        tk.Label(
            subject_card,
            text="Subject Teacher Assignments",
            bg=CARD_BG,
            fg=TEXT_PRIMARY,
            font=(FF, 11, "bold"),
        ).pack(anchor="w", padx=10, pady=(8, 4))
        subj_cols = ("teacher", "subject", "class", "stream")
        subj_tree = ttk.Treeview(
            subject_card,
            columns=subj_cols,
            show="headings",
            style="App.Treeview",
            height=8,
        )
        subj_tree.heading("teacher", text="Teacher")
        subj_tree.heading("subject", text="Subject")
        subj_tree.heading("class", text="Class")
        subj_tree.heading("stream", text="Stream")
        subj_tree.column("teacher", width=180, anchor="w")
        subj_tree.column("subject", width=170, anchor="w")
        subj_tree.column("class", width=140, anchor="w")
        subj_tree.column("stream", width=120, anchor="center")
        subj_tree.pack(fill="both", expand=True, padx=10, pady=6)
        tk.Button(
            subject_card,
            text="Remove Selected",
            bg="#dc2626",
            fg="white",
            font=(FF, 9, "bold"),
            padx=10,
            pady=5,
            command=lambda: self._remove_assignment_from_tree(subj_tree, refresh_all),
        ).pack(anchor="e", padx=10, pady=(0, 8))

        class_card = tk.Frame(assignments_wrap, bg=CARD_BG, relief="flat", bd=1)
        class_card.grid(row=0, column=1, sticky="nsew", padx=(6, 0))
        tk.Label(
            class_card,
            text="Grade Facilitator Assignments",
            bg=CARD_BG,
            fg=TEXT_PRIMARY,
            font=(FF, 11, "bold"),
        ).pack(anchor="w", padx=10, pady=(8, 4))
        class_cols = ("teacher", "class", "stream")
        class_tree = ttk.Treeview(
            class_card,
            columns=class_cols,
            show="headings",
            style="App.Treeview",
            height=8,
        )
        class_tree.heading("teacher", text="Teacher")
        class_tree.heading("class", text="Class")
        class_tree.heading("stream", text="Stream")
        class_tree.column("teacher", width=210, anchor="w")
        class_tree.column("class", width=180, anchor="w")
        class_tree.column("stream", width=120, anchor="center")
        class_tree.pack(fill="both", expand=True, padx=10, pady=6)
        tk.Button(
            class_card,
            text="Remove Selected",
            bg="#dc2626",
            fg="white",
            font=(FF, 9, "bold"),
            padx=10,
            pady=5,
            command=lambda: self._remove_assignment_from_tree(class_tree, refresh_all),
        ).pack(anchor="e", padx=10, pady=(0, 8))

        def refresh_all():
            self._load_teachers_tree(tree, search_var.get())
            self._load_teacher_assignments_trees(subj_tree, class_tree)

        search_entry.bind("<Return>", lambda _e: refresh_all())
        search_var.trace_add("write", lambda *_: refresh_all())
        refresh_all()

    def _load_teachers_tree(self, tree, search_query=""):
        for item in tree.get_children():
            tree.delete(item)

        query = str(search_query or "").strip().lower()
        for teacher in db.get_all_teachers():
            role_label = (
                "Subject Teacher"
                if teacher.get("role") == "teacher"
                else "Grade Facilitator"
            )
            if query:
                haystack = " ".join(
                    [
                        teacher.get("id", "")[:8],
                        teacher.get("abbreviation", "")
                        or self._generate_short_label(
                            teacher.get("full_name", ""), "teacher"
                        ),
                        teacher.get("full_name", ""),
                        teacher.get("username", ""),
                        role_label,
                    ]
                ).lower()
                if query not in haystack:
                    continue
            tree.insert(
                "",
                "end",
                iid=teacher.get("id", ""),
                values=(
                    teacher.get("id", "")[:8],
                    teacher.get("abbreviation", "")
                    or self._generate_short_label(
                        teacher.get("full_name", ""), "teacher"
                    ),
                    teacher.get("full_name", ""),
                    teacher.get("username", ""),
                    role_label,
                ),
            )

    def _load_teacher_assignments_trees(self, subj_tree, class_tree):
        for item in subj_tree.get_children():
            subj_tree.delete(item)
        for item in class_tree.get_children():
            class_tree.delete(item)

        for assignment in db.get_subject_teacher_assignments():
            subj_tree.insert(
                "",
                "end",
                iid=assignment.get("id", str(uuid.uuid4())),
                values=(
                    self._get_teacher_label(assignment),
                    self._get_subject_label(
                        assignment.get("subject", ""), assignment.get("class_name", "")
                    ),
                    self._get_class_label(assignment.get("class_name", "")),
                    assignment.get("stream_name", "") or "Whole Class",
                ),
            )

        for assignment in db.get_class_teacher_assignments():
            class_tree.insert(
                "",
                "end",
                iid=assignment.get("id", str(uuid.uuid4())),
                values=(
                    self._get_teacher_label(assignment),
                    self._get_class_label(assignment.get("class_name", "")),
                    assignment.get("stream_name", "") or "Whole Class",
                ),
            )

    def _remove_assignment_from_tree(self, tree, reload_callback=None):
        selected = tree.selection()
        if not selected:
            self._show_notice(
                "Select Assignment",
                "Please select an assignment to remove.",
                kind="info",
            )
            return
        if not self._confirm_delete_action("assignment"):
            return
        assignment_id = selected[0]
        if db.remove_assignment(assignment_id):
            if callable(reload_callback):
                reload_callback()
            self._show_delete_result_notice("assignment", 1, 0)
        else:
            self._show_delete_result_notice("assignment", 0, 1)

    def _normalize_teacher_import_role(self, value):
        role_key = self._normalize_key(value)
        if role_key in {"classteacher", "class_teacher", "class"}:
            return "class_teacher"
        return "teacher"

    def _build_teacher_import_lookup(self):
        lookup = {}
        teachers = db.get_all_teachers()
        for teacher in teachers:
            candidates = [
                teacher.get("username", ""),
                teacher.get("full_name", ""),
                teacher.get("abbreviation", ""),
                self._get_teacher_label(teacher),
            ]
            for candidate in candidates:
                key = self._normalize_key(candidate)
                if key and key not in lookup:
                    lookup[key] = teacher
        return lookup

    def _find_teacher_for_import(self, reference, lookup):
        key = self._normalize_key(reference)
        return lookup.get(key) if key else None

    def _register_teacher_import_lookup(self, lookup, teacher):
        candidates = [
            teacher.get("username", ""),
            teacher.get("full_name", ""),
            teacher.get("abbreviation", ""),
            self._get_teacher_label(teacher),
        ]
        for candidate in candidates:
            key = self._normalize_key(candidate)
            if key:
                lookup[key] = teacher

    def _build_teacher_template_seed_data(self, class_names):
        teacher_names = [
            "Ms. Orna Bogeta",
            "Mr. Erick Onyango",
            "Ms. Janet Omolo",
            "Ms. Milcah Bosibo Ri",
            "Mr. William Siwa",
            "Ms. Linner Atieno",
            "Mr. James Misiko",
            "Ms. Mercy Niatha",
            "Ms. Jully Joy",
            "Ms. Berryl Odhiambo",
            "Mr. Joseph Wambaa",
            "Ms. Dorice Akoth",
            "Mr. Wako Dida",
            "Ms. Isab Ella Kemunto",
            "Ms. Diana Anunda",
            "Mr. John Otieno",
            "Ms. Lilian Otieno",
            "Ms. Sarah Orina",
            "Ms. Elean Nyamora",
            "Mr. Geoffry Asadhi",
            "Mr. Enock Bunde",
            "Ms. Irean Imtembo",
            "Mr. Paul",
            "Mrs. Phelister",
        ]

        def make_username(full_name, used):
            cleaned = re.sub(
                r"^(mr|mrs|ms|miss|dr)\.?\s+",
                "",
                full_name.strip(),
                flags=re.IGNORECASE,
            )
            parts = [re.sub(r"[^a-z0-9]", "", p.lower()) for p in cleaned.split()]
            parts = [p for p in parts if p]
            base = "".join(parts[:2]) or f"teacher{len(used) + 1}"
            candidate = base
            suffix = 2
            while candidate in used:
                candidate = f"{base}{suffix}"
                suffix += 1
            used.add(candidate)
            return candidate

        if not class_names:
            class_names = ["Grade 1"]

        rng = random.Random(42)
        class_cycle = list(class_names)
        rng.shuffle(class_cycle)
        class_teacher_classes = class_cycle[: min(len(class_cycle), len(teacher_names))]
        class_teacher_usernames = set()
        used_usernames = set()
        teacher_rows = []
        subject_assignment_rows = []
        class_assignment_rows = []

        for idx, full_name in enumerate(teacher_names):
            username = make_username(full_name, used_usernames)
            role = "class_teacher" if idx < len(class_teacher_classes) else "teacher"
            abbreviation = self._generate_short_label(full_name, "teacher")
            assigned_class = rng.choice(class_names)
            stream_options = self._get_stream_names_for_class(assigned_class)
            assigned_stream = rng.choice(stream_options) if stream_options else ""
            subject_options = (
                self._get_subjects_for_selected_class(assigned_class, TERMS[0])
                or self.get_current_subjects()
            )
            assigned_subject = rng.choice(subject_options) if subject_options else ""
            class_teacher_stream_options = (
                self._get_stream_names_for_class(class_teacher_classes[idx])
                if idx < len(class_teacher_classes)
                else []
            )
            class_teacher_stream = (
                rng.choice(class_teacher_stream_options)
                if class_teacher_stream_options
                else ""
            )

            teacher_rows.append(
                {
                    "full_name": full_name,
                    "username": username,
                    "role": role,
                    "abbreviation": abbreviation,
                    "password (optional)": username,
                }
            )
            subject_assignment_rows.append(
                {
                    "teacher_username": username,
                    "class": assigned_class,
                    "stream": assigned_stream,
                    "subject": assigned_subject,
                }
            )

            if role == "class_teacher" and username not in class_teacher_usernames:
                class_assignment_rows.append(
                    {
                        "teacher_username": username,
                        "class": class_teacher_classes[idx],
                        "stream": class_teacher_stream,
                    }
                )
                class_teacher_usernames.add(username)

        return teacher_rows, subject_assignment_rows, class_assignment_rows

    def _show_teacher_import_preview_dialog(self, preview_rows, stats):
        top = tk.Toplevel(self.root)
        top.title("Teacher Import Preview")
        top.geometry("1120x660")
        top.configure(bg=CONTENT_BG)
        top.transient(self.root)
        top.grab_set()
        top.resizable(True, True)
        top.minsize(980, 560)

        shell_bo, shell_bi = _card_colors("mint")
        outer = tk.Frame(top, bg=shell_bo)
        outer.pack(fill="both", expand=True, padx=16, pady=16)
        body = tk.Frame(outer, bg=shell_bi, padx=18, pady=16)
        body.pack(fill="both", expand=True, padx=1, pady=1)
        body.grid_columnconfigure(0, weight=1)
        body.grid_rowconfigure(3, weight=1)

        tk.Label(
            body,
            text="Teacher Workbook Import Preview",
            bg=shell_bi,
            fg=TEXT_PRIMARY,
            font=(FF, 13, "bold"),
        ).grid(row=0, column=0, sticky="w")

        tk.Label(
            body,
            text="Review valid rows and skipped rows before saving teachers and assignments.",
            bg=shell_bi,
            fg=TEXT_SECONDARY,
            font=(FF, 10),
        ).grid(row=1, column=0, sticky="w", pady=(4, 12))

        stats_row = tk.Frame(body, bg=shell_bi)
        stats_row.grid(row=2, column=0, sticky="ew", pady=(0, 12))
        stat_items = [
            ("Rows Found", str(stats.get("total_rows", 0)), "#EEF3FF", BLUE),
            ("Ready To Import", str(stats.get("valid_rows", 0)), "#EAF4EC", GREEN),
            ("Skipped", str(stats.get("skipped_rows", 0)), "#FFF1F0", "#dc2626"),
            ("Sheets", stats.get("sheet_summary", "-"), "#FFF6E8", ORANGE),
        ]
        for label_text, value_text, chip_bg, chip_fg in stat_items:
            chip = tk.Frame(stats_row, bg=chip_bg, padx=12, pady=8)
            chip.pack(side="left", padx=(0, 10))
            tk.Label(
                chip,
                text=label_text,
                bg=chip_bg,
                fg=TEXT_SECONDARY,
                font=(FF, 9, "bold"),
            ).pack(anchor="w")
            tk.Label(
                chip, text=value_text, bg=chip_bg, fg=chip_fg, font=(FF, 11, "bold")
            ).pack(anchor="w")

        frame = tk.Frame(body, bg=CARD_BG)
        frame.grid(row=3, column=0, sticky="nsew", pady=(0, 12))
        cols = ("sheet", "row", "action", "target", "status", "details")
        tv = ttk.Treeview(
            frame, columns=cols, show="headings", style="App.Treeview", height=16
        )
        headings = {
            "sheet": "Sheet",
            "row": "Row",
            "action": "Action",
            "target": "Target",
            "status": "Status",
            "details": "Details",
        }
        widths = {
            "sheet": 190,
            "row": 70,
            "action": 170,
            "target": 250,
            "status": 110,
            "details": 410,
        }
        for col in cols:
            tv.heading(col, text=headings[col])
            tv.column(
                col,
                width=widths[col],
                minwidth=70,
                anchor="w" if col not in ("row", "status") else "center",
            )

        ysb = ttk.Scrollbar(
            frame, orient="vertical", command=tv.yview, style="App.Vertical.TScrollbar"
        )
        xsb = ttk.Scrollbar(
            frame,
            orient="horizontal",
            command=tv.xview,
            style="App.Vertical.TScrollbar",
        )
        tv.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        tv.grid(row=0, column=0, sticky="nsew", padx=(0, 4), pady=4)
        ysb.grid(row=0, column=1, sticky="ns", pady=4)
        xsb.grid(row=1, column=0, sticky="ew", padx=(0, 4))
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)
        tv.tag_configure("ready", background="#eaf7ee", foreground="#1f5130")
        tv.tag_configure("skip", background="#fff1f0", foreground="#9f1239")

        for item in preview_rows:
            status_key = str(item.get("status", "")).strip().lower()
            row_tag = (
                "ready"
                if status_key == "ready"
                else "skip"
                if status_key == "skip"
                else ""
            )
            tv.insert(
                "",
                "end",
                values=(
                    item.get("sheet", ""),
                    item.get("row", ""),
                    item.get("action", ""),
                    item.get("target", ""),
                    item.get("status", ""),
                    item.get("details", ""),
                ),
                tags=(row_tag,) if row_tag else (),
            )

        decision = {"ok": False}
        footer = tk.Frame(body, bg="#f7fbe8", padx=14, pady=12)
        footer.grid(row=4, column=0, sticky="ew")
        footer.grid_columnconfigure(0, weight=1)

        footer_text = (
            f"Ready rows: {stats.get('valid_rows', 0)}   "
            f"Skipped rows: {stats.get('skipped_rows', 0)}"
        )
        tk.Label(
            footer,
            text=footer_text,
            bg="#f7fbe8",
            fg=TEXT_SECONDARY,
            font=(FF, 10, "bold"),
        ).grid(row=0, column=0, sticky="w")

        btn_row = tk.Frame(footer, bg="#f7fbe8")
        btn_row.grid(row=0, column=1, sticky="e")
        tk.Button(
            btn_row,
            text="Cancel",
            bg=LEMON_SOFT,
            fg=TEXT_PRIMARY,
            font=(FF, 10, "bold"),
            padx=18,
            pady=10,
            command=top.destroy,
        ).pack(side="left", padx=(0, 10))
        tk.Button(
            btn_row,
            text="Import Now",
            bg=GREEN,
            fg="white",
            font=(FF, 11, "bold"),
            padx=22,
            pady=10,
            command=lambda: (decision.__setitem__("ok", True), top.destroy()),
        ).pack(side="right")

        top.bind(
            "<Return>", lambda _e: (decision.__setitem__("ok", True), top.destroy())
        )
        top.bind("<Escape>", lambda _e: top.destroy())

        self.root.wait_window(top)
        return decision["ok"]

    def download_teacher_import_template(self):
        file_path = filedialog.asksaveasfilename(
            title="Save Teacher Import Template",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="teacher_import_template.xlsx",
        )
        if not file_path:
            return

        try:
            class_names = db.get_class_progression_order() or self.get_current_classes()
            if not class_names:
                class_names = ["Grade 1"]
            teacher_rows, subject_assignment_rows, class_assignment_rows = (
                self._build_teacher_template_seed_data(class_names)
            )

            subject_rows = []
            stream_rows = []
            for class_name in class_names:
                subjects = (
                    self._get_subjects_for_selected_class(class_name, TERMS[0])
                    or self.get_current_subjects()
                )
                for subject in subjects:
                    subject_rows.append({"class": class_name, "subject": subject})
                stream_names = self._get_stream_names_for_class(class_name)
                if stream_names:
                    for stream_name in stream_names:
                        stream_rows.append({"class": class_name, "stream": stream_name})
                else:
                    stream_rows.append({"class": class_name, "stream": ""})

            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                pd.DataFrame(
                    {
                        "Instructions": [
                            "Use the Teachers sheet to add or update teachers by username.",
                            "For new teachers, leave password blank to use the username as the starting password.",
                            "Use teacher_username in assignment sheets so the app can match teachers correctly.",
                            "Use the stream column when assigning a teacher to a specific stream such as Green or Yellow.",
                            "Leave stream blank to assign the teacher to the whole class.",
                            "Subject Assignments adds or updates subject-teacher links.",
                            "Grade Facilitator Assignments replaces the Grade Facilitator for that exact class/stream target.",
                        ]
                    }
                ).to_excel(writer, sheet_name="Instructions", index=False)

                pd.DataFrame(teacher_rows).to_excel(
                    writer, sheet_name="Teachers", index=False
                )
                pd.DataFrame(subject_assignment_rows).to_excel(
                    writer, sheet_name="Subject Assignments", index=False
                )
                pd.DataFrame(class_assignment_rows).to_excel(
                    writer, sheet_name="Grade Facilitator Assignments", index=False
                )

                pd.DataFrame({"class": class_names}).to_excel(
                    writer, sheet_name="Classes Reference", index=False
                )
                pd.DataFrame(stream_rows).to_excel(
                    writer, sheet_name="Streams Reference", index=False
                )
                pd.DataFrame(
                    subject_rows or [{"class": class_names[0], "subject": ""}]
                ).to_excel(writer, sheet_name="Subjects Reference", index=False)

            messagebox.showinfo(
                "Template Ready",
                f"Teacher import template saved to:\n{file_path}\n\n"
                "Use the Teachers screen to import the filled workbook.",
            )
        except Exception as exc:
            messagebox.showerror(
                "Template Error", f"Failed to create teacher template:\n{exc}"
            )

    def export_teachers_excel(self):
        file_path = filedialog.asksaveasfilename(
            title="Export Teachers Workbook",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="teachers_export.xlsx",
        )
        if not file_path:
            return

        try:
            teachers = db.get_all_teachers()
            subject_assignments = db.get_subject_teacher_assignments()
            class_assignments = db.get_class_teacher_assignments()
            class_names = db.get_class_progression_order() or self.get_current_classes()
            if not class_names:
                class_names = ["Grade 1"]

            subject_rows = []
            stream_rows = []
            for class_name in class_names:
                subjects = (
                    self._get_subjects_for_selected_class(class_name, TERMS[0])
                    or self.get_current_subjects()
                )
                for subject in subjects:
                    subject_rows.append({"class": class_name, "subject": subject})
                stream_names = self._get_stream_names_for_class(class_name)
                if stream_names:
                    for stream_name in stream_names:
                        stream_rows.append({"class": class_name, "stream": stream_name})
                else:
                    stream_rows.append({"class": class_name, "stream": ""})

            teacher_rows = [
                {
                    "full_name": teacher.get("full_name", ""),
                    "username": teacher.get("username", ""),
                    "role": teacher.get("role", "teacher"),
                    "abbreviation": teacher.get("abbreviation", ""),
                    "password (optional)": "",
                }
                for teacher in teachers
            ]

            subject_assignment_rows = [
                {
                    "teacher_username": assignment.get("username", ""),
                    "teacher_name": assignment.get("full_name", ""),
                    "class": assignment.get("class_name", ""),
                    "stream": assignment.get("stream_name", ""),
                    "subject": assignment.get("subject", ""),
                }
                for assignment in subject_assignments
            ]

            class_assignment_rows = [
                {
                    "teacher_username": assignment.get("username", ""),
                    "teacher_name": assignment.get("full_name", ""),
                    "class": assignment.get("class_name", ""),
                    "stream": assignment.get("stream_name", ""),
                }
                for assignment in class_assignments
            ]

            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                pd.DataFrame(
                    {
                        "Instructions": [
                            "This workbook was exported from Teachers Management.",
                            "Edit the Teachers, Subject Assignments, and Grade Facilitator Assignments sheets as needed.",
                            "Use the stream column for class-specific streams like Green or Yellow.",
                            "Leave password blank if you do not want to change an existing teacher password.",
                            "Import the edited workbook from the same Teachers screen using Import Excel.",
                        ]
                    }
                ).to_excel(writer, sheet_name="Instructions", index=False)

                pd.DataFrame(
                    teacher_rows
                    or [
                        {
                            "full_name": "",
                            "username": "",
                            "role": "teacher",
                            "abbreviation": "",
                            "password (optional)": "",
                        }
                    ]
                ).to_excel(writer, sheet_name="Teachers", index=False)

                pd.DataFrame(
                    subject_assignment_rows
                    or [
                        {
                            "teacher_username": "",
                            "teacher_name": "",
                            "class": class_names[0],
                            "stream": "",
                            "subject": "",
                        }
                    ]
                ).to_excel(writer, sheet_name="Subject Assignments", index=False)

                pd.DataFrame(
                    class_assignment_rows
                    or [
                        {
                            "teacher_username": "",
                            "teacher_name": "",
                            "class": class_names[0],
                            "stream": "",
                        }
                    ]
                ).to_excel(
                    writer, sheet_name="Grade Facilitator Assignments", index=False
                )

                pd.DataFrame({"class": class_names}).to_excel(
                    writer, sheet_name="Classes Reference", index=False
                )
                pd.DataFrame(stream_rows).to_excel(
                    writer, sheet_name="Streams Reference", index=False
                )
                pd.DataFrame(
                    subject_rows or [{"class": class_names[0], "subject": ""}]
                ).to_excel(writer, sheet_name="Subjects Reference", index=False)

            messagebox.showinfo(
                "Export Complete", f"Teacher workbook exported to:\n{file_path}"
            )
        except Exception as exc:
            messagebox.showerror(
                "Export Error", f"Failed to export teacher workbook:\n{exc}"
            )

    def import_teachers_excel(self, on_complete=None):
        file_path = filedialog.askopenfilename(
            title="Import Teachers Workbook",
            filetypes=[("Excel files", "*.xlsx *.xls")],
        )
        if not file_path:
            return

        progress_dialog = None
        try:
            workbook = pd.ExcelFile(file_path)
            sheet_lookup = {
                self._normalize_key(name): name for name in workbook.sheet_names
            }

            teachers_sheet = next(
                (
                    sheet_lookup[key]
                    for key in ("teachers", "teacher", "staff")
                    if key in sheet_lookup
                ),
                None,
            )
            subject_sheet = next(
                (
                    sheet_lookup[key]
                    for key in (
                        "subjectassignments",
                        "subjectassignment",
                        "subject_assignment",
                    )
                    if key in sheet_lookup
                ),
                None,
            )
            class_sheet = next(
                (
                    sheet_lookup[key]
                    for key in (
                        "classteacherassignments",
                        "classteacherassignment",
                        "class_teacher_assignments",
                    )
                    if key in sheet_lookup
                ),
                None,
            )

            if not any([teachers_sheet, subject_sheet, class_sheet]):
                messagebox.showerror(
                    "Import Error",
                    "Workbook not recognized.\n\nExpected at least one of these sheets:\n"
                    "Teachers\nSubject Assignments\nGrade Facilitator Assignments",
                )
                return

            teachers_df = workbook.parse(teachers_sheet) if teachers_sheet else None
            subject_df = workbook.parse(subject_sheet) if subject_sheet else None
            class_df = workbook.parse(class_sheet) if class_sheet else None
            for df in (teachers_df, subject_df, class_df):
                if df is not None:
                    df.columns = [self._normalize_key(col) for col in df.columns]

            total_input_rows = sum(
                len(df.index)
                for df in (teachers_df, subject_df, class_df)
                if df is not None
            )
            progress_dialog, status_label, percent_label, progress = (
                self._open_progress_dialog(
                    "Importing Teachers",
                    "Preparing teacher import...",
                    allow_cancel=True,
                )
            )

            def ensure_not_cancelled():
                if self._progress_cancel_requested(progress_dialog):
                    raise ImportCancelledError(
                        "Teacher import cancelled.\n\nAny records already imported before cancellation were kept."
                    )

            def build_teacher_progress_details(
                current_sheet="",
                analyzed_rows=0,
                total_rows_count=0,
                ready_rows=0,
                skipped_rows=0,
                teachers_added=0,
                teachers_updated=0,
                subject_links=0,
                class_links=0,
            ):
                detail_lines = []
                if current_sheet:
                    detail_lines.append(f"Sheet: {current_sheet}")
                if total_rows_count:
                    detail_lines.append(
                        f"Rows analyzed: {analyzed_rows}/{total_rows_count}"
                    )
                detail_lines.append(f"Ready: {ready_rows}   Skipped: {skipped_rows}")
                detail_lines.append(
                    f"Teachers added: {teachers_added}   Updated: {teachers_updated}"
                )
                detail_lines.append(
                    f"Subject links: {subject_links}   Grade Facilitators: {class_links}"
                )
                return "\n".join(detail_lines)

            def update_teacher_import_progress(percent_value, message, details):
                self._update_progress_dialog(
                    progress_dialog,
                    status_label,
                    percent_label,
                    progress,
                    max(0, min(100, int(percent_value))),
                    100,
                    message,
                    details,
                )

            def clean(value):
                if pd.isna(value):
                    return ""
                text = str(value).strip()
                return "" if text.lower() == "nan" else text

            def pick(row, *keys):
                for key in keys:
                    if key in row and clean(row.get(key)):
                        return clean(row.get(key))
                return ""

            class_names = db.get_class_progression_order() or self.get_current_classes()
            known_classes = set(class_names)
            preview_rows = []
            teacher_actions = []
            subject_actions = []
            class_actions = []
            total_rows = 0
            analyzed_rows = 0
            ready_rows = 0
            skipped_rows = 0

            teacher_lookup = self._build_teacher_import_lookup()

            if teachers_df is not None:
                for idx, row in teachers_df.iterrows():
                    total_rows += 1
                    analyzed_rows += 1
                    update_teacher_import_progress(
                        int((analyzed_rows / max(1, total_input_rows)) * 35),
                        f"Analyzing teacher row {idx + 2}...",
                        build_teacher_progress_details(
                            current_sheet=teachers_sheet,
                            analyzed_rows=analyzed_rows,
                            total_rows_count=total_input_rows,
                            ready_rows=ready_rows,
                            skipped_rows=skipped_rows,
                        ),
                    )
                    ensure_not_cancelled()
                    full_name = pick(row, "fullname", "name")
                    username = pick(row, "username", "teacherusername")
                    role = self._normalize_teacher_import_role(
                        row.get("role", "teacher")
                    )
                    abbreviation = pick(row, "abbreviation", "abbr")
                    password = pick(row, "passwordoptional", "password", "temppassword")

                    if not any([full_name, username, abbreviation]):
                        continue
                    if not full_name or not username:
                        preview_rows.append(
                            {
                                "sheet": teachers_sheet,
                                "row": idx + 2,
                                "action": "Teacher",
                                "target": username or full_name or "-",
                                "status": "Skip",
                                "details": "full_name and username are required",
                            }
                        )
                        skipped_rows += 1
                        continue

                    existing = self._find_teacher_for_import(username, teacher_lookup)
                    action = "Update Teacher" if existing else "Add Teacher"
                    preview_rows.append(
                        {
                            "sheet": teachers_sheet,
                            "row": idx + 2,
                            "action": action,
                            "target": f"{full_name} ({username})",
                            "status": "Ready",
                            "details": f"Role: {role}, Label: {abbreviation or '-'}",
                        }
                    )
                    ready_rows += 1
                    teacher_actions.append(
                        {
                            "mode": "update" if existing else "add",
                            "row": idx + 2,
                            "sheet": teachers_sheet,
                            "teacher_id": existing.get("id", "") if existing else "",
                            "full_name": full_name,
                            "username": username,
                            "role": role,
                            "abbreviation": abbreviation,
                            "password": password or username,
                        }
                    )
                    self._register_teacher_import_lookup(
                        teacher_lookup,
                        {
                            "id": existing.get("id", "") if existing else "",
                            "full_name": full_name,
                            "username": username,
                            "role": role,
                            "abbreviation": abbreviation,
                        },
                    )

            if subject_df is not None:
                for idx, row in subject_df.iterrows():
                    total_rows += 1
                    analyzed_rows += 1
                    update_teacher_import_progress(
                        int((analyzed_rows / max(1, total_input_rows)) * 35),
                        f"Analyzing subject assignment row {idx + 2}...",
                        build_teacher_progress_details(
                            current_sheet=subject_sheet,
                            analyzed_rows=analyzed_rows,
                            total_rows_count=total_input_rows,
                            ready_rows=ready_rows,
                            skipped_rows=skipped_rows,
                        ),
                    )
                    ensure_not_cancelled()
                    teacher_ref = pick(row, "teacherusername", "username", "teacher")
                    raw_class = pick(row, "class", "classname")
                    raw_stream = pick(row, "stream", "classstream")
                    raw_subject = pick(row, "subject", "learningarea")

                    if not any([teacher_ref, raw_class, raw_stream, raw_subject]):
                        continue
                    if not all([teacher_ref, raw_class, raw_subject]):
                        preview_rows.append(
                            {
                                "sheet": subject_sheet,
                                "row": idx + 2,
                                "action": "Subject Assignment",
                                "target": teacher_ref or raw_subject or "-",
                                "status": "Skip",
                                "details": "teacher_username, class, and subject are required",
                            }
                        )
                        skipped_rows += 1
                        continue

                    teacher = self._find_teacher_for_import(teacher_ref, teacher_lookup)
                    class_name = self._match_known_class_name(raw_class) or raw_class
                    stream_name = (
                        self._match_known_stream_name(raw_stream, class_name)
                        if raw_stream
                        else ""
                    )
                    if class_name not in known_classes:
                        preview_rows.append(
                            {
                                "sheet": subject_sheet,
                                "row": idx + 2,
                                "action": "Subject Assignment",
                                "target": f"{teacher_ref} -> {raw_subject}",
                                "status": "Skip",
                                "details": f'Unknown class "{raw_class}"',
                            }
                        )
                        skipped_rows += 1
                        continue
                    if not teacher:
                        preview_rows.append(
                            {
                                "sheet": subject_sheet,
                                "row": idx + 2,
                                "action": "Subject Assignment",
                                "target": f"{teacher_ref} -> {raw_subject}",
                                "status": "Skip",
                                "details": f'Teacher "{teacher_ref}" not found',
                            }
                        )
                        skipped_rows += 1
                        continue
                    if raw_stream and not stream_name:
                        preview_rows.append(
                            {
                                "sheet": subject_sheet,
                                "row": idx + 2,
                                "action": "Subject Assignment",
                                "target": f"{teacher_ref} -> {class_name}",
                                "status": "Skip",
                                "details": f'Stream "{raw_stream}" not found in {class_name}',
                            }
                        )
                        skipped_rows += 1
                        continue

                    subject_name = self._map_sheet_subject(raw_subject, class_name)
                    valid_subjects = (
                        self._get_subjects_for_selected_class(class_name, TERMS[0])
                        or self.get_current_subjects()
                    )
                    if subject_name not in valid_subjects:
                        preview_rows.append(
                            {
                                "sheet": subject_sheet,
                                "row": idx + 2,
                                "action": "Subject Assignment",
                                "target": f"{teacher.get('username', teacher_ref)} -> {class_name}",
                                "status": "Skip",
                                "details": f'Subject "{raw_subject}" not valid for {class_name}',
                            }
                        )
                        skipped_rows += 1
                        continue

                    preview_rows.append(
                        {
                            "sheet": subject_sheet,
                            "row": idx + 2,
                            "action": "Subject Assignment",
                            "target": f"{teacher.get('username', teacher_ref)} -> {self._format_class_stream_label(class_name, stream_name)}",
                            "status": "Ready",
                            "details": subject_name,
                        }
                    )
                    ready_rows += 1
                    subject_actions.append(
                        {
                            "sheet": subject_sheet,
                            "row": idx + 2,
                            "teacher_username": teacher.get("username", teacher_ref),
                            "class_name": class_name,
                            "stream_name": stream_name,
                            "subject_name": subject_name,
                        }
                    )

            if class_df is not None:
                for idx, row in class_df.iterrows():
                    total_rows += 1
                    analyzed_rows += 1
                    update_teacher_import_progress(
                        int((analyzed_rows / max(1, total_input_rows)) * 35),
                        f"Analyzing Grade Facilitator row {idx + 2}...",
                        build_teacher_progress_details(
                            current_sheet=class_sheet,
                            analyzed_rows=analyzed_rows,
                            total_rows_count=total_input_rows,
                            ready_rows=ready_rows,
                            skipped_rows=skipped_rows,
                        ),
                    )
                    ensure_not_cancelled()
                    teacher_ref = pick(row, "teacherusername", "username", "teacher")
                    raw_class = pick(row, "class", "classname")
                    raw_stream = pick(row, "stream", "classstream")

                    if not any([teacher_ref, raw_class, raw_stream]):
                        continue
                    if not all([teacher_ref, raw_class]):
                        preview_rows.append(
                            {
                                "sheet": class_sheet,
                                "row": idx + 2,
                                "action": "Grade Facilitator",
                                "target": teacher_ref or raw_class or "-",
                                "status": "Skip",
                                "details": "teacher_username and class are required",
                            }
                        )
                        skipped_rows += 1
                        continue

                    teacher = self._find_teacher_for_import(teacher_ref, teacher_lookup)
                    class_name = self._match_known_class_name(raw_class) or raw_class
                    stream_name = (
                        self._match_known_stream_name(raw_stream, class_name)
                        if raw_stream
                        else ""
                    )
                    if class_name not in known_classes:
                        preview_rows.append(
                            {
                                "sheet": class_sheet,
                                "row": idx + 2,
                                "action": "Grade Facilitator",
                                "target": f"{teacher_ref} -> {raw_class}",
                                "status": "Skip",
                                "details": f'Unknown class "{raw_class}"',
                            }
                        )
                        skipped_rows += 1
                        continue
                    if not teacher:
                        preview_rows.append(
                            {
                                "sheet": class_sheet,
                                "row": idx + 2,
                                "action": "Grade Facilitator",
                                "target": f"{teacher_ref} -> {class_name}",
                                "status": "Skip",
                                "details": f'Teacher "{teacher_ref}" not found',
                            }
                        )
                        skipped_rows += 1
                        continue
                    if raw_stream and not stream_name:
                        preview_rows.append(
                            {
                                "sheet": class_sheet,
                                "row": idx + 2,
                                "action": "Grade Facilitator",
                                "target": f"{teacher_ref} -> {class_name}",
                                "status": "Skip",
                                "details": f'Stream "{raw_stream}" not found in {class_name}',
                            }
                        )
                        skipped_rows += 1
                        continue

                    preview_rows.append(
                        {
                            "sheet": class_sheet,
                            "row": idx + 2,
                            "action": "Grade Facilitator",
                            "target": f"{teacher.get('username', teacher_ref)} -> {self._format_class_stream_label(class_name, stream_name)}",
                            "status": "Ready",
                            "details": "Will replace current Grade Facilitator for this class/stream",
                        }
                    )
                    ready_rows += 1
                    class_actions.append(
                        {
                            "sheet": class_sheet,
                            "row": idx + 2,
                            "teacher_username": teacher.get("username", teacher_ref),
                            "class_name": class_name,
                            "stream_name": stream_name,
                        }
                    )

            valid_rows = len(
                [row for row in preview_rows if row.get("status") == "Ready"]
            )
            skipped_rows = len(
                [row for row in preview_rows if row.get("status") == "Skip"]
            )
            sheet_names = [
                name for name in [teachers_sheet, subject_sheet, class_sheet] if name
            ]
            if not preview_rows:
                if progress_dialog is not None:
                    try:
                        progress_dialog.destroy()
                    except Exception:
                        pass
                messagebox.showwarning(
                    "No Data",
                    "No teacher rows or assignment rows were found in the workbook.",
                )
                return

            total_actions = max(
                1, len(teacher_actions) + len(subject_actions) + len(class_actions)
            )
            processed = 0
            teachers_added = 0
            teachers_updated = 0
            subject_links = 0
            class_links = 0
            runtime_skipped = []

            for action in teacher_actions:
                processed += 1
                update_teacher_import_progress(
                    35 + int((processed / total_actions) * 65),
                    f"Processing teacher row {action['row']} ({processed}/{total_actions})",
                    build_teacher_progress_details(
                        current_sheet=action["sheet"],
                        analyzed_rows=analyzed_rows,
                        total_rows_count=total_input_rows,
                        ready_rows=valid_rows,
                        skipped_rows=skipped_rows,
                        teachers_added=teachers_added,
                        teachers_updated=teachers_updated,
                        subject_links=subject_links,
                        class_links=class_links,
                    ),
                )
                ensure_not_cancelled()
                if action["mode"] == "update":
                    success, msg = db.update_teacher(
                        action["teacher_id"],
                        action["full_name"],
                        action["username"],
                        action["role"],
                        action["abbreviation"],
                        action["password"],
                    )
                    if success:
                        teachers_updated += 1
                    else:
                        runtime_skipped.append(
                            f"{action['sheet']} row {action['row']}: {msg}"
                        )
                else:
                    success, msg = db.add_teacher(
                        action["full_name"],
                        action["username"],
                        action["password"],
                        action["role"],
                        action["abbreviation"],
                    )
                    if success:
                        teachers_added += 1
                    else:
                        runtime_skipped.append(
                            f"{action['sheet']} row {action['row']}: {msg}"
                        )

            teacher_lookup = self._build_teacher_import_lookup()

            for action in subject_actions:
                processed += 1
                update_teacher_import_progress(
                    35 + int((processed / total_actions) * 65),
                    f"Processing subject assignment row {action['row']} ({processed}/{total_actions})",
                    build_teacher_progress_details(
                        current_sheet=action["sheet"],
                        analyzed_rows=analyzed_rows,
                        total_rows_count=total_input_rows,
                        ready_rows=valid_rows,
                        skipped_rows=skipped_rows,
                        teachers_added=teachers_added,
                        teachers_updated=teachers_updated,
                        subject_links=subject_links,
                        class_links=class_links,
                    ),
                )
                ensure_not_cancelled()
                teacher = self._find_teacher_for_import(
                    action["teacher_username"], teacher_lookup
                )
                if not teacher:
                    runtime_skipped.append(
                        f'{action["sheet"]} row {action["row"]}: teacher "{action["teacher_username"]}" not found at save time'
                    )
                    continue
                if db.assign_subject_teacher(
                    teacher["id"],
                    action["class_name"],
                    action["subject_name"],
                    action.get("stream_name", ""),
                ):
                    subject_links += 1
                else:
                    runtime_skipped.append(
                        f"{action['sheet']} row {action['row']}: could not save assignment"
                    )

            for action in class_actions:
                processed += 1
                update_teacher_import_progress(
                    35 + int((processed / total_actions) * 65),
                    f"Processing Grade Facilitator row {action['row']} ({processed}/{total_actions})",
                    build_teacher_progress_details(
                        current_sheet=action["sheet"],
                        analyzed_rows=analyzed_rows,
                        total_rows_count=total_input_rows,
                        ready_rows=valid_rows,
                        skipped_rows=skipped_rows,
                        teachers_added=teachers_added,
                        teachers_updated=teachers_updated,
                        subject_links=subject_links,
                        class_links=class_links,
                    ),
                )
                ensure_not_cancelled()
                teacher = self._find_teacher_for_import(
                    action["teacher_username"], teacher_lookup
                )
                if not teacher:
                    runtime_skipped.append(
                        f'{action["sheet"]} row {action["row"]}: teacher "{action["teacher_username"]}" not found at save time'
                    )
                    continue
                if db.assign_class_teacher(
                    teacher["id"], action["class_name"], action.get("stream_name", "")
                ):
                    class_links += 1
                else:
                    runtime_skipped.append(
                        f"{action['sheet']} row {action['row']}: could not save assignment"
                    )

            update_teacher_import_progress(
                100,
                "Refreshing teacher view...",
                build_teacher_progress_details(
                    analyzed_rows=analyzed_rows,
                    total_rows_count=total_input_rows,
                    ready_rows=valid_rows,
                    skipped_rows=skipped_rows,
                    teachers_added=teachers_added,
                    teachers_updated=teachers_updated,
                    subject_links=subject_links,
                    class_links=class_links,
                ),
            )
            ensure_not_cancelled()

            try:
                progress_dialog.destroy()
            except Exception:
                pass

            if callable(on_complete):
                on_complete()
            else:
                self.show_teachers()

            msg = (
                "Teacher workbook imported successfully.\n\n"
                f"Teachers added: {teachers_added}\n"
                f"Teachers updated: {teachers_updated}\n"
                f"Subject assignments saved: {subject_links}\n"
                f"Grade Facilitator assignments saved: {class_links}"
            )
            all_skipped = [row for row in preview_rows if row.get("status") == "Skip"]
            skipped_messages = [
                f"{row.get('sheet', '')} row {row.get('row', '')}: {row.get('details', '')}"
                for row in all_skipped
            ] + runtime_skipped
            if skipped_messages:
                preview = "\n".join(skipped_messages[:12])
                extra = (
                    ""
                    if len(skipped_messages) <= 12
                    else f"\n...and {len(skipped_messages) - 12} more"
                )
                msg += f"\n\nSkipped rows: {len(skipped_messages)}\n{preview}{extra}"
            messagebox.showinfo("Import Complete", msg)
        except ImportCancelledError as exc:
            if progress_dialog is not None:
                try:
                    progress_dialog.destroy()
                except Exception:
                    pass
            messagebox.showinfo("Import Cancelled", str(exc))
        except Exception as exc:
            if progress_dialog is not None:
                try:
                    progress_dialog.destroy()
                except Exception:
                    pass
            messagebox.showerror(
                "Import Error", f"Failed to import teacher workbook:\n{exc}"
            )

    def _open_teacher_dialog(self, teacher_row=None):
        dialog = tk.Toplevel(self.root)
        dialog.title("Edit Teacher" if teacher_row else "Add Teacher")
        dialog.geometry("430x500")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)

        tk.Label(dialog, text="Full Name:", font=(FF, 11)).pack(pady=(20, 5))
        name_entry = tk.Entry(dialog, font=(FF, 11))
        name_entry.pack(fill="x", padx=20)

        tk.Label(dialog, text="Short Label / Abbreviation:", font=(FF, 11)).pack(
            pady=(15, 5)
        )
        abbr_entry = tk.Entry(dialog, font=(FF, 11))
        abbr_entry.pack(fill="x", padx=20)

        tk.Label(dialog, text="Username:", font=(FF, 11)).pack(pady=(15, 5))
        username_entry = tk.Entry(dialog, font=(FF, 11))
        username_entry.pack(fill="x", padx=20)

        tk.Label(dialog, text="Password:", font=(FF, 11)).pack(pady=(15, 5))
        password_entry = tk.Entry(dialog, font=(FF, 11), show="*")
        password_entry.pack(fill="x", padx=20)

        tk.Label(dialog, text="Role:", font=(FF, 11)).pack(pady=(15, 5))
        role_var = tk.StringVar(value="teacher")
        role_frame = tk.Frame(dialog)
        role_frame.pack()
        tk.Radiobutton(
            role_frame,
            text="Subject Teacher",
            variable=role_var,
            value="teacher",
            font=(FF, 10),
        ).pack(side="left", padx=10)
        tk.Radiobutton(
            role_frame,
            text="Grade Facilitator",
            variable=role_var,
            value="class_teacher",
            font=(FF, 10),
        ).pack(side="left", padx=10)

        if teacher_row:
            name_entry.insert(0, teacher_row.get("full_name", ""))
            abbr_entry.insert(0, teacher_row.get("abbreviation", ""))
            username_entry.insert(0, teacher_row.get("username", ""))
            role_var.set(teacher_row.get("role", "teacher"))

        def save():
            name = name_entry.get().strip()
            abbreviation = abbr_entry.get().strip()
            username = username_entry.get().strip()
            password = password_entry.get().strip()
            role = role_var.get()

            if not name or not username or (not teacher_row and not password):
                messagebox.showerror(
                    "Error",
                    "Name, username, and password are required for new teachers",
                )
                return

            if teacher_row:
                success, msg = db.update_teacher(
                    teacher_row["id"], name, username, role, abbreviation, password
                )
            else:
                success, msg = db.add_teacher(
                    name, username, password, role, abbreviation
                )
            if success:
                dialog.destroy()
                self.show_settings()
            else:
                messagebox.showerror("Error", msg)

        tk.Button(
            dialog,
            text="Save",
            bg=GREEN,
            fg="white",
            font=(FF, 11),
            padx=20,
            pady=8,
            command=save,
        ).pack(pady=(16, 24))

    def _edit_teacher_dialog(self, tree):
        selected = tree.selection()
        if len(selected) != 1:
            messagebox.showwarning(
                "Select One Teacher", "Please select exactly one teacher to edit"
            )
            return
        teacher_id = selected[0]
        teacher_row = next(
            (row for row in db.get_all_teachers() if row.get("id") == teacher_id), None
        )
        if not teacher_row:
            messagebox.showerror("Error", "Could not load the selected teacher")
            return
        self._open_teacher_dialog(teacher_row)

    def _build_grading_scale_tab(self, parent):
        toolbar = tk.Frame(parent, bg=CONTENT_BG)
        toolbar.pack(fill="x", pady=10)

        tk.Label(
            toolbar, text="Class:", bg=CONTENT_BG, fg=TEXT_SECONDARY, font=(FF, 10)
        ).pack(side="left", padx=(5, 4))
        class_options = [
            row.get("name") for row in db.get_all_classes()
        ] or self.get_current_classes()
        class_var = tk.StringVar(value=class_options[0] if class_options else "")
        class_cb = ttk.Combobox(
            toolbar,
            textvariable=class_var,
            values=class_options,
            state="readonly",
            style="App.TCombobox",
            width=22,
        )
        class_cb.pack(side="left", padx=(0, 10))

        tree = ttk.Treeview(
            parent,
            columns=("code", "name", "min", "max", "order"),
            show="headings",
            style="App.Treeview",
        )
        tree.heading("code", text="Grade")
        tree.heading("name", text="Description")
        tree.heading("min", text="Min")
        tree.heading("max", text="Max")
        tree.heading("order", text="Order")
        tree.column("code", width=80, anchor="center")
        tree.column("name", width=220)
        tree.column("min", width=80, anchor="center")
        tree.column("max", width=80, anchor="center")
        tree.column("order", width=80, anchor="center")
        tree.pack(fill="both", expand=True, padx=10, pady=10)

        actions = tk.Frame(parent, bg=CONTENT_BG)
        actions.pack(fill="x", pady=(0, 10))
        tk.Button(
            actions,
            text="+ Add Band",
            bg=GREEN,
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: self._open_grading_scale_dialog(
                class_var.get(),
                refresh=lambda: self._load_grading_scale_tree(tree, class_var.get()),
            ),
        ).pack(side="left", padx=5)
        tk.Button(
            actions,
            text="Edit Selected",
            bg=BLUE,
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: self._edit_grading_scale_dialog(tree, class_var.get()),
        ).pack(side="left", padx=5)
        tk.Button(
            actions,
            text="Delete Selected",
            bg="#e74c3c",
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: self._delete_grading_scale(tree, class_var.get()),
        ).pack(side="left", padx=5)
        tk.Button(
            actions,
            text="Load 8-Band Template",
            bg="#7c3aed",
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: self._apply_grade_band_template(
                class_var.get(),
                refresh=lambda: self._load_grading_scale_tree(tree, class_var.get()),
            ),
        ).pack(side="left", padx=5)
        tk.Button(
            actions,
            text="Refresh",
            bg="#666",
            fg="white",
            font=(FF, 10),
            padx=12,
            pady=5,
            command=lambda: self._load_grading_scale_tree(tree, class_var.get()),
        ).pack(side="left", padx=5)

        class_cb.bind(
            "<<ComboboxSelected>>",
            lambda e: self._load_grading_scale_tree(tree, class_var.get()),
        )
        tree.bind(
            "<Double-1>",
            lambda e: self._edit_grading_scale_dialog(tree, class_var.get()),
        )
        self._load_grading_scale_tree(tree, class_var.get())

    def _load_grading_scale_tree(self, tree, class_name):
        for item in tree.get_children():
            tree.delete(item)
        for scale in db.get_grading_scales(class_name):
            tree.insert(
                "",
                "end",
                iid=scale.get("id", ""),
                values=(
                    scale.get("grade_code", ""),
                    scale.get("grade_name", "")
                    or GRADE_LABELS.get(
                        scale.get("grade_code", ""), scale.get("grade_code", "")
                    ),
                    scale.get("min_mark", ""),
                    scale.get("max_mark", ""),
                    scale.get("sort_order", 0),
                ),
            )

    def _open_grading_scale_dialog(self, class_name, scale_row=None, refresh=None):
        dialog = tk.Toplevel(self.root)
        dialog.title("Edit Grade Band" if scale_row else "Add Grade Band")
        dialog.geometry("430x500")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)

        class_options = [
            row.get("name") for row in db.get_all_classes()
        ] or self.get_current_classes()

        tk.Label(dialog, text="Class:", font=(FF, 11)).pack(pady=(20, 5))
        class_var = tk.StringVar(value=class_name)
        class_cb = ttk.Combobox(
            dialog,
            textvariable=class_var,
            values=class_options,
            state="readonly",
            style="App.TCombobox",
        )
        class_cb.pack(fill="x", padx=20)

        tk.Label(dialog, text="Grade Code:", font=(FF, 11)).pack(pady=(15, 5))
        code_var = tk.StringVar()
        grade_code_options = [
            "EE",
            "EE1",
            "EE2",
            "ME",
            "ME1",
            "ME2",
            "AE",
            "AE1",
            "AE2",
            "BE",
            "BE1",
            "BE2",
            "IE",
        ]
        code_cb = ttk.Combobox(
            dialog,
            textvariable=code_var,
            values=grade_code_options,
            style="App.TCombobox",
        )
        code_cb.pack(fill="x", padx=20)
        tk.Label(
            dialog,
            text="Tip: pick from list or type your own code (e.g. EE1, ME2, BE1).",
            font=(FF, 9),
            fg=TEXT_SECONDARY,
            bg=dialog.cget("bg"),
        ).pack(anchor="w", padx=22, pady=(4, 0))

        tk.Label(dialog, text="Grade Name:", font=(FF, 11)).pack(pady=(15, 5))
        name_entry = tk.Entry(dialog, font=(FF, 11))
        name_entry.pack(fill="x", padx=20)

        tk.Label(dialog, text="Minimum Mark:", font=(FF, 11)).pack(pady=(15, 5))
        min_entry = tk.Entry(dialog, font=(FF, 11))
        min_entry.pack(fill="x", padx=20)

        tk.Label(dialog, text="Maximum Mark:", font=(FF, 11)).pack(pady=(15, 5))
        max_entry = tk.Entry(dialog, font=(FF, 11))
        max_entry.pack(fill="x", padx=20)

        tk.Label(dialog, text="Display Order:", font=(FF, 11)).pack(pady=(15, 5))
        order_entry = tk.Entry(dialog, font=(FF, 11))
        order_entry.pack(fill="x", padx=20)

        if scale_row:
            class_var.set(scale_row.get("class_name", class_name))
            code_var.set(scale_row.get("grade_code", ""))
            name_entry.insert(0, scale_row.get("grade_name", ""))
            min_entry.insert(0, str(scale_row.get("min_mark", "")))
            max_entry.insert(0, str(scale_row.get("max_mark", "")))
            order_entry.insert(0, str(scale_row.get("sort_order", 0)))
        else:
            code_var.set("EE")
            name_entry.insert(0, GRADE_LABELS.get("EE", "Exceeding Expectations"))
            order_entry.insert(0, "1")

        def sync_grade_name(event=None):
            current = name_entry.get().strip()
            code = code_var.get().strip()
            if (
                not current
                or current in GRADE_LABELS.values()
                or current == GRADE_LABELS.get(grade_base_code(code), "")
            ):
                name_entry.delete(0, tk.END)
                name_entry.insert(
                    0, GRADE_LABELS.get(grade_base_code(code), code or "")
                )

        code_cb.bind("<<ComboboxSelected>>", sync_grade_name)

        def save():
            try:
                min_mark = float(min_entry.get().strip())
                max_mark = float(max_entry.get().strip())
                sort_order = int(order_entry.get().strip() or "0")
            except ValueError:
                messagebox.showerror("Error", "Min, max, and order must be numeric")
                return

            if not class_var.get().strip() or not code_var.get().strip():
                messagebox.showerror("Error", "Class and grade code are required")
                return

            grade_code = code_var.get().strip()
            grade_name = name_entry.get().strip() or GRADE_LABELS.get(
                grade_base_code(grade_code), grade_code
            )
            if scale_row:
                success, msg = db.update_grading_scale(
                    scale_row["id"],
                    class_var.get().strip(),
                    min_mark,
                    max_mark,
                    grade_code,
                    grade_name,
                    sort_order,
                )
            else:
                success, msg = db.add_grading_scale(
                    class_var.get().strip(),
                    min_mark,
                    max_mark,
                    grade_code,
                    grade_name,
                    sort_order,
                )

            if success:
                dialog.destroy()
                if refresh:
                    refresh()
            else:
                messagebox.showerror("Error", msg)

        btn_row = tk.Frame(dialog, bg=dialog.cget("bg"))
        btn_row.pack(fill="x", pady=(18, 20), padx=20)
        tk.Button(
            btn_row,
            text="Cancel",
            bg=LEMON_SOFT,
            fg=TEXT_PRIMARY,
            font=(FF, 10, "bold"),
            padx=18,
            pady=8,
            command=dialog.destroy,
        ).pack(side="left", padx=(0, 8))
        tk.Button(
            btn_row,
            text="Save",
            bg=GREEN,
            fg="white",
            font=(FF, 11),
            padx=20,
            pady=8,
            command=save,
        ).pack(side="left")

    def _apply_grade_band_template(self, class_name, refresh=None):
        class_name = str(class_name or "").strip()
        if not class_name:
            messagebox.showwarning("Missing Class", "Please select a class first.")
            return

        if not messagebox.askyesno(
            "Apply Template",
            f"Apply 8-band template to {class_name}?\n\n"
            "This will replace existing grading bands for this class.",
        ):
            return

        default_bands = [
            (90, 100, "EE1", "Exceeding Expectation 1", 1),
            (76, 89, "EE2", "Exceeding Expectation 2", 2),
            (60, 75, "ME1", "Meeting Expectation 1", 3),
            (40, 59, "ME2", "Meeting Expectation 2", 4),
            (30, 39, "AE1", "Approaching Expectation 1", 5),
            (20, 29, "AE2", "Approaching Expectation 2", 6),
            (10, 19, "BE1", "Below Expectation 1", 7),
            (0, 9, "BE2", "Below Expectation 2", 8),
        ]

        for scale in db.get_grading_scales(class_name):
            db.delete_grading_scale(scale.get("id", ""))

        for min_mark, max_mark, code, name, order in default_bands:
            ok, msg = db.add_grading_scale(
                class_name, min_mark, max_mark, code, name, order
            )
            if not ok:
                messagebox.showerror(
                    "Template Failed", f"Could not add {code} band.\n\n{msg}"
                )
                if refresh:
                    refresh()
                return

        if refresh:
            refresh()
        messagebox.showinfo(
            "Template Applied", f"8-band grading template applied to {class_name}."
        )

    def _edit_grading_scale_dialog(self, tree, class_name):
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("Select", "Please select a grading band to edit")
            return
        scale_row = db.get_grading_scale(selected[0])
        if not scale_row:
            messagebox.showerror("Error", "Could not load the selected grading band")
            return
        self._open_grading_scale_dialog(
            class_name,
            scale_row,
            refresh=lambda: self._load_grading_scale_tree(tree, class_name),
        )

    def _delete_grading_scale(self, tree, class_name):
        selected = tree.selection()
        if not selected:
            self._show_notice(
                "Select Grading Band",
                "Please select a grading band to delete.",
                kind="info",
            )
            return
        if not self._confirm_delete_action("grading band"):
            return
        if db.delete_grading_scale(selected[0]):
            self._load_grading_scale_tree(tree, class_name)
            self._show_delete_result_notice("grading band", 1, 0)
        else:
            self._show_delete_result_notice("grading band", 0, 1)

    def _build_assignments_tab(self, parent):
        """Build teacher assignments tab"""
        # This reuses the teachers page functionality
        container = tk.Frame(parent, bg=CONTENT_BG)
        container.pack(fill="both", expand=True)

        # Subject Teacher Assignments
        tk.Label(
            container,
            text="Subject Teacher Assignments",
            font=(FF, 14, "bold"),
            bg=CONTENT_BG,
            fg=TEXT_PRIMARY,
        ).pack(pady=10)

        subj_frame = tk.Frame(container, bg=CARD_BG, relief="flat", bd=1)
        subj_frame.pack(fill="x", padx=10, pady=5)

        cols = ("teacher", "subject", "class", "stream")
        subj_tree = ttk.Treeview(subj_frame, columns=cols, show="headings", height=8)
        subj_tree.heading("teacher", text="Teacher")
        subj_tree.heading("subject", text="Subject")
        subj_tree.heading("class", text="Class")
        subj_tree.heading("stream", text="Stream")
        subj_tree.column("teacher", width=150)
        subj_tree.column("subject", width=150)
        subj_tree.column("class", width=150)
        subj_tree.column("stream", width=110)
        subj_tree.pack(fill="x", padx=10, pady=10)

        assignments = db.get_subject_teacher_assignments()
        for a in assignments:
            subj_tree.insert(
                "",
                "end",
                values=(
                    self._get_teacher_label(a),
                    self._get_subject_label(
                        a.get("subject", ""), a.get("class_name", "")
                    ),
                    self._get_class_label(a.get("class_name", "")),
                    a.get("stream_name", "") or "Whole Class",
                ),
            )

        # Grade Facilitator Assignments
        tk.Label(
            container,
            text="Grade Facilitator Assignments",
            font=(FF, 14, "bold"),
            bg=CONTENT_BG,
            fg=TEXT_PRIMARY,
        ).pack(pady=10)

        class_frame = tk.Frame(container, bg=CARD_BG, relief="flat", bd=1)
        class_frame.pack(fill="x", padx=10, pady=5)

        cols = ("teacher", "class", "stream")
        class_tree = ttk.Treeview(class_frame, columns=cols, show="headings", height=5)
        class_tree.heading("teacher", text="Teacher")
        class_tree.heading("class", text="Class")
        class_tree.heading("stream", text="Stream")
        class_tree.column("teacher", width=200)
        class_tree.column("class", width=200)
        class_tree.column("stream", width=110)
        class_tree.pack(fill="x", padx=10, pady=10)

        class_assignments = db.get_class_teacher_assignments()
        for a in class_assignments:
            class_tree.insert(
                "",
                "end",
                values=(
                    self._get_teacher_label(a),
                    self._get_class_label(a.get("class_name", "")),
                    a.get("stream_name", "") or "Whole Class",
                ),
            )

        # Buttons
        btn_frame = tk.Frame(container, bg=CONTENT_BG)
        btn_frame.pack(pady=10)

        tk.Button(
            btn_frame,
            text="Assign Subject Teacher",
            bg=BLUE,
            fg="white",
            font=(FF, 10),
            padx=10,
            pady=5,
            command=self._assign_subject_dialog,
        ).pack(side="left", padx=5)

        tk.Button(
            btn_frame,
            text="Assign Grade Facilitator",
            bg=PURPLE,
            fg="white",
            font=(FF, 10),
            padx=10,
            pady=5,
            command=self._assign_class_teacher_dialog,
        ).pack(side="left", padx=5)

    # ==================== DASHBOARD ====================
    def show_dashboard(self):
        self.clear_frame()
        self._set_nav("Dashboard")

        stats = db.get_statistics("One", DEFAULT_EXAM_TYPE)
        dashboard_classes = self.get_current_classes()
        all_students = db.get_all_students()
        recent_students = sorted(
            all_students, key=lambda row: row.get("created_at", ""), reverse=True
        )[:8]
        level = self.current_level

        # Top title and breadcrumb row
        top_row = tk.Frame(self.content_frame, bg=CONTENT_BG)
        top_row.pack(fill="x", pady=(2, 12))
        title_block = tk.Frame(top_row, bg=CONTENT_BG)
        title_block.pack(side="left")
        tk.Label(
            title_block,
            text="Dashboard",
            bg=CONTENT_BG,
            fg=TEXT_PRIMARY,
            font=(FF, 22, "bold"),
        ).pack(anchor="w")
        tk.Label(
            title_block,
            text="Live academic pulse across learners, performance, and enrollment.",
            bg=CONTENT_BG,
            fg=TEXT_SECONDARY,
            font=(FF, 10),
        ).pack(anchor="w", pady=(2, 0))
        crumb = tk.Frame(top_row, bg="#ecefe6", padx=10, pady=6)
        crumb.pack(side="right")
        tk.Label(
            crumb, text="Home", bg="#ecefe6", fg=TEXT_SECONDARY, font=(FF, 9)
        ).pack(side="left")
        tk.Label(crumb, text=" / ", bg="#ecefe6", fg=TEXT_SECONDARY, font=(FF, 9)).pack(
            side="left"
        )
        tk.Label(
            crumb, text="Dashboard", bg="#ecefe6", fg=TEXT_PRIMARY, font=(FF, 9, "bold")
        ).pack(side="left")

        # KPI cards row (preserve olive/lemon tone with RGB variation)
        cards_row = tk.Frame(self.content_frame, bg=CONTENT_BG)
        cards_row.pack(fill="x", pady=(0, 12))
        for i in range(4):
            cards_row.columnconfigure(i, weight=1)

        # Real KPI calculations (recent growth windows from student created_at)
        now = datetime.now()
        recent_28 = 0
        prev_28 = 0
        for student in all_students:
            created_at = str(student.get("created_at", "") or "").strip()
            if not created_at:
                continue
            dt = None
            try:
                dt = datetime.fromisoformat(created_at.replace("Z", "+00:00")).replace(
                    tzinfo=None
                )
            except Exception:
                for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                    try:
                        dt = datetime.strptime(created_at[:19], fmt)
                        break
                    except Exception:
                        continue
            if not dt:
                continue
            days = (now - dt).days
            if 0 <= days <= 27:
                recent_28 += 1
            elif 28 <= days <= 55:
                prev_28 += 1
        growth_pct = (
            int(((recent_28 - prev_28) / prev_28) * 100)
            if prev_28 > 0
            else (100 if recent_28 > 0 else 0)
        )

        metrics = [
            (
                "Total Students",
                str(stats.get("students", 0)),
                f"{growth_pct:+d}% vs previous 28 days",
                "#58c7a6",
                "#2ea886",
                "👥",
            ),
            (
                "New Students",
                str(recent_28),
                "Joined in the last 28 days",
                "#e3c060",
                "#d39f2d",
                "🧑",
            ),
            (
                "Current Classes",
                str(len(dashboard_classes)),
                "Active classes",
                "#67a8e8",
                "#3f7ec1",
                "🏫",
            ),
            (
                "Avg Score",
                str(stats.get("avg_score", 0)),
                f"Term One / {DEFAULT_EXAM_TYPE}",
                "#db6f8b",
                "#c6456d",
                "🎯",
            ),
        ]

        for idx, (title, value, note, soft, strong, icon) in enumerate(metrics):
            outer = tk.Frame(cards_row, bg=strong)
            outer.grid(row=0, column=idx, padx=6, sticky="nsew")
            card = tk.Frame(outer, bg=soft, padx=14, pady=12)
            card.pack(fill="both", expand=True, padx=1, pady=1)

            head = tk.Frame(card, bg=soft)
            head.pack(fill="x")
            tk.Label(
                head,
                text=icon,
                bg=strong,
                fg="white",
                font=(FF, 10, "bold"),
                padx=8,
                pady=4,
            ).pack(side="left")
            tk.Label(head, text=title, bg=soft, fg="white", font=(FF, 10, "bold")).pack(
                side="left", padx=8
            )

            tk.Label(card, text=value, bg=soft, fg="white", font=(FF, 18, "bold")).pack(
                anchor="w", pady=(10, 2)
            )
            tk.Label(card, text=note, bg=soft, fg="#f8fdf9", font=(FF, 8)).pack(
                anchor="w"
            )

            progress = tk.Canvas(card, bg=soft, height=8, highlightthickness=0)
            progress.pack(fill="x", pady=(8, 0))
            progress.create_line(
                2, 4, 240, 4, fill=_mix_hex(strong, "#0f172a", 0.42), width=2
            )
            progress.create_line(2, 4, 130 + (idx * 20), 4, fill="#f5f5f5", width=2)

        # Analytics panels row (like reference: two large charts)
        charts_row = tk.Frame(self.content_frame, bg=CONTENT_BG)
        charts_row.pack(fill="both", expand=True, pady=(2, 4))
        charts_row.columnconfigure(0, weight=1)
        charts_row.columnconfigure(1, weight=1)
        charts_row.rowconfigure(0, weight=1)

        def make_chart_panel(parent, title, theme="sky"):
            bo, bi = _card_colors(theme)
            outer = tk.Frame(parent, bg=bo)
            body = tk.Frame(outer, bg=bi, padx=12, pady=10)
            body.pack(fill="both", expand=True, padx=1, pady=1)
            hdr = tk.Frame(body, bg=bi)
            hdr.pack(fill="x", pady=(0, 6))
            tk.Label(
                hdr, text=title, bg=bi, fg=TEXT_PRIMARY, font=(FF, 12, "bold")
            ).pack(side="left")
            tk.Label(hdr, text="• • •", bg=bi, fg=TEXT_SECONDARY, font=(FF, 10)).pack(
                side="right"
            )
            plot_holder = tk.Frame(body, bg=bi)
            plot_holder.pack(fill="both", expand=True)
            return outer, plot_holder, bi

        panel_left, left_holder, left_bg = make_chart_panel(
            charts_row, "Performance Trend", "azure"
        )
        panel_left.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        panel_right, right_holder, right_bg = make_chart_panel(
            charts_row, "Enrollment Movement", "mint"
        )
        panel_right.grid(row=0, column=1, sticky="nsew", padx=(6, 0))

        # Chart 1: real exam-session performance trend
        exam_points = {}
        results_cache = {}
        for class_name in dashboard_classes:
            history_rows = db.get_class_exam_history(class_name) or []
            for row in history_rows:
                term = str(row.get("term", "") or "")
                exam_type = str(row.get("exam_type", "") or DEFAULT_EXAM_TYPE)
                academic_year = str(row.get("academic_year", "") or datetime.now().year)
                created_at = str(row.get("created_at", "") or "").strip()
                point_key = (academic_year, term, exam_type)
                key_dt = datetime.min
                if created_at:
                    try:
                        key_dt = datetime.fromisoformat(
                            created_at.replace("Z", "+00:00")
                        ).replace(tzinfo=None)
                    except Exception:
                        pass
                cache_key = (class_name, academic_year, term, exam_type)
                if cache_key not in results_cache:
                    class_results = self._get_ranked_results(
                        class_name, term, exam_type, academic_year
                    )
                    class_avg = (
                        round(
                            sum(r.get("average", 0) for r in class_results)
                            / len(class_results),
                            1,
                        )
                        if class_results
                        else None
                    )
                    results_cache[cache_key] = class_avg
                class_avg = results_cache[cache_key]
                if class_avg is None:
                    continue
                if point_key not in exam_points:
                    exam_points[point_key] = {"dt": key_dt, "values": []}
                exam_points[point_key]["dt"] = max(exam_points[point_key]["dt"], key_dt)
                exam_points[point_key]["values"].append(class_avg)

        ordered_points = sorted(
            [
                (key, info["dt"], round(sum(info["values"]) / len(info["values"]), 1))
                for key, info in exam_points.items()
                if info["values"]
            ],
            key=lambda x: x[1],
        )
        if len(ordered_points) > 8:
            ordered_points = ordered_points[-8:]

        if ordered_points:
            labels = [
                f"{item[0][0]} T{self._format_term_number(item[0][1])}-{str(item[0][2]).split('-')[0][:3].upper()}"
                for item in ordered_points
            ]
            y1 = [item[2] for item in ordered_points]
        else:
            labels = ["No Data"]
            y1 = [float(stats.get("avg_score", 0) or 0)]
        y2 = [round((y1[i - 1] if i > 0 else y1[i]), 1) for i in range(len(y1))]
        if (
            ordered_points
            and ordered_points[0][1] != datetime.min
            and ordered_points[-1][1] != datetime.min
        ):
            trend_window = f"{ordered_points[0][1].strftime('%b %Y')} - {ordered_points[-1][1].strftime('%b %Y')}"
        else:
            trend_window = "No dated exam sessions"
        trend_scope = (
            f"Scope: {len(dashboard_classes)} classes | Window: {trend_window}"
        )
        fig1, ax1 = plt.subplots(figsize=(5.6, 3.0))
        fig1.patch.set_facecolor(left_bg)
        ax1.set_facecolor(_mix_hex(left_bg, "#ffffff", 0.45))
        ax1.plot(
            labels,
            y1,
            color="#5561d8",
            linewidth=2.2,
            marker="o",
            markersize=4,
            label="Current",
        )
        ax1.plot(
            labels,
            y2,
            color="#8d96a3",
            linewidth=2.2,
            marker="o",
            markersize=4,
            label="Previous",
        )
        ax1.grid(axis="y", color="#d9dee8", linewidth=0.7)
        ax1.spines["top"].set_visible(False)
        ax1.spines["right"].set_visible(False)
        ax1.tick_params(axis="x", labelsize=8, colors=TEXT_SECONDARY)
        ax1.tick_params(axis="y", labelsize=8, colors=TEXT_SECONDARY)
        ax1.set_ylim(0, 100)
        ax1.set_title(
            "Average Score Trend by Exam Session",
            fontsize=10,
            color=TEXT_PRIMARY,
            pad=10,
        )
        ax1.text(
            0.01,
            1.005,
            trend_scope,
            transform=ax1.transAxes,
            fontsize=8,
            color=TEXT_SECONDARY,
        )
        ax1.legend(
            loc="upper left", bbox_to_anchor=(0.0, 0.98), fontsize=8, frameon=False
        )
        fig1.subplots_adjust(top=0.83, left=0.09, right=0.98, bottom=0.14)
        self.dashboard_chart_1 = FigureCanvasTkAgg(fig1, master=left_holder)
        self.dashboard_chart_1.draw()
        self.dashboard_chart_1.get_tk_widget().pack(fill="both", expand=True)
        plt.close(fig1)

        # Chart 2: real enrollment movement by month
        month_map = {}
        for student in all_students:
            created_at = str(student.get("created_at", "") or "").strip()
            if not created_at:
                continue
            dt = None
            try:
                dt = datetime.fromisoformat(created_at.replace("Z", "+00:00")).replace(
                    tzinfo=None
                )
            except Exception:
                for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                    try:
                        dt = datetime.strptime(created_at[:19], fmt)
                        break
                    except Exception:
                        continue
            if not dt:
                continue
            key = dt.strftime("%Y-%m")
            month_map[key] = month_map.get(key, 0) + 1

        ordered_months = sorted(month_map.items(), key=lambda x: x[0])
        if len(ordered_months) > 8:
            ordered_months = ordered_months[-8:]

        if ordered_months:
            labels2 = [
                datetime.strptime(k, "%Y-%m").strftime("%b") for k, _ in ordered_months
            ]
            y3 = [v for _, v in ordered_months]  # new students
            cumulative = 0
            y4 = []
            for value in y3:
                cumulative += value
                y4.append(cumulative)
            enroll_window = f"{datetime.strptime(ordered_months[0][0], '%Y-%m').strftime('%b %Y')} - {datetime.strptime(ordered_months[-1][0], '%Y-%m').strftime('%b %Y')}"
        else:
            # fallback: class distribution
            labels2 = [self._get_class_label(c) for c in dashboard_classes[:7]] or [
                "No Data"
            ]
            y3 = [len(db.get_students_by_class(c)) for c in dashboard_classes[:7]] or [
                0
            ]
            cumulative = 0
            y4 = []
            for value in y3:
                cumulative += value
                y4.append(cumulative)
            enroll_window = "Fallback class snapshot"
        enroll_scope = f"Scope: {len(all_students)} students | Window: {enroll_window}"
        fig2, ax2 = plt.subplots(figsize=(5.6, 3.0))
        fig2.patch.set_facecolor(right_bg)
        ax2.set_facecolor(_mix_hex(right_bg, "#ffffff", 0.45))
        ax2.plot(labels2, y3, color="#58c7a6", linewidth=2.0, label="New Students")
        ax2.fill_between(labels2, y3, color="#58c7a6", alpha=0.25)
        ax2.plot(labels2, y4, color="#7c8593", linewidth=2.0, label="Cumulative")
        ax2.fill_between(labels2, y4, color="#7c8593", alpha=0.18)
        ax2.grid(axis="y", color="#d9dee8", linewidth=0.7)
        ax2.spines["top"].set_visible(False)
        ax2.spines["right"].set_visible(False)
        ax2.tick_params(axis="x", labelsize=8, colors=TEXT_SECONDARY)
        ax2.tick_params(axis="y", labelsize=8, colors=TEXT_SECONDARY)
        upper = max(y3 + y4 + [5]) + 5
        ax2.set_ylim(0, upper)
        ax2.set_title("Enrollment Movement", fontsize=10, color=TEXT_PRIMARY, pad=10)
        ax2.text(
            0.01,
            1.005,
            enroll_scope,
            transform=ax2.transAxes,
            fontsize=8,
            color=TEXT_SECONDARY,
        )
        ax2.legend(
            loc="upper left", bbox_to_anchor=(0.0, 0.98), fontsize=8, frameon=False
        )
        fig2.subplots_adjust(top=0.83, left=0.09, right=0.98, bottom=0.14)
        self.dashboard_chart_2 = FigureCanvasTkAgg(fig2, master=right_holder)
        self.dashboard_chart_2.draw()
        self.dashboard_chart_2.get_tk_widget().pack(fill="both", expand=True)
        plt.close(fig2)

    # ==================== ENHANCED DASHBOARD ====================
    def show_dashboard(self):
        self.clear_frame()
        self._set_nav("Dashboard")

        stats = db.get_statistics("One", DEFAULT_EXAM_TYPE)
        dashboard_classes = self.get_current_classes()
        all_students = db.get_all_students()
        recent_students = sorted(
            all_students, key=lambda row: row.get("created_at", ""), reverse=True
        )[:8]
        level = self.current_level

        # ── Top title and breadcrumb ──────────────────────────────────────
        top_row = tk.Frame(self.content_frame, bg=CONTENT_BG)
        top_row.pack(fill="x", pady=(2, 12))
        title_block = tk.Frame(top_row, bg=CONTENT_BG)
        title_block.pack(side="left")
        tk.Label(
            title_block,
            text="Dashboard",
            bg=CONTENT_BG,
            fg=TEXT_PRIMARY,
            font=(FF, 22, "bold"),
        ).pack(anchor="w")
        tk.Label(
            title_block,
            text="Live academic pulse across learners, performance, and enrollment.",
            bg=CONTENT_BG,
            fg=TEXT_SECONDARY,
            font=(FF, 10),
        ).pack(anchor="w", pady=(2, 0))
        crumb = tk.Frame(top_row, bg="#ecefe6", padx=10, pady=6)
        crumb.pack(side="right")
        tk.Label(
            crumb, text="Home", bg="#ecefe6", fg=TEXT_SECONDARY, font=(FF, 9)
        ).pack(side="left")
        tk.Label(crumb, text=" / ", bg="#ecefe6", fg=TEXT_SECONDARY, font=(FF, 9)).pack(
            side="left"
        )
        tk.Label(
            crumb, text="Dashboard", bg="#ecefe6", fg=TEXT_PRIMARY, font=(FF, 9, "bold")
        ).pack(side="left")

        # ── KPI CARDS ROW ───────────────────────────────────────────────────
        def add_section_label(text, subtext=""):
            row = tk.Frame(self.content_frame, bg=CONTENT_BG)
            row.pack(fill="x", pady=(0, 6))
            tk.Label(
                row,
                text=text,
                bg=CONTENT_BG,
                fg=TEXT_PRIMARY,
                font=(FF, 11, "bold"),
            ).pack(side="left")
            if subtext:
                tk.Label(
                    row,
                    text=subtext,
                    bg=CONTENT_BG,
                    fg=TEXT_SECONDARY,
                    font=(FF, 8),
                ).pack(side="left", padx=(8, 0))

        add_section_label("Snapshot Metrics", "Current learner, teacher, and performance totals")
        kpi_row = tk.Frame(self.content_frame, bg=CONTENT_BG)
        kpi_row.pack(fill="x", pady=(0, 8))
        for i in range(6):
            kpi_row.columnconfigure(i, weight=1)

        # Calculate growth metrics
        now = datetime.now()
        recent_28 = 0
        prev_28 = 0
        for student in all_students:
            created_at = str(student.get("created_at", "") or "").strip()
            if not created_at:
                continue
            dt = None
            try:
                dt = datetime.fromisoformat(created_at.replace("Z", "+00:00")).replace(
                    tzinfo=None
                )
            except Exception:
                for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                    try:
                        dt = datetime.strptime(created_at[:19], fmt)
                        break
                    except Exception:
                        continue
            if not dt:
                continue
            days = (now - dt).days
            if 0 <= days <= 27:
                recent_28 += 1
            elif 28 <= days <= 55:
                prev_28 += 1

        growth_pct = (
            int(((recent_28 - prev_28) / prev_28) * 100)
            if prev_28 > 0
            else (100 if recent_28 > 0 else 0)
        )

        teacher_stats = db.get_teacher_stats()
        class_stats = db.get_class_stats()
        stream_stats = db.get_stream_stats()
        demo_stats = db.get_student_demographics()

        total_classes = len(class_stats)
        total_streams = stream_stats["total_streams"]
        total_teachers = teacher_stats["total_teachers"]

        perf_stats = db.get_performance_stats()
        pass_rate = perf_stats.get("pass_rate", 0)

        kpi_cards = [
            (
                "Total Students",
                str(stats.get("students", 0)),
                f"{growth_pct:+d}% vs previous 28d",
                "#58c7a6",
                "#2ea886",
                "👥",
            ),
            (
                "New Students",
                str(recent_28),
                "Joined last 28 days",
                "#e3c060",
                "#d39f2d",
                "🧑",
            ),
            (
                "Teachers",
                str(total_teachers),
                f"{teacher_stats['subject_teachers']} subject | {teacher_stats['class_teachers']} class",
                "#67a8e8",
                "#3f7ec1",
                "👨‍🏫",
            ),
            (
                "Classes",
                str(total_classes),
                f"Active classes | {total_streams} streams",
                "#db6f8b",
                "#c6456d",
                "🏫",
            ),
            (
                "Avg Score",
                f"{stats.get('avg_score', 0)}%",
                f"Term One / {DEFAULT_EXAM_TYPE}",
                "#8b7fc7",
                "#6b5ab3",
                "📊",
            ),
            (
                "Pass Rate",
                f"{pass_rate:.1f}%",
                f"{perf_stats.get('pass_count', 0)}/{perf_stats.get('total_students', 0)} students",
                "#f5995a",
                "#d9823e",
                "✅",
            ),
        ]

        for idx, (title, value, note, soft, strong, icon) in enumerate(kpi_cards):
            outer = tk.Frame(kpi_row, bg=strong)
            outer.grid(row=0, column=idx, padx=4, sticky="nsew")
            card = tk.Frame(outer, bg=soft, padx=12, pady=12)
            card.pack(fill="both", expand=True, padx=1, pady=1)

            head = tk.Frame(card, bg=soft)
            head.pack(fill="x")
            tk.Label(
                head,
                text=icon,
                bg=strong,
                fg="white",
                font=(FF, 10, "bold"),
                padx=6,
                pady=2,
            ).pack(side="left")
            tk.Label(
                head,
                text=title.upper(),
                bg=soft,
                fg="#f7fbf8",
                font=(FF, 8, "bold"),
            ).pack(side="left", padx=6)

            tk.Label(
                card, text=value, bg=soft, fg="white", font=(FF, 19, "bold")
            ).pack(
                anchor="w", pady=(8, 2)
            )
            tk.Label(
                card,
                text=note,
                bg=soft,
                fg="#f8fdf9",
                font=(FF, 8),
                wraplength=150,
                justify="left",
            ).pack(anchor="w")
            tk.Label(
                card,
                text="Updated from current school records",
                bg=soft,
                fg=_mix_hex("#ffffff", strong, 0.35),
                font=(FF, 7),
            ).pack(anchor="w", pady=(5, 0))
            prog = tk.Canvas(card, bg=soft, height=10, highlightthickness=0)
            prog.pack(fill="x", pady=(8, 0))
            prog.create_line(
                4, 5, 168, 5, fill=_mix_hex(strong, "#0f172a", 0.48), width=3
            )
            prog.create_line(
                4,
                5,
                min(168, 58 + (idx * 19)),
                5,
                fill="#ffffff",
                width=3,
            )

        # ── SHORTCUT NAVIGATION CARD ROW ────────────────────────────────────
        add_section_label("Quick Actions", "Jump straight into the core school workflows")
        shortcuts_row = tk.Frame(self.content_frame, bg=CONTENT_BG)
        shortcuts_row.pack(fill="x", pady=(0, 8))
        for i in range(5):
            shortcuts_row.columnconfigure(i, weight=1)

        shortcuts = [
            ("Teachers", "👨‍🏫", self.show_teachers, "#3498db"),
            ("Classes", "🏫", self.show_classes, "#2ecc71"),
            ("Students", "👤", self.show_students, "#9b59b6"),
            ("Streams", "📚", self.show_settings_streams, "#e67e22"),
            ("Grade Facilitators", "🎓", self.show_teachers, "#1abc9c"),
        ]

        shortcut_notes = {
            "Teachers": "Staff roster and assignments",
            "Classes": "Levels, classes, and structure",
            "Students": "Admissions and learner records",
            "Streams": "Class stream configuration",
            "Grade Facilitators": "Leadership and academic oversight",
        }

        for idx, (label, icon, command, color) in enumerate(shortcuts):
            outer = tk.Frame(shortcuts_row, bg=color, padx=4, pady=4)
            outer.grid(row=0, column=idx, padx=4, sticky="nsew")
            shell = tk.Frame(outer, bg=_mix_hex(color, "#ffffff", 0.18), padx=10, pady=10)
            shell.pack(fill="both", expand=True, padx=1, pady=1)
            icon_row = tk.Frame(shell, bg=shell.cget("bg"))
            icon_row.pack(fill="x")
            tk.Label(
                icon_row,
                text=icon,
                bg=color,
                fg="white",
                font=(FF, 10, "bold"),
                padx=8,
                pady=3,
            ).pack(side="left")
            tk.Label(
                icon_row,
                text="OPEN",
                bg=shell.cget("bg"),
                fg="#f7fbf8",
                font=(FF, 8, "bold"),
            ).pack(side="right")
            btn = tk.Button(
                shell,
                text=label,
                bg=shell.cget("bg"),
                fg="white",
                activebackground=_mix_hex(color, "#000000", 0.15),
                activeforeground="white",
                font=(FF, 10, "bold"),
                relief="flat",
                anchor="w",
                cursor="hand2",
                command=command,
            )
            btn.pack(fill="x", pady=(10, 2))
            tk.Label(
                shell,
                text=shortcut_notes.get(label, "Quick access"),
                bg=shell.cget("bg"),
                fg=_mix_hex("#ffffff", color, 0.38),
                font=(FF, 8),
                anchor="w",
            ).pack(fill="x")

        # ── MAIN ANALYTICS PANELS ────────────────────────────────────────────
        add_section_label("Academic Signals", "Trend cards built from live exam and enrollment data")
        charts_row = tk.Frame(self.content_frame, bg=CONTENT_BG)
        charts_row.pack(fill="both", expand=True, pady=(2, 4))
        charts_row.columnconfigure(0, weight=3)
        charts_row.columnconfigure(1, weight=2)
        charts_row.rowconfigure(0, weight=1)

        def make_chart_panel(parent, title, theme="sky"):
            bo, bi = _card_colors(theme)
            outer = tk.Frame(parent, bg=bo, highlightthickness=1, highlightbackground=bo)
            body = tk.Frame(outer, bg=bi, padx=14, pady=14)
            body.pack(fill="both", expand=True, padx=1, pady=1)
            hdr = tk.Frame(body, bg=bi)
            hdr.pack(fill="x", pady=(0, 8))
            theme_labels = {
                "azure": "Performance",
                "mint": "Breakdown",
                "sky": "Admissions",
                "peach": "Demographics",
            }
            tk.Label(
                hdr, text=title, bg=bi, fg=TEXT_PRIMARY, font=(FF, 12, "bold")
            ).pack(side="left")
            tk.Label(
                hdr,
                text=theme_labels.get(theme, "Live"),
                bg=_mix_hex(bo, "#ffffff", 0.2),
                fg=TEXT_SECONDARY,
                font=(FF, 8, "bold"),
                padx=8,
                pady=2,
            ).pack(side="right")
            meta_var = tk.StringVar(value="")
            tk.Label(
                body,
                textvariable=meta_var,
                bg=bi,
                fg=TEXT_SECONDARY,
                font=(FF, 8),
                anchor="w",
                justify="left",
                wraplength=460,
            ).pack(fill="x", pady=(0, 8))
            plot_holder = tk.Frame(body, bg=bi, height=250)
            plot_holder.pack(fill="both", expand=True)
            plot_holder.pack_propagate(False)
            return outer, plot_holder, bi, meta_var

        def style_chart_axis(ax, panel_bg, rotation=0):
            ax.set_facecolor(_mix_hex(panel_bg, "#ffffff", 0.56))
            ax.grid(axis="y", color=_mix_hex(panel_bg, "#8b95a7", 0.78), linewidth=0.7)
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
            ax.spines["left"].set_color(_mix_hex(panel_bg, "#5b6472", 0.5))
            ax.spines["bottom"].set_color(_mix_hex(panel_bg, "#5b6472", 0.5))
            ax.tick_params(axis="x", labelsize=8, colors=TEXT_SECONDARY, rotation=rotation)
            ax.tick_params(axis="y", labelsize=8, colors=TEXT_SECONDARY)
            ax.margins(x=0.05)

        # LEFT PANEL: Performance trend + grade distribution
        left_panel = tk.Frame(charts_row, bg=CONTENT_BG)
        left_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 4))
        left_panel.columnconfigure(0, weight=1)
        left_panel.rowconfigure(0, weight=1)
        left_panel.rowconfigure(1, weight=1)

        trend_panel, trend_holder, trend_bg, trend_meta_var = make_chart_panel(
            left_panel, "Score Trend by Exam Session", "azure"
        )
        trend_panel.grid(row=0, column=0, sticky="nsew", pady=(0, 4))

        grade_panel, grade_holder, grade_bg, grade_meta_var = make_chart_panel(
            left_panel, "Grade Distribution (Latest Exam)", "mint"
        )
        grade_panel.grid(row=1, column=0, sticky="nsew")

        # RIGHT PANEL: Enrollment + demographics
        right_panel = tk.Frame(charts_row, bg=CONTENT_BG)
        right_panel.grid(row=0, column=1, sticky="nsew", padx=(4, 0))
        right_panel.columnconfigure(0, weight=1)
        right_panel.rowconfigure(0, weight=3)
        right_panel.rowconfigure(1, weight=2)

        enroll_panel, enroll_holder, enroll_bg, enroll_meta_var = make_chart_panel(
            right_panel, "Enrollment Movement", "sky"
        )
        enroll_panel.grid(row=0, column=0, sticky="nsew", pady=(0, 4))

        gender_panel, gender_holder, gender_bg, gender_meta_var = make_chart_panel(
            right_panel, "Student Gender Distribution", "peach"
        )
        gender_panel.grid(row=1, column=0, sticky="nsew")

        # ── CHART 1: Performance trend (keeping existing logic) ──────────────
        exam_points = {}
        results_cache = {}
        for class_name in dashboard_classes:
            history_rows = db.get_class_exam_history(class_name) or []
            for row in history_rows:
                term = str(row.get("term", "") or "")
                exam_type = str(row.get("exam_type", "") or DEFAULT_EXAM_TYPE)
                academic_year = str(row.get("academic_year", "") or datetime.now().year)
                created_at = str(row.get("created_at", "") or "").strip()
                point_key = (academic_year, term, exam_type)
                key_dt = datetime.min
                if created_at:
                    try:
                        key_dt = datetime.fromisoformat(
                            created_at.replace("Z", "+00:00")
                        ).replace(tzinfo=None)
                    except Exception:
                        pass
                cache_key = (class_name, academic_year, term, exam_type)
                if cache_key not in results_cache:
                    class_results = self._get_ranked_results(
                        class_name, term, exam_type, academic_year
                    )
                    class_avg = (
                        round(
                            sum(r.get("average", 0) for r in class_results)
                            / len(class_results),
                            1,
                        )
                        if class_results
                        else None
                    )
                    results_cache[cache_key] = class_avg
                class_avg = results_cache[cache_key]
                if class_avg is None:
                    continue
                if point_key not in exam_points:
                    exam_points[point_key] = {"dt": key_dt, "values": []}
                exam_points[point_key]["dt"] = max(exam_points[point_key]["dt"], key_dt)
                exam_points[point_key]["values"].append(class_avg)

        ordered_points = sorted(
            [
                (key, info["dt"], round(sum(info["values"]) / len(info["values"]), 1))
                for key, info in exam_points.items()
                if info["values"]
            ],
            key=lambda x: x[1],
        )
        if len(ordered_points) > 8:
            ordered_points = ordered_points[-8:]

        if ordered_points:
            labels1 = [
                f"{item[0][0]} T{self._format_term_number(item[0][1])}-{str(item[0][2]).split('-')[0][:3].upper()}"
                for item in ordered_points
            ]
            y1 = [item[2] for item in ordered_points]
        else:
            labels1 = ["No Data"]
            y1 = [float(stats.get("avg_score", 0) or 0)]

        y_prev = [round((y1[i - 1] if i > 0 else y1[i]), 1) for i in range(len(y1))]

        if (
            ordered_points
            and ordered_points[0][1] != datetime.min
            and ordered_points[-1][1] != datetime.min
        ):
            trend_window = f"{ordered_points[0][1].strftime('%b %Y')} - {ordered_points[-1][1].strftime('%b %Y')}"
        else:
            trend_window = "No dated sessions"
        trend_meta_var.set(
            f"Scope: {len(dashboard_classes)} classes   |   Window: {trend_window}"
        )

        fig1, ax1 = plt.subplots(figsize=(6.4, 3.2))
        fig1.patch.set_facecolor(trend_bg)
        ax1.plot(
            labels1,
            y1,
            color="#4361ee",
            linewidth=2.4,
            marker="o",
            markersize=5,
            label="Current",
        )
        ax1.plot(
            labels1,
            y_prev,
            color="#9d9d9d",
            linewidth=2.0,
            marker="o",
            markersize=4,
            label="Previous",
        )
        style_chart_axis(ax1, trend_bg, rotation=0 if len(labels1) <= 4 else 18)
        ax1.set_ylim(0, 100)
        ax1.set_ylabel("Average (%)", fontsize=9, color=TEXT_SECONDARY)
        ax1.legend(loc="upper left", fontsize=8, frameon=False, ncol=2, handlelength=1.8)
        fig1.subplots_adjust(top=0.9, left=0.1, right=0.98, bottom=0.22)
        self.dashboard_chart_1 = FigureCanvasTkAgg(fig1, master=trend_holder)
        self.dashboard_chart_1.draw()
        self.dashboard_chart_1.get_tk_widget().pack(fill="both", expand=True)
        plt.close(fig1)

        # ── CHART 2: Grade Distribution (bar chart) ───────────────────────────
        grade_dist = db.get_grade_distribution()
        grades_order = ["EE", "ME", "AE", "BE", "IE"]
        grade_counts = [grade_dist["distribution"].get(g, 0) for g in grades_order]
        grade_colors = [GRADE_COLORS[g] for g in grades_order]
        grade_meta_var.set(
            f"Showing {grade_dist['term']} / {grade_dist['exam_type']}   |   Total learners: {sum(grade_counts)}"
        )

        fig2, ax2 = plt.subplots(figsize=(6.1, 2.9))
        fig2.patch.set_facecolor(grade_bg)
        bars = ax2.bar(
            grades_order,
            grade_counts,
            color=grade_colors,
            edgecolor=grade_bg,
            linewidth=1.5,
        )
        for bar, count in zip(bars, grade_counts):
            ax2.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height() + 0.3,
                str(count),
                ha="center",
                va="bottom",
                fontsize=8,
                color=TEXT_SECONDARY,
            )
        ax2.set_xlabel("Grade", fontsize=9, color=TEXT_SECONDARY)
        ax2.set_ylabel("Students", fontsize=9, color=TEXT_SECONDARY)
        ax2.set_title(
            f"Grades – {grade_dist['term']}/{grade_dist['exam_type']}",
            fontsize=11,
            color=TEXT_PRIMARY,
            pad=6,
        )
        style_chart_axis(ax2, grade_bg)
        ax2.set_ylim(0, max(grade_counts + [1]) + 2)
        fig2.subplots_adjust(top=0.9, left=0.1, right=0.98, bottom=0.18)
        self.dashboard_chart_2 = FigureCanvasTkAgg(fig2, master=grade_holder)
        self.dashboard_chart_2.draw()
        self.dashboard_chart_2.get_tk_widget().pack(fill="both", expand=True)
        plt.close(fig2)

        # ── CHART 3: Enrollment Movement ──────────────────────────────────────
        month_map = {}
        for student in all_students:
            created_at = str(student.get("created_at", "") or "").strip()
            if not created_at:
                continue
            dt = None
            try:
                dt = datetime.fromisoformat(created_at.replace("Z", "+00:00")).replace(
                    tzinfo=None
                )
            except Exception:
                for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                    try:
                        dt = datetime.strptime(created_at[:19], fmt)
                        break
                    except Exception:
                        continue
            if not dt:
                continue
            key = dt.strftime("%Y-%m")
            month_map[key] = month_map.get(key, 0) + 1

        ordered_months = sorted(month_map.items(), key=lambda x: x[0])
        if len(ordered_months) > 8:
            ordered_months = ordered_months[-8:]

        if ordered_months:
            labels3 = [
                datetime.strptime(k, "%Y-%m").strftime("%b") for k, _ in ordered_months
            ]
            y_new = [v for _, v in ordered_months]
            cumulative = 0
            y_cum = []
            for val in y_new:
                cumulative += val
                y_cum.append(cumulative)
            window_label = f"{datetime.strptime(ordered_months[0][0], '%Y-%m').strftime('%b %Y')} - {datetime.strptime(ordered_months[-1][0], '%Y-%m').strftime('%b %Y')}"
        else:
            labels3 = [c[:3] for c in dashboard_classes[:7]] or ["No Data"]
            y_new = [
                len(db.get_students_by_class(c)) for c in dashboard_classes[:7]
            ] or [0]
            cumulative = 0
            y_cum = []
            for val in y_new:
                cumulative += val
                y_cum.append(cumulative)
            window_label = "Class distribution fallback"
        enroll_meta_var.set(
            f"Scope: {len(all_students)} students   |   Window: {window_label}"
        )

        fig3, ax3 = plt.subplots(figsize=(5.8, 3.2))
        fig3.patch.set_facecolor(enroll_bg)
        ax3.plot(labels3, y_new, color="#2ecc71", linewidth=2.2, label="New")
        ax3.fill_between(labels3, y_new, color="#2ecc71", alpha=0.2)
        ax3.plot(labels3, y_cum, color="#7f8c8d", linewidth=2.2, label="Total")
        ax3.fill_between(labels3, y_cum, color="#7f8c8d", alpha=0.12)
        style_chart_axis(ax3, enroll_bg)
        upper = max(y_new + y_cum + [5]) + 5
        ax3.set_ylim(0, upper)
        ax3.set_ylabel("Students", fontsize=9, color=TEXT_SECONDARY)
        ax3.legend(loc="upper left", fontsize=8, frameon=False, ncol=2, handlelength=1.8)
        fig3.subplots_adjust(top=0.9, left=0.1, right=0.98, bottom=0.18)
        self.dashboard_chart_3 = FigureCanvasTkAgg(fig3, master=enroll_holder)
        self.dashboard_chart_3.draw()
        self.dashboard_chart_3.get_tk_widget().pack(fill="both", expand=True)
        plt.close(fig3)

        # ── CHART 4: Gender pie chart ─────────────────────────────────────────
        if demo_stats["total"] > 0:
            gender_meta_var.set(
                f"Male: {demo_stats['male']}   |   Female: {demo_stats['female']}   |   Total: {demo_stats['total']}"
            )
            gender_fig, gender_ax = plt.subplots(figsize=(5.0, 2.7))
            gender_fig.patch.set_facecolor(gender_bg)
            labels_gender = ["Male", "Female"]
            sizes = [demo_stats["male"], demo_stats["female"]]
            colors_g = ["#3498db", "#e91e63"]
            wedges, texts, autotexts = gender_ax.pie(
                sizes,
                labels=labels_gender,
                colors=colors_g,
                startangle=90,
                autopct="%1.1f%%",
                textprops={"fontsize": 8, "color": TEXT_PRIMARY},
                wedgeprops={"linewidth": 1, "edgecolor": "white"},
            )
            gender_ax.axis("equal")
            gender_fig.subplots_adjust(top=0.94, left=0.02, right=0.98, bottom=0.08)
            self.dashboard_chart_4 = FigureCanvasTkAgg(gender_fig, master=gender_holder)
            self.dashboard_chart_4.draw()
            self.dashboard_chart_4.get_tk_widget().pack(fill="both", expand=True)
            plt.close(gender_fig)
        else:
            gender_meta_var.set("No student data available for gender split.")
            empty_label = tk.Label(
                gender_holder,
                text="No student data available",
                bg=gender_bg,
                fg=TEXT_SECONDARY,
                font=(FF, 12),
            )
            empty_label.pack(expand=True)

        # ── CLASS SIZE BAR CHART (in Gender panel as stacked?) ──────────────
        # Display class size as horizontal bar in gender panel footer
        if class_stats:
            smallest_class = min(class_stats, key=lambda c: c["student_count"])
            largest_class = max(class_stats, key=lambda c: c["student_count"])
            avg_class_size = sum(c["student_count"] for c in class_stats) / len(
                class_stats
            )

            stats_text = (
                f"Avg class size: {avg_class_size:.1f} students  |  "
                f"Smallest: {smallest_class['name']} ({smallest_class['student_count']})  |  "
                f"Largest: {largest_class['name']} ({largest_class['student_count']})"
            )
            current_meta = str(gender_meta_var.get() or "").strip()
            gender_meta_var.set(
                f"{current_meta}\n{stats_text}" if current_meta else stats_text
            )

    # ==================== QUICK NAVIGATION SHORTCUTS =======================

    def show_teachers(self):
        """Show teacher management page"""
        self.clear_frame()
        self._set_nav("Teachers")
        self._page_header(
            "Teachers Management", "Manage teachers and their subject/class assignments"
        )

        # Toolbar
        toolbar = tk.Frame(self.content_frame, bg=CONTENT_BG)
        toolbar.pack(fill="x", pady=(0, 10))

        tk.Button(
            toolbar,
            text="+ Add Teacher",
            bg=GREEN,
            fg="white",
            font=(FF, 10),
            padx=15,
            pady=5,
            command=lambda: self._open_teacher_dialog(),
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Edit Selected",
            bg=BLUE,
            fg="white",
            font=(FF, 10),
            padx=15,
            pady=5,
            command=lambda: self._edit_teacher_dialog(tree),
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Delete Selected",
            bg="#d9534f",
            fg="white",
            font=(FF, 10),
            padx=15,
            pady=5,
            command=lambda: self._delete_teacher(
                tree, reload_callback=self.show_teachers
            ),
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Select All Filtered",
            bg="#0f766e",
            fg="white",
            font=(FF, 10),
            padx=15,
            pady=5,
            command=lambda: self.teachers_table.select_all_filtered(),
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Clear Selection",
            bg="#64748b",
            fg="white",
            font=(FF, 10),
            padx=15,
            pady=5,
            command=lambda: self.teachers_table.clear_selection(),
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Template",
            bg=ORANGE,
            fg="white",
            font=(FF, 10),
            padx=15,
            pady=5,
            command=self.download_teacher_import_template,
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Export Excel",
            bg=GREEN,
            fg="white",
            font=(FF, 10),
            padx=15,
            pady=5,
            command=self.export_teachers_excel,
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Import Excel",
            bg="#1d4ed8",
            fg="white",
            font=(FF, 10),
            padx=15,
            pady=5,
            command=lambda: self.import_teachers_excel(on_complete=self.show_teachers),
        ).pack(side="left", padx=5)

        tk.Button(
            toolbar,
            text="🗑️ Delete All Teachers",
            bg="#e74c3c",
            fg="white",
            font=(FF, 10),
            padx=15,
            pady=5,
            command=self._delete_all_teachers_batch,
        ).pack(side="left", padx=5)
        tk.Button(
            toolbar,
            text="Refresh",
            bg="#666",
            fg="white",
            font=(FF, 10),
            padx=15,
            pady=5,
            command=self.show_teachers,
        ).pack(side="left", padx=5)

        # Teachers list
        tf_bo, tf_bi = _card_colors("azure")
        tf_outer = tk.Frame(self.content_frame, bg=tf_bo)
        tf_outer.pack(fill="both", expand=True, pady=4)
        teachers_frame = tk.Frame(tf_outer, bg=tf_bi)
        teachers_frame.pack(fill="both", expand=True, padx=1, pady=1)

        self.teachers_table = AdvancedDataTable(
            teachers_frame,
            columns=[
                {"key": "id", "title": "ID", "width": 90, "anchor": "center"},
                {
                    "key": "abbr",
                    "title": "Short Label",
                    "width": 110,
                    "anchor": "center",
                },
                {"key": "name", "title": "Full Name", "width": 200, "anchor": "w"},
                {"key": "username", "title": "Username", "width": 130, "anchor": "w"},
                {"key": "role", "title": "Role", "width": 130, "anchor": "center"},
                {
                    "key": "assignments",
                    "title": "Assignments",
                    "width": 420,
                    "anchor": "w",
                },
            ],
            page_size=12,
            search_label="Search teachers",
            selectmode="extended",
            enable_select_all=True,
        )
        tree = self.teachers_table.tree

        # Load teachers
        teachers = db.get_all_teachers()
        rows = []
        for teacher in teachers:
            # Get assignments
            assignments = []
            subject_assignments = db.get_subject_teacher_assignments()
            class_assignments = db.get_class_teacher_assignments()

            for sa in subject_assignments:
                if sa["teacher_id"] == teacher["id"]:
                    assignments.append(
                        f"{self._get_subject_label(sa['subject'], sa['class_name'])} ({self._format_class_stream_label(sa['class_name'], sa.get('stream_name', ''))})"
                    )

            for ca in class_assignments:
                if ca["teacher_id"] == teacher["id"]:
                    assignments.append(
                        f"Grade Facilitator: {self._format_class_stream_label(ca['class_name'], ca.get('stream_name', ''))}"
                    )

            role_label = (
                "Subject Teacher"
                if teacher.get("role") == "teacher"
                else "Grade Facilitator"
            )
            assignment_text = (
                ", ".join(assignments) if assignments else "No assignments"
            )

            row_values = (
                teacher.get("id", "")[:8],
                teacher.get("abbreviation", "")
                or self._generate_short_label(teacher.get("full_name", ""), "teacher"),
                teacher.get("full_name", ""),
                teacher.get("username", ""),
                role_label,
                assignment_text,
            )
            rows.append(
                {
                    "iid": teacher.get("id", str(uuid.uuid4())),
                    "values": row_values,
                    "value_map": {
                        "id": row_values[0],
                        "abbr": row_values[1],
                        "name": row_values[2],
                        "username": row_values[3],
                        "role": row_values[4],
                        "assignments": row_values[5],
                    },
                    "search": " ".join(str(v) for v in row_values),
                }
            )
        self.teachers_table.set_rows(rows)
        tree.bind("<Double-1>", lambda e: self._edit_teacher_dialog(tree))

        # Action buttons
        action_frame = tk.Frame(teachers_frame, bg=tf_bi)
        action_frame.pack(fill="x", padx=10, pady=(0, 10))

        tk.Button(
            action_frame,
            text="Assign Subject",
            bg=BLUE,
            fg="white",
            font=(FF, 10),
            padx=10,
            pady=5,
            command=self._assign_subject_dialog,
        ).pack(side="left", padx=5)

        tk.Button(
            action_frame,
            text="Assign Grade Facilitator",
            bg=PURPLE,
            fg="white",
            font=(FF, 10),
            padx=10,
            pady=5,
            command=self._assign_class_teacher_dialog,
        ).pack(side="left", padx=5)

        tk.Button(
            action_frame,
            text="Delete",
            bg="#e74c3c",
            fg="white",
            font=(FF, 10),
            padx=10,
            pady=5,
            command=lambda: self._delete_teacher(
                tree, reload_callback=self.show_teachers
            ),
        ).pack(side="left", padx=5)

    def _add_teacher_dialog(self):
        self._open_teacher_dialog()

    def _assign_subject_dialog(self, on_saved=None):
        """Dialog to assign a subject to a teacher"""
        teachers = db.get_all_teachers()
        if not teachers:
            messagebox.showwarning("No Teachers", "Please add teachers first")
            return
        classes = [row.get("name") for row in db.get_all_classes()]
        if not classes:
            classes = self.get_current_classes()

        dialog = tk.Toplevel(self.root)
        dialog.title("Assign Subject to Teacher")
        dialog.geometry("430x450")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)

        tk.Label(dialog, text="Select Teacher:", font=(FF, 11)).pack(pady=(20, 5))
        teacher_names = [
            f"{self._get_teacher_label(t)} - {t['full_name']} ({t['username']})"
            for t in teachers
        ]
        teacher_var = tk.StringVar()
        teacher_cb = ttk.Combobox(
            dialog,
            textvariable=teacher_var,
            values=teacher_names,
            state="readonly",
            font=(FF, 10),
        )
        teacher_cb.pack(fill="x", padx=20)

        tk.Label(dialog, text="Select Class:", font=(FF, 11)).pack(pady=(15, 5))
        class_var = tk.StringVar()
        class_cb = ttk.Combobox(
            dialog,
            textvariable=class_var,
            values=classes,
            state="readonly",
            font=(FF, 10),
        )
        class_cb.pack(fill="x", padx=20)

        tk.Label(dialog, text="Select Stream:", font=(FF, 11)).pack(pady=(15, 5))
        stream_var = tk.StringVar(value="Whole Class")
        stream_cb = ttk.Combobox(
            dialog,
            textvariable=stream_var,
            values=["Whole Class"],
            state="readonly",
            font=(FF, 10),
        )
        stream_cb.pack(fill="x", padx=20)

        tk.Label(dialog, text="Select Subject:", font=(FF, 11)).pack(pady=(15, 5))
        subject_var = tk.StringVar()
        subject_cb = ttk.Combobox(
            dialog,
            textvariable=subject_var,
            values=self.get_current_subjects(),
            state="readonly",
            font=(FF, 10),
        )
        subject_cb.pack(fill="x", padx=20)

        def refresh_subject_options(event=None):
            selected_class = class_var.get().strip()
            values = (
                self._get_subjects_for_selected_class(selected_class, TERMS[0])
                if selected_class
                else self.get_current_subjects()
            )
            subject_cb.configure(values=values)
            if subject_var.get() not in values:
                subject_var.set(values[0] if values else "")
            stream_values = self._get_assignment_stream_options(selected_class)
            stream_cb.configure(values=stream_values)
            if stream_var.get() not in stream_values:
                stream_var.set(stream_values[0] if stream_values else "Whole Class")

        class_cb.bind("<<ComboboxSelected>>", refresh_subject_options)

        def save_assignment():
            if not all([teacher_var.get(), class_var.get(), subject_var.get()]):
                messagebox.showerror("Error", "All fields are required")
                return

            # Get selected teacher ID
            selected_idx = teacher_cb.current()
            teacher_id = teachers[selected_idx]["id"]
            stream_name = (
                "" if stream_var.get() == "Whole Class" else stream_var.get().strip()
            )

            success = db.assign_subject_teacher(
                teacher_id, class_var.get(), subject_var.get(), stream_name
            )
            if success:
                messagebox.showinfo("Success", "Subject assigned successfully!")
                dialog.destroy()
                if callable(on_saved):
                    on_saved()
                else:
                    self.show_teachers()
            else:
                messagebox.showerror("Error", "Failed to assign subject")

        btn_row = tk.Frame(dialog, bg=dialog.cget("bg"))
        btn_row.pack(fill="x", pady=(18, 20), padx=20)
        tk.Button(
            btn_row,
            text="Cancel",
            bg=LEMON_SOFT,
            fg=TEXT_PRIMARY,
            font=(FF, 10, "bold"),
            padx=18,
            pady=8,
            command=dialog.destroy,
        ).pack(side="left", padx=(0, 8))
        tk.Button(
            btn_row,
            text="Assign Subject",
            bg=BLUE,
            fg="white",
            font=(FF, 11),
            padx=20,
            pady=8,
            command=save_assignment,
        ).pack(side="left")

    def _assign_class_teacher_dialog(self, on_saved=None):
        """Dialog to assign a Grade Facilitator"""
        teachers = db.get_all_teachers()
        if not teachers:
            messagebox.showwarning("No Teachers", "Please add teachers first")
            return
        classes = [row.get("name") for row in db.get_all_classes()]
        if not classes:
            classes = self.get_current_classes()

        dialog = tk.Toplevel(self.root)
        dialog.title("Assign Grade Facilitator")
        dialog.geometry("430x390")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)

        tk.Label(dialog, text="Select Teacher:", font=(FF, 11)).pack(pady=(20, 5))
        teacher_names = [
            f"{self._get_teacher_label(t)} - {t['full_name']} ({t['username']})"
            for t in teachers
        ]
        teacher_var = tk.StringVar()
        teacher_cb = ttk.Combobox(
            dialog,
            textvariable=teacher_var,
            values=teacher_names,
            state="readonly",
            font=(FF, 10),
        )
        teacher_cb.pack(fill="x", padx=20)

        tk.Label(dialog, text="Select Class:", font=(FF, 11)).pack(pady=(15, 5))
        class_var = tk.StringVar()
        class_cb = ttk.Combobox(
            dialog,
            textvariable=class_var,
            values=classes,
            state="readonly",
            font=(FF, 10),
        )
        class_cb.pack(fill="x", padx=20)

        tk.Label(dialog, text="Select Stream:", font=(FF, 11)).pack(pady=(15, 5))
        stream_var = tk.StringVar(value="Whole Class")
        stream_cb = ttk.Combobox(
            dialog,
            textvariable=stream_var,
            values=["Whole Class"],
            state="readonly",
            font=(FF, 10),
        )
        stream_cb.pack(fill="x", padx=20)

        def refresh_stream_options(event=None):
            stream_values = self._get_assignment_stream_options(class_var.get().strip())
            stream_cb.configure(values=stream_values)
            if stream_var.get() not in stream_values:
                stream_var.set(stream_values[0] if stream_values else "Whole Class")

        class_cb.bind("<<ComboboxSelected>>", refresh_stream_options)

        def save_assignment():
            if not all([teacher_var.get(), class_var.get()]):
                messagebox.showerror("Error", "All fields are required")
                return

            selected_idx = teacher_cb.current()
            teacher_id = teachers[selected_idx]["id"]
            stream_name = (
                "" if stream_var.get() == "Whole Class" else stream_var.get().strip()
            )

            success = db.assign_class_teacher(teacher_id, class_var.get(), stream_name)
            if success:
                messagebox.showinfo(
                    "Success", "Grade Facilitator assigned successfully!"
                )
                dialog.destroy()
                if callable(on_saved):
                    on_saved()
                else:
                    self.show_teachers()
            else:
                messagebox.showerror("Error", "Failed to assign Grade Facilitator")

        refresh_stream_options()

        btn_row = tk.Frame(dialog, bg=dialog.cget("bg"))
        btn_row.pack(fill="x", pady=(18, 20), padx=20)
        tk.Button(
            btn_row,
            text="Cancel",
            bg=LEMON_SOFT,
            fg=TEXT_PRIMARY,
            font=(FF, 10, "bold"),
            padx=18,
            pady=8,
            command=dialog.destroy,
        ).pack(side="left", padx=(0, 8))
        tk.Button(
            btn_row,
            text="Assign Grade Facilitator",
            bg=PURPLE,
            fg="white",
            font=(FF, 11),
            padx=20,
            pady=8,
            command=save_assignment,
        ).pack(side="left")

    def _delete_teacher(self, tree, reload_callback=None):
        selected = (
            self.teachers_table.get_selected_iids()
            if hasattr(self, "teachers_table")
            else list(tree.selection())
        )
        if not selected:
            self._show_notice(
                "Select Teachers",
                "Please select one or more teachers to delete.",
                kind="info",
            )
            return

        if not self._confirm_delete_action(
            "teacher",
            len(selected),
            scope="selected",
            details="This will also remove the selected teachers' assignments.",
        ):
            return

        failed = 0
        for teacher_id in selected:
            if not db.delete_user(teacher_id):
                failed += 1

        if reload_callback:
            reload_callback()
        else:
            self.show_teachers()

        self._show_delete_result_notice(
            "teacher", len(selected) - failed, failed, duration_ms=4200
        )

    # ─────────────────────────── BATCH DELETION METHODS ───────────────────────────
    def _delete_all_marks_for_term(self):
        """Delete all marks for selected term and exam type with confirmation."""
        if not hasattr(self, "marks_term_cb") or not hasattr(self, "marks_exam_cb"):
            messagebox.showwarning("Not Available", "Navigate to Enter Marks first")
            return

        term = self.marks_term_cb.get()
        exam_type = self.marks_exam_cb.get()

        year = self.marks_year_cb.get() if hasattr(self, "marks_year_cb") else str(datetime.now().year)
        if not term or not exam_type:
            messagebox.showwarning("Select Values", "Please select Year, Term and Exam Type")
            return

        if not self._confirm_delete_action(
            "mark record",
            1,
            scope="all",
            details=f"Year: {year}\nTerm: {term}\nExam Type: {exam_type}",
        ):
            return

        success, message = db.clear_all_marks(
            term, exam_type, academic_year=year
        )
        if success:
            self._show_notice(
                "Marks Deleted", message, kind="success", duration_ms=4200
            )
            self._load_marks_table()
        else:
            self._show_notice("Delete Failed", message, kind="error", duration_ms=4200)

    def _delete_all_teachers_batch(self):
        """Delete ALL teachers with confirmation."""
        teachers = db.get_all_teachers()
        if not teachers:
            self._show_notice(
                "No Teachers", "No teachers to delete.", kind="info", duration_ms=2800
            )
            return

        teacher_count = len(teachers)
        if not self._confirm_delete_action(
            "teacher",
            teacher_count,
            scope="all",
            details="This will remove all teacher assignments.",
        ):
            return

        success, message = db.delete_all_teachers()
        if success:
            self._show_notice(
                "Teachers Deleted", message, kind="success", duration_ms=4200
            )
            self.show_teachers()
        else:
            self._show_notice("Delete Failed", message, kind="error", duration_ms=4200)

    def _delete_all_subjects_batch(self):
        """Delete ALL subjects with confirmation."""
        subjects = db.get_subjects_by_level()
        if not subjects:
            self._show_notice(
                "No Subjects", "No subjects to delete.", kind="info", duration_ms=2800
            )
            return

        subject_count = len(subjects)
        if not self._confirm_delete_action(
            "subject",
            subject_count,
            scope="all",
            details="This will remove associated marks and teacher assignments.",
        ):
            return

        success, message = db.delete_all_subjects()
        if success:
            self._show_notice(
                "Subjects Deleted", message, kind="success", duration_ms=4200
            )
            self.show_settings(
                initial_tab="subjects", nav_label="Subjects", show_tabs=False
            )
        else:
            self._show_notice("Delete Failed", message, kind="error", duration_ms=4200)

    def _delete_all_students_batch(self):
        """Delete ALL students with confirmation."""
        students = db.get_all_students()
        if not students:
            self._show_notice(
                "No Students", "No students to delete.", kind="info", duration_ms=2800
            )
            return

        student_count = len(students)
        if not self._confirm_delete_action(
            "student",
            student_count,
            scope="all",
            details="This will also remove all marks linked to those students.",
        ):
            return

        success, message = db.delete_all_students()
        if success:
            self._show_notice(
                "Students Deleted", message, kind="success", duration_ms=4200
            )
            self.students_tab.refresh_students()
        else:
            self._show_notice("Delete Failed", message, kind="error", duration_ms=4200)

    def _delete_all_classes_batch(self):
        """Delete ALL classes with confirmation."""
        classes = db.get_all_classes()
        if not classes:
            self._show_notice(
                "No Classes", "No classes to delete.", kind="info", duration_ms=2800
            )
            return

        class_count = len(classes)
        if not self._confirm_delete_action(
            "class",
            class_count,
            scope="all",
            details="This will remove all students, marks, streams, and assignments tied to those classes.",
        ):
            return

        success, message = db.delete_all_classes()
        if success:
            self._show_notice(
                "Classes Deleted", message, kind="success", duration_ms=4200
            )
            self.show_settings(
                initial_tab="classes", nav_label="Classes", show_tabs=False
            )
        else:
            self._show_notice("Delete Failed", message, kind="error", duration_ms=4200)

    def _reset_all_data_batch(self):
        """Complete database reset with multiple confirmations."""
        if not messagebox.askyesno(
            "⚠️ WARNING - Complete Reset",
            "This will DELETE EVERYTHING:\n\n• ALL students\n• ALL marks\n• ALL classes & streams\n• ALL subjects\n• ALL teachers & assignments\n\n(Admin account will remain)",
        ):
            return

        if not messagebox.askyesno(
            "Final Confirmation", "Are you ABSOLUTELY sure?\n\nThis cannot be undone."
        ):
            return

        success, message = db.reset_all_data()
        if success:
            messagebox.showinfo("Success", message)
            self.show_dashboard()
        else:
            messagebox.showerror("Error", message)

    # ==================== Grade Facilitator VIEWS ====================
    def show_class_students(self):
        """Show Grade Facilitator their assigned students"""
        if not self.current_user:
            messagebox.showerror("Error", "Please login first")
            return

        teacher_id = self.current_user.get("id")
        classes = db.get_teacher_classes(teacher_id)

        if not classes:
            messagebox.showwarning(
                "No Class", "You are not assigned as a Grade Facilitator for any class"
            )
            return

        self.clear_frame()
        self._set_nav("My Students")
        self._page_header("My Students", f"Grade Facilitator for: {classes[0]}")

        # Class selector
        ctrl = tk.Frame(self.content_frame, bg=CONTENT_BG)
        ctrl.pack(fill="x", pady=(0, 10))

        tk.Label(
            ctrl, text="Class:", bg=CONTENT_BG, fg=TEXT_SECONDARY, font=(FF, 10)
        ).pack(side="left", padx=5)
        class_var = tk.StringVar(value=classes[0])
        class_cb = ttk.Combobox(
            ctrl,
            textvariable=class_var,
            values=classes,
            state="readonly",
            font=(FF, 10),
        )
        class_cb.pack(side="left", padx=5)
        class_cb.bind(
            "<<ComboboxSelected>>",
            lambda e: self._update_class_streams_and_students(
                class_var.get(), stream_var, stream_cb
            ),
        )

        tk.Label(
            ctrl,
            text="Stream:",
            bg=CONTENT_BG,
            fg=TEXT_SECONDARY,
            font=(FF, 10),
        ).pack(side="left", padx=(16, 5))
        stream_var = tk.StringVar(value="All Streams")
        stream_cb = ttk.Combobox(
            ctrl,
            textvariable=stream_var,
            values=["All Streams"],
            state="readonly",
            font=(FF, 10),
            width=18,
        )
        stream_cb.pack(side="left", padx=5)
        stream_cb.bind(
            "<<ComboboxSelected>>",
            lambda e: self._load_class_students(class_var.get(), stream_var.get()),
        )

        print_btn = tk.Button(
            ctrl,
            text="🖨️ Print Class List",
            bg=GREEN,
            fg="white",
            font=(FF, 10, "bold"),
            padx=14,
            pady=6,
            command=lambda: self._print_class_students_pdf(
                class_var.get(), stream_var.get()
            ),
            cursor="hand2",
        )
        print_btn.pack(side="left", padx=(8, 0))

        # Students list
        list_frame = tk.Frame(self.content_frame, bg=CARD_BG, relief="flat", bd=1)
        list_frame.pack(fill="both", expand=True)

        self.class_students_table = AdvancedDataTable(
            list_frame,
            columns=[
                {
                    "key": "adm",
                    "title": "Admission No",
                    "width": 140,
                    "anchor": "center",
                },
                {"key": "name", "title": "Name", "width": 240, "anchor": "w"},
                {"key": "gender", "title": "Gender", "width": 100, "anchor": "center"},
            ],
            page_size=15,
            search_label="Search learners",
        )
        self.class_students_tree = self.class_students_table.tree
        self._update_class_streams_and_students(classes[0], stream_var, stream_cb)

    def _update_class_streams_and_students(self, class_name, stream_var, stream_cb):
        stream_var.set("All Streams")
        stream_values = ["All Streams"]
        class_row = db.get_class_by_name(class_name)
        if class_row:
            streams = db.get_streams_for_class(class_row["id"])
            stream_values.extend([s.get("name", "") for s in streams if s.get("name")])
        stream_cb["values"] = stream_values
        if stream_var.get() not in stream_values:
            stream_var.set("All Streams")
        self._load_class_students(class_name, stream_var.get())

    def _load_class_students(self, class_name, stream_name=""):
        """Load students for a specific class"""
        rows = []
        if stream_name and stream_name != "All Streams":
            students = db.get_students_by_class_and_stream(class_name, stream_name)
        else:
            students = db.get_students_by_class(class_name)
        for s in students:
            values = (s.get("admission_no", ""), s.get("name", ""), s.get("gender", ""))
            rows.append(
                {
                    "iid": s.get("id", str(uuid.uuid4())),
                    "values": values,
                    "value_map": {
                        "adm": values[0],
                        "name": values[1],
                        "gender": values[2],
                    },
                    "search": " ".join(str(v) for v in values),
                }
            )
        if hasattr(self, "class_students_table"):
            self.class_students_table.set_rows(rows)

    def _print_class_students_pdf(self, class_name, stream_name=""):
        if not class_name:
            messagebox.showwarning(
                "Print Error",
                "Please select a class before printing the student list.",
            )
            return

        timestamp = datetime.now().strftime("%Y%m%d-%H%M")
        filename = f"Student_List_{class_name.replace(' ', '_')}"
        if stream_name and stream_name != "All Streams":
            filename += f"_{stream_name.replace(' ', '_')}"
        filename += f"_{timestamp}.pdf"

        file_path = filedialog.asksaveasfilename(
            title="Save Class Student List PDF",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=filename,
        )
        if not file_path:
            return

        if self.generate_student_list_pdf(
            class_name, stream_name if stream_name != "All Streams" else "", file_path
        ):
            messagebox.showinfo("Saved", f"Class student list saved to {file_path}")

    def show_add_comments(self):
        """Show page for adding student comments"""
        if not self.current_user:
            messagebox.showerror("Error", "Please login first")
            return

        teacher_id = self.current_user.get("id")
        classes = db.get_teacher_classes(teacher_id)

        if not classes:
            messagebox.showwarning(
                "No Class", "You are not assigned as a Grade Facilitator for any class"
            )
            return

        self.clear_frame()
        self._set_nav("Add Comments")
        self._page_header("Add Comments", "Add comments for students in your class")

        # Controls
        ctrl = tk.Frame(self.content_frame, bg=CONTENT_BG)
        ctrl.pack(fill="x", pady=(0, 10))

        tk.Label(
            ctrl, text="Class:", bg=CONTENT_BG, fg=TEXT_SECONDARY, font=(FF, 10)
        ).pack(side="left", padx=5)
        class_var = tk.StringVar(value=classes[0])
        class_cb = ttk.Combobox(
            ctrl,
            textvariable=class_var,
            values=classes,
            state="readonly",
            font=(FF, 10),
        )
        class_cb.pack(side="left", padx=5)

        tk.Label(
            ctrl, text="Term:", bg=CONTENT_BG, fg=TEXT_SECONDARY, font=(FF, 10)
        ).pack(side="left", padx=15)
        term_var = tk.StringVar(value="One")
        term_cb = ttk.Combobox(
            ctrl, textvariable=term_var, values=TERMS, state="readonly", font=(FF, 10)
        )
        term_cb.pack(side="left", padx=5)

        tk.Label(
            ctrl, text="Year:", bg=CONTENT_BG, fg=TEXT_SECONDARY, font=(FF, 10)
        ).pack(side="left", padx=15)
        year_var = tk.StringVar(value=str(datetime.now().year))
        year_cb = ttk.Combobox(
            ctrl,
            textvariable=year_var,
            values=self._get_year_options(),
            state="readonly",
            font=(FF, 10),
            width=10,
        )
        year_cb.pack(side="left", padx=5)

        tk.Button(
            ctrl,
            text="Load Students",
            bg=BLUE,
            fg="white",
            font=(FF, 10),
            command=lambda: self._load_students_for_comments(
                class_var.get(), term_var.get(), year_var.get()
            ),
        ).pack(side="left", padx=15)

        # Students with comments
        list_frame = tk.Frame(self.content_frame, bg=CARD_BG, relief="flat", bd=1)
        list_frame.pack(fill="both", expand=True)

        self.comments_table = AdvancedDataTable(
            list_frame,
            columns=[
                {
                    "key": "adm",
                    "title": "Admission No",
                    "width": 140,
                    "anchor": "center",
                },
                {"key": "name", "title": "Name", "width": 220, "anchor": "w"},
                {
                    "key": "comment",
                    "title": "Current Comment",
                    "width": 450,
                    "anchor": "w",
                },
            ],
            page_size=15,
            search_label="Search comments",
        )
        self.comments_tree = self.comments_table.tree
        self.comments_class_var = class_var
        self.comments_term_var = term_var
        self.comments_year_var = year_var

    def _load_students_for_comments(self, class_name, term, academic_year=None):
        """Load students for adding comments"""
        academic_year = str(academic_year or datetime.now().year)
        rows = []
        students = db.get_students_by_class(class_name)
        for s in students:
            # Get existing comment
            comment_data = db.get_student_comment(s["id"], term, academic_year)
            comment = comment_data.get("comment_text", "") if comment_data else ""

            values = (s.get("admission_no", ""), s.get("name", ""), comment)
            rows.append(
                {
                    "iid": s.get("id", str(uuid.uuid4())),
                    "values": values,
                    "tags": (s.get("id", ""),),
                    "value_map": {
                        "adm": values[0],
                        "name": values[1],
                        "comment": values[2],
                    },
                    "search": " ".join(str(v) for v in values),
                }
            )
        if hasattr(self, "comments_table"):
            self.comments_table.set_rows(rows)

        # Add double-click to edit
        self.comments_tree.bind(
            "<Double-1>",
            lambda e: self._edit_comment(
                self.comments_tree, class_name, term, academic_year
            ),
        )

    def _edit_comment(self, tree, class_name, term, academic_year=None):
        """Edit comment for selected student"""
        academic_year = str(academic_year or datetime.now().year)
        selected = tree.selection()
        if not selected:
            return

        item = tree.item(selected)
        student_id = item["tags"][0]
        current_comment = item["values"][2]

        # Dialog to edit comment
        dialog = tk.Toplevel(self.root)
        dialog.title("Add/Edit Comment")
        dialog.geometry("540x360")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)

        tk.Label(dialog, text="Comment:", font=(FF, 11)).pack(pady=(20, 5))
        comment_text = tk.Text(dialog, font=(FF, 11), height=8)
        comment_text.pack(fill="both", expand=True, padx=20, pady=10)
        comment_text.insert("1.0", current_comment)

        def save():
            text = comment_text.get("1.0", "end").strip()
            if db.save_comment(
                student_id, self.current_user["id"], term, text, academic_year
            ):
                messagebox.showinfo("Success", "Comment saved!")
                dialog.destroy()
                self._load_students_for_comments(class_name, term, academic_year)
            else:
                messagebox.showerror("Error", "Failed to save comment")

        btn_row = tk.Frame(dialog, bg=dialog.cget("bg"))
        btn_row.pack(fill="x", pady=(0, 14), padx=20)
        tk.Button(
            btn_row,
            text="Cancel",
            bg=LEMON_SOFT,
            fg=TEXT_PRIMARY,
            font=(FF, 10, "bold"),
            padx=18,
            pady=8,
            command=dialog.destroy,
        ).pack(side="left", padx=(0, 8))
        tk.Button(
            btn_row,
            text="Save Comment",
            bg=GREEN,
            fg="white",
            font=(FF, 11),
            padx=20,
            pady=8,
            command=save,
        ).pack(side="left")

    # ==================== PROMOTIONS ====================
    def show_promotions(self):
        """Show student promotions management interface."""
        self.clear_frame()
        self._set_nav("Promotions")
        self._page_header(
            "Student Promotions", "Manage student promotions to next class/grade"
        )

        # Main container with two columns
        main_container = tk.Frame(self.content_frame, bg=CONTENT_BG)
        main_container.pack(fill="both", expand=True, padx=10, pady=10)
        main_container.columnconfigure(0, weight=1)
        main_container.columnconfigure(1, weight=1)

        # Left column - Promotion Settings
        left_frame = tk.Frame(main_container, bg=CARD_BG, relief="flat", bd=1)
        left_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5))

        tk.Label(
            left_frame,
            text="Promotion Settings",
            bg=CARD_BG,
            fg=TEXT_PRIMARY,
            font=(FF, 14, "bold"),
        ).pack(anchor="w", padx=15, pady=(15, 10))

        # Settings form
        settings_frame = tk.Frame(left_frame, bg=CARD_BG)
        settings_frame.pack(fill="x", padx=15, pady=5)

        # Get current settings
        settings = promotion_manager.get_settings()

        # Promotion date
        tk.Label(
            settings_frame,
            text="Promotion Date (MM-DD):",
            bg=CARD_BG,
            fg=TEXT_SECONDARY,
            font=(FF, 10),
        ).pack(anchor="w", pady=(10, 2))
        self.promo_date_var = tk.StringVar(
            value=settings.get("promotion_date", "12-01")
        )
        promo_date_entry = tk.Entry(
            settings_frame, textvariable=self.promo_date_var, font=(FF, 10), width=20
        )
        promo_date_entry.pack(anchor="w", pady=(0, 10))

        # Minimum passing average
        tk.Label(
            settings_frame,
            text="Minimum Passing Average (%):",
            bg=CARD_BG,
            fg=TEXT_SECONDARY,
            font=(FF, 10),
        ).pack(anchor="w", pady=(10, 2))
        self.min_avg_var = tk.StringVar(
            value=settings.get("min_passing_average", "50.0")
        )
        min_avg_entry = tk.Entry(
            settings_frame, textvariable=self.min_avg_var, font=(FF, 10), width=20
        )
        min_avg_entry.pack(anchor="w", pady=(0, 10))

        tk.Label(
            settings_frame,
            text="Academic Year Override (optional):",
            bg=CARD_BG,
            fg=TEXT_SECONDARY,
            font=(FF, 10),
        ).pack(anchor="w", pady=(10, 2))
        self.promo_year_var = tk.StringVar(
            value=settings.get("promotion_academic_year", "")
        )
        promo_year_entry = tk.Entry(
            settings_frame, textvariable=self.promo_year_var, font=(FF, 10), width=20
        )
        promo_year_entry.pack(anchor="w", pady=(0, 10))

        # Auto-promote enabled
        self.auto_promote_var = tk.BooleanVar(
            value=settings.get("auto_promote_enabled", "false").lower() == "true"
        )
        auto_promote_cb = tk.Checkbutton(
            settings_frame,
            text="Enable Auto-Promotion",
            variable=self.auto_promote_var,
            bg=CARD_BG,
            fg=TEXT_PRIMARY,
            font=(FF, 10),
            selectcolor=CARD_BG,
        )
        auto_promote_cb.pack(anchor="w", pady=(10, 10))

        # Save settings button
        tk.Button(
            settings_frame,
            text="Save Settings",
            bg=GREEN,
            fg="white",
            font=(FF, 10, "bold"),
            padx=15,
            pady=5,
            command=self._save_promotion_settings,
        ).pack(anchor="w", pady=(10, 0))

        # Promotion status
        status_frame = tk.Frame(left_frame, bg=CARD_BG)
        status_frame.pack(fill="x", padx=15, pady=15)

        is_due = promotion_manager.is_promotion_due()
        academic_year = promotion_manager.get_current_academic_year()

        tk.Label(
            status_frame,
            text=f"Current Academic Year:",
            bg=CARD_BG,
            fg=TEXT_SECONDARY,
            font=(FF, 10),
        ).pack(anchor="w")
        tk.Label(
            status_frame,
            text=academic_year,
            bg=CARD_BG,
            fg=TEXT_PRIMARY,
            font=(FF, 11, "bold"),
        ).pack(anchor="w", pady=(0, 10))

        tk.Label(
            status_frame,
            text=f"Promotion Due:",
            bg=CARD_BG,
            fg=TEXT_SECONDARY,
            font=(FF, 10),
        ).pack(anchor="w")
        status_color = "#2ecc71" if is_due else "#e74c3c"
        tk.Label(
            status_frame,
            text="Yes" if is_due else "No",
            bg=CARD_BG,
            fg=status_color,
            font=(FF, 11, "bold"),
        ).pack(anchor="w", pady=(0, 10))

        tk.Label(
            status_frame,
            text="Configured Trigger Date:",
            bg=CARD_BG,
            fg=TEXT_SECONDARY,
            font=(FF, 10),
        ).pack(anchor="w")
        tk.Label(
            status_frame,
            text=settings.get("promotion_date", "12-01"),
            bg=CARD_BG,
            fg=TEXT_PRIMARY,
            font=(FF, 11, "bold"),
        ).pack(anchor="w", pady=(0, 10))

        # Right column - Promotion Actions
        right_frame = tk.Frame(main_container, bg=CARD_BG, relief="flat", bd=1)
        right_frame.grid(row=0, column=1, sticky="nsew", padx=(5, 0))

        tk.Label(
            right_frame,
            text="Promotion Actions",
            bg=CARD_BG,
            fg=TEXT_PRIMARY,
            font=(FF, 14, "bold"),
        ).pack(anchor="w", padx=15, pady=(15, 10))

        # Action buttons
        actions_frame = tk.Frame(right_frame, bg=CARD_BG)
        actions_frame.pack(fill="x", padx=15, pady=5)

        tk.Button(
            actions_frame,
            text="Preview Promotions",
            bg=BLUE,
            fg="white",
            font=(FF, 10, "bold"),
            padx=15,
            pady=8,
            command=self._preview_promotions,
        ).pack(fill="x", pady=(0, 10))

        tk.Button(
            actions_frame,
            text="Execute Promotions",
            bg=GREEN,
            fg="white",
            font=(FF, 10, "bold"),
            padx=15,
            pady=8,
            command=self._execute_promotions,
        ).pack(fill="x", pady=(0, 10))

        tk.Button(
            actions_frame,
            text="View Promotion History",
            bg=PURPLE,
            fg="white",
            font=(FF, 10, "bold"),
            padx=15,
            pady=8,
            command=self._view_promotion_history,
        ).pack(fill="x", pady=(0, 10))

        tk.Button(
            actions_frame,
            text="View Audit Log",
            bg=ORANGE,
            fg="white",
            font=(FF, 10, "bold"),
            padx=15,
            pady=8,
            command=self._view_promotion_audit_log,
        ).pack(fill="x", pady=(0, 10))

        # Statistics
        stats_frame = tk.Frame(right_frame, bg=CARD_BG)
        stats_frame.pack(fill="x", padx=15, pady=15)

        stats = promotion_manager.get_promotion_statistics()

        tk.Label(
            stats_frame,
            text="Promotion Statistics",
            bg=CARD_BG,
            fg=TEXT_PRIMARY,
            font=(FF, 12, "bold"),
        ).pack(anchor="w", pady=(0, 10))

        tk.Label(
            stats_frame,
            text=f"Total Promotions: {stats.get('total_promotions', 0)}",
            bg=CARD_BG,
            fg=TEXT_SECONDARY,
            font=(FF, 10),
        ).pack(anchor="w")
        tk.Label(
            stats_frame,
            text=f"Promoted: {stats.get('promoted_count', 0)}",
            bg=CARD_BG,
            fg=TEXT_SECONDARY,
            font=(FF, 10),
        ).pack(anchor="w")
        tk.Label(
            stats_frame,
            text=f"Repeating: {stats.get('repeating_count', 0)}",
            bg=CARD_BG,
            fg=TEXT_SECONDARY,
            font=(FF, 10),
        ).pack(anchor="w")

    def _save_promotion_settings(self):
        """Save promotion settings."""
        settings = {
            "promotion_date": self.promo_date_var.get().strip(),
            "min_passing_average": self.min_avg_var.get().strip(),
            "promotion_academic_year": self.promo_year_var.get().strip(),
            "auto_promote_enabled": "true" if self.auto_promote_var.get() else "false",
        }

        success, message = promotion_manager.update_settings(settings)
        if success:
            messagebox.showinfo("Success", message)
            self.show_promotions()
        else:
            messagebox.showerror("Error", message)

    def _preview_promotions(self):
        """Preview which students will be promoted or repeat."""
        preview = promotion_manager.get_promotion_preview()

        # Create preview window
        preview_window = tk.Toplevel(self.root)
        preview_window.title("Promotion Preview")
        preview_window.geometry("800x600")
        preview_window.configure(bg=CONTENT_BG)

        # Title
        tk.Label(
            preview_window,
            text="Promotion Preview",
            bg=CONTENT_BG,
            fg=TEXT_PRIMARY,
            font=(FF, 16, "bold"),
        ).pack(pady=(20, 10))

        # Summary
        summary_frame = tk.Frame(preview_window, bg=CARD_BG, relief="flat", bd=1)
        summary_frame.pack(fill="x", padx=20, pady=10)

        tk.Label(
            summary_frame,
            text="Summary",
            bg=CARD_BG,
            fg=TEXT_PRIMARY,
            font=(FF, 12, "bold"),
        ).pack(anchor="w", padx=15, pady=(10, 5))

        tk.Label(
            summary_frame,
            text=f"Eligible for Promotion: {len(preview['eligible'])}",
            bg=CARD_BG,
            fg=TEXT_SECONDARY,
            font=(FF, 10),
        ).pack(anchor="w", padx=15)
        tk.Label(
            summary_frame,
            text=f"Will Repeat: {len(preview['repeating'])}",
            bg=CARD_BG,
            fg=TEXT_SECONDARY,
            font=(FF, 10),
        ).pack(anchor="w", padx=15)
        tk.Label(
            summary_frame,
            text=f"No Exam Data: {len(preview['no_data'])}",
            bg=CARD_BG,
            fg=TEXT_SECONDARY,
            font=(FF, 10),
        ).pack(anchor="w", padx=15, pady=(0, 10))
        tk.Label(
            summary_frame,
            text=f"Final Class / No Next Class: {len(preview['terminal'])}",
            bg=CARD_BG,
            fg=TEXT_SECONDARY,
            font=(FF, 10),
        ).pack(anchor="w", padx=15)
        tk.Label(
            summary_frame,
            text=f"Already Processed This Year: {len(preview['already_processed'])}",
            bg=CARD_BG,
            fg=TEXT_SECONDARY,
            font=(FF, 10),
        ).pack(anchor="w", padx=15, pady=(0, 10))

        # Notebook for tabs
        notebook = ttk.Notebook(preview_window)
        notebook.pack(fill="both", expand=True, padx=20, pady=10)

        # Eligible students tab
        eligible_frame = tk.Frame(notebook, bg=CONTENT_BG)
        notebook.add(eligible_frame, text=f"Eligible ({len(preview['eligible'])})")

        if preview["eligible"]:
            cols = ("name", "admission_no", "current_class", "next_class", "average")
            tree = ttk.Treeview(
                eligible_frame, columns=cols, show="headings", style="App.Treeview"
            )
            tree.heading("name", text="Name")
            tree.heading("admission_no", text="Admission No")
            tree.heading("current_class", text="Current Class")
            tree.heading("next_class", text="Next Class")
            tree.heading("average", text="Average")

            for student in preview["eligible"]:
                tree.insert(
                    "",
                    "end",
                    values=(
                        student.get("name", ""),
                        student.get("admission_no", ""),
                        student.get("current_class", ""),
                        student.get("next_class", ""),
                        f"{student.get('average_marks', 0):.1f}%",
                    ),
                )

            tree.pack(fill="both", expand=True, padx=10, pady=10)
        else:
            tk.Label(
                eligible_frame,
                text="No students eligible for promotion",
                bg=CONTENT_BG,
                fg=TEXT_SECONDARY,
                font=(FF, 10),
            ).pack(pady=50)

        # Repeating students tab
        repeating_frame = tk.Frame(notebook, bg=CONTENT_BG)
        notebook.add(repeating_frame, text=f"Repeating ({len(preview['repeating'])})")

        if preview["repeating"]:
            cols = ("name", "admission_no", "current_class", "average")
            tree = ttk.Treeview(
                repeating_frame, columns=cols, show="headings", style="App.Treeview"
            )
            tree.heading("name", text="Name")
            tree.heading("admission_no", text="Admission No")
            tree.heading("current_class", text="Current Class")
            tree.heading("average", text="Average")

            for student in preview["repeating"]:
                tree.insert(
                    "",
                    "end",
                    values=(
                        student.get("name", ""),
                        student.get("admission_no", ""),
                        student.get("current_class", ""),
                        f"{student.get('average_marks', 0):.1f}%",
                    ),
                )

            tree.pack(fill="both", expand=True, padx=10, pady=10)
        else:
            tk.Label(
                repeating_frame,
                text="No students repeating",
                bg=CONTENT_BG,
                fg=TEXT_SECONDARY,
                font=(FF, 10),
            ).pack(pady=50)

        no_data_frame = tk.Frame(notebook, bg=CONTENT_BG)
        notebook.add(no_data_frame, text=f"No Data ({len(preview['no_data'])})")

        if preview["no_data"]:
            cols = ("name", "admission_no", "current_class")
            tree = ttk.Treeview(
                no_data_frame, columns=cols, show="headings", style="App.Treeview"
            )
            tree.heading("name", text="Name")
            tree.heading("admission_no", text="Admission No")
            tree.heading("current_class", text="Current Class")

            for student in preview["no_data"]:
                tree.insert(
                    "",
                    "end",
                    values=(
                        student.get("name", ""),
                        student.get("admission_no", ""),
                        student.get("current_class", ""),
                    ),
                )

            tree.pack(fill="both", expand=True, padx=10, pady=10)
        else:
            tk.Label(
                no_data_frame,
                text="All students have marks for the latest exam session",
                bg=CONTENT_BG,
                fg=TEXT_SECONDARY,
                font=(FF, 10),
            ).pack(pady=50)

        terminal_frame = tk.Frame(notebook, bg=CONTENT_BG)
        notebook.add(terminal_frame, text=f"Final Class ({len(preview['terminal'])})")

        if preview["terminal"]:
            cols = ("name", "admission_no", "current_class", "average", "reason")
            tree = ttk.Treeview(
                terminal_frame, columns=cols, show="headings", style="App.Treeview"
            )
            tree.heading("name", text="Name")
            tree.heading("admission_no", text="Admission No")
            tree.heading("current_class", text="Current Class")
            tree.heading("average", text="Average")
            tree.heading("reason", text="Reason")

            for student in preview["terminal"]:
                tree.insert(
                    "",
                    "end",
                    values=(
                        student.get("name", ""),
                        student.get("admission_no", ""),
                        student.get("current_class", ""),
                        f"{float(student.get('average_marks') or 0):.1f}%",
                        student.get("reason", ""),
                    ),
                )

            tree.pack(fill="both", expand=True, padx=10, pady=10)
        else:
            tk.Label(
                terminal_frame,
                text="Every passing class has a configured next class",
                bg=CONTENT_BG,
                fg=TEXT_SECONDARY,
                font=(FF, 10),
            ).pack(pady=50)

        processed_frame = tk.Frame(notebook, bg=CONTENT_BG)
        notebook.add(
            processed_frame, text=f"Processed ({len(preview['already_processed'])})"
        )

        if preview["already_processed"]:
            cols = (
                "name",
                "admission_no",
                "current_class",
                "status",
                "to_class",
                "promotion_date",
            )
            tree = ttk.Treeview(
                processed_frame, columns=cols, show="headings", style="App.Treeview"
            )
            tree.heading("name", text="Name")
            tree.heading("admission_no", text="Admission No")
            tree.heading("current_class", text="Current Class")
            tree.heading("status", text="Status")
            tree.heading("to_class", text="Processed To")
            tree.heading("promotion_date", text="Processed On")

            for student in preview["already_processed"]:
                tree.insert(
                    "",
                    "end",
                    values=(
                        student.get("name", ""),
                        student.get("admission_no", ""),
                        student.get("current_class", ""),
                        student.get("status", ""),
                        student.get("to_class", ""),
                        (student.get("promotion_date", "") or "")[:10],
                    ),
                )

            tree.pack(fill="both", expand=True, padx=10, pady=10)
        else:
            tk.Label(
                processed_frame,
                text="No students have been processed for this academic year yet",
                bg=CONTENT_BG,
                fg=TEXT_SECONDARY,
                font=(FF, 10),
            ).pack(pady=50)

        # Close button
        tk.Button(
            preview_window,
            text="Close",
            bg=GREEN,
            fg="white",
            font=(FF, 10, "bold"),
            padx=20,
            pady=8,
            command=preview_window.destroy,
        ).pack(pady=20)

    def _execute_promotions(self):
        """Execute student promotions."""
        preview = promotion_manager.get_promotion_preview()
        total_to_process = len(preview["eligible"]) + len(preview["repeating"])

        if total_to_process == 0:
            messagebox.showinfo(
                "Nothing To Process",
                "No students are ready for batch promotion.\n\n"
                f"No data: {len(preview['no_data'])}\n"
                f"Final class / no next class: {len(preview['terminal'])}\n"
                f"Already processed this year: {len(preview['already_processed'])}",
            )
            return

        # Confirm execution
        if not messagebox.askyesno(
            "Confirm Promotion",
            "Are you sure you want to execute promotions?\n\n"
            f"Promote: {len(preview['eligible'])}\n"
            f"Repeat: {len(preview['repeating'])}\n"
            f"No data (skipped): {len(preview['no_data'])}\n"
            f"Final class / no next class (skipped): {len(preview['terminal'])}\n"
            f"Already processed this year (skipped): {len(preview['already_processed'])}\n\n"
            "This will update student classes and cannot be easily undone.",
        ):
            return

        # Get current user ID
        user_id = self.current_user.get("id") if self.current_user else None

        # Execute promotion
        success, message, results = promotion_manager.execute_promotion(
            class_name=None, performed_by=user_id, dry_run=False
        )

        if success:
            messagebox.showinfo(
                "Success",
                f"Promotions completed successfully!\n\n"
                f"Promoted: {results.get('promoted', 0)}\n"
                f"Repeating: {results.get('repeating', 0)}\n"
                f"No data skipped: {results.get('no_data', 0)}\n"
                f"Final class skipped: {results.get('terminal', 0)}\n"
                f"Already processed skipped: {results.get('already_processed', 0)}\n"
                f"Failed: {results.get('failed', 0)}\n"
                f"Batch ID: {results.get('batch_id', 'n/a')}",
            )
            # Refresh the page
            self.show_promotions()
        else:
            errors = results.get("errors", [])[:5] if isinstance(results, dict) else []
            error_text = "\n".join(errors)
            messagebox.showerror(
                "Error",
                f"Promotion failed: {message}"
                + (f"\n\nDetails:\n{error_text}" if error_text else ""),
            )

    def _view_promotion_history(self):
        """View promotion history."""
        history = promotion_manager.get_promotion_history()

        # Create history window
        history_window = tk.Toplevel(self.root)
        history_window.title("Promotion History")
        history_window.geometry("900x600")
        history_window.configure(bg=CONTENT_BG)

        # Title
        tk.Label(
            history_window,
            text="Promotion History",
            bg=CONTENT_BG,
            fg=TEXT_PRIMARY,
            font=(FF, 16, "bold"),
        ).pack(pady=(20, 10))

        # Treeview
        cols = (
            "student_name",
            "admission_no",
            "from_class",
            "to_class",
            "status",
            "academic_year",
            "promotion_date",
            "reason",
        )
        tree = ttk.Treeview(
            history_window, columns=cols, show="headings", style="App.Treeview"
        )
        tree.heading("student_name", text="Student Name")
        tree.heading("admission_no", text="Admission No")
        tree.heading("from_class", text="From Class")
        tree.heading("to_class", text="To Class")
        tree.heading("status", text="Status")
        tree.heading("academic_year", text="Academic Year")
        tree.heading("promotion_date", text="Promotion Date")
        tree.heading("reason", text="Reason")

        for record in history:
            tree.insert(
                "",
                "end",
                values=(
                    record.get("student_name", ""),
                    record.get("admission_no", ""),
                    record.get("from_class", ""),
                    record.get("to_class", ""),
                    record.get("status", ""),
                    record.get("academic_year", ""),
                    record.get("promotion_date", "")[:10]
                    if record.get("promotion_date")
                    else "",
                    record.get("reason", ""),
                ),
            )

        tree.pack(fill="both", expand=True, padx=20, pady=10)

        # Close button
        tk.Button(
            history_window,
            text="Close",
            bg=GREEN,
            fg="white",
            font=(FF, 10, "bold"),
            padx=20,
            pady=8,
            command=history_window.destroy,
        ).pack(pady=20)

    def _view_promotion_audit_log(self):
        """View promotion audit log."""
        audit_log = promotion_manager.get_promotion_audit_log(limit=100)

        # Create audit log window
        audit_window = tk.Toplevel(self.root)
        audit_window.title("Promotion Audit Log")
        audit_window.geometry("900x600")
        audit_window.configure(bg=CONTENT_BG)

        # Title
        tk.Label(
            audit_window,
            text="Promotion Audit Log",
            bg=CONTENT_BG,
            fg=TEXT_PRIMARY,
            font=(FF, 16, "bold"),
        ).pack(pady=(20, 10))

        # Treeview
        cols = ("batch_id", "action", "details", "performed_by", "performed_at")
        tree = ttk.Treeview(
            audit_window, columns=cols, show="headings", style="App.Treeview"
        )
        tree.heading("batch_id", text="Batch ID")
        tree.heading("action", text="Action")
        tree.heading("details", text="Details")
        tree.heading("performed_by", text="Performed By")
        tree.heading("performed_at", text="Performed At")

        for log in audit_log:
            tree.insert(
                "",
                "end",
                values=(
                    log.get("promotion_batch_id", ""),
                    log.get("action", ""),
                    log.get("details", ""),
                    log.get("performed_by_name", "System"),
                    log.get("performed_at", "")[:19] if log.get("performed_at") else "",
                ),
            )

        tree.pack(fill="both", expand=True, padx=20, pady=10)

        # Close button
        tk.Button(
            audit_window,
            text="Close",
            bg=GREEN,
            fg="white",
            font=(FF, 10, "bold"),
            padx=20,
            pady=8,
            command=audit_window.destroy,
        ).pack(pady=20)

    # ==================== EXAM ANALYTICS ====================
    def show_exam_analytics(self):
        """Show exam analytics and comparison interface - Enhanced UI with full visibility."""
        self.clear_frame()
        self._set_nav("Exam Analytics")
        self._page_header("Exam Analytics", "Compare exam types and analyze deviations")

        # Main container
        main_container = tk.Frame(self.content_frame, bg=CONTENT_BG)
        main_container.pack(fill="both", expand=True, padx=10, pady=10)

        # ============ FILTERS SECTION ============
        filter_frame = tk.Frame(main_container, bg=CARD_BG, relief="flat", bd=1)
        filter_frame.pack(fill="x", pady=(0, 12))

        tk.Label(
            filter_frame,
            text="⚙ Analysis Filters",
            bg=CARD_BG,
            fg=TEXT_PRIMARY,
            font=(FF, 12, "bold"),
        ).pack(anchor="w", padx=15, pady=(12, 10))

        # Filter controls row
        controls_frame = tk.Frame(filter_frame, bg=CARD_BG)
        controls_frame.pack(fill="x", padx=15, pady=5)

        # Class filter
        tk.Label(
            controls_frame, text="Class:", bg=CARD_BG, fg=TEXT_SECONDARY, font=(FF, 9)
        ).grid(row=0, column=0, sticky="w", padx=(0, 8))
        self.analytics_class_var = tk.StringVar(value="All Classes")
        class_options = ["All Classes"] + [c["name"] for c in db.get_all_classes()]
        self.analytics_class_cb = ttk.Combobox(
            controls_frame,
            textvariable=self.analytics_class_var,
            values=class_options,
            state="readonly",
            width=18,
            style="App.TCombobox",
        )
        self.analytics_class_cb.grid(row=0, column=1, padx=(0, 20))

        # Stream filter
        tk.Label(
            controls_frame, text="Stream:", bg=CARD_BG, fg=TEXT_SECONDARY, font=(FF, 9)
        ).grid(row=0, column=2, sticky="w", padx=(0, 8))
        self.analytics_stream_var = tk.StringVar(value="All Streams")
        self.analytics_stream_cb = ttk.Combobox(
            controls_frame,
            textvariable=self.analytics_stream_var,
            values=["All Streams"],
            state="readonly",
            width=14,
            style="App.TCombobox",
        )
        self.analytics_stream_cb.grid(row=0, column=3, padx=(0, 20))

        # Term filter
        tk.Label(
            controls_frame, text="Term:", bg=CARD_BG, fg=TEXT_SECONDARY, font=(FF, 9)
        ).grid(row=0, column=4, sticky="w", padx=(0, 8))
        self.analytics_term_var = tk.StringVar(value="All Terms")
        term_options = ["All Terms", "One", "Two", "Three"]
        ttk.Combobox(
            controls_frame,
            textvariable=self.analytics_term_var,
            values=term_options,
            state="readonly",
            width=14,
            style="App.TCombobox",
        ).grid(row=0, column=5, padx=(0, 20))

        # Exam type filter
        tk.Label(
            controls_frame,
            text="Exam Type:",
            bg=CARD_BG,
            fg=TEXT_SECONDARY,
            font=(FF, 9),
        ).grid(row=0, column=6, sticky="w", padx=(0, 8))
        self.analytics_exam_type_var = tk.StringVar(value="All Types")
        exam_type_options = ["All Types"] + self._get_ordered_exam_type_options(
            include_available_sessions=True
        )
        ttk.Combobox(
            controls_frame,
            textvariable=self.analytics_exam_type_var,
            values=exam_type_options,
            state="readonly",
            width=14,
            style="App.TCombobox",
        ).grid(row=0, column=7, padx=(0, 20))

        # Action buttons
        btn_frame = tk.Frame(filter_frame, bg=CARD_BG)
        btn_frame.pack(fill="x", padx=15, pady=(10, 12))
        self.analytics_class_cb.bind(
            "<<ComboboxSelected>>", lambda e: self._refresh_exam_analytics_streams()
        )
        self._refresh_exam_analytics_streams()

        tk.Button(
            btn_frame,
            text="▶ Run Analysis",
            bg=GREEN,
            fg="white",
            font=(FF, 10, "bold"),
            padx=18,
            pady=6,
            command=self._run_exam_analysis,
        ).pack(side="left", padx=(0, 12))
        tk.Button(
            btn_frame,
            text="📊 Generate Report",
            bg=BLUE,
            fg="white",
            font=(FF, 10, "bold"),
            padx=18,
            pady=6,
            command=self._generate_analytics_report,
        ).pack(side="left", padx=(0, 12))
        tk.Button(
            btn_frame,
            text="💾 Export Charts",
            bg=PURPLE,
            fg="white",
            font=(FF, 10, "bold"),
            padx=18,
            pady=6,
            command=self._export_analytics_charts,
        ).pack(side="left")

        # ============ KPI CARDS SECTION ============
        kpi_frame = tk.Frame(main_container, bg=CONTENT_BG)
        kpi_frame.pack(fill="x", pady=(0, 12))

        # KPI Cards (will be populated after analysis)
        self.analytics_kpi_frame = tk.Frame(kpi_frame, bg=CONTENT_BG)
        self.analytics_kpi_frame.pack(fill="x")

        # Placeholder KPI cards
        placeholder_text = tk.Label(
            self.analytics_kpi_frame,
            text="Run analysis to see key metrics",
            bg=CONTENT_BG,
            fg=TEXT_SECONDARY,
            font=(FF, 10),
            pady=15,
        )
        placeholder_text.pack(fill="x")

        # ============ SCROLLABLE RESULTS SECTION ============
        # Create scrollable canvas for all results
        canvas_frame = tk.Frame(main_container, bg=CONTENT_BG)
        canvas_frame.pack(fill="both", expand=True)

        canvas = tk.Canvas(canvas_frame, bg=CONTENT_BG, highlightthickness=0)
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=CONTENT_BG)

        scrollable_frame.bind(
            "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        analytics_window = canvas.create_window(
            (0, 0), window=scrollable_frame, anchor="nw"
        )
        canvas.configure(yscrollcommand=scrollbar.set)

        def _sync_analytics_width(event):
            canvas.itemconfigure(analytics_window, width=event.width)

        def _scroll_widget(widget, event, horizontal=False):
            delta = int(-1 * (event.delta / 120))
            if horizontal:
                widget.xview_scroll(delta, "units")
            else:
                widget.yview_scroll(delta, "units")
            return "break"

        _install_canvas_mousewheel(canvas)
        canvas.bind("<Configure>", _sync_analytics_width)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # ============ RESULTS TABS ============
        tabs_outer = tk.Frame(scrollable_frame, bg="#d1d5db", padx=1, pady=1)
        tabs_outer.pack(fill="x", padx=0, pady=(0, 0))
        tabs_bar = tk.Frame(tabs_outer, bg="#e5e7eb")
        tabs_bar.pack(fill="x", padx=1, pady=1)

        content_outer = tk.Frame(scrollable_frame, bg="#d1d5db", padx=1, pady=1)
        content_outer.pack(fill="x", padx=0, pady=(0, 10))
        content_stack = tk.Frame(content_outer, bg=CONTENT_BG, height=480)
        content_stack.pack(fill="x", padx=1, pady=1)
        content_stack.pack_propagate(False)

        analytics_tab_defs = [
            ("summary", "📋 Analysis Summary", "#34495e"),
            ("metrics", "📈 Deviation Metrics", "#2563eb"),
            ("subjects", "🎯 Subject Deviation Matrix", "#0f766e"),
        ]
        analytics_tab_buttons = {}
        analytics_tab_frames = {}
        neutral_bg = "#475569"

        for key, _label, _color in analytics_tab_defs:
            frame = tk.Frame(content_stack, bg=CARD_BG)
            frame.place(relx=0, rely=0, relwidth=1, relheight=1)
            frame.grid_rowconfigure(1, weight=1)
            frame.grid_columnconfigure(0, weight=1)
            analytics_tab_frames[key] = frame

        # ---- Summary Tab ----
        summary_frame = analytics_tab_frames["summary"]
        summary_header = tk.Frame(summary_frame, bg=CARD_BG)
        summary_header.grid(row=0, column=0, sticky="ew", padx=12, pady=(12, 4))
        tk.Label(
            summary_header,
            text="📋 Analysis Summary",
            bg=CARD_BG,
            fg=TEXT_PRIMARY,
            font=(FF, 11, "bold"),
        ).pack(anchor="w")
        tk.Label(
            summary_header,
            text="Overall findings and patterns",
            bg=CARD_BG,
            fg=TEXT_SECONDARY,
            font=(FF, 9),
        ).pack(anchor="w")

        summary_body = tk.Frame(summary_frame, bg=CARD_BG)
        summary_body.grid(row=1, column=0, sticky="nsew", padx=12, pady=(4, 12))
        summary_body.grid_rowconfigure(0, weight=1)
        summary_body.grid_columnconfigure(0, weight=1)

        self.analytics_summary_text = tk.Text(
            summary_body,
            wrap="word",
            font=(FF, 10),
            bg="white",
            fg=TEXT_PRIMARY,
            height=22,
            bd=0,
            relief="flat",
            padx=12,
            pady=10,
        )
        summary_scroll = ttk.Scrollbar(
            summary_body, orient="vertical", command=self.analytics_summary_text.yview
        )
        self.analytics_summary_text.configure(yscrollcommand=summary_scroll.set)
        self.analytics_summary_text.grid(row=0, column=0, sticky="nsew")
        summary_scroll.grid(row=0, column=1, sticky="ns", padx=(8, 0))
        self.analytics_summary_text.bind(
            "<MouseWheel>",
            lambda event: _scroll_widget(self.analytics_summary_text, event),
        )
        self.analytics_summary_text.insert(
            "1.0", 'Click "Run Analysis" to generate exam analytics...'
        )
        self.analytics_summary_text.config(state="disabled")

        # Add text tags for better formatting
        self.analytics_summary_text.tag_configure(
            "header", font=(FF, 11, "bold"), foreground="#2c3e50"
        )
        self.analytics_summary_text.tag_configure(
            "section",
            font=(FF, 10, "bold"),
            foreground="#34495e",
            spacing1=6,
            spacing2=2,
        )
        self.analytics_summary_text.tag_configure(
            "positive", foreground="#27ae60", font=(FF, 10, "bold")
        )
        self.analytics_summary_text.tag_configure(
            "warning", foreground="#e67e22", font=(FF, 10, "bold")
        )
        self.analytics_summary_text.tag_configure(
            "critical", foreground="#e74c3c", font=(FF, 10, "bold")
        )

        # ---- Deviation Metrics Tab ----
        metrics_frame = analytics_tab_frames["metrics"]
        metric_header = tk.Frame(metrics_frame, bg=CARD_BG)
        metric_header.grid(row=0, column=0, sticky="ew", padx=12, pady=(12, 4))
        tk.Label(
            metric_header,
            text="📈 Deviation Metrics",
            bg=CARD_BG,
            fg=TEXT_PRIMARY,
            font=(FF, 11, "bold"),
        ).pack(anchor="w")
        tk.Label(
            metric_header,
            text="Key performance indicators",
            bg=CARD_BG,
            fg=TEXT_SECONDARY,
            font=(FF, 9),
        ).pack(anchor="w")

        cols = ("metric", "value", "severity")
        metrics_body = tk.Frame(metrics_frame, bg=CARD_BG)
        metrics_body.grid(row=1, column=0, sticky="nsew", padx=12, pady=(4, 12))
        metrics_body.grid_rowconfigure(0, weight=1)
        metrics_body.grid_columnconfigure(0, weight=1)

        self.analytics_tree = ttk.Treeview(
            metrics_body, columns=cols, show="headings", style="App.Treeview", height=18
        )
        self.analytics_tree.heading("metric", text="Metric")
        self.analytics_tree.heading("value", text="Value")
        self.analytics_tree.heading("severity", text="Status")
        self.analytics_tree.column("metric", width=220, stretch=True)
        self.analytics_tree.column("value", width=110, anchor="center", stretch=False)
        self.analytics_tree.column(
            "severity", width=140, anchor="center", stretch=False
        )

        metrics_scroll = ttk.Scrollbar(
            metrics_body, orient="vertical", command=self.analytics_tree.yview
        )
        self.analytics_tree.configure(yscrollcommand=metrics_scroll.set)
        self.analytics_tree.grid(row=0, column=0, sticky="nsew")
        metrics_scroll.grid(row=0, column=1, sticky="ns", padx=(8, 0))
        self.analytics_tree.bind(
            "<MouseWheel>", lambda event: _scroll_widget(self.analytics_tree, event)
        )
        self.analytics_tree.tag_configure(
            "low", background="#eef9f1", foreground="#1f6f43"
        )
        self.analytics_tree.tag_configure(
            "medium", background="#fff6e6", foreground="#9a6100"
        )
        self.analytics_tree.tag_configure(
            "high", background="#fff0ea", foreground="#b64926"
        )
        self.analytics_tree.tag_configure(
            "critical", background="#fdecec", foreground="#9f1f1f"
        )

        # ---- Subject Deviation Matrix Tab ----
        subject_frame = analytics_tab_frames["subjects"]
        subject_header = tk.Frame(subject_frame, bg=CARD_BG)
        subject_header.grid(row=0, column=0, sticky="ew", padx=12, pady=(12, 4))
        tk.Label(
            subject_header,
            text="🎯 Subject Deviation Matrix",
            bg=CARD_BG,
            fg=TEXT_PRIMARY,
            font=(FF, 11, "bold"),
        ).pack(anchor="w")
        tk.Label(
            subject_header,
            text="Performance changes across subjects and exam types",
            bg=CARD_BG,
            fg=TEXT_SECONDARY,
            font=(FF, 9),
        ).pack(anchor="w")

        self.analytics_chart_frame = tk.Frame(subject_frame, bg=CARD_BG)
        self.analytics_chart_frame.grid(
            row=1, column=0, sticky="nsew", padx=12, pady=(4, 12)
        )
        self.analytics_chart_frame.grid_rowconfigure(0, weight=1)
        self.analytics_chart_frame.grid_columnconfigure(0, weight=1)

        subject_cols = (
            "subject",
            "pair",
            "baseline",
            "comparison",
            "deviation",
            "pass_change",
        )
        self.analytics_subject_tree = ttk.Treeview(
            self.analytics_chart_frame,
            columns=subject_cols,
            show="headings",
            style="App.Treeview",
            height=18,
        )
        self.analytics_subject_tree.heading("subject", text="Subject")
        self.analytics_subject_tree.heading("pair", text="Exam Pair")
        self.analytics_subject_tree.heading("baseline", text="Baseline")
        self.analytics_subject_tree.heading("comparison", text="Compared")
        self.analytics_subject_tree.heading("deviation", text="Deviation (△)")
        self.analytics_subject_tree.heading("pass_change", text="Pass Shift")
        self.analytics_subject_tree.column("subject", width=140, stretch=True)
        self.analytics_subject_tree.column("pair", width=180, stretch=True)
        self.analytics_subject_tree.column(
            "baseline", width=90, anchor="center", stretch=False
        )
        self.analytics_subject_tree.column(
            "comparison", width=90, anchor="center", stretch=False
        )
        self.analytics_subject_tree.column(
            "deviation", width=110, anchor="center", stretch=False
        )
        self.analytics_subject_tree.column(
            "pass_change", width=100, anchor="center", stretch=False
        )
        self.analytics_subject_tree.tag_configure(
            "gain", background="#eef9f1", foreground="#1f6f43"
        )
        self.analytics_subject_tree.tag_configure(
            "drop", background="#fff1ee", foreground="#9f2f1f"
        )
        self.analytics_subject_tree.tag_configure(
            "stable", background="#f4f6f8", foreground="#52606d"
        )
        subject_y_scroll = ttk.Scrollbar(
            self.analytics_chart_frame,
            orient="vertical",
            command=self.analytics_subject_tree.yview,
        )
        subject_x_scroll = ttk.Scrollbar(
            self.analytics_chart_frame,
            orient="horizontal",
            command=self.analytics_subject_tree.xview,
        )
        self.analytics_subject_tree.configure(
            yscrollcommand=subject_y_scroll.set, xscrollcommand=subject_x_scroll.set
        )
        self.analytics_subject_tree.grid(row=0, column=0, sticky="nsew")
        subject_y_scroll.grid(row=0, column=1, sticky="ns", padx=(8, 0))
        subject_x_scroll.grid(row=1, column=0, sticky="ew", pady=(8, 0))
        self.analytics_subject_tree.bind(
            "<MouseWheel>",
            lambda event: _scroll_widget(self.analytics_subject_tree, event),
        )
        self.analytics_subject_tree.bind(
            "<Shift-MouseWheel>",
            lambda event: _scroll_widget(
                self.analytics_subject_tree, event, horizontal=True
            ),
        )

        # ---- Tab switching logic ----
        def activate_analytics_tab(active_key):
            for key, _label, color in analytics_tab_defs:
                if key == active_key:
                    analytics_tab_frames[key].lift()
                    analytics_tab_buttons[key].configure(
                        bg=color, fg="white", relief="sunken", bd=2
                    )
                else:
                    analytics_tab_buttons[key].configure(
                        bg=neutral_bg, fg="white", relief="raised", bd=1
                    )

        for key, label, color in analytics_tab_defs:
            btn = tk.Button(
                tabs_bar,
                text=f"  {label}  ",
                bg=neutral_bg,
                fg="white",
                activebackground=color,
                activeforeground="white",
                font=(FF, 10, "bold"),
                padx=10,
                pady=8,
                relief="raised",
                bd=1,
                cursor="hand2",
                command=lambda k=key: activate_analytics_tab(k),
            )
            btn.pack(side="left", padx=2, pady=2)
            analytics_tab_buttons[key] = btn

        activate_analytics_tab("summary")

    def _run_exam_analysis(self):
        """Run exam analysis based on selected filters."""
        # Get filter values
        class_name = self.analytics_class_var.get()
        term = self.analytics_term_var.get()
        exam_type = self.analytics_exam_type_var.get()

        # Convert 'All' options to None
        class_name = None if class_name == "All Classes" else class_name
        term = None if term == "All Terms" else term
        exam_type = None if exam_type == "All Types" else exam_type
        stream = self.analytics_stream_var.get()
        stream = None if stream == "All Streams" else stream

        # Get exam sessions
        sessions = exam_analytics.get_exam_sessions(
            class_name=class_name, term=term, exam_type=exam_type, stream=stream
        )

        if len(sessions) < 2:
            messagebox.showwarning(
                "Insufficient Data",
                "At least 2 exam sessions are required for comparison analysis.",
            )
            return

        # Run comparison
        comparison = exam_analytics.compare_exam_sessions(sessions)

        # ============ UPDATE KPI CARDS ============
        # Clear previous KPI cards
        for widget in self.analytics_kpi_frame.winfo_children():
            widget.destroy()

        # Create KPI cards container
        kpi_container = tk.Frame(self.analytics_kpi_frame, bg=CONTENT_BG)
        kpi_container.pack(fill="x")

        # Get severity colors
        severity_colors = {
            "low": "#27ae60",  # Green
            "medium": "#f39c12",  # Orange
            "high": "#e74c3c",  # Red
            "critical": "#c0392b",  # Dark red
        }

        # KPI 1: Sessions Analyzed
        self._create_kpi_card(
            kpi_container, "📊 Sessions", f"{len(sessions)}", "#3498db", 0
        )

        # KPI 2: Similarity Score
        similarity = comparison.overall_similarity_score
        similarity_color = (
            "#27ae60"
            if similarity > 75
            else "#f39c12"
            if similarity > 50
            else "#e74c3c"
        )
        self._create_kpi_card(
            kpi_container, "🎯 Similarity", f"{similarity:.1f}%", similarity_color, 1
        )

        # KPI 3: Critical Issues
        critical_count = len(
            [d for d in comparison.deviations if d.severity == "critical"]
        )
        critical_color = "#e74c3c" if critical_count > 0 else "#27ae60"
        self._create_kpi_card(
            kpi_container, "⚠ Critical", f"{critical_count}", critical_color, 2
        )

        # KPI 4: High Deviations
        high_count = len([d for d in comparison.deviations if d.severity == "high"])
        high_color = "#e67e22" if high_count > 0 else "#95a5a6"
        self._create_kpi_card(
            kpi_container, "⚡ High Dev.", f"{high_count}", high_color, 3
        )

        # ============ UPDATE SUMMARY TEXT ============
        self.analytics_summary_text.config(state="normal")
        self.analytics_summary_text.delete("1.0", "end")

        # Add formatted summary content
        idx = 1.0

        # Header
        self.analytics_summary_text.insert(idx, "Analysis Results\n", "header")
        idx = self.analytics_summary_text.index(f"{idx}+1line")
        self.analytics_summary_text.insert(idx, "═" * 50 + "\n\n")
        idx = self.analytics_summary_text.index(f"{idx}+1line")

        # Key Metrics Section
        self.analytics_summary_text.insert(idx, "KEY METRICS\n", "section")
        idx = self.analytics_summary_text.index(f"{idx}+1line")
        self.analytics_summary_text.insert(
            idx, f"• Sessions Analyzed: {len(sessions)}\n"
        )
        self.analytics_summary_text.insert(
            idx := self.analytics_summary_text.index(f"{idx}+1line"),
            f"• Similarity Score: {comparison.overall_similarity_score:.1f}/100\n",
        )
        idx = self.analytics_summary_text.index(f"{idx}+1line")

        if comparison.anova_results:
            status = comparison.anova_results.get("interpretation", "N/A")
            self.analytics_summary_text.insert(idx, f"• Statistical Test: {status}\n")
            idx = self.analytics_summary_text.index(f"{idx}+1line")

        self.analytics_summary_text.insert(idx, "\n")
        idx = self.analytics_summary_text.index(f"{idx}+1line")

        # Exam Type Performance
        if comparison.exam_type_summaries:
            self.analytics_summary_text.insert(
                idx, "EXAM TYPE PERFORMANCE\n", "section"
            )
            idx = self.analytics_summary_text.index(f"{idx}+1line")
            for item in comparison.exam_type_summaries:
                perf_text = (
                    f"  {item['exam_type']}\n"
                    f"    Mean: {item['mean_score']:.1f} | Pass: {item['pass_rate']:.1f}% | "
                    f"Difficulty: {item['difficulty_index']:.2f}\n"
                )
                self.analytics_summary_text.insert(idx, perf_text)
                idx = self.analytics_summary_text.index(f"{idx}+1line")
            self.analytics_summary_text.insert(idx, "\n")
            idx = self.analytics_summary_text.index(f"{idx}+1line")

        # Patterns Identified
        if comparison.patterns:
            self.analytics_summary_text.insert(idx, "PATTERNS IDENTIFIED\n", "section")
            idx = self.analytics_summary_text.index(f"{idx}+1line")
            for pattern in comparison.patterns:
                self.analytics_summary_text.insert(idx, f"  ✓ {pattern}\n")
                idx = self.analytics_summary_text.index(f"{idx}+1line")
            self.analytics_summary_text.insert(idx, "\n")
            idx = self.analytics_summary_text.index(f"{idx}+1line")

        # Anomalies
        self.analytics_summary_text.insert(idx, "ANOMALIES DETECTED\n", "section")
        idx = self.analytics_summary_text.index(f"{idx}+1line")
        if comparison.anomalies:
            for anomaly in comparison.anomalies:
                tag = "critical" if "critical" in anomaly.lower() else "warning"
                self.analytics_summary_text.insert(idx, f"  ! {anomaly}\n", tag)
                idx = self.analytics_summary_text.index(f"{idx}+1line")
        else:
            self.analytics_summary_text.insert(
                idx, "  ✓ No significant anomalies detected\n", "positive"
            )
            idx = self.analytics_summary_text.index(f"{idx}+1line")

        self.analytics_summary_text.insert(idx, "\n")
        idx = self.analytics_summary_text.index(f"{idx}+1line")

        # Recommendations
        self.analytics_summary_text.insert(idx, "RECOMMENDATIONS\n", "section")
        idx = self.analytics_summary_text.index(f"{idx}+1line")
        if comparison.recommendations:
            for i, rec in enumerate(comparison.recommendations, 1):
                tag = (
                    "critical"
                    if "critical" in rec.lower()
                    else "warning"
                    if "review" in rec.lower()
                    else "positive"
                )
                self.analytics_summary_text.insert(idx, f"  {i}. {rec}\n", tag)
                idx = self.analytics_summary_text.index(f"{idx}+1line")

        self.analytics_summary_text.config(state="disabled")

        # ============ UPDATE DEVIATION METRICS TABLE ============
        for item in self.analytics_tree.get_children():
            self.analytics_tree.delete(item)

        severity_symbols = {"low": "✓", "medium": "⚠", "high": "!", "critical": "✕"}

        for dev in comparison.deviations:
            metric_name = dev.metric.value.replace("_", " ").title()
            severity_symbol = severity_symbols.get(dev.severity, "?")
            severity_display = f"{severity_symbol} {dev.severity.upper()}"

            self.analytics_tree.insert(
                "",
                "end",
                values=(metric_name, f"{dev.value:.3f}", severity_display),
                tags=(dev.severity,),
            )

        # ============ UPDATE SUBJECT DEVIATION TABLE ============
        for item in self.analytics_subject_tree.get_children():
            self.analytics_subject_tree.delete(item)

        for row in comparison.subject_deviation_rows:
            deviation = row.get("score_deviation", 0)
            deviation_symbol = "▲" if deviation > 0 else "▼" if deviation < 0 else "—"
            row_tag = "gain" if deviation > 0 else "drop" if deviation < 0 else "stable"

            self.analytics_subject_tree.insert(
                "",
                "end",
                values=(
                    row.get("subject", ""),
                    f"{row.get('baseline_exam_type', '')} → {row.get('comparison_exam_type', '')}",
                    f"{row.get('baseline_mean', 0):.1f}",
                    f"{row.get('comparison_mean', 0):.1f}",
                    f"{deviation_symbol} {deviation:+.1f}",
                    f"{row.get('pass_rate_deviation', 0):+.1f}%",
                ),
                tags=(row_tag,),
            )

        self.analytics_summary_text.yview_moveto(0)
        self.analytics_tree.yview_moveto(0)
        self.analytics_subject_tree.yview_moveto(0)

        # Store comparison for later use
        self.current_comparison = comparison

        messagebox.showinfo(
            "Analysis Complete",
            f"Analysis completed successfully!\n\n"
            f"Sessions analyzed: {len(sessions)}\n"
            f"Similarity score: {comparison.overall_similarity_score:.1f}/100",
        )

    def _create_kpi_card(self, parent, title, value, color, column):
        """Create a KPI card widget."""
        card_frame = tk.Frame(parent, bg=color, relief="flat", bd=0)
        card_frame.grid(
            row=0, column=column, padx=8, pady=8, sticky="ew", ipadx=15, ipady=12
        )
        parent.grid_columnconfigure(column, weight=1)

        tk.Label(
            card_frame, text=title, bg=color, fg="white", font=(FF, 9), wraplength=80
        ).pack(anchor="w")
        tk.Label(
            card_frame, text=value, bg=color, fg="white", font=(FF, 24, "bold")
        ).pack(anchor="w", pady=(4, 0))

    def _generate_analytics_report(self):
        """Generate and display a detailed analytics report."""
        if not hasattr(self, "current_comparison"):
            messagebox.showwarning("No Analysis", "Please run an analysis first.")
            return

        # Generate report
        report = exam_analytics.generate_comparison_report(self.current_comparison)

        # Create report window
        report_window = tk.Toplevel(self.root)
        report_window.title("Exam Analytics Report")
        report_window.geometry("800x600")
        report_window.configure(bg=CONTENT_BG)

        # Title
        tk.Label(
            report_window,
            text="Exam Analytics Report",
            bg=CONTENT_BG,
            fg=TEXT_PRIMARY,
            font=(FF, 16, "bold"),
        ).pack(pady=(20, 10))

        # Report text
        report_text = tk.Text(report_window, wrap="word", font=(FF, 10))
        report_text.pack(fill="both", expand=True, padx=20, pady=10)
        report_text.insert("1.0", report)
        report_text.config(state="disabled")

        # Scrollbar
        scrollbar = ttk.Scrollbar(
            report_window, orient="vertical", command=report_text.yview
        )
        scrollbar.pack(side="right", fill="y", padx=(0, 20), pady=10)
        report_text.config(yscrollcommand=scrollbar.set)

        # Close button
        tk.Button(
            report_window,
            text="Close",
            bg=GREEN,
            fg="white",
            font=(FF, 10, "bold"),
            padx=20,
            pady=8,
            command=report_window.destroy,
        ).pack(pady=20)

    def _export_analytics_charts(self):
        """Export analytics charts to files."""
        if not hasattr(self, "current_comparison"):
            messagebox.showwarning("No Analysis", "Please run an analysis first.")
            return

        # Ask for export directory
        export_dir = filedialog.askdirectory(title="Select Export Directory")
        if not export_dir:
            return

        try:
            # Create dashboard charts
            charts = exam_analytics.create_comparison_dashboard(
                self.current_comparison, export_dir
            )

            messagebox.showinfo(
                "Export Complete",
                f"Charts exported successfully to:\n{export_dir}\n\n"
                f"Files created:\n" + "\n".join(charts.values()),
            )
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export charts: {str(e)}")

    # ==================== STUDENTS ====================
    def show_students(self):
        self.clear_frame()
        self._set_nav("Students")
        self._page_header(
            "Students",
            "View all students with class/stream filters. Print lists with header/footer.",
        )
        self.students_tab = StudentsTab(self.content_frame, self)

    def _student_dialog(
        self,
        title,
        adm="",
        name="",
        cls=CLASSES[0],
        gender="Male",
        photo_path="",
        guardian_name="",
        parent_email="",
        stream="",
        on_save=None,
        nav_ids=None,
        nav_index=0,
        nav_save_fn=None,
    ):
        dlg = tk.Toplevel(self.root)
        dlg.title(title)
        dlg.configure(bg=CONTENT_BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.resizable(False, False)

        # Centre on screen
        DLG_W, DLG_H = 540, 660
        dlg.update_idletasks()
        sw = dlg.winfo_screenwidth()
        sh = dlg.winfo_screenheight()
        dlg.geometry(f"{DLG_W}x{DLG_H}+{(sw - DLG_W) // 2}+{(sh - DLG_H) // 2}")

        # ── Header ──────────────────────────────────────────────────
        hdr = tk.Frame(dlg, bg=OLIVE_PRIMARY, height=72)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)

        tk.Label(hdr, text="👤", bg=OLIVE_PRIMARY, fg=LEMON_ACCENT, font=(FF, 22)).pack(
            side="left", padx=(20, 10)
        )

        hdr_txt = tk.Frame(hdr, bg=OLIVE_PRIMARY)
        hdr_txt.pack(side="left", fill="y")
        tk.Label(
            hdr_txt,
            text=title,
            bg=OLIVE_PRIMARY,
            fg=LEMON_ACCENT,
            font=(FF, 14, "bold"),
        ).pack(anchor="w", pady=(14, 0))
        hdr_subtitle_lbl = tk.Label(
            hdr_txt,
            text=f"Editing: {name}" if name else "Fill in the student details below",
            bg=OLIVE_PRIMARY,
            fg="#c8d9a4",
            font=(FF, 9),
        )
        hdr_subtitle_lbl.pack(anchor="w")

        # ── Body ────────────────────────────────────────────────────
        BODY_BG = "#f8f9f3"
        body = tk.Frame(dlg, bg=BODY_BG)
        body.pack(fill="both", expand=True)

        # ── Photo row ───────────────────────────────────────────────
        photo_row = tk.Frame(body, bg=BODY_BG)
        photo_row.pack(fill="x", padx=24, pady=(18, 8))

        AV = 72  # avatar size px
        self.temp_photo_path = photo_path
        self._dlg_avatar_ref = None

        av_cv = tk.Canvas(
            photo_row, width=AV, height=AV, bg=BODY_BG, highlightthickness=0
        )
        av_cv.pack(side="left")

        def _draw_placeholder():
            av_cv.delete("all")
            av_cv.create_oval(
                1, 1, AV - 1, AV - 1, fill="#d1d5db", outline=BORDER_CLR, width=2
            )
            cx = AV // 2
            av_cv.create_oval(cx - 9, 12, cx + 9, 30, fill="#9ca3af", outline="")
            av_cv.create_arc(
                cx - 18,
                34,
                cx + 18,
                AV - 2,
                start=0,
                extent=180,
                fill="#9ca3af",
                outline="",
            )

        def _draw_photo(p):
            if p and os.path.exists(p):
                try:
                    img = Image.open(p).resize(
                        (AV - 4, AV - 4), Image.Resampling.LANCZOS
                    )
                    ref = ImageTk.PhotoImage(img)
                    av_cv.delete("all")
                    av_cv.create_oval(
                        1, 1, AV - 1, AV - 1, fill="white", outline=BORDER_CLR, width=2
                    )
                    av_cv.create_image(AV // 2, AV // 2, image=ref)
                    self._dlg_avatar_ref = ref
                    return
                except Exception:
                    pass
            _draw_placeholder()

        _draw_photo(photo_path)

        photo_meta = tk.Frame(photo_row, bg=BODY_BG)
        photo_meta.pack(side="left", padx=(14, 0), fill="y")
        tk.Label(
            photo_meta,
            text="Student Photo",
            bg=BODY_BG,
            fg=TEXT_PRIMARY,
            font=(FF, 10, "bold"),
        ).pack(anchor="w")
        tk.Label(
            photo_meta,
            text="JPG or PNG  •  Optional",
            bg=BODY_BG,
            fg=TEXT_SECONDARY,
            font=(FF, 8),
        ).pack(anchor="w", pady=(2, 8))

        def pick_photo():
            p = filedialog.askopenfilename(
                filetypes=[("Image files", "*.jpg *.png *.jpeg")]
            )
            if p:
                self.temp_photo_path = p
                _draw_photo(p)

        tk.Button(
            photo_meta,
            text="📷  Change Photo",
            command=pick_photo,
            bg=OLIVE_PRIMARY,
            fg="white",
            activebackground=OLIVE_DARK,
            activeforeground="white",
            font=(FF, 9, "bold"),
            padx=12,
            pady=5,
            relief="flat",
            cursor="hand2",
        ).pack(anchor="w")

        # ── Divider ─────────────────────────────────────────────────
        tk.Frame(body, bg=BORDER_CLR, height=1).pack(fill="x", padx=24, pady=(0, 10))

        # ── Form grid ───────────────────────────────────────────────
        form = tk.Frame(body, bg=BODY_BG)
        form.pack(fill="x", padx=24)
        form.columnconfigure(0, weight=1, uniform="col")
        form.columnconfigure(1, weight=1, uniform="col")

        LBL_FONT = (FF, 9, "bold")
        LBL_KW = dict(bg=BODY_BG, fg=TEXT_SECONDARY, font=LBL_FONT, anchor="w")
        GAP = 8  # gap between columns

        def _lbl2(text, row, col, span=1, pady_top=10):
            px = (0, GAP) if col == 0 and span == 1 else (GAP, 0) if col == 1 else 0
            tk.Label(form, text=text, **LBL_KW).grid(
                row=row,
                column=col,
                columnspan=span,
                sticky="w",
                padx=px,
                pady=(pady_top, 2),
            )

        def _entry2(row, col, default="", span=1):
            px = (0, GAP) if col == 0 and span == 1 else (GAP, 0) if col == 1 else 0
            e = ttk.Entry(form, style="App.TEntry")
            e.insert(0, default)
            e.grid(
                row=row,
                column=col,
                columnspan=span,
                sticky="ew",
                padx=px,
                ipady=6,
                pady=(0, 2),
            )
            return e

        def _combo2(widget, row, col, span=1):
            px = (0, GAP) if col == 0 and span == 1 else (GAP, 0) if col == 1 else 0
            widget.grid(
                row=row,
                column=col,
                columnspan=span,
                sticky="ew",
                padx=px,
                ipady=6,
                pady=(0, 2),
            )

        # Full Name – full width
        _lbl2("Full Name *", 0, 0, span=2, pady_top=0)
        name_e = _entry2(1, 0, name, span=2)

        # Admission No | Gender
        _lbl2("Admission No", 2, 0)
        _lbl2("Gender", 2, 1)
        adm_e = _entry2(3, 0, adm)
        gen_cb = ttk.Combobox(
            form, values=["Male", "Female"], state="readonly", style="App.TCombobox"
        )
        gen_cb.set(gender)
        _combo2(gen_cb, 3, 1)

        # Class | Stream
        _lbl2("Class", 4, 0)
        _lbl2("Stream", 4, 1)
        cls_cb = ttk.Combobox(
            form,
            values=self.get_current_classes(),
            state="readonly",
            style="App.TCombobox",
        )
        cls_cb.set(cls)
        _combo2(cls_cb, 5, 0)
        stream_cb = ttk.Combobox(form, style="App.TCombobox")
        _combo2(stream_cb, 5, 1)

        # Guardian Name – full width
        _lbl2("Guardian Name", 6, 0, span=2)
        guardian_e = _entry2(7, 0, guardian_name, span=2)

        # Parent Email – full width
        _lbl2("Parent Email", 8, 0, span=2)
        parent_email_e = _entry2(9, 0, parent_email, span=2)

        # ── Stream / admission helpers ───────────────────────────────
        last_suggested_adm = {"value": ""}

        def refresh_streams():
            current = stream_cb.get().strip()
            values = self._get_stream_names_for_class(cls_cb.get())
            stream_cb["values"] = values
            if current and (current in values or not values):
                stream_cb.set(current)
            elif stream and stream in values:
                stream_cb.set(stream)
            elif values:
                stream_cb.set(values[0])
            else:
                stream_cb.set(current or stream)

        def refresh_admission_no(force=False):
            class_name = cls_cb.get().strip()
            if not class_name:
                return
            current = adm_e.get().strip()
            suggested = db.get_next_class_admission_no(class_name)
            if force or not current or current == last_suggested_adm["value"]:
                adm_e.delete(0, "end")
                adm_e.insert(0, suggested)
            last_suggested_adm["value"] = suggested

        refresh_streams()
        if not str(adm or "").strip():
            refresh_admission_no(force=True)

        def on_class_selected(_event=None):
            refresh_streams()
            if not str(adm or "").strip():
                refresh_admission_no()

        cls_cb.bind("<<ComboboxSelected>>", on_class_selected)

        # ── Navigation state ─────────────────────────────────────────
        _nav_ids = list(nav_ids) if nav_ids else []
        _state = {"index": nav_index}
        _use_nav = len(_nav_ids) > 1

        # ── Field loader (in-place refresh for navigation) ───────────
        def _load_fields(student):
            hdr_subtitle_lbl.config(text=f"Editing: {student.get('name', '')}")
            if _use_nav:
                idx = _state["index"]
                _nav_counter_lbl.config(text=f"Student  {idx + 1}  of  {len(_nav_ids)}")
                _prev_btn.config(state="normal" if idx > 0 else "disabled")
                _next_btn.config(
                    state="normal" if idx < len(_nav_ids) - 1 else "disabled"
                )
            name_e.delete(0, "end")
            name_e.insert(0, student.get("name", ""))
            adm_e.delete(0, "end")
            adm_e.insert(0, student.get("admission_no", ""))
            gen_cb.set(student.get("gender", "Male"))
            _prev_cls = cls_cb.get()
            cls_cb.set(student.get("class", CLASSES[0] if CLASSES else ""))
            if cls_cb.get() != _prev_cls:
                refresh_streams()
            stream_cb.set(student.get("stream", ""))
            guardian_e.delete(0, "end")
            guardian_e.insert(0, student.get("guardian_name", ""))
            parent_email_e.delete(0, "end")
            parent_email_e.insert(0, student.get("parent_email", ""))
            self.temp_photo_path = student.get("photo_path", "")
            _draw_photo(self.temp_photo_path)
            name_e.focus_set()

        # ── Save helpers ─────────────────────────────────────────────
        def _collect_fields():
            adm_val = adm_e.get().strip() or self._generate_admission_no(
                cls_cb.get(), name_e.get().strip()
            )
            adm_e.delete(0, "end")
            adm_e.insert(0, adm_val)
            return (
                adm_val,
                name_e.get().strip(),
                cls_cb.get(),
                gen_cb.get(),
                self.temp_photo_path,
                guardian_e.get().strip(),
                parent_email_e.get().strip(),
                stream_cb.get().strip(),
            )

        def _save_current(close_after=True):
            if not name_e.get().strip():
                messagebox.showerror("Error", "Name is required", parent=dlg)
                return False
            fields = _collect_fields()
            # nav_save_fn takes precedence regardless of _use_nav (also works for 1-item list)
            if nav_save_fn and _nav_ids:
                current_sid = _nav_ids[_state["index"]]
                result = nav_save_fn(current_sid, *fields)
            elif on_save:
                result = on_save(*fields)
            else:
                result = True
            if result is not False and close_after:
                dlg.destroy()
            return result is not False

        def do_save():
            student_name = name_e.get().strip()
            if _save_current(close_after=True):
                self._show_notice(
                    "✓ Student Saved",
                    f"{student_name} — record updated successfully.",
                    kind="success",
                    duration_ms=2800,
                )

        def on_cancel():
            self._show_notice(
                "Cancelled",
                "No changes were saved.",
                kind="info",
                duration_ms=2000,
            )
            dlg.destroy()

        # Mutable ref so do_save_next can update the button label after it is created
        _btn_nav_ref = {}

        def do_save_next():
            """Save the current student and advance to the next one (or close if last)."""
            student_name = name_e.get().strip()
            if not _save_current(close_after=False):
                return
            new_idx = _state["index"] + 1
            if new_idx < len(_nav_ids):
                _state["index"] = new_idx
                student = db.get_student(_nav_ids[new_idx]) or {}
                _load_fields(student)
                # Update button label based on whether this is now the last student
                btn = _btn_nav_ref.get("save_next")
                if btn:
                    is_now_last = new_idx >= len(_nav_ids) - 1
                    btn.config(
                        text="✅  Save & Done" if is_now_last else "Save & Next  ▶"
                    )
                self._show_notice(
                    "✓ Saved",
                    f"{student_name} saved — editing {new_idx + 1} of {len(_nav_ids)}.",
                    kind="success",
                    duration_ms=2000,
                )
            else:
                self._show_notice(
                    "✅ All Done",
                    f"All {len(_nav_ids)} students updated successfully.",
                    kind="success",
                    duration_ms=3000,
                )
                dlg.destroy()

        def _navigate(direction):
            if not _save_current(close_after=False):
                return
            new_idx = _state["index"] + direction
            if 0 <= new_idx < len(_nav_ids):
                _state["index"] = new_idx
                student = db.get_student(_nav_ids[new_idx]) or {}
                _load_fields(student)
                btn = _btn_nav_ref.get("save_next")
                if btn:
                    is_now_last = new_idx >= len(_nav_ids) - 1
                    btn.config(
                        text="✅  Save & Done" if is_now_last else "Save & Next  ▶"
                    )

        # ── Action buttons ───────────────────────────────────────────
        tk.Frame(body, bg=BORDER_CLR, height=1).pack(fill="x", padx=24, pady=(14, 0))

        BODY_BG = "#f8f9f3"

        # Navigation row (only when multiple students)
        if _use_nav:
            nav_bar = tk.Frame(body, bg=BODY_BG)
            nav_bar.pack(fill="x", padx=24, pady=(10, 0))

            _prev_btn = tk.Button(
                nav_bar,
                text="◀  Prev Student",
                command=lambda: _navigate(-1),
                bg="#eef2ff",
                fg="#3730a3",
                activebackground="#dde4ff",
                font=(FF, 9, "bold"),
                padx=12,
                pady=5,
                relief="flat",
                cursor="hand2",
                state="normal" if nav_index > 0 else "disabled",
            )
            _prev_btn.pack(side="left")

            _nav_counter_lbl = tk.Label(
                nav_bar,
                text=f"Student  {nav_index + 1}  of  {len(_nav_ids)}",
                bg=BODY_BG,
                fg=TEXT_SECONDARY,
                font=(FF, 9, "bold"),
            )
            _nav_counter_lbl.pack(side="left", padx=12)

            _next_btn = tk.Button(
                nav_bar,
                text="Next Student  ▶",
                command=lambda: _navigate(1),
                bg="#eef2ff",
                fg="#3730a3",
                activebackground="#dde4ff",
                font=(FF, 9, "bold"),
                padx=12,
                pady=5,
                relief="flat",
                cursor="hand2",
                state="normal" if nav_index < len(_nav_ids) - 1 else "disabled",
            )
            _next_btn.pack(side="left")
        else:
            # placeholders so _load_fields references don't fail
            _prev_btn = None
            _next_btn = None
            _nav_counter_lbl = None

        btn_bar = tk.Frame(body, bg=BODY_BG)
        btn_bar.pack(fill="x", padx=24, pady=(8, 18))

        tk.Button(
            btn_bar,
            text="Cancel",
            command=on_cancel,
            bg="#e5e7eb",
            fg=TEXT_PRIMARY,
            activebackground="#d1d5db",
            activeforeground=TEXT_PRIMARY,
            font=(FF, 10, "bold"),
            padx=20,
            pady=8,
            relief="flat",
            cursor="hand2",
        ).pack(side="left", padx=(0, 10))

        save_label = "💾  Save & Close" if _use_nav else "💾  Save Student"
        tk.Button(
            btn_bar,
            text=save_label,
            command=do_save,
            bg=OLIVE_PRIMARY,
            fg="white",
            activebackground=OLIVE_DARK,
            activeforeground="white",
            font=(FF, 10, "bold"),
            padx=24,
            pady=8,
            relief="flat",
            cursor="hand2",
        ).pack(side="left")

        if _use_nav:
            _save_next_btn = tk.Button(
                btn_bar,
                text="Save & Next  ▶",
                command=do_save_next,
                bg="#1d4ed8",
                fg="white",
                activebackground="#1e40af",
                activeforeground="white",
                font=(FF, 10, "bold"),
                padx=20,
                pady=8,
                relief="flat",
                cursor="hand2",
            )
            _save_next_btn.pack(side="left", padx=(10, 0))
            # Store reference so do_save_next can update the label
            _btn_nav_ref["save_next"] = _save_next_btn
            # Set initial label (in case we open starting at the last student)
            if nav_index >= len(_nav_ids) - 1:
                _save_next_btn.config(text="✅  Save & Done")

        # Enter key → save; Escape key → cancel; focus first field
        dlg.bind("<Return>", lambda _e: do_save())
        dlg.bind("<Escape>", lambda _e: on_cancel())
        name_e.focus_set()

    def add_student(self):
        def on_save(adm, name, cls, gender, photo, guardian_name, parent_email, stream):
            existing = db.get_student_by_admission_no(adm, cls)
            if existing:
                messagebox.showerror(
                    "Duplicate Admission No",
                    f"Admission No {adm} already exists in {cls}.",
                    parent=self.root,
                )
                return False
            db.add_student(
                name, cls, gender, adm, photo, guardian_name, parent_email, stream
            )
            self.students_tab.refresh_students()
            self._show_notice(
                "✓ Student Added",
                f"{name} has been enrolled in {cls}.",
                kind="success",
                duration_ms=2800,
            )
            return True

        self._student_dialog("Add Student", on_save=on_save)

    def edit_student(self):
        # Use the StudentsTab table if available
        if (
            hasattr(self, "students_tab")
            and self.students_tab
            and self.students_tab.students_table
        ):
            selected_iids = self.students_tab.students_table.get_selected_iids()
        else:
            selected_iids = []

        if not selected_iids:
            messagebox.showwarning(
                "Select Student", "Please select one or more students to edit"
            )
            return

        # ── Build nav_ids ─────────────────────────────────────────────
        # Single selection → navigate through ALL currently filtered students
        # Multiple selection → navigate through only the selected students
        if len(selected_iids) == 1:
            if (
                hasattr(self, "students_tab")
                and self.students_tab
                and self.students_tab.students_table
            ):
                nav_ids = [
                    row["iid"]
                    for row in self.students_tab.students_table.filtered_rows
                    if not str(row.get("iid", "")).startswith("grp_")
                ]
            else:
                nav_ids = selected_iids
            sid = selected_iids[0]
            nav_index = nav_ids.index(sid) if sid in nav_ids else 0
            title = "Edit Student"
        else:
            nav_ids = list(selected_iids)
            nav_index = 0
            title = f"Edit Students  ({len(nav_ids)} selected)"

        # ── Shared save function ──────────────────────────────────────
        def nav_save_fn(sid, a, n, c, g, p, guardian_name, parent_email, stream):
            existing = db.get_student_by_admission_no(a, c)
            if existing and existing.get("id") != sid:
                messagebox.showerror(
                    "Duplicate Admission No",
                    f"Admission No {a} already exists in {c}.",
                    parent=self.root,
                )
                return False
            db.update_student(sid, n, c, g, a, p, guardian_name, parent_email, stream)
            self.students_tab.refresh_students()
            return True

        first_sid = nav_ids[nav_index]
        first_student = db.get_student(first_sid) or {}
        self._student_dialog(
            title,
            first_student.get("admission_no", ""),
            first_student.get("name", ""),
            first_student.get("class", CLASSES[0] if CLASSES else ""),
            first_student.get("gender", "Male"),
            first_student.get("photo_path", ""),
            first_student.get("guardian_name", ""),
            first_student.get("parent_email", ""),
            first_student.get("stream", ""),
            nav_ids=nav_ids,
            nav_index=nav_index,
            nav_save_fn=nav_save_fn,
        )

    def delete_student(self):
        # Use the StudentsTab table if available
        if (
            hasattr(self, "students_tab")
            and self.students_tab
            and self.students_tab.students_table
        ):
            selected_iids = self.students_tab.students_table.get_selected_iids()
        else:
            selected_iids = []

        if not selected_iids:
            self._show_notice(
                "Select Students",
                "Please select one or more students to delete.",
                kind="info",
            )
            return
        if not self._confirm_delete_action(
            "student",
            len(selected_iids),
            scope="selected",
            details="This will also remove marks linked to the selected students.",
        ):
            return
        failed = 0
        for sid in selected_iids:
            if not sid or not db.delete_student(sid):
                failed += 1
        self.students_tab.refresh_students()
        self._show_delete_result_notice(
            "student", len(selected_iids) - failed, failed, duration_ms=4200
        )

    def import_excel(self):
        """Import students from an Excel file"""
        import_target_class = ""
        progress_dialog = None
        if hasattr(self, "students_import_class_var"):
            selected_target = (self.students_import_class_var.get() or "").strip()
            if selected_target and selected_target != "Use Class From File":
                import_target_class = (
                    self._match_known_class_name(selected_target) or selected_target
                )

        file_path = filedialog.askopenfilename(
            title="Select Excel File", filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if not file_path:
            return

        try:
            workbook = pd.read_excel(file_path, sheet_name=None)
            if not isinstance(workbook, dict):
                workbook = {"Sheet1": workbook}

            col_aliases = {
                "admission_no": {
                    "admission_no",
                    "admission no",
                    "adm_no",
                    "adm no",
                    "adm",
                    "reg_no",
                    "reg no",
                },
                "name": {
                    "name",
                    "student name",
                    "student_name",
                    "full name",
                    "full_name",
                    "learner",
                },
                "class": {"class", "grade", "class_name", "student_class"},
                "stream": {"stream", "student_stream", "class_stream"},
                "gender": {"gender", "sex"},
                "photo_path": {
                    "photo",
                    "photo_path",
                    "photo path",
                    "image",
                    "image_path",
                },
                "subject": {
                    "subject",
                    "subject_name",
                    "subj",
                    "course",
                    "course_name",
                },
            }

            def clean_value(value):
                value = str(value or "").strip()
                return "" if value.lower() == "nan" else value

            prepared_sheets = []
            total_rows = 0
            for sheet_name, raw_df in workbook.items():
                if raw_df is None or raw_df.empty:
                    continue
                sheet_context = self._get_sheet_context(str(sheet_name))
                if sheet_context.get("is_summary"):
                    continue
                df = raw_df.copy()
                df.columns = [self._normalize_text(c) for c in df.columns]

                def find_col(alias_key):
                    aliases = col_aliases[alias_key]
                    return next((col for col in df.columns if col in aliases), None)

                name_col = find_col("name")
                if not name_col:
                    continue

                prepared_sheets.append(
                    {
                        "sheet_name": str(sheet_name),
                        "df": df,
                        "name_col": name_col,
                        "adm_col": find_col("admission_no"),
                        "class_col": find_col("class"),
                        "stream_col": find_col("stream"),
                        "gender_col": find_col("gender"),
                        "photo_col": find_col("photo_path"),
                        "sheet_default_class": sheet_context.get("class_name", ""),
                        "sheet_default_stream": sheet_context.get("stream_name", ""),
                    }
                )
                total_rows += len(df.index)

            if not prepared_sheets:
                messagebox.showerror(
                    "Error",
                    "No valid worksheets found.\nEach sheet should include at least a learner name column.",
                )
                return

            current_classes = self.get_current_classes()
            default_class = current_classes[0] if current_classes else "Grade 1"
            default_class = self._match_known_class_name(default_class) or default_class

            def show_import_preview_dialog():
                top = tk.Toplevel(self.root)
                top.title("Import Preview")
                top.geometry("1040x620")
                top.configure(bg=CONTENT_BG)
                top.transient(self.root)
                top.grab_set()
                top.resizable(True, True)
                top.minsize(920, 540)

                shell_bo, shell_bi = _card_colors("mint")
                outer = tk.Frame(top, bg=shell_bo)
                outer.pack(fill="both", expand=True, padx=16, pady=16)
                body = tk.Frame(outer, bg=shell_bi, padx=18, pady=16)
                body.pack(fill="both", expand=True, padx=1, pady=1)

                tk.Label(
                    body,
                    text="Workbook Import Preview",
                    bg=shell_bi,
                    fg=TEXT_PRIMARY,
                    font=(FF, 13, "bold"),
                ).pack(anchor="w")

                scope_txt = (
                    import_target_class
                    if import_target_class
                    else "Use class from each sheet/row"
                )
                tk.Label(
                    body,
                    text=f"Selected import scope: {scope_txt}",
                    bg=shell_bi,
                    fg=TEXT_SECONDARY,
                    font=(FF, 10),
                ).pack(anchor="w", pady=(4, 12))

                stats = tk.Frame(body, bg=shell_bi)
                stats.pack(fill="x", pady=(0, 12))
                stat_items = [
                    ("Sheets Ready", str(len(prepared_sheets)), "#EAF4EC", GREEN),
                    ("Rows Found", str(total_rows), "#EEF3FF", BLUE),
                    ("Fallback Class", default_class, "#FFF6E8", ORANGE),
                ]
                for label_text, value_text, chip_bg, chip_fg in stat_items:
                    chip = tk.Frame(stats, bg=chip_bg, padx=12, pady=8)
                    chip.pack(side="left", padx=(0, 10))
                    tk.Label(
                        chip,
                        text=label_text,
                        bg=chip_bg,
                        fg=TEXT_SECONDARY,
                        font=(FF, 9, "bold"),
                    ).pack(anchor="w")
                    tk.Label(
                        chip,
                        text=value_text,
                        bg=chip_bg,
                        fg=chip_fg,
                        font=(FF, 11, "bold"),
                    ).pack(anchor="w")

                tk.Label(
                    body,
                    text="Review how each worksheet will be mapped before importing.",
                    bg=shell_bi,
                    fg=TEXT_SECONDARY,
                    font=(FF, 9),
                ).pack(anchor="w", pady=(0, 8))

                frame = tk.Frame(body, bg=CARD_BG)
                frame.pack(fill="both", expand=True, pady=(0, 12))
                cols = ("sheet", "rows", "name_col", "class_col", "mapped")
                tv = ttk.Treeview(
                    frame,
                    columns=cols,
                    show="headings",
                    style="App.Treeview",
                    height=15,
                )
                tv.heading("sheet", text="Sheet")
                tv.heading("rows", text="Rows")
                tv.heading("name_col", text="Name Column")
                tv.heading("class_col", text="Class Column")
                tv.heading("mapped", text="Mapped Class")
                tv.column("sheet", width=260, minwidth=180, anchor="w")
                tv.column("rows", width=80, minwidth=70, anchor="center")
                tv.column("name_col", width=170, minwidth=130, anchor="w")
                tv.column("class_col", width=170, minwidth=130, anchor="w")
                tv.column("mapped", width=320, minwidth=220, anchor="w")
                sb = ttk.Scrollbar(
                    frame,
                    orient="vertical",
                    command=tv.yview,
                    style="App.Vertical.TScrollbar",
                )
                xsb = ttk.Scrollbar(
                    frame,
                    orient="horizontal",
                    command=tv.xview,
                    style="App.Vertical.TScrollbar",
                )
                tv.configure(yscrollcommand=sb.set, xscrollcommand=xsb.set)
                tv.grid(row=0, column=0, sticky="nsew", padx=(0, 4), pady=4)
                sb.grid(row=0, column=1, sticky="ns", pady=4)
                xsb.grid(row=1, column=0, sticky="ew", padx=(0, 4))
                frame.grid_rowconfigure(0, weight=1)
                frame.grid_columnconfigure(0, weight=1)

                for pack in prepared_sheets:
                    mapped_class = (
                        import_target_class
                        or pack.get("sheet_default_class")
                        or f"From '{pack.get('class_col') or 'class'}' column / fallback {default_class}"
                    )
                    if pack.get("sheet_default_stream"):
                        mapped = f"{mapped_class} [{pack['sheet_default_stream']}]"
                    else:
                        mapped = mapped_class
                    class_col_label = pack.get("class_col") or "-"
                    tv.insert(
                        "",
                        "end",
                        values=(
                            pack.get("sheet_name", ""),
                            len(pack.get("df", [])),
                            pack.get("name_col", ""),
                            class_col_label,
                            mapped,
                        ),
                    )

                decision = {"ok": False}
                btn_row = tk.Frame(body, bg=shell_bi)
                btn_row.pack(fill="x")
                tk.Button(
                    btn_row,
                    text="Cancel",
                    bg=LEMON_SOFT,
                    fg=TEXT_PRIMARY,
                    font=(FF, 10, "bold"),
                    padx=18,
                    pady=8,
                    command=top.destroy,
                ).pack(side="left")
                tk.Button(
                    btn_row,
                    text="Proceed Import",
                    bg=GREEN,
                    fg="white",
                    font=(FF, 10, "bold"),
                    padx=18,
                    pady=8,
                    command=lambda: (decision.__setitem__("ok", True), top.destroy()),
                ).pack(side="right")

                self.root.wait_window(top)
                return decision["ok"]

            # Import directly with realtime progress instead of blocking on a preview dialog.

            progress_dialog, status_label, percent_label, progress = (
                self._open_progress_dialog(
                    "Importing Students",
                    "Preparing student rows...",
                    allow_cancel=True,
                )
            )

            def ensure_not_cancelled():
                if self._progress_cancel_requested(progress_dialog):
                    raise ImportCancelledError(
                        "Student import cancelled.\n\nAny learners imported before cancellation were kept."
                    )

            def build_student_progress_details(
                current_sheet="",
                processed_rows=0,
                total_rows_count=0,
                added=0,
                updated=0,
                generated=0,
                subjects_added_count=0,
                classes_added_count=0,
                sheets_processed=0,
                total_sheet_count=0,
            ):
                detail_lines = []
                if current_sheet:
                    detail_lines.append(f"Sheet: {current_sheet}")
                if total_rows_count:
                    detail_lines.append(
                        f"Rows processed: {processed_rows}/{total_rows_count}"
                    )
                if total_sheet_count:
                    detail_lines.append(
                        f"Sheets processed: {sheets_processed}/{total_sheet_count}"
                    )
                detail_lines.append(
                    f"Added: {added}   Updated: {updated}   Generated adm: {generated}"
                )
                detail_lines.append(
                    f"New subjects: {subjects_added_count}   New classes: {classes_added_count}"
                )
                return "\n".join(detail_lines)

            # Collect unique subjects and classes from Excel data
            unique_subjects = set()
            unique_classes = set()
            total_sheets = len(prepared_sheets)

            for sheet_index, sheet_pack in enumerate(prepared_sheets, start=1):
                self._update_progress_dialog(
                    progress_dialog,
                    status_label,
                    percent_label,
                    progress,
                    0,
                    max(1, total_rows),
                    f"Scanning sheet {sheet_index} of {total_sheets}: {sheet_pack['sheet_name']}",
                    build_student_progress_details(
                        current_sheet=sheet_pack["sheet_name"],
                        total_rows_count=total_rows,
                        sheets_processed=sheet_index - 1,
                        total_sheet_count=total_sheets,
                    ),
                )
                ensure_not_cancelled()
                df = sheet_pack["df"]
                subject_col = next(
                    (col for col in df.columns if col in col_aliases["subject"]), None
                )
                class_col = sheet_pack["class_col"]

                # Collect subjects
                if subject_col:
                    subjects_in_sheet = df[subject_col].dropna().unique()
                    for subj in subjects_in_sheet:
                        subj_clean = clean_value(subj)
                        if subj_clean:
                            unique_subjects.add(subj_clean)

                # Collect classes
                if class_col:
                    classes_in_sheet = df[class_col].dropna().unique()
                    for cls in classes_in_sheet:
                        cls_clean = clean_value(cls)
                        if cls_clean:
                            unique_classes.add(cls_clean)
                elif sheet_pack["sheet_default_class"]:
                    unique_classes.add(sheet_pack["sheet_default_class"])

            # Import classes without redundancy
            subjects_added = 0
            classes_added = 0

            # Import classes
            for class_name in unique_classes:
                ensure_not_cancelled()
                if not db.get_class_by_name(class_name):
                    # Determine level from class name
                    level = self._determine_class_level(class_name)
                    success, _ = db.add_class(class_name, level)
                    if success:
                        classes_added += 1

            imported_count = 0
            updated_count = 0
            generated_count = 0
            processed = 0
            class_import_summary = {}
            for sheet_index, sheet_pack in enumerate(prepared_sheets, start=1):
                sheet_name = sheet_pack["sheet_name"]
                df = sheet_pack["df"]
                name_col = sheet_pack["name_col"]
                adm_col = sheet_pack["adm_col"]
                class_col = sheet_pack["class_col"]
                stream_col = sheet_pack["stream_col"]
                gender_col = sheet_pack["gender_col"]
                photo_col = sheet_pack["photo_col"]
                sheet_default_class = sheet_pack["sheet_default_class"]
                sheet_default_stream = sheet_pack.get("sheet_default_stream", "")

                for index, (_, row) in enumerate(df.iterrows(), start=1):
                    self._update_progress_dialog(
                        progress_dialog,
                        status_label,
                        percent_label,
                        progress,
                        processed,
                        total_rows,
                        f"[{sheet_name}] Processing row {index}...",
                        build_student_progress_details(
                            current_sheet=sheet_name,
                            processed_rows=processed,
                            total_rows_count=total_rows,
                            added=imported_count,
                            updated=updated_count,
                            generated=generated_count,
                            subjects_added_count=subjects_added,
                            classes_added_count=classes_added,
                            sheets_processed=sheet_index - 1,
                            total_sheet_count=total_sheets,
                        ),
                    )
                    ensure_not_cancelled()
                    processed += 1

                    name = clean_value(row.get(name_col, ""))
                    admission_no = clean_value(row.get(adm_col, "")) if adm_col else ""
                    cls = clean_value(row.get(class_col, "")) if class_col else ""
                    stream = clean_value(row.get(stream_col, "")) if stream_col else ""
                    gender = clean_value(row.get(gender_col, "")) if gender_col else ""
                    photo_path = (
                        clean_value(row.get(photo_col, "")) if photo_col else ""
                    )

                    if not name:
                        continue
                    if import_target_class:
                        cls = import_target_class
                    elif not cls and sheet_default_class:
                        cls = sheet_default_class
                    cls = self._match_known_class_name(cls) or cls
                    if not stream and sheet_default_stream:
                        stream = sheet_default_stream
                    if not cls:
                        cls = default_class
                    if gender not in ("Male", "Female"):
                        gender = "Male"

                    if cls not in class_import_summary:
                        class_import_summary[cls] = {
                            "rows": 0,
                            "added": 0,
                            "updated": 0,
                            "generated": 0,
                        }
                    class_import_summary[cls]["rows"] += 1

                    if not admission_no:
                        admission_no = self._generate_admission_no(cls, name)
                        generated_count += 1
                        class_import_summary[cls]["generated"] += 1

                    # Upsert: update existing student (preserving photo if not provided),
                    # or add new student.
                    existing = db.get_student_by_admission_no(admission_no, cls)
                    if existing:
                        db.update_student(
                            existing["id"],
                            name,
                            cls,
                            gender,
                            admission_no,
                            photo_path,
                            existing.get("guardian_name", ""),
                            existing.get("parent_email", ""),
                            stream or existing.get("stream", ""),
                        )
                        updated_count += 1
                        class_import_summary[cls]["updated"] += 1
                    else:
                        db.add_student(
                            name, cls, gender, admission_no, photo_path, "", "", stream
                        )
                        imported_count += 1
                        class_import_summary[cls]["added"] += 1

            self._update_progress_dialog(
                progress_dialog,
                status_label,
                percent_label,
                progress,
                total_rows,
                total_rows,
                "Refreshing student list...",
                build_student_progress_details(
                    processed_rows=total_rows,
                    total_rows_count=total_rows,
                    added=imported_count,
                    updated=updated_count,
                    generated=generated_count,
                    subjects_added_count=subjects_added,
                    classes_added_count=classes_added,
                    sheets_processed=total_sheets,
                    total_sheet_count=total_sheets,
                ),
            )
            ensure_not_cancelled()
            refresh_dynamic_school_config()
            self.students_tab.refresh_students()
            progress_dialog.destroy()
            scope_note = (
                f"\nImport target class: {import_target_class}"
                if import_target_class
                else "\nImport target class: From file values"
            )
            class_order_lookup = {
                name: index
                for index, name in enumerate(db.get_class_progression_order())
            }
            ordered_classes = sorted(
                class_import_summary.items(),
                key=lambda item: (class_order_lookup.get(item[0], 999), item[0]),
            )
            breakdown_lines = []
            for class_name, stats in ordered_classes:
                breakdown_lines.append(
                    f"{class_name}: rows {stats['rows']}, added {stats['added']}, updated {stats['updated']}, generated adm {stats['generated']}"
                )
            messagebox.showinfo(
                "Import Complete",
                f"Done!  {imported_count} new student(s) added, "
                f"{updated_count} updated.\n"
                f"Generated admission numbers: {generated_count}\n"
                f"New classes added: {classes_added}\n"
                f"Sheets processed: {len(prepared_sheets)}\n"
                f"Existing photos were preserved where no new photo was supplied."
                f"\n\nClass breakdown:\n"
                + "\n".join(breakdown_lines)
                + f"{scope_note}",
            )

        except ImportCancelledError as e:
            try:
                progress_dialog.destroy()
            except Exception:
                pass
            messagebox.showinfo("Import Cancelled", str(e))
        except Exception as e:
            try:
                progress_dialog.destroy()
            except Exception:
                pass
            messagebox.showerror(
                "Import Error", f"Failed to import Excel file: {str(e)}"
            )

    def export_excel(self):
        """Export students to a workbook with one sheet per class."""
        file_path = filedialog.asksaveasfilename(
            title="Export Students",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="students_list.xlsx",
        )
        if not file_path:
            return

        try:
            students = db.get_all_students()
            export_columns = [
                "admission_no",
                "name",
                "class",
                "stream",
                "gender",
                "photo_path",
                "guardian_name",
                "parent_email",
            ]

            def build_df(rows):
                cleaned_rows = []
                for student in rows:
                    cleaned_rows.append(
                        {column: student.get(column, "") for column in export_columns}
                    )
                return pd.DataFrame(cleaned_rows, columns=export_columns)

            all_df = build_df(students)
            class_order = db.get_class_progression_order()
            grouped = {class_name: [] for class_name in class_order}
            for student in students:
                grouped.setdefault(student.get("class", "") or "Unassigned", []).append(
                    student
                )

            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                all_df.to_excel(writer, sheet_name="All Students", index=False)
                written_sheet_count = 1
                for class_name, class_students in grouped.items():
                    if not class_students:
                        continue
                    sheet_title = self._safe_excel_sheet_name(
                        class_name, f"Class {written_sheet_count}"
                    )
                    build_df(class_students).to_excel(
                        writer, sheet_name=sheet_title, index=False
                    )
                    written_sheet_count += 1

            messagebox.showinfo(
                "Export Complete",
                f"Exported {len(students)} students to:\n{file_path}\n\n"
                f"Workbook structure: 1 'All Students' sheet plus {max(0, written_sheet_count - 1)} class sheet(s).",
            )
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export: {str(e)}")

    def download_template(self):
        """Download a student-import template workbook with multiple class sheets."""
        file_path = filedialog.asksaveasfilename(
            title="Save Template",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="student_template.xlsx",
        )
        if not file_path:
            return

        try:
            template_columns = [
                "name",
                "class",
                "stream",
                "admission_no (optional)",
                "gender (optional)",
                "photo_path",
            ]

            class_names = db.get_class_progression_order() or self.get_current_classes()
            if not class_names:
                class_names = ["Grade 1"]

            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                instructions_df = pd.DataFrame(
                    {
                        "Instructions": [
                            "Fill learner rows in any class sheet.",
                            "You can leave admission_no empty and the app will auto-generate it.",
                            "If a class sheet already identifies the class, the class column can stay blank.",
                            "Use the stream column when a class has multiple streams.",
                        ]
                    }
                )
                instructions_df.to_excel(writer, sheet_name="Instructions", index=False)

                all_students_df = pd.DataFrame(
                    [
                        {
                            "name": "John Doe",
                            "class": class_names[0],
                            "stream": "",
                            "admission_no (optional)": "001",
                            "gender (optional)": "Male",
                            "photo_path": "",
                        },
                        {
                            "name": "Jane Smith",
                            "class": class_names[min(1, len(class_names) - 1)],
                            "stream": "",
                            "admission_no (optional)": "",
                            "gender (optional)": "Female",
                            "photo_path": "",
                        },
                    ],
                    columns=template_columns,
                )
                all_students_df.to_excel(writer, sheet_name="All Students", index=False)

                for class_name in class_names:
                    stream_names = self._get_known_stream_names(class_name)
                    sample_stream = stream_names[0] if stream_names else ""
                    class_df = pd.DataFrame(
                        [
                            {
                                "name": "",
                                "class": class_name,
                                "stream": sample_stream,
                                "admission_no (optional)": "",
                                "gender (optional)": "",
                                "photo_path": "",
                            }
                        ],
                        columns=template_columns,
                    )
                    class_df.to_excel(
                        writer,
                        sheet_name=self._safe_excel_sheet_name(class_name, class_name),
                        index=False,
                    )

            messagebox.showinfo(
                "Template Created",
                f"Template saved to:\n{file_path}\n\n"
                "Workbook structure: Instructions, All Students, and one sheet per class.",
            )
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create template: {str(e)}")

    # ==================== ENTER MARKS ====================
    def show_marks_entry(self):
        """Show class selection page with all school classes and their streams."""
        self.clear_frame()
        self._set_nav("Enter Marks")
        self._page_header("Enter Marks", "Select a class and stream to enter marks")

        def _stream_palette(stream_name):
            """Return (bg, fg, hover_bg, border) colors for a stream button."""
            name = (stream_name or "").strip().lower()
            if "green" in name:
                return ("#2E7D32", "#FFFFFF", "#1B5E20", "#A5D6A7")
            if "yellow" in name:
                return ("#F9A825", "#1F1F1F", "#F57F17", "#FFE082")
            if "blue" in name:
                return ("#1976D2", "#FFFFFF", "#0D47A1", "#90CAF9")
            if "red" in name:
                return ("#D32F2F", "#FFFFFF", "#B71C1C", "#EF9A9A")
            return (OLIVE_PRIMARY, "#FFFFFF", OLIVE_DARK, "#C5D29A")

        actions = tk.Frame(self.content_frame, bg=CONTENT_BG)
        actions.pack(fill="x", padx=10, pady=(0, 6))

        tk.Label(
            actions,
            text="Need to load a workbook for the whole school? Import one multi-sheet results file here.",
            bg=CONTENT_BG,
            fg=TEXT_SECONDARY,
            font=(FF, 10),
        ).pack(side="left")
        self._toolbar_btn(
            actions,
            "\U0001f4e5  Import Whole School Results",
            self.import_whole_school_marks_excel,
            bg=PURPLE,
        ).pack(side="right")
        self._toolbar_btn(
            actions,
            "\U0001f4c4  Download Whole School Template",
            self.download_whole_school_marks_template,
            bg=ORANGE,
        ).pack(side="right", padx=(0, 8))

        # Create scrollable container for class cards
        canvas_outer = tk.Frame(self.content_frame, bg=CONTENT_BG)
        canvas_outer.pack(fill="both", expand=True, padx=10, pady=10)

        canvas = tk.Canvas(canvas_outer, bg=CONTENT_BG, highlightthickness=0)
        scrollbar = ttk.Scrollbar(
            canvas_outer,
            orient="vertical",
            command=canvas.yview,
            style="App.Vertical.TScrollbar",
        )
        scrollable_frame = tk.Frame(canvas, bg=CONTENT_BG)

        scrollable_frame.bind(
            "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas_window = canvas.create_window(
            (0, 0), window=scrollable_frame, anchor="nw"
        )
        canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        _install_canvas_mousewheel(canvas)

        # Keep scrollable content width synced with canvas so responsive grids can use full page width
        def _sync_canvas_width(event):
            canvas.itemconfigure(canvas_window, width=event.width)

        canvas.bind("<Configure>", _sync_canvas_width)

        grade_levels = {
            level_name: list(CLASSES_BY_LEVEL.get(level_name, []))
            for level_name in LEVELS
            if CLASSES_BY_LEVEL.get(level_name)
        }

        card_themes_list = list(CARD_THEMES.keys())
        theme_idx = 0

        for level_name, classes in grade_levels.items():
            # Level header
            level_frame = tk.Frame(scrollable_frame, bg=CONTENT_BG)
            level_frame.pack(fill="x", pady=(10, 5), padx=10)

            tk.Label(
                level_frame,
                text=level_name,
                bg=CONTENT_BG,
                fg=OLIVE_PRIMARY,
                font=(FF, 14, "bold"),
            ).pack(anchor="w", pady=(0, 10))

            # Responsive cards grid (1-3 columns) per level
            cards_grid = tk.Frame(scrollable_frame, bg=CONTENT_BG)
            cards_grid.pack(fill="x", padx=10, pady=(0, 6))
            level_card_outers = []

            def _reflow_level_cards(
                event=None, frame=cards_grid, cards=level_card_outers
            ):
                width = frame.winfo_width()
                if width and width < 760:
                    cols = 1
                elif width and width < 1180:
                    cols = 2
                else:
                    cols = 3
                for c in range(4):
                    frame.grid_columnconfigure(c, weight=0)
                for c in range(cols):
                    frame.grid_columnconfigure(
                        c, weight=1, uniform=f"level_{level_name}"
                    )
                for i, card_outer in enumerate(cards):
                    card_outer.grid_forget()
                    row = i // cols
                    col = i % cols
                    card_outer.grid(row=row, column=col, padx=6, pady=6, sticky="nsew")

            # Create cards for each class in this level
            for class_name in classes:
                # Get class info and streams
                class_row = db.get_class_by_name(class_name)
                streams = []
                if class_row:
                    streams = db.get_streams_for_class(class_row["id"])

                # Get student count
                students = db.get_students_by_class(class_name)
                student_count = len(
                    [
                        s
                        for s in students
                        if not self._is_summary_student_name(s.get("name"))
                    ]
                )

                # Create class card
                theme_key = card_themes_list[theme_idx % len(card_themes_list)]
                theme_idx += 1
                border_clr, card_bg = _card_colors(theme_key)

                card_outer = tk.Frame(cards_grid, bg=border_clr, padx=2, pady=2)
                level_card_outers.append(card_outer)

                card = tk.Frame(card_outer, bg=card_bg, padx=20, pady=15)
                card.pack(fill="both", expand=True)

                # Class header
                header_frame = tk.Frame(card, bg=card_bg)
                header_frame.pack(fill="x", pady=(0, 10))

                tk.Label(
                    header_frame,
                    text=f"📚 {class_name}",
                    bg=card_bg,
                    fg=OLIVE_PRIMARY,
                    font=(FF, 13, "bold"),
                ).pack(side="left")

                tk.Label(
                    header_frame,
                    text=f"{student_count} students",
                    bg=card_bg,
                    fg=TEXT_SECONDARY,
                    font=(FF, 10),
                ).pack(side="right", padx=10)

                # Streams section
                if streams:
                    streams_frame = tk.Frame(card, bg=card_bg)
                    streams_frame.pack(fill="x", pady=5)

                    tk.Label(
                        streams_frame,
                        text="Select Stream",
                        bg=card_bg,
                        fg=TEXT_SECONDARY,
                        font=(FF, 10, "bold"),
                    ).pack(anchor="w", pady=(0, 5))

                    # Create buttons for each stream
                    buttons_frame = tk.Frame(streams_frame, bg=card_bg)
                    buttons_frame.pack(fill="x", pady=(0, 4))

                    visible_idx = 0
                    for stream in streams:
                        stream_name = stream.get("name", "").strip()
                        if stream_name:
                            # Get student count for this stream
                            stream_students = db.get_students_by_class_and_stream(
                                class_name, stream_name
                            )
                            stream_count = len(
                                [
                                    s
                                    for s in stream_students
                                    if not self._is_summary_student_name(s.get("name"))
                                ]
                            )
                            stream_bg, stream_fg, stream_hover_bg, stream_border = (
                                _stream_palette(stream_name)
                            )

                            stream_chip = tk.Frame(
                                buttons_frame, bg=stream_border, padx=0, pady=0
                            )
                            stream_chip.pack(fill="x", pady=4)

                            btn = tk.Button(
                                stream_chip,
                                text=f"Stream {stream_name} — {stream_count} students",
                                bg=stream_bg,
                                fg=stream_fg,
                                font=(FF, 10, "bold"),
                                relief="flat",
                                bd=0,
                                padx=10,
                                pady=10,
                                cursor="hand2",
                                command=lambda c=class_name, s=stream_name: (
                                    self._show_marks_entry_form(c, s)
                                ),
                            )
                            btn.pack(fill="x", expand=True)

                            # Hover effects
                            def on_enter(e, b=btn, c=stream_hover_bg):
                                b.configure(bg=c)

                            def on_leave(e, b=btn, c=stream_bg):
                                b.configure(bg=c)

                            btn.bind("<Enter>", on_enter)
                            btn.bind("<Leave>", on_leave)
                            visible_idx += 1
                else:
                    # No streams - show button for entire class
                    btn_frame = tk.Frame(card, bg=card_bg)
                    btn_frame.pack(fill="x", pady=5)

                    btn = tk.Button(
                        btn_frame,
                        text=f"Enter Marks for {class_name}",
                        bg=OLIVE_PRIMARY,
                        fg="white",
                        font=(FF, 11, "bold"),
                        relief="flat",
                        padx=20,
                        pady=10,
                        cursor="hand2",
                        command=lambda c=class_name: self._show_marks_entry_form(c, ""),
                    )
                    btn.pack(anchor="w")

                    # Hover effects
                    def on_enter(e, b=btn):
                        b.configure(bg=OLIVE_DARK)

                    def on_leave(e, b=btn):
                        b.configure(bg=OLIVE_PRIMARY)

                    btn.bind("<Enter>", on_enter)
                    btn.bind("<Leave>", on_leave)

            cards_grid.bind("<Configure>", _reflow_level_cards)
            _reflow_level_cards()

    def _show_marks_entry_form(self, class_name, stream=""):
        """Show the actual marks entry form for a specific class and stream"""
        self.clear_frame()
        self._set_nav("Enter Marks")

        # Build header text
        header_text = f"Enter Marks - {class_name}"
        if stream:
            header_text += f" (Stream {stream})"

        self._page_header("Enter Marks", header_text)

        # Back button
        back_frame = tk.Frame(self.content_frame, bg=CONTENT_BG)
        back_frame.pack(fill="x", pady=(0, 10))

        back_btn = tk.Button(
            back_frame,
            text="← Back to Class Selection",
            bg=OLIVE_MID,
            fg="white",
            font=(FF, 10, "bold"),
            relief="flat",
            padx=15,
            pady=8,
            cursor="hand2",
            command=self.show_marks_entry,
        )
        back_btn.pack(side="left", padx=10)

        # controls
        ctrl = tk.Frame(self.content_frame, bg=CONTENT_BG)
        ctrl.pack(fill="x", pady=(0, 12))

        def lbl(text):
            tk.Label(
                ctrl, text=text, bg=CONTENT_BG, fg=TEXT_SECONDARY, font=(FF, 10)
            ).pack(side="left", padx=(10, 4))

        lbl("Class:")
        self.marks_class_cb = ttk.Combobox(
            ctrl,
            values=self.get_current_classes(),
            state="readonly",
            style="App.TCombobox",
            width=12,
        )
        self.marks_class_cb.set(class_name)
        self.marks_class_cb.pack(side="left", ipady=4)

        lbl("Stream:")
        self.marks_stream_cb = ttk.Combobox(
            ctrl,
            state="readonly",
            style="App.TCombobox",
            width=14,
        )
        self.marks_stream_cb.pack(side="left", ipady=4)

        lbl("Year:")
        year_options = [
            str(datetime.now().year - i) for i in range(0, 6)
        ]
        self.marks_year_cb = ttk.Combobox(
            ctrl,
            values=year_options,
            state="readonly",
            style="App.TCombobox",
            width=10,
        )
        self.marks_year_cb.set(str(datetime.now().year))
        self.marks_year_cb.pack(side="left", ipady=4)

        # Store the selected stream
        self._selected_marks_stream = stream
        self._refresh_marks_streams(reload_results=False)

        lbl("Term:")
        self.marks_term_cb = ttk.Combobox(
            ctrl, values=TERMS, state="readonly", style="App.TCombobox", width=10
        )
        self.marks_term_cb.set(TERMS[0])
        self.marks_term_cb.pack(side="left", ipady=4)

        lbl("Exam:")
        self.marks_exam_cb = ttk.Combobox(
            ctrl, values=EXAM_TYPES, state="readonly", style="App.TCombobox", width=12
        )
        self.marks_exam_cb.set(DEFAULT_EXAM_TYPE)
        self.marks_exam_cb.pack(side="left", ipady=4)

        self.marks_class_cb.bind(
            "<<ComboboxSelected>>",
            lambda e: self._refresh_marks_streams(reload_results=True),
        )
        self.marks_stream_cb.bind(
            "<<ComboboxSelected>>", lambda e: self._load_marks_table()
        )
        self.marks_year_cb.bind(
            "<<ComboboxSelected>>", lambda e: self._load_marks_table()
        )
        self.marks_term_cb.bind(
            "<<ComboboxSelected>>", lambda e: self._load_marks_table()
        )
        self.marks_exam_cb.bind(
            "<<ComboboxSelected>>", lambda e: self._load_marks_table()
        )

        self._toolbar_btn(ctrl, "\U0001f4be  Save All Marks", self.save_marks).pack(
            side="left", padx=16
        )
        self._toolbar_btn(
            ctrl, "\U0001f4cb  Template", self.download_marks_template, bg=ORANGE
        ).pack(side="left", padx=4)
        self._toolbar_btn(
            ctrl, "\U0001f4e5  Import Marks", self.import_marks_excel, bg=PURPLE
        ).pack(side="left", padx=4)
        self._toolbar_btn(
            ctrl,
            "\U0001f5d1  Clear Term Marks",
            self._delete_all_marks_for_term,
            bg="#e74c3c",
        ).pack(side="left", padx=4)

        def validate_mark(event, sid, sub):
            val = event.widget.get().strip()
            if val == "":
                return True
            try:
                num = int(val)
                if num < 0 or num > 100:
                    event.widget.delete(0, tk.END)
                    event.widget.insert(0, "0")
                return True
            except ValueError:
                event.widget.delete(0, tk.END)
                return False

        self._validate_mark = validate_mark

        # editable marks grid card
        mk_bo, mk_bi = _card_colors("sand")
        tc_outer = tk.Frame(self.content_frame, bg=mk_bo)
        tc_outer.pack(fill="both", expand=True, pady=4)
        self.marks_card = tk.Frame(tc_outer, bg=mk_bi, padx=1, pady=1)
        self.marks_card.pack(fill="both", expand=True)
        self.marks_panel_bg = mk_bi

        # sticky header + scrollable rows grid
        self.marks_header_canvas = tk.Canvas(
            self.marks_card, bg=mk_bi, highlightthickness=0, height=118
        )
        self.marks_header_canvas.pack(fill="x", side="top")

        rows_outer = tk.Frame(self.marks_card, bg=mk_bi)
        rows_outer.pack(fill="both", expand=True, side="top")

        self.marks_scroll_canvas = tk.Canvas(rows_outer, bg=mk_bi, highlightthickness=0)
        marks_scroll_sb = ttk.Scrollbar(
            rows_outer,
            orient="vertical",
            command=self.marks_scroll_canvas.yview,
            style="App.Vertical.TScrollbar",
        )
        self.marks_scroll_x = ttk.Scrollbar(
            self.marks_card,
            orient="horizontal",
            command=self._marks_xview,
            style="App.Vertical.TScrollbar",
        )

        self.marks_scroll_canvas.configure(
            yscrollcommand=marks_scroll_sb.set, xscrollcommand=self._marks_xscroll_set
        )
        self.marks_header_canvas.configure(xscrollcommand=self._marks_xscroll_set)

        marks_scroll_sb.pack(side="right", fill="y")
        self.marks_scroll_canvas.pack(side="left", fill="both", expand=True)
        self.marks_scroll_x.pack(fill="x", side="bottom")

        self.marks_header_inner = tk.Frame(self.marks_header_canvas, bg=mk_bi)
        self.marks_inner = tk.Frame(self.marks_scroll_canvas, bg=mk_bi)

        self._marks_header_win = self.marks_header_canvas.create_window(
            (0, 0), window=self.marks_header_inner, anchor="nw"
        )
        self._marks_body_win = self.marks_scroll_canvas.create_window(
            (0, 0), window=self.marks_inner, anchor="nw"
        )

        def _update_header_scrollregion(_event=None):
            self.marks_header_canvas.configure(
                scrollregion=self.marks_header_canvas.bbox("all")
            )

        def _update_body_scrollregion(_event=None):
            self.marks_scroll_canvas.configure(
                scrollregion=self.marks_scroll_canvas.bbox("all")
            )

        def _on_header_canvas_resize(e):
            min_width = self.marks_header_inner.winfo_reqwidth()
            self.marks_header_canvas.itemconfig(
                self._marks_header_win, width=max(e.width, min_width)
            )
            _update_header_scrollregion()

        def _on_body_canvas_resize(e):
            min_width = self.marks_inner.winfo_reqwidth()
            self.marks_scroll_canvas.itemconfig(
                self._marks_body_win, width=max(e.width, min_width)
            )
            _update_body_scrollregion()

        self.marks_header_canvas.bind("<Configure>", _on_header_canvas_resize)
        self.marks_scroll_canvas.bind("<Configure>", _on_body_canvas_resize)
        self.marks_header_inner.bind("<Configure>", _update_header_scrollregion)
        self.marks_inner.bind("<Configure>", _update_body_scrollregion)

        self.marks_scroll_canvas.bind(
            "<MouseWheel>",
            lambda e: self.marks_scroll_canvas.yview_scroll(
                int(-1 * (e.delta / 120)), "units"
            ),
        )
        self.marks_scroll_canvas.bind(
            "<Shift-MouseWheel>",
            lambda e: self._marks_xview("scroll", int(-1 * (e.delta / 120)), "units"),
        )

        self.marks_entries = {}  # sid -> {subject: Entry}
        self.student_widgets = []  # for double-click

        self._load_marks_table()

    def _is_summary_student_name(self, name):
        key = re.sub(r"[^a-z0-9]+", "", str(name or "").lower())
        return key in {
            "average",
            "avg",
            "mean",
            "overallaverage",
            "classaverage",
            "total",
            "totals",
            "grandtotal",
            "position",
            "psn",
            "rank",
        }

    def _style_marks_entry(self, entry, row_index):
        base_bg = "#ffffff" if row_index % 2 == 0 else "#f6fbf6"
        entry.configure(
            bg=base_bg,
            fg="#163d19",
            insertbackground="#163d19",
            relief="flat",
            highlightthickness=1,
            highlightbackground="#cfe7d1",
            highlightcolor=SIDEBAR_ACTIVE,
            disabledbackground=base_bg,
            disabledforeground="#163d19",
            readonlybackground=base_bg,
            font=(FF, 10, "bold"),
            selectbackground=SIDEBAR_ACTIVE,
            selectforeground="white",
        )
        return base_bg

    def _marks_xview(self, *args):
        """Scroll marks header and body horizontally in sync."""
        if hasattr(self, "marks_header_canvas"):
            self.marks_header_canvas.xview(*args)
        if hasattr(self, "marks_scroll_canvas"):
            self.marks_scroll_canvas.xview(*args)

    def _marks_xscroll_set(self, *args):
        """Mirror horizontal scroll state to shared scrollbar and header canvas."""
        if hasattr(self, "marks_scroll_x"):
            self.marks_scroll_x.set(*args)
        if hasattr(self, "marks_header_canvas"):
            self.marks_header_canvas.xview_moveto(args[0])

    def _load_marks_table(self):
        # clear existing
        if hasattr(self, "marks_header_inner"):
            for w in self.marks_header_inner.winfo_children():
                w.destroy()
        for w in self.marks_inner.winfo_children():
            w.destroy()
        self.marks_entries.clear()
        self.student_widgets.clear()

        cls = self.marks_class_cb.get()
        year = self.marks_year_cb.get() or str(datetime.now().year)
        term = self.marks_term_cb.get()
        exam_type = self.marks_exam_cb.get() or DEFAULT_EXAM_TYPE

        # Get students based on stream selection
        stream = self._get_selected_marks_stream()
        self._selected_marks_stream = stream
        if stream:
            students = [
                s
                for s in db.get_students_by_class_and_stream(cls, stream)
                if not self._is_summary_student_name(s.get("name"))
            ]
        else:
            students = [
                s
                for s in db.get_students_by_class(cls)
                if not self._is_summary_student_name(s.get("name"))
            ]
        subjects = self._get_subjects_for_selected_class(cls, term, exam_type)
        student_col_width = 260
        subject_col_width = 96
        table_width = student_col_width + (len(subjects) * subject_col_width)

        mp_bg = getattr(self, "marks_panel_bg", CARD_BG)
        if not students:
            if hasattr(self, "marks_header_inner"):
                tk.Frame(self.marks_header_inner, bg=mp_bg, height=1).pack(fill="x")
            tk.Label(
                self.marks_inner,
                text="No students in this class",
                bg=mp_bg,
                fg=TEXT_SECONDARY,
                font=(FF, 12),
            ).pack(pady=40)
            return

        header_wrap = tk.Frame(self.marks_header_inner, bg=mp_bg, width=table_width)
        header_wrap.pack(anchor="nw", pady=(0, 2))

        summary = tk.Frame(
            header_wrap,
            bg="#edf7ee",
            highlightthickness=1,
            highlightbackground="#cfe7d1",
        )
        summary.pack(fill="x", pady=(0, 10))
        tk.Label(
            summary,
            text=f"{self._get_class_label(cls)}  •  Term {term}  •  Year {year}",
            bg="#edf7ee",
            fg=SIDEBAR_BG,
            font=(FF, 11, "bold"),
        ).pack(side="left", padx=14, pady=10)
        tk.Label(
            summary,
            text=f"{len(students)} learners",
            bg="#edf7ee",
            fg=GREEN,
            font=(FF, 10, "bold"),
        ).pack(side="left", padx=(0, 14))
        tk.Label(
            summary,
            text=f"{len(subjects)} subjects",
            bg="#edf7ee",
            fg=GREEN,
            font=(FF, 10, "bold"),
        ).pack(side="left")

        # header row
        hdr = tk.Frame(header_wrap, bg=SIDEBAR_BG, width=table_width, height=62)
        hdr.pack(anchor="nw", pady=(0, 8))
        hdr.pack_propagate(False)

        rows_wrap = tk.Frame(self.marks_inner, bg=mp_bg, width=table_width)
        rows_wrap.pack(anchor="nw", pady=(0, 2))

        student_hdr_cell = tk.Frame(
            hdr, width=student_col_width, bg=SIDEBAR_BG, height=62
        )
        student_hdr_cell.pack(side="left", fill="y")
        student_hdr_cell.pack_propagate(False)
        tk.Label(
            student_hdr_cell,
            text="Student",
            bg=SIDEBAR_BG,
            fg="white",
            font=(FF, 11, "bold"),
            anchor="w",
            padx=14,
            pady=12,
        ).pack(fill="both", expand=True)

        for sub in subjects:
            subject_style = self._get_subject_colors(sub, cls)
            sub_hdr_cell = tk.Frame(
                hdr, width=subject_col_width, bg=subject_style["base"], height=62
            )
            sub_hdr_cell.pack(side="left", fill="y")
            sub_hdr_cell.pack_propagate(False)
            tk.Label(
                sub_hdr_cell,
                text=self._get_subject_label(sub, cls, multiline=True),
                bg=subject_style["base"],
                fg=subject_style["text"],
                font=(FF, 9, "bold"),
                anchor="center",
                justify="center",
                wraplength=subject_col_width - 10,
                pady=8,
            ).pack(fill="both", expand=True)

        # student rows
        for row_index, s in enumerate(students):
            sid = s["id"]
            row_bg = "#ffffff" if row_index % 2 == 0 else "#f6fbf6"
            row_frame = tk.Frame(
                rows_wrap,
                bg=row_bg,
                width=table_width,
                height=44,
                highlightthickness=1,
                highlightbackground="#e3efe4",
            )
            row_frame.pack(anchor="nw", pady=2)
            row_frame.pack_propagate(False)

            # student name (clickable)
            name_cell = tk.Frame(
                row_frame, width=student_col_width, bg=row_bg, height=44
            )
            name_cell.pack(side="left", fill="y")
            name_cell.pack_propagate(False)
            name_btn = tk.Label(
                name_cell,
                text=s["name"],
                bg=row_bg,
                fg=TEXT_PRIMARY,
                font=(FF, 10, "bold"),
                anchor="w",
                cursor="hand2",
                padx=14,
                pady=10,
            )
            name_btn.pack(fill="both", expand=True)
            name_btn.bind(
                "<Button-1>", lambda e, sid=sid: self._edit_student_marks(sid)
            )
            self.student_widgets.append(name_btn)

            # mark entries
            self.marks_entries[sid] = {}
            m = db.get_student_marks(sid, term, exam_type, year)
            for sub in subjects:
                e_frame = tk.Frame(
                    row_frame, bg=row_bg, width=subject_col_width, height=44
                )
                e_frame.pack(side="left", fill="y")
                e_frame.pack_propagate(False)

                e = tk.Entry(
                    e_frame, width=6, justify="center", bd=1, bg="white", fg="black"
                )
                e.pack(fill="x", padx=10, pady=8)
                val = m.get(sub, "")
                e.insert(0, "" if val in (None, "") else str(val))
                e.bind(
                    "<KeyRelease>",
                    lambda ev, s=sub, sid=sid: self._validate_mark(ev, sid, s),
                )

                self.marks_entries[sid][sub] = e

        # Refresh scroll regions and keep header/body horizontally synchronized.
        self.marks_header_inner.update_idletasks()
        self.marks_inner.update_idletasks()
        self.marks_header_canvas.configure(
            scrollregion=self.marks_header_canvas.bbox("all")
        )
        self.marks_scroll_canvas.configure(
            scrollregion=self.marks_scroll_canvas.bbox("all")
        )
        self._marks_xview("moveto", "0")
        self._update_marks_page_header()

    def _edit_student_marks(self, sid=None):
        if sid is None:
            # Fallback for old Treeview binding
            sel = getattr(self, "marks_tree", None) and self.marks_tree.selection()
            if sel:
                item = self.marks_tree.item(sel[0])
                sid = item["tags"][0]
        if not sid:
            return

        students = db.get_students_by_class(self.marks_class_cb.get())
        name = next((s["name"] for s in students if s["id"] == sid), "Unknown")
        year = self.marks_year_cb.get() or str(datetime.now().year)
        term = self.marks_term_cb.get()
        exam_type = self.marks_exam_cb.get() or DEFAULT_EXAM_TYPE
        cur = db.get_student_marks(sid, term, exam_type, year)
        subjects = self._get_subjects_for_selected_class(
            self.marks_class_cb.get(), term, exam_type
        )

        dlg = tk.Toplevel(self.root)
        dlg.title(f"Marks – {name}")
        dlg.geometry("460x420")
        dlg.configure(bg=CONTENT_BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.resizable(False, False)

        em_bo, em_bi = _card_colors("mint")
        outer = tk.Frame(dlg, bg=em_bo)
        outer.place(relx=0.5, rely=0.5, anchor="center")
        card = tk.Frame(outer, bg=em_bi, padx=30, pady=22)
        card.pack(padx=1, pady=1)

        tk.Label(
            card,
            text=f"Marks for {name}  (Term {term} - {exam_type})",
            bg=em_bi,
            fg=TEXT_PRIMARY,
            font=(FF, 13, "bold"),
        ).pack(anchor="w", pady=(0, 16))

        entries: dict = {}
        grid = tk.Frame(card, bg=em_bi)
        grid.pack(fill="x")
        for i, sub in enumerate(subjects):
            r, c = divmod(i, 3)
            subject_style = self._get_subject_colors(sub, self.marks_class_cb.get())
            tk.Label(
                grid,
                text=sub,
                bg=subject_style["soft"],
                fg=subject_style["dark_text"],
                font=(FF, 10, "bold"),
                anchor="w",
                width=9,
            ).grid(row=r * 2, column=c, padx=6, pady=(6, 2), sticky="w")
            e = tk.Entry(grid, width=8, justify="center", bd=1, bg="white", fg="black")
            e.insert(0, "" if cur.get(sub, "") in (None, "") else str(cur.get(sub, "")))
            e.grid(row=r * 2 + 1, column=c, padx=6, pady=(0, 6), ipady=2, sticky="ew")
            entries[sub] = e
        for c in range(3):
            grid.columnconfigure(c, weight=1)

        btn_row = tk.Frame(card, bg=em_bi)
        btn_row.pack(fill="x", pady=(14, 0))
        ca = tk.Label(
            btn_row,
            text="Cancel",
            bg="#e8f5e9",
            fg=TEXT_PRIMARY,
            font=(FF, 10, "bold"),
            padx=18,
            pady=8,
            cursor="hand2",
        )
        ca.pack(side="left", padx=(0, 8))
        ca.bind("<Button-1>", lambda e: dlg.destroy())

        def do_save():
            marks = {}
            for sub, ent in entries.items():
                val = ent.get().strip()
                if val:
                    try:
                        marks[sub] = min(100, max(0, int(val)))
                    except ValueError:
                        pass
            db.save_student_marks(sid, marks, term, exam_type, year)
            self._load_marks_table()
            dlg.destroy()

        sv = tk.Label(
            btn_row,
            text="Save Marks",
            bg=BLUE,
            fg="white",
            font=(FF, 10, "bold"),
            padx=18,
            pady=8,
            cursor="hand2",
        )
        sv.pack(side="left")
        sv.bind("<Button-1>", lambda e: do_save())

    def save_marks(self):
        cls = self.marks_class_cb.get()
        year = self.marks_year_cb.get() or str(datetime.now().year)
        term = self.marks_term_cb.get()
        exam_type = self.marks_exam_cb.get() or DEFAULT_EXAM_TYPE

        saved_count = 0
        for sid, subject_entries in self.marks_entries.items():
            marks = {}
            for sub, entry in subject_entries.items():
                val = entry.get().strip()
                if val:
                    try:
                        mark = min(100, max(0, int(val)))
                        marks[sub] = mark
                        saved_count += 1
                    except ValueError:
                        pass
            if marks:
                db.save_student_marks(sid, marks, term, exam_type, year)

        messagebox.showinfo(
            "Success",
            f"Marks saved successfully for {term} - {exam_type}! ({saved_count} values updated)",
        )

    # ── Marks import / template helpers ──────────────────────────────────────

    def download_marks_template(self):
        """Export a clean, colorful marks template (Student + subjects)."""
        cls = self.marks_class_cb.get()
        term = self.marks_term_cb.get()
        exam_type = self.marks_exam_cb.get() or DEFAULT_EXAM_TYPE
        subjects = self._get_subjects_for_selected_class(cls, term, exam_type)
        students = db.get_students_by_class(cls)

        year = self.marks_year_cb.get() if hasattr(self, "marks_year_cb") else str(datetime.now().year)
        file_path = filedialog.asksaveasfilename(
            title="Save Marks Template",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"marks_template_{cls.replace(' ', '_')}_{year}_T{term}_{exam_type.replace('-', '_')}.xlsx",
        )
        if not file_path:
            return

        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = cls.upper().replace(" ", "_")
            ws.sheet_view.showGridLines = False

            # Header theme inspired by your Enter Marks table
            base_fill = PatternFill(
                "solid", fgColor="6F7C4A"
            )  # olive for student column
            data_fill = PatternFill("solid", fgColor="F8FAFC")
            zebra_fill = PatternFill("solid", fgColor="EEF3EF")
            meta_fill = PatternFill("solid", fgColor="EEF3EF")
            hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
            data_font = Font(name="Calibri", size=10)
            ctr = Alignment(horizontal="center", vertical="center")
            left = Alignment(horizontal="left", vertical="center")
            thin = Side(style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # Meta row (for human readability only)
            last_col = 2 + len(subjects)
            ws.merge_cells(
                start_row=1, start_column=1, end_row=1, end_column=max(2, last_col)
            )
            meta = ws.cell(
                row=1,
                column=1,
                value=f"{cls} | Year {year} | Term {term} | {exam_type} | Fill marks 0-100 (blanks are allowed)",
            )
            meta.fill = meta_fill
            meta.font = Font(bold=True, color="2F3B1A", name="Calibri", size=10)
            meta.alignment = left
            meta.border = border

            # Row 2 headers: optional admission_no + student + subjects
            ws.cell(row=2, column=1, value="admission_no").fill = base_fill
            ws.cell(row=2, column=1).font = hdr_font
            ws.cell(row=2, column=1).alignment = ctr
            ws.cell(row=2, column=1).border = border

            ws.cell(row=2, column=2, value="Student").fill = base_fill
            ws.cell(row=2, column=2).font = hdr_font
            ws.cell(row=2, column=2).alignment = left
            ws.cell(row=2, column=2).border = border

            subject_palette = [
                "F57C00",  # orange
                "E65100",  # deep orange
                "D81B60",  # pink
                "6D4C41",  # brown
                "FB8C00",  # amber
                "8E24AA",  # purple
                "3949AB",  # indigo
                "FF8F00",  # amber dark
                "9E9D24",  # olive
                "2E7D32",  # green
                "1E88E5",  # blue
                "C62828",  # red
            ]

            for i, subject in enumerate(subjects, start=3):
                fill = PatternFill(
                    "solid", fgColor=subject_palette[(i - 3) % len(subject_palette)]
                )
                cell = ws.cell(row=2, column=i, value=str(subject).strip().upper())
                cell.fill = fill
                cell.font = hdr_font
                cell.alignment = ctr
                cell.border = border

            # Data rows
            start_row = 3
            for row_offset, student in enumerate(students):
                ri = start_row + row_offset
                existing = db.get_student_marks(student["id"], term, exam_type, year)
                row_fill = zebra_fill if row_offset % 2 else data_fill

                adm = (student.get("admission_no") or "").strip()
                ws.cell(row=ri, column=1, value=adm)
                ws.cell(row=ri, column=2, value=student.get("name", ""))

                for col_idx in range(1, last_col + 1):
                    cell = ws.cell(row=ri, column=col_idx)
                    cell.fill = row_fill
                    cell.font = data_font
                    cell.border = border
                    cell.alignment = left if col_idx == 2 else ctr

                for col_idx, subject in enumerate(subjects, start=3):
                    mark_val = existing.get(subject, "")
                    ws.cell(
                        row=ri, column=col_idx, value=mark_val if mark_val != "" else ""
                    )
                    ws.cell(row=ri, column=col_idx).alignment = ctr

            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["B"].width = 34
            for col_idx in range(3, last_col + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 11
            ws.freeze_panes = "C3"

            wb.save(file_path)
            messagebox.showinfo(
                "Template Saved",
                f"Template saved to:\n{file_path}\n\n"
                "Format: admission_no (optional), Student, then subject columns.\n"
                "You can leave any mark cell blank and import will skip blanks.",
            )
        except Exception as exc:
            messagebox.showerror("Error", f"Could not create template:\n{exc}")

    def _write_whole_school_results_template(self, file_path):
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        workbook = openpyxl.Workbook()
        instructions_sheet = workbook.active
        instructions_sheet.title = "SUMMARY - Instructions"
        instructions_sheet.sheet_view.showGridLines = False

        olive_fill = PatternFill("solid", fgColor="6F7C4A")
        soft_fill = PatternFill("solid", fgColor="EEF3EF")
        header_fill = PatternFill("solid", fgColor="D9E4D0")
        white_fill = PatternFill("solid", fgColor="FFFFFF")
        white_font = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
        dark_font = Font(bold=True, color="2F3B1A", name="Calibri", size=11)
        normal_font = Font(name="Calibri", size=10)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="AAB58A")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        subject_palette = [
            "F57C00",
            "E65100",
            "D81B60",
            "6D4C41",
            "FB8C00",
            "8E24AA",
            "3949AB",
            "FF8F00",
            "9E9D24",
            "2E7D32",
            "1E88E5",
            "C62828",
        ]

        instructions_sheet.merge_cells("A1:H1")
        title_cell = instructions_sheet["A1"]
        title_cell.value = "WHOLE SCHOOL RESULTS IMPORT TEMPLATE"
        title_cell.fill = olive_fill
        title_cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=14)
        title_cell.alignment = center

        instruction_rows = [
            (
                "What this file is for",
                'Use this workbook with the "Import Whole School Results" button. Each class is on its own sheet.',
            ),
            (
                "Term and exam",
                "Choose the correct term and exam type in the app before importing this workbook.",
            ),
            (
                "Sheet names",
                "Keep class sheet names as they are: PP1, PP2, Grade 1 ... Grade 9.",
            ),
            (
                "Required columns",
                "Keep the header row with No, Admission No, Learner Name, then subject columns.",
            ),
            (
                "Marks format",
                "Enter marks from 0 to 100. Leave cells blank where a learner has no mark.",
            ),
            (
                "Streams",
                'If you use streams, you may rename a sheet like "Grade 4 Blue". The class will still be detected.',
            ),
            (
                "Summary sheets",
                "Any sheet with words like SUMMARY or ANALYSIS is skipped automatically during import.",
            ),
            (
                "Important",
                "Do not delete the Learner Name column. The importer uses it to identify each learner row.",
            ),
        ]

        for row_idx, (label, text) in enumerate(instruction_rows, start=3):
            for col_idx in range(1, 9):
                cell = instructions_sheet.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.alignment = left
                cell.font = normal_font
            label_cell = instructions_sheet.cell(row=row_idx, column=1, value=label)
            label_cell.fill = header_fill
            label_cell.font = dark_font
            text_cell = instructions_sheet.cell(row=row_idx, column=2, value=text)
            text_cell.font = normal_font

        note_row = 14
        instructions_sheet.merge_cells(
            start_row=note_row, start_column=1, end_row=note_row, end_column=8
        )
        note_cell = instructions_sheet.cell(
            row=note_row,
            column=1,
            value="Class sheets below are ready for direct data entry. Fill learner rows only; blank rows are safe.",
        )
        note_cell.fill = soft_fill
        note_cell.font = dark_font
        note_cell.alignment = left
        note_cell.border = border

        for col, width in {
            "A": 24,
            "B": 70,
            "C": 14,
            "D": 14,
            "E": 14,
            "F": 14,
            "G": 14,
            "H": 14,
        }.items():
            instructions_sheet.column_dimensions[col].width = width

        for level_name, class_names in CLASSES_BY_LEVEL.items():
            for class_name in class_names:
                subjects = self._get_whole_school_template_subject_headers(class_name)
                stream_names = self._get_known_stream_names(class_name)
                sheet_targets = (
                    [(class_name, stream_name) for stream_name in stream_names]
                    if stream_names
                    else [(class_name, "")]
                )

                for _, stream_name in sheet_targets:
                    sheet_title_text = (
                        f"{class_name} {stream_name}".strip()
                        if stream_name
                        else class_name
                    )
                    sheet = workbook.create_sheet(
                        title=self._safe_excel_sheet_name(sheet_title_text, class_name)
                    )
                    sheet.sheet_view.showGridLines = False
                    headers = ["No", "Admission No", "Learner Name"] + list(subjects)
                    total_cols = len(headers)

                    sheet.merge_cells(
                        start_row=1, start_column=1, end_row=1, end_column=total_cols
                    )
                    sheet_title = sheet.cell(
                        row=1,
                        column=1,
                        value=f"{sheet_title_text} | Whole School Import Template",
                    )
                    sheet_title.fill = olive_fill
                    sheet_title.font = Font(
                        bold=True, color="FFFFFF", name="Calibri", size=13
                    )
                    sheet_title.alignment = center
                    sheet_title.border = border

                    sheet.merge_cells(
                        start_row=2, start_column=1, end_row=2, end_column=total_cols
                    )
                    subtitle_text = (
                        f"Fill learner marks below for stream {stream_name}. Keep the header row unchanged. Blanks are allowed."
                        if stream_name
                        else "Fill learner marks below. Keep the header row unchanged. Blanks are allowed."
                    )
                    subtitle = sheet.cell(
                        row=2,
                        column=1,
                        value=subtitle_text,
                    )
                    subtitle.fill = soft_fill
                    subtitle.font = Font(
                        bold=True, color="2F3B1A", name="Calibri", size=10
                    )
                    subtitle.alignment = left
                    subtitle.border = border

                    for col_idx, header in enumerate(headers, start=1):
                        cell = sheet.cell(row=4, column=col_idx, value=header)
                        cell.border = border
                        cell.alignment = left if col_idx == 3 else center
                        if col_idx <= 3:
                            cell.fill = olive_fill
                            cell.font = white_font
                        else:
                            cell.fill = PatternFill(
                                "solid",
                                fgColor=subject_palette[
                                    (col_idx - 4) % len(subject_palette)
                                ],
                            )
                            cell.font = white_font

                    for row_idx in range(5, 45):
                        row_fill = soft_fill if row_idx % 2 == 0 else white_fill
                        for col_idx in range(1, total_cols + 1):
                            cell = sheet.cell(row=row_idx, column=col_idx)
                            cell.fill = row_fill
                            cell.font = normal_font
                            cell.border = border
                            cell.alignment = left if col_idx == 3 else center
                        sheet.cell(row=row_idx, column=1, value=row_idx - 4)

                    sheet.column_dimensions["A"].width = 8
                    sheet.column_dimensions["B"].width = 16
                    sheet.column_dimensions["C"].width = 30
                    for col_idx in range(4, total_cols + 1):
                        sheet.column_dimensions[get_column_letter(col_idx)].width = 15
                    sheet.freeze_panes = "D5"
                    sheet.auto_filter.ref = f"A4:{get_column_letter(total_cols)}44"

        workbook.save(file_path)

    def _open_whole_school_template_guide(self):
        guide_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "WHOLE_SCHOOL_IMPORT_TEMPLATE_GUIDE.md",
        )
        if not os.path.exists(guide_path):
            return False
        try:
            os.startfile(guide_path)
            return True
        except Exception:
            return False

    def download_whole_school_marks_template(self):
        """Export a multi-sheet marks template covering all classes in the school."""
        file_path = filedialog.asksaveasfilename(
            title="Save Whole School Results Template",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="whole_school_results_template.xlsx",
        )
        if not file_path:
            return

        try:
            self._write_whole_school_results_template(file_path)
            messagebox.showinfo(
                "Template Saved",
                f"Whole-school results template saved to:\n{file_path}\n\n"
                "It includes one sheet for each class, and one sheet per stream where streams already exist.\n"
                "The guide will open automatically.",
            )
            self._open_whole_school_template_guide()
        except Exception as exc:
            messagebox.showerror(
                "Template Error",
                f"Could not create whole-school template:\n{exc}",
            )

    def import_whole_school_marks_excel(self):
        """Import a multi-sheet workbook that contains results for many classes."""
        dlg = tk.Toplevel(self.root)
        dlg.title("Import Whole School Results")
        dlg.geometry("430x250")
        dlg.configure(bg=CONTENT_BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.resizable(False, False)

        card_bo, card_bi = _card_colors("mint")
        outer = tk.Frame(dlg, bg=card_bo)
        outer.place(relx=0.5, rely=0.5, anchor="center")
        card = tk.Frame(outer, bg=card_bi, padx=24, pady=20)
        card.pack(padx=1, pady=1)

        tk.Label(
            card,
            text="Import Whole School Results",
            bg=card_bi,
            fg=TEXT_PRIMARY,
            font=(FF, 12, "bold"),
        ).pack(anchor="w")
        tk.Label(
            card,
            text="Choose the term and exam, then import a workbook where each class has its own sheet. Summary sheets are skipped automatically.",
            bg=card_bi,
            fg=TEXT_SECONDARY,
            font=(FF, 9),
            justify="left",
            wraplength=360,
        ).pack(anchor="w", pady=(6, 14))

        term_var = tk.StringVar(value=TERMS[0])
        exam_var = tk.StringVar(value=DEFAULT_EXAM_TYPE)
        year_var = tk.StringVar(value=str(datetime.now().year))

        row = tk.Frame(card, bg=card_bi)
        row.pack(fill="x", pady=(0, 10))
        tk.Label(row, text="Term:", bg=card_bi, fg=TEXT_SECONDARY, font=(FF, 10)).pack(
            side="left"
        )
        ttk.Combobox(
            row,
            textvariable=term_var,
            values=TERMS,
            state="readonly",
            width=12,
            style="App.TCombobox",
        ).pack(side="left", padx=(8, 14), ipady=3)
        tk.Label(row, text="Exam:", bg=card_bi, fg=TEXT_SECONDARY, font=(FF, 10)).pack(
            side="left"
        )
        ttk.Combobox(
            row,
            textvariable=exam_var,
            values=EXAM_TYPES,
            state="readonly",
            width=12,
            style="App.TCombobox",
        ).pack(side="left", padx=(8, 14), ipady=3)
        tk.Label(row, text="Year:", bg=card_bi, fg=TEXT_SECONDARY, font=(FF, 10)).pack(
            side="left"
        )
        ttk.Combobox(
            row,
            textvariable=year_var,
            values=[str(datetime.now().year - i) for i in range(0, 6)],
            state="readonly",
            width=10,
            style="App.TCombobox",
        ).pack(side="left", padx=(8, 0), ipady=3)

        result = {"ok": False}

        def proceed():
            dlg.destroy()
            file_path = filedialog.askopenfilename(
                title="Select Whole School Results File",
                filetypes=[("Excel files", "*.xlsx *.xls")],
            )
            if not file_path:
                return
            result["ok"] = self._import_marks_workbook(
                file_path,
                term_var.get(),
                exam_var.get(),
                academic_year=year_var.get() or str(datetime.now().year),
            )

        btn_row = tk.Frame(card, bg=card_bi)
        btn_row.pack(fill="x", pady=(8, 0))
        tk.Button(
            btn_row,
            text="Cancel",
            bg=LEMON_SOFT,
            fg=TEXT_PRIMARY,
            font=(FF, 10, "bold"),
            relief="flat",
            padx=16,
            pady=8,
            command=dlg.destroy,
        ).pack(side="left")
        tk.Button(
            btn_row,
            text="Choose Workbook",
            bg=PURPLE,
            fg="white",
            font=(FF, 10, "bold"),
            relief="flat",
            padx=16,
            pady=8,
            command=proceed,
        ).pack(side="right")

        self.root.wait_window(dlg)
        if result["ok"]:
            self.show_marks_entry()

    def _import_marks_workbook(
        self,
        file_path,
        term,
        exam_type,
        class_name="",
        academic_year: str = None,
    ):
        """Import marks from either a class workbook or a whole-school workbook."""
        academic_year = academic_year or str(datetime.now().year)
        progress_dialog = None
        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path, data_only=True)
            progress_dialog, status_label, percent_label, progress = (
                self._open_progress_dialog(
                    "Importing Marks",
                    "Scanning workbook sheets...",
                    allow_cancel=True,
                )
            )

            parsed_sheets = []
            skipped_sheets = []
            total_sheets = len(wb.worksheets)

            def ensure_not_cancelled():
                if self._progress_cancel_requested(progress_dialog):
                    raise ImportCancelledError(
                        "Marks import cancelled.\n\nAny rows imported before cancellation were kept."
                    )

            def build_marks_progress_details(
                current_sheet="",
                processed_students=0,
                total_students=0,
                updated=0,
                created=0,
                skipped=0,
                unmatched=0,
                sheets_ready=0,
                sheets_skipped=0,
                subjects_added_count=0,
                classes_added_count=0,
            ):
                detail_lines = []
                if current_sheet:
                    detail_lines.append(f"Sheet: {current_sheet}")
                if total_students:
                    detail_lines.append(
                        f"Rows processed: {processed_students}/{total_students}"
                    )
                detail_lines.append(
                    f"Updated: {updated}   Created: {created}   Skipped: {skipped}"
                )
                if unmatched:
                    detail_lines.append(f"Unmatched rows: {unmatched}")
                detail_lines.append(
                    f"Sheets ready: {sheets_ready}   Sheets skipped: {sheets_skipped}"
                )
                if subjects_added_count or classes_added_count:
                    detail_lines.append(
                        f"New subjects: {subjects_added_count}   New classes: {classes_added_count}"
                    )
                return "\n".join(detail_lines)

            # Collect and import subjects and classes from mark sheets
            unique_subjects = set()
            unique_classes = set()

            for sheet_index, ws in enumerate(wb.worksheets, start=1):
                self._update_progress_dialog(
                    progress_dialog,
                    status_label,
                    percent_label,
                    progress,
                    sheet_index - 1,
                    total_sheets,
                    f"Scanning sheet {sheet_index} of {total_sheets}: {ws.title}",
                    build_marks_progress_details(
                        current_sheet=ws.title,
                        sheets_ready=len(parsed_sheets),
                        sheets_skipped=len(skipped_sheets),
                    ),
                )
                ensure_not_cancelled()
                parsed = self._parse_assessment_sheet(ws)
                if parsed:
                    parsed_sheets.append(parsed)
                    # Collect subjects
                    for student_data in parsed.get("students", []):
                        for subject in student_data.get("marks", {}).keys():
                            unique_subjects.add(subject)
                    # Collect class
                    class_name_from_sheet = parsed.get("class_name", "")
                    if class_name_from_sheet:
                        unique_classes.add(class_name_from_sheet)
                elif not self._get_sheet_context(ws.title, ws).get("is_summary"):
                    skipped_sheets.append(ws.title)

            subjects_added = 0

            # Import classes without redundancy
            classes_added = 0
            for class_name_from_data in unique_classes:
                ensure_not_cancelled()
                if not db.get_class_by_name(class_name_from_data):
                    level = self._determine_class_level(class_name_from_data)
                    success, _ = db.add_class(class_name_from_data, level)
                    if success:
                        classes_added += 1

            if parsed_sheets:
                total_updated = 0
                total_created = 0
                affected_targets = []
                total_students = sum(
                    len(parsed["students"]) for parsed in parsed_sheets
                )
                processed_students = 0
                unmatched_students = 0
                skipped_students = 0
                class_import_summary = {}

                for parsed in parsed_sheets:
                    sheet_class = parsed["class_name"]
                    stream_name = parsed.get("stream_name", "")
                    sheet_class, stream_name = self._ensure_import_class_setup(
                        sheet_class, stream_name
                    )
                    target_label = (
                        f"{sheet_class} [{stream_name}]" if stream_name else sheet_class
                    )
                    affected_targets.append(target_label)
                    if sheet_class not in class_import_summary:
                        class_import_summary[sheet_class] = {
                            "updated": 0,
                            "created": 0,
                            "sheets": set(),
                        }
                    class_import_summary[sheet_class]["sheets"].add(
                        parsed.get("sheet_name", target_label)
                    )
                    if stream_name:
                        existing_students = db.get_students_by_class_and_stream(
                            sheet_class, stream_name
                        )
                        if not existing_students:
                            existing_students = db.get_students_by_class(sheet_class)
                    else:
                        existing_students = db.get_students_by_class(sheet_class)
                    name_to_student = {
                        self._normalize_key(student["name"]): student
                        for student in existing_students
                    }

                    for item in parsed["students"]:
                        processed_students += 1
                        self._update_progress_dialog(
                            progress_dialog,
                            status_label,
                            percent_label,
                            progress,
                            processed_students,
                            total_students,
                            f"Importing {item['name'].strip()} into {sheet_class} ({processed_students}/{total_students})",
                            build_marks_progress_details(
                                current_sheet=parsed.get("sheet_name", target_label),
                                processed_students=processed_students,
                                total_students=total_students,
                                updated=total_updated,
                                created=total_created,
                                skipped=skipped_students,
                                unmatched=unmatched_students,
                                sheets_ready=len(parsed_sheets),
                                sheets_skipped=len(skipped_sheets),
                                subjects_added_count=subjects_added,
                                classes_added_count=classes_added,
                            ),
                        )
                        ensure_not_cancelled()
                        name_key = self._normalize_key(item["name"])
                        student = name_to_student.get(name_key)
                        if not student:
                            admission_no = self._generate_admission_no(
                                sheet_class, item["name"]
                            )
                            student = db.add_student(
                                item["name"].strip(),
                                sheet_class,
                                "Male",
                                admission_no,
                                "",
                                "",
                                "",
                                stream_name,
                            )
                            name_to_student[name_key] = student
                            total_created += 1
                            class_import_summary[sheet_class]["created"] += 1

                        current_marks = db.get_student_marks(
                            student["id"], term, exam_type, academic_year
                        )
                        current_marks.update(item["marks"])
                        db.save_student_marks(
                            student["id"], current_marks, term, exam_type, academic_year
                        )
                        total_updated += 1
                        class_import_summary[sheet_class]["updated"] += 1

                self._update_progress_dialog(
                    progress_dialog,
                    status_label,
                    percent_label,
                    progress,
                    total_students,
                    total_students,
                    "Refreshing marks grid...",
                    build_marks_progress_details(
                        processed_students=total_students,
                        total_students=total_students,
                        updated=total_updated,
                        created=total_created,
                        skipped=skipped_students,
                        unmatched=unmatched_students,
                        sheets_ready=len(parsed_sheets),
                        sheets_skipped=len(skipped_sheets),
                        subjects_added_count=subjects_added,
                        classes_added_count=classes_added,
                    ),
                )
                ensure_not_cancelled()
                refresh_dynamic_school_config()
                if (
                    getattr(self, "marks_class_cb", None)
                    and self.marks_class_cb.winfo_exists()
                ):
                    self._load_marks_table()
                progress_dialog.destroy()
                affected_targets = sorted(set(affected_targets))
                class_order_lookup = {
                    name: index
                    for index, name in enumerate(db.get_class_progression_order())
                }
                ordered_classes = sorted(
                    class_import_summary.items(),
                    key=lambda item: (class_order_lookup.get(item[0], 999), item[0]),
                )
                breakdown_lines = []
                for class_name, stats in ordered_classes:
                    breakdown_lines.append(
                        f"{class_name}: sheets {len(stats['sheets'])}, updated {stats['updated']}, created {stats['created']}"
                    )
                messagebox.showinfo(
                    "Import Complete",
                    f"Assessment workbook imported successfully.\n\n"
                    f"Term: {term}\n"
                    f"Exam: {exam_type}\n"
                    f"Sheets imported: {len(parsed_sheets)}\n"
                    f"Sheets skipped: {len(skipped_sheets)}\n"
                    f"Targets: {', '.join(affected_targets)}\n"
                    f"Student records updated: {total_updated}\n"
                    f"New students created with auto admission numbers: {total_created}\n"
                    f"New classes added: {classes_added}\n\n"
                    f"Class breakdown:\n"
                    + "\n".join(breakdown_lines)
                    + (
                        "\n\nSkipped sheets:\n"
                        + "\n".join(skipped_sheets[:12])
                        + (
                            f"\n... and {len(skipped_sheets) - 12} more"
                            if len(skipped_sheets) > 12
                            else ""
                        )
                        if skipped_sheets
                        else ""
                    ),
                )
                return True

            if not class_name:
                progress_dialog.destroy()
                messagebox.showwarning(
                    "Unsupported Layout",
                    "Whole-school import needs a multi-sheet workbook where each class is on its own sheet.\n\n"
                    "Tip: use the per-class Import Marks button for flat single-class tables.",
                )
                return False

            subjects = self._get_subjects_for_selected_class(
                class_name, term, exam_type
            )
            df = pd.read_excel(file_path)
            total_rows = len(df.index)
            self._update_progress_dialog(
                progress_dialog,
                status_label,
                percent_label,
                progress,
                0,
                max(1, total_rows),
                "Workbook did not match assessment layout. Using flat table import...",
                build_marks_progress_details(
                    current_sheet="Flat table import",
                    processed_students=0,
                    total_students=total_rows,
                    sheets_ready=len(parsed_sheets),
                    sheets_skipped=len(skipped_sheets),
                    subjects_added_count=subjects_added,
                    classes_added_count=classes_added,
                ),
            )
            ensure_not_cancelled()

            df.columns = [str(c).strip().lower() for c in df.columns]

            adm_aliases = {
                "admission_no",
                "adm_no",
                "admission no",
                "adm no",
                "admission",
                "adm",
                "reg_no",
                "reg no",
                "regno",
            }
            name_aliases = {
                "name",
                "student_name",
                "student name",
                "learner",
                "learner name",
                "full_name",
                "fullname",
                "pupil",
                "student",
            }

            adm_col = next((c for c in df.columns if c in adm_aliases), None)
            name_col = next((c for c in df.columns if c in name_aliases), None)

            if not adm_col and not name_col:
                progress_dialog.destroy()
                messagebox.showerror(
                    "No Identifier Column Found",
                    "The file needs at least one column to identify students.\n\n"
                    "Accepted column names for admission number:\n"
                    "  admission_no, adm_no, admission, reg_no\n\n"
                    "Accepted column names for student name:\n"
                    "  name, learner, student_name, full_name",
                )
                return False

            subj_col_map = {}
            for subj in subjects:
                subj_lower = subj.strip().lower()
                if subj_lower in df.columns:
                    subj_col_map[subj] = subj_lower

            if not subj_col_map:
                progress_dialog.destroy()
                messagebox.showwarning(
                    "No Subject Columns Found",
                    f"No matching subject columns were found in this file.\n\nExpected subject columns:\n"
                    + ", ".join(subjects),
                )
                return False

            students = db.get_students_by_class(class_name)
            adm_to_sid = {s["admission_no"].strip(): s["id"] for s in students}
            name_to_sid = {s["name"].strip().lower(): s["id"] for s in students}

            updated = 0
            skipped = 0
            not_found = []

            for index, (_, row) in enumerate(df.iterrows(), start=1):
                self._update_progress_dialog(
                    progress_dialog,
                    status_label,
                    percent_label,
                    progress,
                    index - 1,
                    total_rows,
                    f"Importing marks row {index} of {total_rows}...",
                    build_marks_progress_details(
                        current_sheet="Flat table import",
                        processed_students=index - 1,
                        total_students=total_rows,
                        updated=updated,
                        skipped=skipped,
                        unmatched=len(not_found),
                        sheets_ready=len(parsed_sheets),
                        sheets_skipped=len(skipped_sheets),
                        subjects_added_count=subjects_added,
                        classes_added_count=classes_added,
                    ),
                )
                ensure_not_cancelled()

                def _clean(val):
                    value = str(val).strip()
                    return "" if value.lower() == "nan" else value

                adm = _clean(row[adm_col]) if adm_col else ""
                name_key = _clean(row[name_col]).lower() if name_col else ""

                sid = adm_to_sid.get(adm) or name_to_sid.get(name_key)
                if not sid:
                    label = adm or name_key
                    if label:
                        not_found.append(label)
                    continue

                marks = {}
                for subj, col in subj_col_map.items():
                    raw = row.get(col)
                    if raw is not None and str(raw).strip() not in ("", "nan"):
                        try:
                            marks[subj] = min(100, max(0, int(float(raw))))
                        except (ValueError, TypeError):
                            pass

                if marks:
                    db.save_student_marks(
                        sid, marks, term, exam_type, academic_year
                    )
                    updated += 1
                else:
                    skipped += 1

            self._update_progress_dialog(
                progress_dialog,
                status_label,
                percent_label,
                progress,
                total_rows,
                total_rows,
                "Refreshing marks grid...",
                build_marks_progress_details(
                    current_sheet="Flat table import",
                    processed_students=total_rows,
                    total_students=total_rows,
                    updated=updated,
                    skipped=skipped,
                    unmatched=len(not_found),
                    sheets_ready=len(parsed_sheets),
                    sheets_skipped=len(skipped_sheets),
                    subjects_added_count=subjects_added,
                    classes_added_count=classes_added,
                ),
            )
            ensure_not_cancelled()
            if (
                getattr(self, "marks_class_cb", None)
                and self.marks_class_cb.winfo_exists()
            ):
                self._load_marks_table()
            progress_dialog.destroy()

            used_id = f'"{adm_col}"' if adm_col else f'"{name_col}"'
            msg = f"Import complete! (matched by {used_id})\n\nUpdated students: {updated}"
            if skipped:
                msg += f"\nSkipped rows with no valid marks: {skipped}"
            if not_found:
                msg += (
                    f"\nIdentifiers not found in {class_name}: {len(not_found)}\n"
                    + ", ".join(not_found[:10])
                )
                if len(not_found) > 10:
                    msg += f" ... and {len(not_found) - 10} more"
            messagebox.showinfo("Import Complete", msg)
            return True

        except ImportCancelledError as exc:
            if progress_dialog is not None:
                try:
                    progress_dialog.destroy()
                except Exception:
                    pass
            messagebox.showinfo("Import Cancelled", str(exc))
            return False
        except Exception as exc:
            if progress_dialog is not None:
                try:
                    progress_dialog.destroy()
                except Exception:
                    pass
            messagebox.showerror("Import Error", f"Failed to import marks:\n{exc}")
            return False

    def import_marks_excel(self):
        """Import marks from an Excel file into the current class & term.

        Both 'admission_no' and 'name' columns are optional — at least one
        identifier is enough. Column names are matched case-insensitively and
        common aliases are accepted.
        """
        cls = self.marks_class_cb.get()
        term = self.marks_term_cb.get()
        exam_type = self.marks_exam_cb.get() or DEFAULT_EXAM_TYPE
        subjects = self._get_subjects_for_selected_class(cls, term, exam_type)

        file_path = filedialog.askopenfilename(
            title="Select Marks File", filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if not file_path:
            return

        year = self.marks_year_cb.get() or str(datetime.now().year)
        self._import_marks_workbook(
            file_path, term, exam_type, cls, academic_year=year
        )
        return

        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path, data_only=True)
            progress_dialog, status_label, percent_label, progress = (
                self._open_progress_dialog(
                    "Importing Marks", "Scanning workbook sheets..."
                )
            )

            parsed_sheets = []
            total_sheets = len(wb.worksheets)
            for sheet_index, ws in enumerate(wb.worksheets, start=1):
                self._update_progress_dialog(
                    progress_dialog,
                    status_label,
                    percent_label,
                    progress,
                    sheet_index - 1,
                    total_sheets,
                    f"Scanning sheet {sheet_index} of {total_sheets}: {ws.title}",
                )
                parsed = self._parse_assessment_sheet(ws)
                if parsed:
                    parsed_sheets.append(parsed)

            if parsed_sheets:
                total_updated = 0
                total_created = 0
                affected_targets = []
                total_students = sum(
                    len(parsed["students"]) for parsed in parsed_sheets
                )
                processed_students = 0

                for parsed in parsed_sheets:
                    class_name = parsed["class_name"]
                    stream_name = parsed.get("stream_name", "")
                    target_label = (
                        f"{class_name} [{stream_name}]" if stream_name else class_name
                    )
                    affected_targets.append(target_label)
                    if stream_name:
                        existing_students = db.get_students_by_class_and_stream(
                            class_name, stream_name
                        )
                        if not existing_students:
                            existing_students = db.get_students_by_class(class_name)
                    else:
                        existing_students = db.get_students_by_class(class_name)
                    name_to_student = {
                        self._normalize_key(student["name"]): student
                        for student in existing_students
                    }

                    for item in parsed["students"]:
                        processed_students += 1
                        self._update_progress_dialog(
                            progress_dialog,
                            status_label,
                            percent_label,
                            progress,
                            processed_students,
                            total_students,
                            f"Importing {item['name'].strip()} into {class_name} ({processed_students}/{total_students})",
                        )
                        name_key = self._normalize_key(item["name"])
                        student = name_to_student.get(name_key)
                        if not student:
                            admission_no = self._generate_admission_no(
                                class_name, item["name"]
                            )
                            student = db.add_student(
                                item["name"].strip(),
                                class_name,
                                "Male",
                                admission_no,
                                "",
                                "",
                                "",
                                stream_name,
                            )
                            name_to_student[name_key] = student
                            total_created += 1

                        current_marks = db.get_student_marks(
                            student["id"], term, exam_type
                        )
                        current_marks.update(item["marks"])
                        db.save_student_marks(
                            student["id"], current_marks, term, exam_type
                        )
                        total_updated += 1

                self._update_progress_dialog(
                    progress_dialog,
                    status_label,
                    percent_label,
                    progress,
                    total_students,
                    total_students,
                    "Refreshing marks grid...",
                )
                self._load_marks_table()
                progress_dialog.destroy()
                affected_targets = sorted(set(affected_targets))
                messagebox.showinfo(
                    "Import Complete",
                    f"Assessment workbook imported successfully.\n\n"
                    f"Sheets imported: {len(parsed_sheets)}\n"
                    f"Targets: {', '.join(affected_targets)}\n"
                    f"Student records updated: {total_updated}\n"
                    f"New students created with auto admission numbers: {total_created}",
                )
                return

            df = pd.read_excel(file_path)
            total_rows = len(df.index)
            self._update_progress_dialog(
                progress_dialog,
                status_label,
                percent_label,
                progress,
                0,
                max(1, total_rows),
                "Workbook did not match assessment layout. Using flat table import...",
            )

            # ── Normalise all column names to lowercase-stripped for matching ──
            orig_cols = list(df.columns)
            df.columns = [str(c).strip().lower() for c in df.columns]

            # Accepted aliases for the two identifier fields
            ADM_ALIASES = {
                "admission_no",
                "adm_no",
                "admission no",
                "adm no",
                "admission",
                "adm",
                "reg_no",
                "reg no",
                "regno",
            }
            NAME_ALIASES = {
                "name",
                "student_name",
                "student name",
                "learner",
                "learner name",
                "full_name",
                "fullname",
                "pupil",
                "student",
            }

            adm_col = next((c for c in df.columns if c in ADM_ALIASES), None)
            name_col = next((c for c in df.columns if c in NAME_ALIASES), None)

            if not adm_col and not name_col:
                messagebox.showerror(
                    "No Identifier Column Found",
                    "The file needs at least one column to identify students.\n\n"
                    "Accepted column names for admission number:\n"
                    "  admission_no, adm_no, admission, reg_no …\n\n"
                    "Accepted column names for student name:\n"
                    "  name, learner, student_name, full_name …\n\n"
                    "Use the Template button to get a ready-made file.",
                )
                return

            # ── Build subject-column lookup (case-insensitive) ─────────────────
            # Map lowercase subject name → actual df column name
            subj_col_map = {}
            for subj in subjects:
                subj_lower = subj.strip().lower()
                if subj_lower in df.columns:
                    subj_col_map[subj] = subj_lower

            if not subj_col_map:
                messagebox.showwarning(
                    "No Subject Columns Found",
                    f"No matching subject columns were found in this file.\n\n"
                    f"Expected subject columns (any capitalisation):\n"
                    + ", ".join(subjects),
                )
                return

            # ── Student lookup tables ──────────────────────────────────────────
            students = db.get_students_by_class(cls)
            adm_to_sid = {s["admission_no"].strip(): s["id"] for s in students}
            name_to_sid = {s["name"].strip().lower(): s["id"] for s in students}

            updated = 0
            skipped = 0
            not_found = []

            for index, (_, row) in enumerate(df.iterrows(), start=1):
                self._update_progress_dialog(
                    progress_dialog,
                    status_label,
                    percent_label,
                    progress,
                    index - 1,
                    total_rows,
                    f"Importing marks row {index} of {total_rows}...",
                )

                # Extract identifiers (treat pandas 'nan' as empty)
                def _clean(val):
                    v = str(val).strip()
                    return "" if v.lower() == "nan" else v

                adm = _clean(row[adm_col]) if adm_col else ""
                name_key = _clean(row[name_col]).lower() if name_col else ""

                # Resolve: admission_no first, name as fallback
                sid = adm_to_sid.get(adm) or name_to_sid.get(name_key)
                if not sid:
                    label = adm or name_key
                    if label:
                        not_found.append(label)
                    continue

                # Collect valid marks from subject columns
                marks = {}
                for subj, col in subj_col_map.items():
                    raw = row.get(col)
                    if raw is not None and str(raw).strip() not in ("", "nan"):
                        try:
                            marks[subj] = min(100, max(0, int(float(raw))))
                        except (ValueError, TypeError):
                            pass

                if marks:
                    db.save_student_marks(sid, marks, term, exam_type, academic_year)
                    updated += 1
                else:
                    skipped += 1

            # Refresh on-screen grid
            self._update_progress_dialog(
                progress_dialog,
                status_label,
                percent_label,
                progress,
                total_rows,
                total_rows,
                "Refreshing marks grid...",
            )
            self._load_marks_table()
            progress_dialog.destroy()

            used_id = f'"{adm_col}"' if adm_col else f'"{name_col}"'
            msg = f"Import complete!  (matched by {used_id})\n\n✅ {updated} student(s) updated."
            if skipped:
                msg += f"\n⚠️  {skipped} row(s) had no valid mark values (skipped)."
            if not_found:
                msg += (
                    f"\n❌ {len(not_found)} identifier(s) not found in {cls}:\n   "
                    + ", ".join(not_found[:10])
                )
                if len(not_found) > 10:
                    msg += f"  … and {len(not_found) - 10} more"
            messagebox.showinfo("Import Complete", msg)

        except Exception as exc:
            try:
                progress_dialog.destroy()
            except Exception:
                pass
            messagebox.showerror("Import Error", f"Failed to import marks:\n{exc}")

    # ==================== RESULTS ====================
    def show_reports(self):
        self.clear_frame()
        self._set_nav("Results")
        self._page_header("Results", "View student performance and rankings")

        ctrl = tk.Frame(self.content_frame, bg=CONTENT_BG)
        ctrl.pack(fill="x", pady=(0, 12))

        report_classes = list(self.get_current_classes())
        default_class = (
            "PP1"
            if "PP1" in report_classes
            else (report_classes[0] if report_classes else "All")
        )
        class_options = list(report_classes)
        if "All" not in class_options:
            class_options.append("All")

        def lbl(t):
            tk.Label(
                ctrl, text=t, bg=CONTENT_BG, fg=TEXT_SECONDARY, font=(FF, 10)
            ).pack(side="left", padx=(10, 4))

        lbl("Class:")
        self.rep_cls_cb = ttk.Combobox(
            ctrl,
            values=class_options,
            state="readonly",
            style="App.TCombobox",
            width=12,
        )
        self.rep_cls_cb.set(default_class)
        self.rep_cls_cb.pack(side="left", ipady=4)

        lbl("Stream:")
        self.rep_stream_cb = ttk.Combobox(
            ctrl, state="readonly", style="App.TCombobox", width=14
        )
        self.rep_stream_cb.pack(side="left", ipady=4)

        lbl("Year:")
        self.rep_year_cb = ttk.Combobox(
            ctrl,
            values=self._get_year_options(),
            state="readonly",
            style="App.TCombobox",
            width=10,
        )
        self.rep_year_cb.set(str(datetime.now().year))
        self.rep_year_cb.pack(side="left", ipady=4)

        lbl("Term:")
        self.rep_term_cb = ttk.Combobox(
            ctrl, values=TERMS, state="readonly", style="App.TCombobox", width=10
        )
        self.rep_term_cb.set(TERMS[0])
        self.rep_term_cb.pack(side="left", ipady=4)

        lbl("Exam:")
        self.rep_exam_cb = ttk.Combobox(
            ctrl, values=EXAM_TYPES, state="readonly", style="App.TCombobox", width=12
        )
        self.rep_exam_cb.set(DEFAULT_EXAM_TYPE)
        self.rep_exam_cb.pack(side="left", ipady=4)

        self.rep_cls_cb.bind(
            "<<ComboboxSelected>>", lambda e: self._refresh_results_streams()
        )
        self.rep_stream_cb.bind("<<ComboboxSelected>>", lambda e: self.load_reports())
        self.rep_year_cb.bind("<<ComboboxSelected>>", lambda e: self.load_reports())
        self.rep_term_cb.bind("<<ComboboxSelected>>", lambda e: self.load_reports())
        self.rep_exam_cb.bind("<<ComboboxSelected>>", lambda e: self.load_reports())

        self._toolbar_btn(
            ctrl, "\u2193  Export CSV", self.export_csv, bg="#475569"
        ).pack(side="left", padx=16)
        self._toolbar_btn(
            ctrl, "\U0001f5a8  Print PDF", self.print_results_pdf, bg="#7c3aed"
        ).pack(side="left", padx=4)
        self._toolbar_btn(
            ctrl,
            "\U0001f4ca  Spotlight Excel",
            self.export_spotlight_excel,
            bg="#1B5E20",
        ).pack(side="left", padx=4)
        self.rep_mode_badge = tk.Label(
            ctrl,
            text="Mode: -",
            bg="#1d4ed8",
            fg="white",
            font=(FF, 9, "bold"),
            padx=10,
            pady=5,
        )
        self.rep_mode_badge.pack(side="right", padx=(10, 8))

        # subject averages strip
        sa_bo, sa_bi = _card_colors("lilac")
        subj_outer = tk.Frame(self.content_frame, bg=sa_bo)
        subj_outer.pack(fill="x", pady=(0, 10))
        subj_card = tk.Frame(subj_outer, bg=sa_bi, padx=16, pady=14)
        subj_card.pack(fill="both", expand=True, padx=1, pady=1)
        tk.Label(
            subj_card,
            text="Subject Averages",
            bg=sa_bi,
            fg=TEXT_PRIMARY,
            font=(FF, 11, "bold"),
        ).pack(anchor="w", pady=(0, 8))
        self.subj_row = tk.Frame(subj_card, bg=sa_bi)
        self.subj_row.pack(fill="x")

        # rankings table
        rt_bo, rt_bi = _card_colors("mint")
        tc_outer = tk.Frame(self.content_frame, bg=rt_bo)
        tc_outer.pack(fill="both", expand=True, pady=4)
        self.report_table_card = tk.Frame(tc_outer, bg=rt_bi)
        self.report_table_card.pack(fill="both", expand=True, padx=1, pady=1)
        self.report_table_bg = rt_bi
        self.report_table = None
        self.report_subject_columns = None

        self._refresh_results_streams(reload_results=False)
        self._update_results_page_header()
        self.load_reports()

    def load_reports(self):
        for w in self.subj_row.winfo_children():
            w.destroy()

        cls = self.rep_cls_cb.get()
        selected_stream = self._get_selected_results_stream()
        academic_year = self.rep_year_cb.get() or str(datetime.now().year)
        term = self.rep_term_cb.get()
        exam_type = self.rep_exam_cb.get() or DEFAULT_EXAM_TYPE
        self._update_results_page_header()
        if hasattr(self, "rep_mode_badge"):
            badge_text, badge_bg, badge_fg = self._get_subject_mode_badge(
                cls, term, exam_type, academic_year
            )
            self.rep_mode_badge.config(text=badge_text, bg=badge_bg, fg=badge_fg)
        results = self._get_results_page_results(cls, term, exam_type, academic_year)
        subjects = self._get_subjects_for_scope(
            cls, term, exam_type, results, academic_year
        )
        show_class_column = cls == "All"

        table_signature = (show_class_column, tuple(subjects))
        if self.report_subject_columns != table_signature:
            self.report_subject_columns = table_signature
            for widget in self.report_table_card.winfo_children():
                widget.destroy()

            columns = [
                {"key": "pos", "title": "Pos", "width": 60, "anchor": "center"},
                {"key": "adm", "title": "Adm No", "width": 105, "anchor": "center"},
                {"key": "name", "title": "Student Name", "width": 220, "anchor": "w"},
            ]
            if show_class_column:
                columns.append(
                    {"key": "class", "title": "Class", "width": 95, "anchor": "center"}
                )
            for s in subjects:
                columns.append(
                    {
                        "key": s,
                        "title": self._get_subject_label(s, cls),
                        "width": 88,
                        "anchor": "center",
                    }
                )
            columns.extend(
                [
                    {"key": "total", "title": "Total", "width": 80, "anchor": "center"},
                    {"key": "avg", "title": "Average", "width": 85, "anchor": "center"},
                    {"key": "grade", "title": "Grade", "width": 70, "anchor": "center"},
                ]
            )

            self.report_table = AdvancedDataTable(
                self.report_table_card,
                columns=columns,
                page_size=20,
                search_label="Search results",
            )
            self.rep_tree = self.report_table.tree

        # subject averages
        subj_totals = {s: [] for s in subjects}
        for r in results:
            for s in subjects:
                if r["marks"].get(s) is not None:
                    subj_totals[s].append(r["marks"][s])

        for s in subjects:
            vals = subj_totals[s]
            avg = round(sum(vals) / len(vals), 1) if vals else 0
            grade = self._get_grade_code_for_class(avg, cls if cls != "All" else "")
            subject_style = self._get_subject_colors(s, cls)
            tile = tk.Frame(self.subj_row, bg=subject_style["base"], padx=10, pady=8)
            tile.pack(side="left", padx=3, expand=True, fill="both")
            tk.Label(
                tile,
                text=self._get_subject_label(s, cls),
                bg=subject_style["base"],
                fg=subject_style["text"],
                font=(FF, 9, "bold"),
            ).pack()
            tk.Label(
                tile,
                text=str(avg),
                bg=subject_style["base"],
                fg=subject_style["text"],
                font=(FF, 13, "bold"),
            ).pack()
            tk.Label(
                tile,
                text=grade,
                bg=subject_style["base"],
                fg=subject_style["text"],
                font=(FF, 8),
            ).pack()

        rows = []
        for r in results:
            vals = [
                r["position"],
                r["student"]["admission_no"],
                r["student"]["name"],
            ]
            value_map = {
                "pos": r["position"],
                "adm": r["student"]["admission_no"],
                "name": r["student"]["name"],
            }
            if show_class_column:
                vals.append(self._get_class_label(r["student"]["class"]))
                value_map["class"] = self._get_class_label(r["student"]["class"])
            for s in subjects:
                mark_val = r["marks"].get(s, "-")
                vals.append(mark_val)
                value_map[s] = mark_val
            vals += [r["total"], r["average"], r["grade"]]
            value_map.update(
                {
                    "total": r["total"],
                    "avg": r["average"],
                    "grade": r["grade"],
                }
            )
            rows.append(
                {
                    "iid": f"result_{r['student']['id']}",
                    "values": tuple(vals),
                    "value_map": value_map,
                    "search": " ".join(str(v) for v in vals),
                }
            )
        if self.report_table:
            self.report_table.set_rows(rows)
        return

        # rows
        for r in results:
            vals = [r["position"], r["student"]["name"], r["student"]["class"]]
            for s in self.get_current_subjects():
                vals.append(r["marks"].get(s, "—"))
            vals += [r["total"], r["average"], r["grade"]]
            self.rep_tree.insert("", "end", values=vals)

    def export_csv(self):
        cls = self.rep_cls_cb.get()
        selected_stream = self._get_selected_results_stream()
        academic_year = self.rep_year_cb.get() or str(datetime.now().year)
        term = self.rep_term_cb.get()
        exam_type = self.rep_exam_cb.get() or DEFAULT_EXAM_TYPE
        results = self._get_results_page_results(cls, term, exam_type, academic_year)
        subjects = self._get_subjects_for_scope(
            cls, term, exam_type, results, academic_year
        )
        subject_headers = [
            self._get_subject_label(s, cls if cls != "All" else "") for s in subjects
        ]
        stream_part = (
            f"_{self._slugify_report_part(selected_stream)}"
            if selected_stream
            else "_all_streams"
        )
        fn = (
            f"report_{cls.replace(' ', '_')}{stream_part}_{academic_year}_term_{term}_"
            f"{exam_type.replace('-', '_')}.csv"
        )
        file_path = filedialog.asksaveasfilename(
            title="Save Results CSV",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialfile=fn,
        )
        if not file_path:
            return

        with open(file_path, "w", newline="") as f:
            w = csv.writer(f)
            header = ["Position", "Adm No", "Name"]
            if cls == "All":
                header.append("Class")
            header += subject_headers + ["Total", "Average", "Grade"]
            w.writerow(header)
            for r in results:
                row = [
                    r["position"],
                    r["student"]["admission_no"],
                    r["student"]["name"],
                ]
                if cls == "All":
                    row.append(r["student"]["class"])
                row += [r["marks"].get(s, "") for s in subjects]
                row += [r["total"], r["average"], r["grade"]]
                w.writerow(row)
        messagebox.showinfo("Exported", f"Report saved to {file_path}")
        return

    def print_results_pdf(self):
        cls = self.rep_cls_cb.get()
        selected_stream = self._get_selected_results_stream()
        academic_year = self.rep_year_cb.get() or str(datetime.now().year)
        term = self.rep_term_cb.get()
        exam_type = self.rep_exam_cb.get() or DEFAULT_EXAM_TYPE

        results = self._get_results_page_results(cls, term, exam_type, academic_year)
        if not results:
            messagebox.showwarning(
                "No Data",
                f"No results found for {cls}, Term {term}, {exam_type}.",
            )
            return

        year_text = academic_year
        base_name = (
            f"results_{self._slugify_report_part(cls)}_"
            f"{self._slugify_report_part(selected_stream or 'all_streams')}_"
            f"{self._slugify_report_part(self._format_report_card_term(term))}_"
            f"{self._slugify_report_part(exam_type)}_{year_text}"
        )
        file_path = filedialog.asksaveasfilename(
            title="Save Results PDF",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=f"{base_name}.pdf",
        )
        if not file_path:
            return

        if self._build_results_pdf(
            cls,
            term,
            exam_type,
            file_path,
            results=results,
            stream_name=selected_stream,
            academic_year=academic_year,
        ):
            messagebox.showinfo("Done", f"Results PDF saved to {file_path}")

    def _build_results_pdf(
        self,
        cls,
        term,
        exam_type,
        file_path,
        results=None,
        include_averages=True,
        stream_name="",
        academic_year=None,
    ):
        """Generate a printable class/all-results PDF with report-style letterhead."""
        try:
            academic_year = str(academic_year or datetime.now().year)
            rows = list(
                results or self._get_ranked_results(cls, term, exam_type, academic_year)
            )
            if not rows:
                return False

            subjects = self._get_subjects_for_scope(
                cls, term, exam_type, rows, academic_year
            )
            show_class_column = cls == "All"
            class_level = (
                self._get_level_for_class(cls)
                if cls not in ("All", ALL_SCHOOL_LEVEL)
                else ALL_SCHOOL_LEVEL
            )
            theme = self._get_level_theme(class_level)
            context = {
                "subjects": subjects,
                "is_pp": self._is_pre_primary_level(class_level),
            }
            layout = self._get_report_layout_profile(context)

            letterhead_assets = get_letterhead_assets()
            header_path = letterhead_assets.get("header_path")
            footer_path = letterhead_assets.get("footer_path")
            footer_lines = letterhead_assets.get("footer_lines", [])
            meta_label_color = OLIVE_DARK
            meta_value_color = theme["title"]

            def _image_height(image_path, width, min_height, max_height):
                if not image_path or not os.path.exists(image_path):
                    return min_height
                try:
                    image_obj = get_processed_letterhead_image(
                        image_path, "header" if image_path == header_path else "footer"
                    )
                    if image_obj is None:
                        return min_height
                    scaled = width * image_obj.height / float(image_obj.width)
                    return max(min_height, min(max_height, scaled))
                except Exception:
                    return min_height

            pagesize = landscape(A4)
            border_inset = 10
            content_inset = 24
            max_content_width = pagesize[0] - (content_inset * 2)

            base_col_widths = [26, 44, 124]
            if show_class_column:
                base_col_widths.append(42)
            base_subject_widths = [36] * len(subjects)
            summary_widths = [44, 48, 36]
            base_all_widths = base_col_widths + base_subject_widths + summary_widths
            base_total = sum(base_all_widths)

            if base_total <= max_content_width:
                width_weights = [0.5, 0.7, 2.4]
                if show_class_column:
                    width_weights.append(0.9)
                width_weights.extend([1.0] * len(subjects))
                width_weights.extend([0.9, 0.9, 0.7])
                extra_space = max_content_width - base_total
                total_weight = sum(width_weights) or 1
                col_widths = [
                    width + (extra_space * weight / total_weight)
                    for width, weight in zip(base_all_widths, width_weights)
                ]
            else:
                col_widths = [24, 40, 110]
                if show_class_column:
                    col_widths.append(38)
                compact_summary_widths = [42, 46, 34]
                remaining_for_subjects = (
                    max_content_width - sum(col_widths) - sum(compact_summary_widths)
                )
                subject_col_width = max(
                    24,
                    remaining_for_subjects / max(1, len(subjects)),
                )
                col_widths.extend([subject_col_width] * len(subjects))
                col_widths.extend(compact_summary_widths)

            results_table_width = sum(col_widths)
            content_block_x = max(
                content_inset, (pagesize[0] - results_table_width) / 2.0
            )
            page_inner_width = results_table_width
            header_height = _image_height(
                header_path,
                page_inner_width,
                layout["pdf_header_min"],
                layout["pdf_header_max"],
            )
            footer_height = _image_height(
                footer_path,
                page_inner_width,
                layout["pdf_footer_min"],
                layout["pdf_footer_max"],
            )

            doc = SimpleDocTemplate(
                file_path,
                pagesize=pagesize,
                rightMargin=content_block_x,
                leftMargin=content_block_x,
                topMargin=header_height + content_inset,
                bottomMargin=footer_height + content_inset,
            )

            styles = getSampleStyleSheet()
            styles.add(
                ParagraphStyle(
                    name="ResultsTitle",
                    parent=styles["Heading1"],
                    fontName="Helvetica-Bold",
                    fontSize=max(11, layout["pdf_title_font"] - 1),
                    leading=layout["pdf_title_font"] + 1,
                    alignment=1,
                    textColor=colors.HexColor(theme["title"]),
                    spaceAfter=4,
                )
            )
            styles.add(
                ParagraphStyle(
                    name="ResultsMeta",
                    parent=styles["BodyText"],
                    fontName="Helvetica",
                    fontSize=max(6, layout["pdf_normal_font"] - 1),
                    leading=layout["pdf_normal_font"] + 1,
                    textColor=colors.HexColor(theme["muted"]),
                )
            )
            styles.add(
                ParagraphStyle(
                    name="ResultsMetaValue",
                    parent=styles["BodyText"],
                    fontName="Helvetica-Bold",
                    fontSize=max(7, layout["pdf_normal_font"]),
                    leading=layout["pdf_normal_font"] + 2,
                    textColor=colors.HexColor(theme["title"]),
                )
            )
            styles.add(
                ParagraphStyle(
                    name="ResultsCell",
                    parent=styles["BodyText"],
                    fontName="Helvetica",
                    fontSize=max(5, layout["pdf_normal_font"] - 2),
                    leading=max(6, layout["pdf_normal_font"]),
                    alignment=1,
                )
            )
            styles.add(
                ParagraphStyle(
                    name="ResultsCellLeft",
                    parent=styles["ResultsCell"],
                    alignment=0,
                )
            )

            elements = []
            if not header_path:
                elements.append(
                    Paragraph(
                        get_school_profile().get(
                            "school_name", DEFAULT_SCHOOL_PROFILE["school_name"]
                        ),
                        styles["ResultsTitle"],
                    )
                )
            elements.append(
                Table(
                    [[""]],
                    colWidths=[doc.width],
                    rowHeights=[2],
                    style=TableStyle(
                        [
                            (
                                "BACKGROUND",
                                (0, 0),
                                (-1, -1),
                                colors.HexColor(theme["title"]),
                            ),
                            ("LEFTPADDING", (0, 0), (-1, -1), 0),
                            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                            ("TOPPADDING", (0, 0), (-1, -1), 0),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                        ]
                    ),
                )
            )
            elements.append(Spacer(1, 4))

            title_text = self._format_results_heading(
                cls, term, exam_type, year=academic_year, stream_name=stream_name
            )
            elements.append(Paragraph(title_text, styles["ResultsTitle"]))

            info_label = "Class" if show_class_column else "Grade"
            info_items = [
                (
                    info_label,
                    cls if show_class_column else self._get_class_label(cls) or cls,
                ),
                ("Term", self._format_report_card_term_display(term)),
                ("Exam", exam_type),
            ]
            if stream_name:
                info_items.append(("Stream", stream_name))
            info_items.append(("Year", academic_year))

            info_rows = [
                [
                    Paragraph(
                        f"<font color='{meta_label_color}'><b>{label}</b></font><br/>"
                        f"<font color='{meta_value_color}'>{value}</font>",
                        styles["ResultsMetaValue"],
                    )
                    for label, value in info_items
                ]
            ]
            card_width = results_table_width / max(1, len(info_items))
            info_table = Table(
                info_rows,
                colWidths=[card_width] * len(info_items),
                hAlign="LEFT",
            )
            info_table.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 8),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                        ("TOPPADDING", (0, 0), (-1, -1), 5),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                        (
                            "BACKGROUND",
                            (0, 0),
                            (-1, -1),
                            colors.HexColor(theme["header_bg"]),
                        ),
                    ]
                    + [
                        ("BOX", (idx, 0), (idx, 0), 0.8, colors.HexColor(theme["line"]))
                        for idx in range(len(info_items))
                    ]
                )
            )
            elements.append(info_table)
            elements.append(Spacer(1, 4))

            table_headers = [
                Paragraph("<b>Pos</b>", styles["ResultsCell"]),
                Paragraph("<b>Adm No</b>", styles["ResultsCell"]),
                Paragraph("<b>Student Name</b>", styles["ResultsCellLeft"]),
            ]
            if show_class_column:
                table_headers.append(Paragraph("<b>Class</b>", styles["ResultsCell"]))
            for subject in subjects:
                table_headers.append(
                    Paragraph(
                        f"<b>{self._get_subject_label(subject, cls if not show_class_column else '')}</b>",
                        styles["ResultsCell"],
                    )
                )
            table_headers.extend(
                [
                    Paragraph("<b>Total</b>", styles["ResultsCell"]),
                    Paragraph("<b>Average</b>", styles["ResultsCell"]),
                    Paragraph("<b>Grade</b>", styles["ResultsCell"]),
                ]
            )

            table_data = [table_headers]
            for result in rows:
                row = [
                    Paragraph(str(result.get("position", "")), styles["ResultsCell"]),
                    Paragraph(
                        str(result.get("student", {}).get("admission_no", "")),
                        styles["ResultsCell"],
                    ),
                    Paragraph(
                        str(result.get("student", {}).get("name", "")),
                        styles["ResultsCellLeft"],
                    ),
                ]
                if show_class_column:
                    row.append(
                        Paragraph(
                            str(
                                self._get_class_label(
                                    result.get("student", {}).get("class", "")
                                )
                            ),
                            styles["ResultsCell"],
                        )
                    )
                for subject in subjects:
                    mark = result.get("marks", {}).get(subject, "")
                    row.append(
                        Paragraph(
                            "" if mark in (None, "") else str(mark),
                            styles["ResultsCell"],
                        )
                    )
                row.extend(
                    [
                        Paragraph(str(result.get("total", "")), styles["ResultsCell"]),
                        Paragraph(
                            str(result.get("average", "")), styles["ResultsCell"]
                        ),
                        Paragraph(str(result.get("grade", "")), styles["ResultsCell"]),
                    ]
                )
                table_data.append(row)

            results_table = Table(
                table_data, colWidths=col_widths, repeatRows=1, hAlign="LEFT"
            )
            table_style = [
                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor(theme["grid"])),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor(theme["grid"])),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(theme["header_bg"])),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor(theme["muted"])),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 1),
                ("RIGHTPADDING", (0, 0), (-1, -1), 1),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("ALIGN", (0, 0), (1, -1), "CENTER"),
                ("ALIGN", (2, 1), (2, -1), "LEFT"),
            ]
            if show_class_column:
                class_col = 3
                table_style.append(("ALIGN", (class_col, 1), (class_col, -1), "CENTER"))
            for row_index in range(1, len(table_data)):
                row_bg = (
                    colors.white
                    if row_index % 2
                    else colors.HexColor(theme["accent_soft"])
                )
                table_style.append(
                    ("BACKGROUND", (0, row_index), (-1, row_index), row_bg)
                )
            results_table.setStyle(TableStyle(table_style))
            elements.append(results_table)

            if include_averages and subjects:
                elements.append(Spacer(1, 10))
                section_gap = 14
                analysis_width = (results_table_width - section_gap) * 0.6
                chart_width = results_table_width - analysis_width - section_gap

                avg_headers = [
                    Paragraph("<b>Subject</b>", styles["ResultsCellLeft"]),
                    Paragraph("<b>Average</b>", styles["ResultsCell"]),
                    Paragraph("<b>Grade</b>", styles["ResultsCell"]),
                ]
                avg_data = [avg_headers]
                avg_grades = []
                for subject in subjects:
                    vals = [
                        result.get("marks", {}).get(subject)
                        for result in rows
                        if result.get("marks", {}).get(subject) is not None
                    ]
                    avg_value = round(sum(vals) / len(vals), 1) if vals else 0
                    avg_grade = self._get_grade_code_for_class(
                        avg_value, "" if show_class_column else cls
                    )
                    avg_grades.append(avg_grade)
                    avg_data.append(
                        [
                            Paragraph(
                                self._get_subject_label(
                                    subject, cls if not show_class_column else ""
                                ),
                                styles["ResultsCellLeft"],
                            ),
                            Paragraph(str(avg_value), styles["ResultsCell"]),
                            Paragraph(avg_grade, styles["ResultsCell"]),
                        ]
                    )
                avg_col_widths = [
                    analysis_width * 0.64,
                    analysis_width * 0.20,
                    analysis_width * 0.16,
                ]
                avg_table = Table(avg_data, colWidths=avg_col_widths, hAlign="LEFT")
                avg_style = [
                    ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor(theme["grid"])),
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor(theme["grid"])),
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        colors.HexColor(theme["header_bg"]),
                    ),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor(theme["muted"])),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN", (1, 1), (-1, -1), "CENTER"),
                ]
                for row_index, subject in enumerate(subjects, start=1):
                    subject_base = self._get_subject_color(
                        subject, cls if not show_class_column else ""
                    )
                    subject_soft = colors.HexColor(
                        _mix_hex(subject_base, "#ffffff", 0.84)
                    )
                    subject_mid = colors.HexColor(
                        _mix_hex(subject_base, "#ffffff", 0.65)
                    )
                    row_bg = subject_soft if row_index % 2 else colors.white
                    avg_style.append(
                        ("BACKGROUND", (0, row_index), (1, row_index), row_bg)
                    )
                    avg_style.append(
                        ("BACKGROUND", (0, row_index), (0, row_index), subject_mid)
                    )
                    avg_style.append(
                        (
                            "TEXTCOLOR",
                            (0, row_index),
                            (0, row_index),
                            colors.HexColor("#1f2937"),
                        )
                    )
                    grade_fill = colors.HexColor(
                        _mix_hex(
                            self._get_grade_color(avg_grades[row_index - 1]),
                            "#ffffff",
                            0.72,
                        )
                    )
                    avg_style.append(
                        ("BACKGROUND", (2, row_index), (2, row_index), grade_fill)
                    )
                    avg_style.append(
                        (
                            "TEXTCOLOR",
                            (2, row_index),
                            (2, row_index),
                            colors.HexColor("#1f2937"),
                        )
                    )
                avg_table.setStyle(TableStyle(avg_style))

                analysis_title = Table(
                    [[Paragraph("<b>SUBJECT ANALYSIS</b>", styles["ResultsCellLeft"])]],
                    colWidths=[analysis_width],
                    hAlign="LEFT",
                )
                analysis_title.setStyle(
                    TableStyle(
                        [
                            (
                                "BACKGROUND",
                                (0, 0),
                                (-1, -1),
                                colors.HexColor(theme["title"]),
                            ),
                            ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
                            (
                                "BOX",
                                (0, 0),
                                (-1, -1),
                                0.8,
                                colors.HexColor(theme["title"]),
                            ),
                            ("LEFTPADDING", (0, 0), (-1, -1), 8),
                            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                            ("TOPPADDING", (0, 0), (-1, -1), 5),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                        ]
                    )
                )

                grade_order = {}
                if not show_class_column:
                    for idx, scale in enumerate(self._get_class_grading_scale(cls)):
                        code = str(scale.get("grade_code", "") or "").strip()
                        if code:
                            grade_order[code] = idx

                distribution = {}
                for result in rows:
                    raw_grade = str(result.get("grade", "") or "").strip()
                    if not raw_grade:
                        continue
                    grade_key = (
                        grade_base_code(raw_grade) if show_class_column else raw_grade
                    )
                    distribution[grade_key] = distribution.get(grade_key, 0) + 1

                if show_class_column:
                    ordered_grades = sorted(
                        distribution.keys(),
                        key=lambda code: (
                            ["EE", "ME", "AE", "BE", "IE"].index(grade_base_code(code))
                            if grade_base_code(code) in ["EE", "ME", "AE", "BE", "IE"]
                            else 99
                        ),
                    )
                else:
                    ordered_grades = sorted(
                        distribution.keys(),
                        key=lambda code: grade_order.get(code, 999),
                    )

                performance_title = Table(
                    [
                        [
                            Paragraph(
                                "<b>STUDENT PERFORMANCE CHART</b>",
                                styles["ResultsCellLeft"],
                            )
                        ]
                    ],
                    colWidths=[chart_width],
                    hAlign="LEFT",
                )
                performance_title.setStyle(
                    TableStyle(
                        [
                            (
                                "BACKGROUND",
                                (0, 0),
                                (-1, -1),
                                colors.HexColor(OLIVE_DARK),
                            ),
                            ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
                            ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor(OLIVE_DARK)),
                            ("LEFTPADDING", (0, 0), (-1, -1), 8),
                            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                            ("TOPPADDING", (0, 0), (-1, -1), 5),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                        ]
                    )
                )

                perf_headers = [
                    Paragraph("<b>Grade</b>", styles["ResultsCell"]),
                    Paragraph("<b>Students</b>", styles["ResultsCell"]),
                    Paragraph("<b>%</b>", styles["ResultsCell"]),
                    Paragraph("<b>Share</b>", styles["ResultsCellLeft"]),
                ]
                perf_data = [perf_headers]
                total_students = max(1, len(rows))
                for grade_code in ordered_grades:
                    count = distribution.get(grade_code, 0)
                    percent = (count / total_students) * 100 if total_students else 0
                    bar_max_width = max(30, chart_width * 0.36)
                    fill_width = (
                        max(8, bar_max_width * (percent / 100.0)) if count else 0
                    )
                    grade_color = self._get_grade_color(grade_code)
                    bar = Drawing(bar_max_width, 12)
                    bar.add(
                        Rect(
                            0,
                            2,
                            bar_max_width,
                            8,
                            fillColor=colors.HexColor(
                                _mix_hex(grade_color, "#ffffff", 0.88)
                            ),
                            strokeColor=colors.HexColor(
                                _mix_hex(grade_color, "#223022", 0.5)
                            ),
                            strokeWidth=0.4,
                        )
                    )
                    if fill_width:
                        bar.add(
                            Rect(
                                0,
                                2,
                                fill_width,
                                8,
                                fillColor=colors.HexColor(grade_color),
                                strokeColor=colors.HexColor(grade_color),
                                strokeWidth=0.4,
                            )
                        )
                    perf_data.append(
                        [
                            Paragraph(str(grade_code), styles["ResultsCell"]),
                            Paragraph(str(count), styles["ResultsCell"]),
                            Paragraph(f"{percent:.1f}%", styles["ResultsCell"]),
                            bar,
                        ]
                    )

                perf_col_widths = [
                    chart_width * 0.18,
                    chart_width * 0.16,
                    chart_width * 0.16,
                    chart_width * 0.50,
                ]
                performance_table = Table(
                    perf_data, colWidths=perf_col_widths, hAlign="LEFT"
                )
                perf_style = [
                    ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor(theme["grid"])),
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor(theme["grid"])),
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        colors.HexColor(theme["header_bg"]),
                    ),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor(theme["muted"])),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN", (0, 0), (2, -1), "CENTER"),
                    ("ALIGN", (3, 1), (3, -1), "LEFT"),
                ]
                for row_index, grade_code in enumerate(ordered_grades, start=1):
                    grade_color = self._get_grade_color(grade_code)
                    soft_grade = colors.HexColor(_mix_hex(grade_color, "#ffffff", 0.84))
                    perf_style.append(
                        ("BACKGROUND", (0, row_index), (2, row_index), soft_grade)
                    )
                    perf_style.append(
                        (
                            "BACKGROUND",
                            (0, row_index),
                            (0, row_index),
                            colors.HexColor(_mix_hex(grade_color, "#ffffff", 0.68)),
                        )
                    )
                    perf_style.append(
                        (
                            "TEXTCOLOR",
                            (0, row_index),
                            (0, row_index),
                            colors.HexColor("#1f2937"),
                        )
                    )
                performance_table.setStyle(TableStyle(perf_style))

                summary_row = Table(
                    [
                        [
                            [analysis_title, Spacer(1, 4), avg_table],
                            "",
                            [performance_title, Spacer(1, 4), performance_table],
                        ]
                    ],
                    colWidths=[analysis_width, section_gap, chart_width],
                    hAlign="LEFT",
                )
                summary_row.setStyle(
                    TableStyle(
                        [
                            ("VALIGN", (0, 0), (-1, -1), "TOP"),
                            ("BACKGROUND", (1, 0), (1, 0), colors.white),
                            ("LEFTPADDING", (0, 0), (-1, -1), 0),
                            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                            ("TOPPADDING", (0, 0), (-1, -1), 0),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                        ]
                    )
                )
                elements.append(summary_row)

            footer_text = " | ".join(footer_lines) if footer_lines else get_school_profile().get(
                "school_motto", DEFAULT_SCHOOL_PROFILE["school_motto"]
            )

            if footer_path or footer_text:
                footer_reserve = max(16, footer_height + 2)
                elements.append(CondPageBreak(footer_reserve))
                elements.append(Spacer(1, footer_reserve))

            def _draw_results_page(canvas_obj, pdf_doc, page_num, total_pages):
                canvas_obj.saveState()
                canvas_obj.setStrokeColor(colors.HexColor("#1b5e20"))
                canvas_obj.setLineWidth(1.6)
                canvas_obj.rect(
                    border_inset,
                    border_inset,
                    pdf_doc.pagesize[0] - (border_inset * 2),
                    pdf_doc.pagesize[1] - (border_inset * 2),
                )

                inner_x = pdf_doc.leftMargin
                inner_width = pdf_doc.width
                page_height = pdf_doc.pagesize[1]

                if page_num == 1 and header_path and os.path.exists(header_path):
                    header_img = get_processed_letterhead_image(header_path, "header")
                    if header_img is not None:
                        canvas_obj.drawImage(
                            ImageReader(header_img),
                            inner_x,
                            page_height - content_inset - header_height,
                            width=inner_width,
                            height=header_height,
                            preserveAspectRatio=False,
                            mask="auto",
                        )

                page_number_y = 16
                if page_num == total_pages:
                    page_number_y = max(page_number_y, 18 + footer_height + 2)

                canvas_obj.setFont("Helvetica", 8)
                canvas_obj.setFillColor(colors.HexColor("#666666"))
                canvas_obj.drawRightString(
                    pdf_doc.pagesize[0] - content_inset,
                    page_number_y,
                    f"Page {page_num} of {total_pages}",
                )

                if page_num == total_pages:
                    if footer_path and os.path.exists(footer_path):
                        footer_img = get_processed_letterhead_image(
                            footer_path, "footer"
                        )
                        if footer_img is not None:
                            canvas_obj.drawImage(
                                ImageReader(footer_img),
                                inner_x,
                                content_inset,
                                width=inner_width,
                                height=footer_height,
                                preserveAspectRatio=False,
                                mask="auto",
                            )
                        else:
                            canvas_obj.drawCentredString(
                                pdf_doc.pagesize[0] / 2, 24, footer_text
                            )
                    else:
                        canvas_obj.drawCentredString(
                            pdf_doc.pagesize[0] / 2, 24, footer_text
                        )

                canvas_obj.restoreState()

            class ResultsPageCanvas(canvas.Canvas):
                def __init__(self, *args, **kwargs):
                    super().__init__(*args, **kwargs)
                    self._saved_page_states = []

                def showPage(self):
                    self._saved_page_states.append(dict(self.__dict__))
                    self._startPage()

                def save(self):
                    total_pages = len(self._saved_page_states)
                    for state in self._saved_page_states:
                        self.__dict__.update(state)
                        _draw_results_page(self, doc, self._pageNumber, total_pages)
                        super().showPage()
                    super().save()

            doc.build(elements, canvasmaker=ResultsPageCanvas)
            return True
        except Exception as exc:
            messagebox.showerror("Error", f"Failed to generate results PDF:\n{exc}")
            return False

    def generate_student_list_pdf(self, class_name, stream_name, file_path):
        """Generate a printable student list PDF with report-style letterhead."""
        try:
            students = db.get_students_by_class_and_stream(class_name, stream_name)
            if not students:
                messagebox.showwarning(
                    "No Data", f"No students found for {class_name} {stream_name}"
                )
                return False

            class_level = self._get_level_for_class(class_name)
            theme = self._get_level_theme(class_level)
            context = {"is_pp": self._is_pre_primary_level(class_level)}
            layout = self._get_report_layout_profile(context)

            letterhead_assets = get_letterhead_assets()
            header_path = letterhead_assets.get("header_path")
            footer_path = letterhead_assets.get("footer_path")
            footer_lines = letterhead_assets.get("footer_lines", [])
            meta_label_color = OLIVE_DARK
            meta_value_color = theme["title"]

            def _image_height(image_path, width, min_height, max_height):
                if not image_path or not os.path.exists(image_path):
                    return min_height
                try:
                    image_obj = get_processed_letterhead_image(
                        image_path, "header" if image_path == header_path else "footer"
                    )
                    if image_obj is None:
                        return min_height
                    scaled = width * image_obj.height / float(image_obj.width)
                    return max(min_height, min(max_height, scaled))
                except Exception:
                    return min_height

            pagesize = A4  # Portrait for student lists usually better
            border_inset = 10
            content_inset = 24
            max_content_width = pagesize[0] - (content_inset * 2)

            # Columns: #, Adm No, Student Name, Gender, Stream, Guardian, Email
            col_widths = [
                max_content_width * 0.05,  # #
                max_content_width * 0.12,  # Adm No
                max_content_width * 0.25,  # Student Name
                max_content_width * 0.08,  # Gender
                max_content_width * 0.10,  # Stream
                max_content_width * 0.18,  # Guardian
                max_content_width * 0.22,  # Email
            ]

            page_inner_width = sum(col_widths)
            header_height = _image_height(
                header_path,
                page_inner_width,
                layout["pdf_header_min"],
                layout["pdf_header_max"],
            )
            footer_height = _image_height(
                footer_path,
                page_inner_width,
                layout["pdf_footer_min"],
                layout["pdf_footer_max"],
            )

            doc = SimpleDocTemplate(
                file_path,
                pagesize=pagesize,
                rightMargin=content_inset,
                leftMargin=content_inset,
                topMargin=header_height + content_inset,
                bottomMargin=footer_height + content_inset,
            )

            styles = getSampleStyleSheet()
            styles.add(
                ParagraphStyle(
                    name="ResultsTitle",
                    parent=styles["Heading1"],
                    fontName="Helvetica-Bold",
                    fontSize=max(12, layout["pdf_title_font"]),
                    leading=layout["pdf_title_font"] + 2,
                    alignment=1,
                    textColor=colors.HexColor(theme["title"]),
                    spaceAfter=6,
                )
            )
            styles.add(
                ParagraphStyle(
                    name="ResultsMetaValue",
                    parent=styles["BodyText"],
                    fontName="Helvetica-Bold",
                    fontSize=max(8, layout["pdf_normal_font"]),
                    leading=layout["pdf_normal_font"] + 2,
                    textColor=colors.HexColor(theme["title"]),
                )
            )
            styles.add(
                ParagraphStyle(
                    name="ResultsCell",
                    parent=styles["BodyText"],
                    fontName="Helvetica",
                    fontSize=max(7, layout["pdf_normal_font"] - 1),
                    leading=max(8, layout["pdf_normal_font"]),
                    alignment=1,
                )
            )
            styles.add(
                ParagraphStyle(
                    name="ResultsCellLeft",
                    parent=styles["ResultsCell"],
                    alignment=0,
                )
            )

            elements = []
            if not header_path:
                elements.append(
                    Paragraph(
                        get_school_profile().get(
                            "school_name", DEFAULT_SCHOOL_PROFILE["school_name"]
                        ),
                        styles["ResultsTitle"],
                    )
                )

            elements.append(
                Table(
                    [[""]],
                    colWidths=[doc.width],
                    rowHeights=[2],
                    style=TableStyle(
                        [
                            (
                                "BACKGROUND",
                                (0, 0),
                                (-1, -1),
                                colors.HexColor(theme["title"]),
                            ),
                            ("LEFTPADDING", (0, 0), (-1, -1), 0),
                            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                            ("TOPPADDING", (0, 0), (-1, -1), 0),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                        ]
                    ),
                )
            )
            elements.append(Spacer(1, 6))

            title_text = f"STUDENT LIST - {class_name.upper()}"
            if stream_name:
                title_text += f" ({stream_name.upper()})"
            elements.append(Paragraph(title_text, styles["ResultsTitle"]))
            elements.append(Spacer(1, 4))

            info_items = [
                ("Class", class_name),
                ("Stream", stream_name if stream_name else "All Streams"),
                ("Date", datetime.now().strftime("%d %b %Y")),
                ("Total Students", str(len(students))),
            ]

            info_rows = [
                [
                    Paragraph(
                        f"<font color='{meta_label_color}'><b>{label}</b></font><br/>"
                        f"<font color='{meta_value_color}'>{value}</font>",
                        styles["ResultsMetaValue"],
                    )
                    for label, value in info_items
                ]
            ]
            info_table = Table(
                info_rows,
                colWidths=[page_inner_width / len(info_items)] * len(info_items),
                hAlign="CENTER",
            )
            info_table.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 8),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                        ("TOPPADDING", (0, 0), (-1, -1), 5),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                        (
                            "BACKGROUND",
                            (0, 0),
                            (-1, -1),
                            colors.HexColor(theme["header_bg"]),
                        ),
                        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor(theme["line"])),
                    ]
                )
            )
            elements.append(info_table)
            elements.append(Spacer(1, 12))

            table_headers = [
                Paragraph("<b>#</b>", styles["ResultsCell"]),
                Paragraph("<b>Adm No</b>", styles["ResultsCell"]),
                Paragraph("<b>Student Name</b>", styles["ResultsCellLeft"]),
                Paragraph("<b>Gnd</b>", styles["ResultsCell"]),
                Paragraph("<b>Stream</b>", styles["ResultsCell"]),
                Paragraph("<b>Guardian</b>", styles["ResultsCellLeft"]),
                Paragraph("<b>Contact Email</b>", styles["ResultsCellLeft"]),
            ]

            table_data = [table_headers]
            for idx, student in enumerate(students, 1):
                row = [
                    Paragraph(str(idx), styles["ResultsCell"]),
                    Paragraph(
                        str(student.get("admission_no", "")), styles["ResultsCell"]
                    ),
                    Paragraph(str(student.get("name", "")), styles["ResultsCellLeft"]),
                    Paragraph(
                        str(student.get("gender", "")[:1]), styles["ResultsCell"]
                    ),
                    Paragraph(str(student.get("stream", "")), styles["ResultsCell"]),
                    Paragraph(
                        str(student.get("guardian_name", "")), styles["ResultsCellLeft"]
                    ),
                    Paragraph(
                        str(student.get("parent_email", "")), styles["ResultsCellLeft"]
                    ),
                ]
                table_data.append(row)

            students_table = Table(
                table_data, colWidths=col_widths, repeatRows=1, hAlign="LEFT"
            )
            table_style = [
                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor(theme["grid"])),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor(theme["grid"])),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(theme["header_bg"])),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor(theme["muted"])),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
            for row_index in range(1, len(table_data)):
                row_bg = (
                    colors.white
                    if row_index % 2
                    else colors.HexColor(theme["accent_soft"])
                )
                table_style.append(
                    ("BACKGROUND", (0, row_index), (-1, row_index), row_bg)
                )

            students_table.setStyle(TableStyle(table_style))
            elements.append(students_table)

            footer_text = " | ".join(footer_lines) if footer_lines else get_school_profile().get(
                "school_motto", DEFAULT_SCHOOL_PROFILE["school_motto"]
            )

            def _draw_student_list_page(canvas_obj, pdf_doc, page_num, total_pages):
                canvas_obj.saveState()
                canvas_obj.setStrokeColor(colors.HexColor(theme["title"]))
                canvas_obj.setLineWidth(1.6)
                canvas_obj.rect(
                    border_inset,
                    border_inset,
                    pdf_doc.pagesize[0] - (border_inset * 2),
                    pdf_doc.pagesize[1] - (border_inset * 2),
                )

                inner_x = pdf_doc.leftMargin
                inner_width = pdf_doc.width
                page_height = pdf_doc.pagesize[1]

                if header_path and os.path.exists(header_path):
                    header_img = get_processed_letterhead_image(header_path, "header")
                    if header_img is not None:
                        canvas_obj.drawImage(
                            ImageReader(header_img),
                            inner_x,
                            page_height - content_inset + 8 - header_height,
                            width=inner_width,
                            height=header_height,
                            preserveAspectRatio=False,
                            mask="auto",
                        )

                page_number_y = 16
                if footer_path or footer_text:
                    page_number_y = max(page_number_y, 18 + footer_height + 2)

                canvas_obj.setFont("Helvetica", 8)
                canvas_obj.setFillColor(colors.HexColor("#666666"))
                canvas_obj.drawRightString(
                    pdf_doc.pagesize[0] - content_inset,
                    page_number_y,
                    f"Page {page_num} of {total_pages}",
                )

                if footer_path and os.path.exists(footer_path):
                    footer_img = get_processed_letterhead_image(footer_path, "footer")
                    if footer_img is not None:
                        canvas_obj.drawImage(
                            ImageReader(footer_img),
                            inner_x,
                            16,
                            width=inner_width,
                            height=footer_height,
                            preserveAspectRatio=False,
                            mask="auto",
                        )
                    else:
                        canvas_obj.drawCentredString(
                            pdf_doc.pagesize[0] / 2, 24, footer_text
                        )
                else:
                    canvas_obj.drawCentredString(
                        pdf_doc.pagesize[0] / 2, 24, footer_text
                    )

                canvas_obj.restoreState()

            class StudentListPageCanvas(canvas.Canvas):
                def __init__(self, *args, **kwargs):
                    super().__init__(*args, **kwargs)
                    self._saved_page_states = []

                def showPage(self):
                    self._saved_page_states.append(dict(self.__dict__))
                    self._startPage()

                def save(self):
                    total_pages = len(self._saved_page_states)
                    for state in self._saved_page_states:
                        self.__dict__.update(state)
                        _draw_student_list_page(
                            self, doc, self._pageNumber, total_pages
                        )
                        super().showPage()
                    super().save()

            doc.build(elements, canvasmaker=StudentListPageCanvas)
            return True
        except Exception as exc:
            messagebox.showerror(
                "Error", f"Failed to generate student list PDF:\n{exc}"
            )
            return False

    # ==================== WESTERN SPOTLIGHT EXPORT ====================
    def export_spotlight_excel(self):
        """Open a dialog to configure and export the Western Spotlight class report."""
        # Pre-populate from Result page if available
        try:
            initial_cls = self.rep_cls_cb.get()
            if initial_cls == "All":
                initial_cls = CLASSES[0]
            initial_term = self.rep_term_cb.get()
            initial_exam = self.rep_exam_cb.get()
            initial_stream = self._get_selected_results_stream() or "GREEN"
        except AttributeError:
            initial_cls = CLASSES[0]
            initial_term = TERMS[0]
            initial_exam = DEFAULT_EXAM_TYPE
            initial_stream = "GREEN"

        dlg = tk.Toplevel(self.root)
        dlg.title("Export – Western Spotlight")
        dlg.configure(bg=CONTENT_BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.resizable(False, False)
        dlg.minsize(440, 380)

        shell = tk.Frame(dlg, bg=CONTENT_BG)
        shell.pack(fill="both", expand=True, padx=18, pady=18)

        ws_bo, ws_bi = _card_colors("sand")
        outer = tk.Frame(shell, bg=ws_bo)
        outer.pack(fill="both", expand=True)
        card = tk.Frame(outer, bg=ws_bi, padx=30, pady=24)
        card.pack(fill="both", expand=True, padx=1, pady=1)
        card.columnconfigure(1, weight=1)

        tk.Label(
            card,
            text="Western Spotlight Export",
            bg=ws_bi,
            fg=TEXT_PRIMARY,
            font=(FF, 13, "bold"),
        ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 18))

        def row_field(r, label, widget):
            tk.Label(
                card,
                text=label,
                bg=ws_bi,
                fg=TEXT_SECONDARY,
                font=(FF, 10, "bold"),
                anchor="w",
            ).grid(row=r, column=0, sticky="w", padx=(0, 12), pady=5)
            widget.grid(row=r, column=1, sticky="ew", pady=5, ipady=4)

        cls_cb = ttk.Combobox(
            card,
            values=self.get_current_classes(),
            state="readonly",
            style="App.TCombobox",
            width=20,
        )
        cls_cb.set(initial_cls)
        row_field(1, "Class:", cls_cb)

        term_cb = ttk.Combobox(
            card, values=TERMS, state="readonly", style="App.TCombobox", width=20
        )
        term_cb.set(initial_term)
        row_field(2, "Term:", term_cb)

        exam_cb = ttk.Combobox(
            card, values=EXAM_TYPES, state="readonly", style="App.TCombobox", width=20
        )
        exam_cb.set(initial_exam)
        row_field(3, "Exam Type:", exam_cb)

        stream_var = tk.StringVar(value=initial_stream)
        stream_e = ttk.Entry(
            card, textvariable=stream_var, style="App.TEntry", width=20
        )
        row_field(4, "Stream / Group:", stream_e)

        assess_cb = ttk.Combobox(
            card,
            values=["MID-TERM", "END-TERM"],
            state="readonly",
            style="App.TCombobox",
            width=20,
        )
        assess_cb.set("MID-TERM")
        row_field(5, "Assessment:", assess_cb)

        year_var = tk.StringVar(value=str(datetime.now().year))
        year_e = ttk.Entry(card, textvariable=year_var, style="App.TEntry", width=20)
        row_field(6, "Year:", year_e)

        btn_row = tk.Frame(card, bg=ws_bi)
        btn_row.grid(row=7, column=0, columnspan=2, sticky="e", pady=(20, 0))

        cancel = tk.Label(
            btn_row,
            text="Cancel",
            bg="#e8f5e9",
            fg=TEXT_PRIMARY,
            font=(FF, 10, "bold"),
            padx=18,
            pady=8,
            cursor="hand2",
        )
        cancel.pack(side="left", padx=(0, 8))
        cancel.bind("<Button-1>", lambda e: dlg.destroy())

        def do_export():
            selected_cls = cls_cb.get()
            selected_term = term_cb.get()
            selected_exam = exam_cb.get()
            selected_stream = stream_var.get().strip() or "GREEN"
            selected_assess = assess_cb.get()
            selected_year = year_var.get().strip() or str(datetime.now().year)
            dlg.destroy()
            self._do_spotlight_export(
                selected_cls,
                selected_term,
                selected_exam,
                selected_stream,
                selected_assess,
                selected_year,
            )

        export_btn = tk.Label(
            btn_row,
            text="Export Excel",
            bg="#1B5E20",
            fg="white",
            font=(FF, 10, "bold"),
            padx=18,
            pady=8,
            cursor="hand2",
        )
        export_btn.pack(side="left")
        export_btn.bind("<Button-1>", lambda e: do_export())

        dlg.update_idletasks()
        req_w = max(440, outer.winfo_reqwidth() + 36)
        req_h = max(380, outer.winfo_reqheight() + 36)
        screen_w = dlg.winfo_screenwidth()
        screen_h = dlg.winfo_screenheight()
        win_w = min(req_w, screen_w - 120)
        win_h = min(req_h, screen_h - 120)
        parent_x = self.root.winfo_rootx()
        parent_y = self.root.winfo_rooty()
        parent_w = self.root.winfo_width() or self.root.winfo_reqwidth()
        parent_h = self.root.winfo_height() or self.root.winfo_reqheight()
        pos_x = parent_x + max(0, (parent_w - win_w) // 2)
        pos_y = parent_y + max(0, (parent_h - win_h) // 2)
        dlg.geometry(f"{win_w}x{win_h}+{pos_x}+{pos_y}")

    def _do_spotlight_export(self, cls, term, exam_type, stream, assess, year):
        """Generate the Western Spotlight Excel workbook and save it."""
        results = self._get_ranked_results(cls, term, exam_type, year)
        if not results:
            messagebox.showwarning(
                "No Data",
                f"No results found for {cls}, Term {term}.\nPlease enter marks first.",
            )
            return

        file_path = filedialog.asksaveasfilename(
            title="Save – Western Spotlight",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"spotlight_{cls.replace(' ', '_')}_T{term}_{exam_type.replace('-', '_')}_{year}.xlsx",
        )
        if not file_path:
            return

        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Spotlight"
            ws.sheet_view.showGridLines = False

            # ── Palette ──────────────────────────────────────────────────────
            BG_TITLE = PatternFill("solid", fgColor="1B5E20")  # dark green – titles
            BG_HDR = PatternFill("solid", fgColor="2E7D32")  # medium green – headers
            BG_DATA = PatternFill("solid", fgColor="4CAF50")  # bright green – data rows
            FNT_YLW = Font(bold=True, color="FFFF00", size=12, name="Calibri")
            FNT_YLW_S = Font(bold=True, color="FFFF00", size=10, name="Calibri")
            FNT_WHT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
            FNT_WHT_S = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
            ALIGN_CTR = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            ALIGN_LFT = Alignment(horizontal="left", vertical="center")
            _thin = Side(style="thin", color="000000")
            BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

            subjects = self._get_subjects_for_scope(
                cls, term, exam_type, results, year
            )
            n_subj = len(subjects)
            total_max = n_subj * 100

            # Column numbers (1-based)
            C_NO = 1
            C_NAME = 2
            C_S0 = 3  # first subject score col
            C_TOTAL = C_S0 + n_subj * 2
            C_AVG = C_TOTAL + 1
            C_PSN = C_AVG + 1
            C_LAST = C_PSN
            ROW_OFFSET = 0

            def cl(c):
                return get_column_letter(c)

            def apply_style(r, c, fill=None, font=None, align=ALIGN_CTR, border=None):
                cell = ws.cell(row=r, column=c)
                if fill:
                    cell.fill = fill
                if font:
                    cell.font = font
                if align:
                    cell.alignment = align
                if border:
                    cell.border = border
                return cell

            def fill_entire_row(r, fill, font, height=None):
                for c in range(1, C_LAST + 1):
                    apply_style(r, c, fill=fill, font=font)
                if height:
                    ws.row_dimensions[r].height = height

            def merged_cell(
                r1, c1, r2, c2, value="", fill=None, font=None, align=ALIGN_CTR
            ):
                ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
                cell = ws.cell(row=r1, column=c1)
                cell.value = value
                if fill:
                    cell.fill = fill
                if font:
                    cell.font = font
                if align:
                    cell.alignment = align
                return cell

            letterhead_path = ensure_letterhead_assets()
            if letterhead_path and os.path.exists(letterhead_path):
                try:
                    xl_img = XLImage(letterhead_path)
                    max_width = 980
                    if xl_img.width > max_width:
                        ratio = max_width / float(xl_img.width)
                        xl_img.width = int(xl_img.width * ratio)
                        xl_img.height = int(xl_img.height * ratio)
                    ws.add_image(xl_img, "A1")
                    ROW_OFFSET = max(6, int(max(120, xl_img.height) / 20) + 1)
                    for row_idx in range(1, ROW_OFFSET + 1):
                        ws.row_dimensions[row_idx].height = 18
                except Exception as img_exc:
                    print(f"Excel letterhead load error: {img_exc}")

            # ── Title section (rows 1-5) ─────────────────────────────────────
            grade_num = cls.replace("Grade ", "")
            grade_words = {
                "1": "ONE",
                "2": "TWO",
                "3": "THREE",
                "4": "FOUR",
                "5": "FIVE",
                "6": "SIX",
                "7": "SEVEN",
                "8": "EIGHT",
                "9": "NINE",
            }.get(grade_num, grade_num)

            title_row = ROW_OFFSET + 1
            blank_row_1 = title_row + 1
            subtitle_row = title_row + 2
            blank_row_2 = title_row + 3
            banner_row = title_row + 4
            header_row_1 = title_row + 5
            header_row_2 = title_row + 6
            first_data_row = title_row + 7

            fill_entire_row(title_row, BG_TITLE, FNT_YLW, height=26)
            merged_cell(
                title_row,
                C_NO,
                title_row,
                C_LAST,
                "MT.  OLIVES ADVENTIST SCHOOL,  NGONG",
                BG_TITLE,
                FNT_YLW,
                ALIGN_CTR,
            )

            fill_entire_row(blank_row_1, BG_TITLE, FNT_YLW, height=8)
            merged_cell(blank_row_1, C_NO, blank_row_1, C_LAST, "", BG_TITLE, FNT_YLW)

            title3 = (
                f"GRADE {grade_words} ({grade_num}) {stream} "
                f"TERM {term.upper()} {assess} ASSESSMENT REPORT {year}"
            )
            fill_entire_row(subtitle_row, BG_TITLE, FNT_YLW, height=22)
            merged_cell(
                subtitle_row,
                C_NO,
                subtitle_row,
                C_LAST,
                title3,
                BG_TITLE,
                FNT_YLW,
                ALIGN_CTR,
            )

            fill_entire_row(blank_row_2, BG_TITLE, FNT_YLW, height=8)
            merged_cell(blank_row_2, C_NO, blank_row_2, C_LAST, "", BG_TITLE, FNT_YLW)

            fill_entire_row(banner_row, BG_TITLE, FNT_YLW, height=22)
            merged_cell(
                banner_row,
                C_NO,
                banner_row,
                C_LAST,
                "THE WESTERN SPOTLIGHT",
                BG_TITLE,
                FNT_YLW,
                ALIGN_CTR,
            )

            # ── Column header rows (6 & 7) ───────────────────────────────────
            ws.row_dimensions[header_row_1].height = 20
            ws.row_dimensions[header_row_2].height = 18
            for c in range(1, C_LAST + 1):
                for r in (header_row_1, header_row_2):
                    apply_style(
                        r, c, fill=BG_HDR, font=FNT_WHT, align=ALIGN_CTR, border=BORDER
                    )

            # NO.  – merged rows 6-7
            merged_cell(
                header_row_1,
                C_NO,
                header_row_2,
                C_NO,
                "NO.",
                BG_HDR,
                FNT_WHT,
                ALIGN_CTR,
            )
            for r in (header_row_1, header_row_2):
                ws.cell(r, C_NO).border = BORDER

            # LEARNER – merged rows 6-7
            merged_cell(
                header_row_1,
                C_NAME,
                header_row_2,
                C_NAME,
                "LEARNER",
                BG_HDR,
                FNT_WHT,
                ALIGN_CTR,
            )
            for r in (header_row_1, header_row_2):
                ws.cell(r, C_NAME).border = BORDER

            # Subject names row 6 (merged score+grade cols), row 7 sub-labels
            for i, subj in enumerate(subjects):
                sc = C_S0 + i * 2  # score col
                gc = sc + 1  # grade col
                lbl = self._get_subject_label(subj, cls).replace("\n", " / ").upper()
                merged_cell(
                    header_row_1, sc, header_row_1, gc, lbl, BG_HDR, FNT_WHT, ALIGN_CTR
                )
                for c in (sc, gc):
                    ws.cell(header_row_1, c).border = BORDER
                # Row 7: '100' under score, blank under grade
                c7 = ws.cell(header_row_2, sc)
                c7.value = "100"
                c7.fill = BG_HDR
                c7.font = FNT_WHT_S
                c7.alignment = ALIGN_CTR
                c7.border = BORDER
                ws.cell(header_row_2, gc).border = BORDER

            # TOTAL
            ws.cell(header_row_1, C_TOTAL).value = "TOTAL"
            ws.cell(header_row_2, C_TOTAL).value = str(total_max)
            for r in (header_row_1, header_row_2):
                apply_style(
                    r,
                    C_TOTAL,
                    fill=BG_HDR,
                    font=FNT_WHT,
                    align=ALIGN_CTR,
                    border=BORDER,
                )

            # AVERAGE
            ws.cell(header_row_1, C_AVG).value = "AVERAGE"
            ws.cell(header_row_2, C_AVG).value = "100%"
            for r in (header_row_1, header_row_2):
                apply_style(
                    r, C_AVG, fill=BG_HDR, font=FNT_WHT, align=ALIGN_CTR, border=BORDER
                )

            # PSN – merged rows 6-7
            merged_cell(
                header_row_1,
                C_PSN,
                header_row_2,
                C_PSN,
                "PSN",
                BG_HDR,
                FNT_WHT,
                ALIGN_CTR,
            )
            for r in (header_row_1, header_row_2):
                ws.cell(r, C_PSN).border = BORDER

            # ── Data rows ────────────────────────────────────────────────────
            for idx, result in enumerate(results):
                r = first_data_row + idx
                ws.row_dimensions[r].height = 16
                s = result["student"]
                mk = result["marks"]

                def dc(col, value, align=ALIGN_CTR):
                    cell = ws.cell(row=r, column=col)
                    cell.value = value
                    cell.fill = BG_DATA
                    cell.font = FNT_YLW_S
                    cell.alignment = align
                    cell.border = BORDER

                dc(C_NO, result["position"])
                dc(C_NAME, s["name"].upper(), ALIGN_LFT)

                for i, subj in enumerate(subjects):
                    sc = C_S0 + i * 2
                    gc = sc + 1
                    raw = mk.get(subj)
                    mark_val = int(raw) if raw else 0
                    dc(sc, mark_val if raw else "")
                    dc(gc, get_cbc_grade_sublevel(mark_val) if raw else "")

                dc(C_TOTAL, result["total"])
                dc(C_AVG, result["average"])
                dc(C_PSN, result["position"])

            # ── Column widths ─────────────────────────────────────────────────
            ws.column_dimensions[cl(C_NO)].width = 5
            ws.column_dimensions[cl(C_NAME)].width = 24
            for i in range(n_subj):
                sc = C_S0 + i * 2
                ws.column_dimensions[cl(sc)].width = 5.5
                ws.column_dimensions[cl(sc + 1)].width = 5.5
            ws.column_dimensions[cl(C_TOTAL)].width = 7
            ws.column_dimensions[cl(C_AVG)].width = 9
            ws.column_dimensions[cl(C_PSN)].width = 5

            # ── Print settings (A4 landscape) ─────────────────────────────────
            ws.page_setup.orientation = "landscape"
            ws.page_setup.paperSize = 9  # A4
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.freeze_panes = f"A{first_data_row}"
            header_lines = get_letterhead_print_lines()
            school_profile = get_school_profile()
            left_header = (
                header_lines[0]
                if header_lines
                else school_profile.get(
                    "school_name", DEFAULT_SCHOOL_PROFILE["school_name"]
                )
            )
            center_header = header_lines[1] if len(header_lines) > 1 else ""
            right_header = header_lines[2] if len(header_lines) > 2 else ""
            footer_line = (
                header_lines[3]
                if len(header_lines) > 3
                else school_profile.get(
                    "school_motto", DEFAULT_SCHOOL_PROFILE["school_motto"]
                )
            )
            ws.oddHeader.left.text = left_header
            ws.oddHeader.center.text = center_header
            ws.oddHeader.right.text = right_header
            ws.oddFooter.center.text = f"{footer_line}    Page &[Page] of &[Pages]"
            ws.oddFooter.right.text = "&[Date] &[Time]"
            ws.page_margins.top = 0.6
            ws.page_margins.header = 0.3
            ws.page_margins.footer = 0.3

            wb.save(file_path)
            messagebox.showinfo(
                "Export Complete", f"Western Spotlight report saved:\n{file_path}"
            )

        except ImportError:
            messagebox.showerror(
                "Missing Library", "openpyxl is required. Run:\n  pip install openpyxl"
            )
        except Exception as exc:
            messagebox.showerror("Export Error", f"Failed to generate file:\n{exc}")

    def generate_pdf_report(self, student_id=None):
        """Generate a PDF report card for a student."""
        if not student_id:
            # Try to get selected student from the StudentsTab if present
            if (
                hasattr(self, "students_tab")
                and self.students_tab
                and self.students_tab.students_table
            ):
                selected_iids = self.students_tab.students_table.get_selected_iids()
                if not selected_iids:
                    messagebox.showwarning("Warning", "Please select a student first!")
                    return
                student_id = selected_iids[0]
            elif hasattr(self, "students_tree"):
                sel = self.students_tree.selection()
                if not sel:
                    messagebox.showwarning("Warning", "Please select a student first!")
                    return
                item = self.students_tree.item(sel[0])
                student_id = item["tags"][0] if item["tags"] else sel[0]
            else:
                messagebox.showwarning("Warning", "Please select a student first!")
                return

        cls = "Grade 7"
        term = "One"
        exam_type = DEFAULT_EXAM_TYPE
        academic_year = str(datetime.now().year)
        if hasattr(self, "rc_exam_cb"):
            exam_type = self.rc_exam_cb.get() or DEFAULT_EXAM_TYPE
            academic_year = self.rc_year_cb.get() or academic_year
        elif hasattr(self, "marks_exam_cb"):
            exam_type = self.marks_exam_cb.get() or DEFAULT_EXAM_TYPE
            academic_year = self.marks_year_cb.get() or academic_year

        results = self._get_ranked_results(cls, term, exam_type, academic_year)
        result = next((r for r in results if r["student"]["id"] == student_id), None)

        if not result:
            for c in self.get_current_classes():
                if c == cls:
                    continue
                results = self._get_ranked_results(c, term, exam_type, academic_year)
                result = next(
                    (r for r in results if r["student"]["id"] == student_id), None
                )
                if result:
                    cls = c
                    break

        if not result:
            messagebox.showerror(
                "Error", "Could not find results for this student in the current term."
            )
            return

        s = result["student"]
        year_text = self._get_report_card_context(result, term, exam_type).get(
            "year", str(datetime.now().year)
        )
        file_path = filedialog.asksaveasfilename(
            title="Save PDF Report",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=f"{self._get_report_card_file_basename(s, term, year_text)}.pdf",
        )
        if not file_path:
            return

        self._build_report_card_pdf(result, len(results), term, exam_type, file_path)

    def _build_report_card_pdf(
        self, result, total_students, term, exam_type, file_path
    ):
        """Build the styled PDF report card used by preview/print flows."""
        try:
            s = result["student"]
            context = self._get_report_card_context(result, term, exam_type)
            layout = dict(self._get_report_layout_profile(context))
            layout["pdf_title_font"] = max(13, layout["pdf_title_font"] - 1)
            layout["pdf_normal_font"] = max(7, layout["pdf_normal_font"] - 1)
            layout["pdf_legend_font"] = max(6, layout["pdf_legend_font"] - 1)
            layout["pdf_row_pad"] = max(3, layout["pdf_row_pad"] - 2)
            layout["pdf_section_gap"] = max(6, layout["pdf_section_gap"] - 4)
            layout["pdf_header_min"] = max(48, layout["pdf_header_min"] - 8)
            layout["pdf_header_max"] = max(
                layout["pdf_header_min"], layout["pdf_header_max"] - 12
            )
            layout["pdf_footer_min"] = max(22, layout["pdf_footer_min"] - 6)
            layout["pdf_footer_max"] = max(
                layout["pdf_footer_min"], layout["pdf_footer_max"] - 10
            )
            cls_level = context["class_level"]
            is_pp = context["is_pp"]
            subjects = context["subjects"]
            facilitator = context.get("class_teacher_name") or ""
            comment_text = context.get("comment_text") or ""
            pp_scales = context.get("grade_scales", [])
            pp_codes = context.get("grade_codes", ["EE", "ME", "AE", "BE"])
            theme = self._get_level_theme(cls_level)
            pdf_title_color = colors.HexColor(theme["title"])
            pdf_grid_color = colors.HexColor(theme["grid"])
            pdf_line_color = colors.HexColor(theme["line"])
            pdf_muted_color = colors.HexColor(theme["muted"])
            pdf_header_bg = colors.HexColor(theme["header_bg"])
            pdf_accent_soft = colors.HexColor(
                theme.get("accent_soft", theme["header_bg"])
            )

            styles = getSampleStyleSheet()

            # Custom styles
            styles.add(
                ParagraphStyle(
                    name="SchoolName",
                    parent=styles["Heading1"],
                    fontName="Helvetica-Bold",
                    fontSize=20,
                    textColor=pdf_title_color,
                    alignment=1,
                )
            )
            styles.add(
                ParagraphStyle(
                    name="SchoolInfo",
                    parent=styles["Normal"],
                    fontSize=layout["pdf_normal_font"],
                    textColor=colors.gray,
                    alignment=1,
                )
            )
            styles.add(
                ParagraphStyle(
                    name="ReportTitle",
                    parent=styles["Heading2"],
                    fontName="Helvetica-Bold",
                    fontSize=layout["pdf_title_font"],
                    textColor=pdf_title_color,
                    alignment=1,
                    spaceBefore=8,
                    spaceAfter=8,
                )
            )
            styles.add(
                ParagraphStyle(
                    name="FooterLine",
                    parent=styles["Normal"],
                    fontSize=8,
                    textColor=colors.HexColor("#666666"),
                    alignment=1,
                )
            )
            styles.add(
                ParagraphStyle(
                    name="LegendText",
                    parent=styles["Normal"],
                    fontSize=layout["pdf_legend_font"],
                    alignment=1,
                )
            )
            styles.add(
                ParagraphStyle(
                    name="ReportBody",
                    parent=styles["Normal"],
                    fontSize=layout["pdf_normal_font"],
                    leading=layout["pdf_normal_font"] + 2,
                    textColor=colors.HexColor("#4f4f4f"),
                )
            )
            styles.add(
                ParagraphStyle(
                    name="SubjectLabel",
                    parent=styles["Normal"],
                    fontSize=max(7, layout["pdf_normal_font"] - 1),
                    leading=max(8, layout["pdf_normal_font"] + 1),
                    textColor=colors.HexColor("#2f2f2f"),
                    alignment=0,
                )
            )

            letterhead_assets = get_letterhead_assets()
            header_path = letterhead_assets["header_path"]
            footer_path = letterhead_assets["footer_path"]
            header_lines = letterhead_assets["header_lines"]
            footer_lines = letterhead_assets["footer_lines"]
            term_marks = self._get_student_term_marks(
                s["id"], term, context.get("year")
            )
            matrix_spec = self._get_report_assessment_matrix_spec(
                context,
                assessment_types=term_marks.keys(),
                include_exam_type=exam_type,
            )
            assessment_titles = matrix_spec["assessment_matrix_titles"]
            assessment_order = matrix_spec["assessment_order"]
            scale_codes_sorted = matrix_spec["scale_codes"]

            def _image_height(image_path, target_width, min_height, max_height):
                if not image_path or not os.path.exists(image_path):
                    return min_height
                try:
                    img = get_processed_letterhead_image(
                        image_path, "header" if image_path == header_path else "footer"
                    )
                    if img is None:
                        return min_height
                    scaled = int(img.height * target_width / img.width)
                    return max(min_height, min(max_height, scaled))
                except Exception:
                    return min_height

            # ── Orientation: switch to landscape when grade columns are too many ──
            _early_codes = [
                c
                for c in (
                    context.get("grade_codes") or pp_codes or ["EE", "ME", "AE", "BE"]
                )
                if c
            ]
            _num_early_cols = len(_early_codes) * max(1, len(assessment_order))
            _preferred_code_w = layout.get("pdf_marks_code_width_pp", 28)
            _portrait_inner = A4[0] - 40  # ~555 pt
            needs_landscape = (
                100 + _preferred_code_w * _num_early_cols
            ) > _portrait_inner
            pagesize = landscape(A4) if needs_landscape else portrait(A4)

            page_inner_width = pagesize[0] - 40
            header_height = _image_height(
                header_path,
                page_inner_width,
                layout["pdf_header_min"],
                layout["pdf_header_max"],
            )
            footer_height = _image_height(
                footer_path,
                page_inner_width,
                layout["pdf_footer_min"],
                layout["pdf_footer_max"],
            )
            top_margin = header_height + 20
            bottom_margin = footer_height + 20

            doc = SimpleDocTemplate(
                file_path,
                pagesize=pagesize,
                rightMargin=20,
                leftMargin=20,
                topMargin=top_margin,
                bottomMargin=bottom_margin,
            )

            elements = []

            if not header_path:
                elements.append(
                    Paragraph(
                        get_school_profile().get(
                            "school_name", DEFAULT_SCHOOL_PROFILE["school_name"]
                        ),
                        styles["SchoolName"],
                    )
                )
            divider = Table([[""]], colWidths=[doc.width], rowHeights=[2])
            divider.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, -1), pdf_title_color),
                        ("LEFTPADDING", (0, 0), (-1, -1), 0),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                        ("TOPPADDING", (0, 0), (-1, -1), 0),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                    ]
                )
            )
            elements.append(divider)
            elements.append(Spacer(1, layout["pdf_section_gap"]))

            rpt_title_text = self._get_report_title_for_level(cls_level, is_pp).replace(
                "\n", "<br/>"
            )
            border_table = Table(
                [[Paragraph(rpt_title_text, styles["ReportTitle"])]],
                colWidths=[layout["pdf_title_width"]],
            )
            border_table.setStyle(
                TableStyle(
                    [
                        ("BOX", (0, 0), (-1, -1), 1, pdf_title_color),
                        ("LEFTPADDING", (0, 0), (-1, -1), 14),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 14),
                        ("TOPPADDING", (0, 0), (-1, -1), 8),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ]
                )
            )
            elements.append(border_table)
            elements.append(Spacer(1, layout["pdf_section_gap"]))

            # Student Info Grid
            stream_display = s.get("stream", "").strip() or s.get("admission_no", "")
            if is_pp:
                info_data = [
                    [
                        "GRADE:",
                        s["class"],
                        "TERM:",
                        self._format_report_card_term_display(term).upper(),
                        "YEAR:",
                        context.get("year", str(datetime.now().year)),
                    ],
                    ["STUDENT NAME:", s["name"].upper(), "", "", "", ""],
                    [
                        "GRADE FACILITATOR:",
                        (facilitator or " ").title(),
                        "",
                        "",
                        "",
                        "",
                    ],
                ]
                info_table = Table(
                    info_data,
                    colWidths=[100, 180, 55, 120, 45, max(doc.width - 500, 60)],
                    hAlign="LEFT",
                )
            else:
                yr = context.get("year", str(datetime.now().year))
                # All 3 rows use 6 columns so colWidths applies uniformly.
                # Rows 1-2 span their value cells across the filler column.
                info_data = [
                    [
                        "GRADE",
                        s["class"],
                        "TERM",
                        self._format_report_card_term_display(term).upper(),
                        "YEAR",
                        yr,
                    ],
                    [
                        "NAME OF THE LEARNER",
                        s["name"].upper(),
                        "",
                        "NAME OF THE FACILITATOR",
                        (facilitator or " ").upper(),
                        "",
                    ],
                    ["STREAM / ADM NO", stream_display, "", "GENDER", s["gender"], ""],
                ]
                _ic1, _ic2, _ic3 = 65, 90, 42  # GRADE lbl, grade val, TERM lbl
                _ic4, _ic5 = 110, 68  # term val, YEAR lbl
                _ic6 = max(10, doc.width - _ic1 - _ic2 - _ic3 - _ic4 - _ic5)
                info_table = Table(
                    info_data,
                    colWidths=[_ic1, _ic2, _ic3, _ic4, _ic5, _ic6],
                    hAlign="LEFT",
                )

            info_style = [
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), layout["pdf_normal_font"]),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), layout["pdf_row_pad"] + 1),
                ("TOPPADDING", (0, 0), (-1, -1), layout["pdf_row_pad"] + 1),
            ]
            if is_pp:
                info_style.extend(
                    [
                        ("LINEBELOW", (1, 0), (1, 0), 0.7, pdf_line_color),
                        ("LINEBELOW", (3, 0), (3, 0), 0.7, pdf_line_color),
                        ("LINEBELOW", (5, 0), (5, 0), 0.7, pdf_line_color),
                        ("SPAN", (1, 1), (5, 1)),
                        ("SPAN", (1, 2), (5, 2)),
                        ("LINEBELOW", (1, 1), (5, 1), 0.7, pdf_line_color),
                        ("LINEBELOW", (1, 2), (5, 2), 0.7, pdf_line_color),
                    ]
                )
            else:
                info_style.extend(
                    [
                        # Row 0: underline grade, term and year values
                        ("LINEBELOW", (1, 0), (1, 0), 0.7, pdf_line_color),
                        ("LINEBELOW", (3, 0), (3, 0), 0.7, pdf_line_color),
                        ("LINEBELOW", (5, 0), (5, 0), 0.7, pdf_line_color),
                        # Row 1: name spans cols 1-2, facilitator spans cols 4-5
                        ("SPAN", (1, 1), (2, 1)),
                        ("SPAN", (4, 1), (5, 1)),
                        ("LINEBELOW", (1, 1), (2, 1), 0.7, pdf_line_color),
                        ("LINEBELOW", (4, 1), (5, 1), 0.7, pdf_line_color),
                        # Row 2: stream spans cols 1-2, gender spans cols 4-5
                        ("SPAN", (1, 2), (2, 2)),
                        ("SPAN", (4, 2), (5, 2)),
                        ("LINEBELOW", (1, 2), (2, 2), 0.7, pdf_line_color),
                        ("LINEBELOW", (4, 2), (5, 2), 0.7, pdf_line_color),
                    ]
                )
            info_table.setStyle(TableStyle(info_style))
            elements.append(info_table)
            elements.append(Spacer(1, layout["pdf_section_gap"]))

            # Preview-style aligned assessment matrix for all report-card PDFs
            marks_data = [["LEARNING AREAS"], [""]]
            for title in assessment_titles:
                marks_data[0].extend([title] + [""] * (len(scale_codes_sorted) - 1))
                marks_data[1].extend(scale_codes_sorted)
            repeat_rows = 2

            for subj in subjects:
                row = [
                    Paragraph(
                        self._get_subject_label(
                            subj, s.get("class", ""), multiline=True
                        ),
                        styles["SubjectLabel"],
                    )
                ]
                for assessment in assessment_order:
                    raw_mark = term_marks.get(assessment, {}).get(subj, "")
                    if str(raw_mark).strip() == "" and assessment == exam_type:
                        raw_mark = result.get("marks", {}).get(subj, "")
                    grade_code = ""
                    try:
                        if str(raw_mark).strip() != "":
                            grade_code = self._get_grade_code_for_class(
                                float(raw_mark), s.get("class", "")
                            )
                    except (TypeError, ValueError):
                        grade_code = ""
                    display_mark = (
                        self._format_report_mark_value(raw_mark)
                        if str(raw_mark).strip() != ""
                        else ""
                    )
                    for code in scale_codes_sorted:
                        row.append(display_mark if grade_code == code else "")
                marks_data.append(row)

            preferred_left = layout.get("pdf_marks_left_width_pp", 155)
            preferred_code = layout.get("pdf_marks_code_width_pp", 28)
            min_left = 100
            min_code = 16
            num_mark_cols = len(scale_codes_sorted) * len(assessment_order)
            left_width = min(preferred_left, max(min_left, doc.width * 0.28))
            remaining_width = max(doc.width - left_width, 0)
            code_width = max(
                min_code, min(preferred_code, remaining_width / max(1, num_mark_cols))
            )
            if left_width + code_width * num_mark_cols > doc.width:
                code_width = max(
                    min_code, (doc.width - min_left) / max(1, num_mark_cols)
                )
                left_width = max(min_left, doc.width - code_width * num_mark_cols)
            col_widths = [left_width] + [code_width] * num_mark_cols
            marks_table = Table(
                marks_data, colWidths=col_widths, repeatRows=repeat_rows, hAlign="LEFT"
            )
            marks_table_style = [
                ("BOX", (0, 0), (-1, -1), 0.8, pdf_grid_color),
                ("LINEBELOW", (0, 0), (-1, 0), 0.6, pdf_grid_color),
                ("LINEBELOW", (0, 1), (-1, 1), 0.5, pdf_grid_color),
                ("FONTNAME", (0, 0), (-1, 1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 1), layout["pdf_normal_font"]),
                ("FONTSIZE", (0, 2), (-1, -1), layout["pdf_normal_font"]),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("ALIGN", (0, 0), (0, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("SPAN", (0, 0), (0, repeat_rows - 1)),
                ("BACKGROUND", (0, 0), (-1, 0), pdf_header_bg),
                ("BACKGROUND", (0, 1), (-1, 1), pdf_accent_soft),
                ("TEXTCOLOR", (0, 0), (-1, 1), pdf_muted_color),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), layout["pdf_row_pad"]),
                ("BOTTOMPADDING", (0, 0), (-1, -1), layout["pdf_row_pad"]),
            ]
            current_col = 1
            for _ in assessment_order:
                marks_table_style.append(
                    (
                        "SPAN",
                        (current_col, 0),
                        (current_col + len(scale_codes_sorted) - 1, 0),
                    )
                )
                current_col += len(scale_codes_sorted)

            group_width = len(scale_codes_sorted)
            for assessment_idx in range(1, len(assessment_order)):
                group_start_col = 1 + (assessment_idx * group_width)
                marks_table_style.append(
                    (
                        "LINEBEFORE",
                        (group_start_col, 0),
                        (group_start_col, -1),
                        0.8,
                        pdf_grid_color,
                    )
                )

            for row_index in range(repeat_rows, len(marks_data)):
                row_bg = colors.white if row_index % 2 == 0 else pdf_accent_soft
                marks_table_style.append(
                    ("BACKGROUND", (0, row_index), (-1, row_index), row_bg)
                )
                marks_table_style.append(
                    ("LINEBELOW", (0, row_index), (-1, row_index), 0.35, pdf_grid_color)
                )

            marks_table.setStyle(TableStyle(marks_table_style))
            elements.append(marks_table)
            elements.append(Spacer(1, layout["pdf_section_gap"]))

            comments_table = Table(
                [
                    [
                        Paragraph(
                            f"<b>General Comments</b><br/><font color='#4f4f4f'>{comment_text or ' '}</font>",
                            styles["ReportBody"],
                        )
                    ]
                ],
                colWidths=[doc.width],
            )
            comments_table.setStyle(
                TableStyle(
                    [
                        ("BOX", (0, 0), (-1, -1), 0.8, pdf_grid_color),
                        ("BACKGROUND", (0, 0), (-1, -1), pdf_accent_soft),
                        ("LEFTPADDING", (0, 0), (-1, -1), 10),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                        ("TOPPADDING", (0, 0), (-1, -1), layout["pdf_row_pad"] + 1),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), layout["pdf_row_pad"] + 2),
                    ]
                )
            )
            elements.append(comments_table)
            elements.append(Spacer(1, layout["pdf_section_gap"]))

            sign_table = Table(
                [
                    [
                        Paragraph(
                            f"<b>Grade Facilitator Signature</b><br/><br/>__________________________<br/>{facilitator or ' '}",
                            styles["ReportBody"],
                        ),
                        Paragraph(
                            "<b>Head Teacher Signature</b><br/><br/>__________________________",
                            styles["ReportBody"],
                        ),
                    ]
                ],
                colWidths=[doc.width / 2.0, doc.width / 2.0],
            )
            sign_table.setStyle(
                TableStyle(
                    [
                        ("BOX", (0, 0), (-1, -1), 0.8, pdf_grid_color),
                        ("BACKGROUND", (0, 0), (0, 0), pdf_accent_soft),
                        ("BACKGROUND", (1, 0), (1, 0), pdf_accent_soft),
                        ("LEFTPADDING", (0, 0), (-1, -1), 10),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                        ("TOPPADDING", (0, 0), (-1, -1), layout["pdf_row_pad"] + 1),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), layout["pdf_row_pad"] + 2),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]
                )
            )
            elements.append(sign_table)

            footer_text = " | ".join(footer_lines) if footer_lines else get_school_profile().get(
                "school_motto", DEFAULT_SCHOOL_PROFILE["school_motto"]
            )

            def add_border(canvas, doc):
                canvas.saveState()
                canvas.setStrokeColor(colors.HexColor("#1b5e20"))
                canvas.setLineWidth(2)
                canvas.rect(10, 10, A4[0] - 20, A4[1] - 20)

                inner_x = doc.leftMargin
                inner_width = doc.width
                page_height = doc.pagesize[1]

                if header_path and os.path.exists(header_path):
                    header_img = get_processed_letterhead_image(header_path, "header")
                    if header_img is not None:
                        canvas.drawImage(
                            ImageReader(header_img),
                            inner_x,
                            page_height - doc.topMargin + 8,
                            width=inner_width,
                            height=header_height,
                            preserveAspectRatio=False,
                            mask="auto",
                        )

                if footer_path and os.path.exists(footer_path):
                    footer_img = get_processed_letterhead_image(footer_path, "footer")
                    if footer_img is not None:
                        canvas.drawImage(
                            ImageReader(footer_img),
                            inner_x,
                            16,
                            width=inner_width,
                            height=footer_height,
                            preserveAspectRatio=False,
                            mask="auto",
                        )
                    else:
                        canvas.setFont("Helvetica", 8)
                        canvas.setFillColor(colors.HexColor("#666666"))
                        canvas.drawCentredString(doc.pagesize[0] / 2, 24, footer_text)
                else:
                    canvas.setFont("Helvetica", 8)
                    canvas.setFillColor(colors.HexColor("#666666"))
                    canvas.drawCentredString(doc.pagesize[0] / 2, 24, footer_text)

                canvas.restoreState()

            doc.build(elements, onFirstPage=add_border, onLaterPages=add_border)
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate PDF: {str(e)}")
            return False

    def _get_email_settings(self):
        settings = db.get_settings(EMAIL_SETTING_KEYS)
        settings["smtp_port"] = settings.get("smtp_port", "") or "465"
        settings["smtp_use_tls"] = settings.get("smtp_use_tls", "0") or "0"
        return settings

    def _set_report_email_placeholder(self, entry, placeholder, show=None):
        entry.delete(0, "end")
        entry.insert(0, placeholder)
        entry.configure(fg="#8a9270", show="")
        entry._placeholder_active = True
        entry._placeholder_text = placeholder
        entry._actual_show = show or ""

    def _clear_report_email_placeholder(self, entry):
        if not getattr(entry, "_placeholder_active", False):
            return
        entry.delete(0, "end")
        entry.configure(fg=TEXT_PRIMARY, show=getattr(entry, "_actual_show", ""))
        entry._placeholder_active = False

    def _get_report_email_entry_value(self, entry):
        if getattr(entry, "_placeholder_active", False):
            return ""
        return entry.get().strip()

    def _set_report_email_settings_status(self, message="", tone="info"):
        label = getattr(self, "_report_email_settings_status", None)
        if not label or not label.winfo_exists():
            return
        palette = {
            "info": ("#eef7c7", "#55603A"),
            "success": ("#e8f7ee", "#1f6f43"),
            "warning": ("#fff4d6", "#9a6700"),
            "error": ("#fde8e8", "#b42318"),
        }
        bg, fg = palette.get(tone, palette["info"])
        text = str(message or "").strip()
        if text:
            label.config(text=text, bg=bg, fg=fg)
            label.pack(fill="x", pady=(0, 12))
        else:
            label.config(text="")
            label.pack_forget()

    def _populate_report_email_settings_form(self, settings=None):
        if settings is None:
            settings = self._get_email_settings()
        if not self._report_email_settings_entries:
            return

        defaults = {
            "smtp_host": ("e.g. smtp.gmail.com", None),
            "smtp_port": ("e.g. 465 for SSL or 587 for STARTTLS", None),
            "smtp_username": ("e.g. schoolresults@example.com", None),
            "smtp_password": ("Paste the email password or app password here", "*"),
            "smtp_sender_name": ("e.g. MT. Olives Adventist School", None),
        }
        for key, entry in self._report_email_settings_entries.items():
            placeholder, mask = defaults.get(key, ("", None))
            value = str(settings.get(key, "") or "").strip()
            if value:
                entry.delete(0, "end")
                entry.insert(0, value)
                entry.configure(fg=TEXT_PRIMARY, show=mask or "")
                entry._placeholder_active = False
                entry._placeholder_text = placeholder
                entry._actual_show = mask or ""
            else:
                self._set_report_email_placeholder(entry, placeholder, show=mask)

        if self._report_email_settings_tls_var is not None:
            self._report_email_settings_tls_var.set(
                settings.get("smtp_use_tls", "0") == "1"
            )

    def _save_report_email_settings_inline(self):
        if not self._report_email_settings_entries:
            return

        host = self._get_report_email_entry_value(
            self._report_email_settings_entries["smtp_host"]
        )
        port = self._get_report_email_entry_value(
            self._report_email_settings_entries["smtp_port"]
        )
        username = self._get_report_email_entry_value(
            self._report_email_settings_entries["smtp_username"]
        )
        password = self._get_report_email_entry_value(
            self._report_email_settings_entries["smtp_password"]
        )
        sender_name = self._get_report_email_entry_value(
            self._report_email_settings_entries["smtp_sender_name"]
        ) or get_school_profile().get(
            "school_name", DEFAULT_SCHOOL_PROFILE["school_name"]
        )

        if not host or not port or not username or not password:
            self._set_report_email_settings_status(
                "Please fill SMTP host, port, username, and password before saving.",
                tone="warning",
            )
            return

        db.set_setting("smtp_host", host)
        db.set_setting("smtp_port", port)
        db.set_setting("smtp_username", username)
        db.set_setting("smtp_password", password)
        db.set_setting("smtp_sender_name", sender_name)
        db.set_setting(
            "smtp_use_tls",
            "1" if self._report_email_settings_tls_var and self._report_email_settings_tls_var.get() else "0",
        )
        self._set_report_email_settings_status(
            "Email settings saved. You can now send result emails from this page.",
            tone="success",
        )

    def _hide_report_email_settings_panel(self):
        panel = getattr(self, "_report_email_settings_panel", None)
        if panel and panel.winfo_exists():
            panel.pack_forget()

    def _show_report_email_settings_panel(self, message="", tone="info"):
        panel = getattr(self, "_report_email_settings_panel", None)
        host = getattr(self, "_report_email_settings_host", None)
        if not host or not host.winfo_exists() or not panel or not panel.winfo_exists():
            return False
        if not panel.winfo_manager():
            panel.pack(fill="x")
        self._populate_report_email_settings_form()
        self._set_report_email_settings_status(message, tone=tone)
        try:
            panel.update_idletasks()
        except Exception:
            pass
        return True

    def _toggle_report_email_settings_panel(self):
        panel = getattr(self, "_report_email_settings_panel", None)
        host = getattr(self, "_report_email_settings_host", None)
        if not host or not host.winfo_exists() or not panel or not panel.winfo_exists():
            return
        if panel.winfo_manager():
            self._hide_report_email_settings_panel()
            return
        self._show_report_email_settings_panel()

    def _build_report_email_settings_panel(self):
        host = getattr(self, "_report_email_settings_host", None)
        if not host or not host.winfo_exists():
            return

        for child in host.winfo_children():
            child.destroy()

        self._report_email_settings_entries = {}
        self._report_email_settings_tls_var = tk.BooleanVar(value=False)

        em_bo, em_bi = _card_colors("peach")
        outer = tk.Frame(host, bg=em_bo)
        card = tk.Frame(outer, bg=em_bi, padx=22, pady=20)
        card.pack(fill="both", expand=True, padx=1, pady=1)

        header = tk.Frame(card, bg=em_bi)
        header.pack(fill="x", pady=(0, 8))

        tk.Label(
            header,
            text="Result Email Settings",
            bg=em_bi,
            fg=TEXT_PRIMARY,
            font=(FF, 14, "bold"),
        ).pack(side="left")

        tk.Button(
            header,
            text="Hide",
            bg=LEMON_SOFT,
            fg=TEXT_PRIMARY,
            font=(FF, 9, "bold"),
            padx=12,
            pady=6,
            relief="flat",
            cursor="hand2",
            command=self._hide_report_email_settings_panel,
        ).pack(side="right")

        tk.Label(
            card,
            text="Set up the sender account once here, then use Email Result or Email All without leaving the page.",
            bg=em_bi,
            fg=TEXT_SECONDARY,
            font=(FF, 9),
            justify="left",
            wraplength=860,
        ).pack(anchor="w", pady=(0, 12))

        status = tk.Label(
            card,
            text="",
            bg=em_bi,
            fg=TEXT_SECONDARY,
            font=(FF, 9),
            anchor="w",
            justify="left",
            padx=10,
            pady=8,
        )
        self._report_email_settings_status = status

        form = tk.Frame(card, bg=em_bi)
        form.pack(fill="x")

        def entry_field(parent, key, label, placeholder="", show=None):
            field = tk.Frame(parent, bg=em_bi)
            tk.Label(
                field, text=label, bg=em_bi, fg=TEXT_SECONDARY, font=(FF, 10, "bold")
            ).pack(anchor="w", pady=(0, 5))
            ent = tk.Entry(
                field,
                relief="solid",
                bd=1,
                bg="#fffdf8",
                fg=TEXT_PRIMARY,
                insertbackground=TEXT_PRIMARY,
                highlightthickness=1,
                highlightbackground="#b8c48f",
                highlightcolor=OLIVE_MID,
                font=(FF, 10),
            )
            ent.pack(fill="x", ipady=7)
            self._set_report_email_placeholder(ent, placeholder, show=show)
            ent.bind(
                "<FocusIn>",
                lambda _e, widget=ent: self._clear_report_email_placeholder(widget),
            )
            ent.bind(
                "<FocusOut>",
                lambda _e, widget=ent, ph=placeholder, mask=show: (
                    self._set_report_email_placeholder(widget, ph, show=mask)
                    if not widget.get().strip()
                    else widget.configure(
                        fg=TEXT_PRIMARY, show=getattr(widget, "_actual_show", "")
                    )
                ),
            )
            self._report_email_settings_entries[key] = ent
            return field

        top_row = tk.Frame(form, bg=em_bi)
        top_row.pack(fill="x", pady=(0, 12))
        top_row.grid_columnconfigure(0, weight=1)
        top_row.grid_columnconfigure(1, weight=1)

        entry_field(
            top_row,
            "smtp_host",
            "SMTP Host",
            placeholder="e.g. smtp.gmail.com",
        ).grid(row=0, column=0, sticky="ew", padx=(0, 10))
        entry_field(
            top_row,
            "smtp_port",
            "SMTP Port",
            placeholder="e.g. 465 for SSL or 587 for STARTTLS",
        ).grid(row=0, column=1, sticky="ew")

        mid_row = tk.Frame(form, bg=em_bi)
        mid_row.pack(fill="x", pady=(0, 12))
        mid_row.grid_columnconfigure(0, weight=1)
        mid_row.grid_columnconfigure(1, weight=1)

        entry_field(
            mid_row,
            "smtp_username",
            "SMTP Username",
            placeholder="e.g. schoolresults@example.com",
        ).grid(row=0, column=0, sticky="ew", padx=(0, 10))
        entry_field(
            mid_row,
            "smtp_password",
            "SMTP Password / App Password",
            placeholder="Paste the email password or app password here",
            show="*",
        ).grid(row=0, column=1, sticky="ew")

        bottom_row = tk.Frame(form, bg=em_bi)
        bottom_row.pack(fill="x", pady=(0, 10))
        bottom_row.grid_columnconfigure(0, weight=1)

        entry_field(
            bottom_row,
            "smtp_sender_name",
            "Sender Name",
            placeholder="e.g. MT. Olives Adventist School",
        ).grid(row=0, column=0, sticky="ew")

        tls_box = tk.Frame(card, bg=em_bi)
        tls_box.pack(fill="x", pady=(4, 10))
        tk.Checkbutton(
            tls_box,
            text="Use STARTTLS instead of SSL",
            variable=self._report_email_settings_tls_var,
            bg=em_bi,
            fg=TEXT_PRIMARY,
            activebackground=em_bi,
            selectcolor=em_bi,
            font=(FF, 9),
        ).pack(anchor="w")
        tk.Label(
            tls_box,
            text="Tip: Gmail commonly uses smtp.gmail.com with port 465 and an app password.",
            bg=em_bi,
            fg=TEXT_SECONDARY,
            font=(FF, 8),
            justify="left",
            wraplength=860,
        ).pack(anchor="w", pady=(6, 0))

        btn_row = tk.Frame(card, bg=em_bi)
        btn_row.pack(fill="x", pady=(8, 0))
        tk.Button(
            btn_row,
            text="Close",
            bg=LEMON_SOFT,
            fg=TEXT_PRIMARY,
            font=(FF, 10, "bold"),
            padx=18,
            pady=8,
            command=self._hide_report_email_settings_panel,
            cursor="hand2",
        ).pack(side="right", padx=(10, 0))
        tk.Button(
            btn_row,
            text="Save Settings",
            bg=BLUE,
            fg="white",
            font=(FF, 10, "bold"),
            padx=18,
            pady=8,
            command=self._save_report_email_settings_inline,
            cursor="hand2",
        ).pack(side="right")

        self._report_email_settings_panel = outer
        self._populate_report_email_settings_form()
        outer.pack_forget()

    def _set_failed_email_logs_status(self, message="", tone="info"):
        label = getattr(self, "_failed_email_logs_status", None)
        if not label or not label.winfo_exists():
            return
        palette = {
            "info": ("#eef2ff", "#334155"),
            "success": ("#e8f7ee", "#1f6f43"),
            "warning": ("#fff4d6", "#9a6700"),
            "error": ("#fde8e8", "#b42318"),
        }
        bg, fg = palette.get(tone, palette["info"])
        text = str(message or "").strip()
        if text:
            label.config(text=text, bg=bg, fg=fg)
            label.pack(fill="x", pady=(0, 10))
        else:
            label.config(text="")
            label.pack_forget()

    def _hide_failed_email_logs_panel(self):
        panel = getattr(self, "_failed_email_logs_panel", None)
        if panel and panel.winfo_exists():
            panel.pack_forget()

    def _show_failed_email_logs_panel(self, message="", tone="info"):
        panel = getattr(self, "_failed_email_logs_panel", None)
        host = getattr(self, "_failed_email_logs_host", None)
        if not host or not host.winfo_exists() or not panel or not panel.winfo_exists():
            return False
        self._refresh_failed_email_logs_panel(show_panel=False)
        if not panel.winfo_manager():
            panel.pack(fill="x")
        self._set_failed_email_logs_status(message, tone=tone)
        try:
            panel.update_idletasks()
        except Exception:
            pass
        return True

    def _retry_selected_failed_email_logs_inline(self):
        tree = getattr(self, "_failed_email_logs_tree", None)
        if not tree or not tree.winfo_exists():
            return
        selected = set(tree.selection())
        picked = [log for log in self._failed_email_logs_rows if str(log.get("id")) in selected]
        if not picked:
            self._set_failed_email_logs_status(
                "Select one or more failed email rows to retry.",
                tone="warning",
            )
            return
        self._retry_failed_email_logs(picked)

    def _retry_all_failed_email_logs_inline(self):
        if not self._failed_email_logs_rows:
            self._set_failed_email_logs_status(
                "There are no failed email records to retry.",
                tone="info",
            )
            return
        self._retry_failed_email_logs(list(self._failed_email_logs_rows))

    def _refresh_failed_email_logs_panel(self, show_panel=True):
        tree = getattr(self, "_failed_email_logs_tree", None)
        title = getattr(self, "_failed_email_logs_title", None)
        panel = getattr(self, "_failed_email_logs_panel", None)
        host = getattr(self, "_failed_email_logs_host", None)
        if (
            not host
            or not host.winfo_exists()
            or not panel
            or not panel.winfo_exists()
            or not tree
            or not tree.winfo_exists()
            or not title
            or not title.winfo_exists()
        ):
            return

        cls = self.rc_cls_cb.get()
        academic_year = self.rc_year_cb.get() or str(datetime.now().year)
        term = self.rc_term_cb.get()
        exam_type = self.rc_exam_cb.get() or DEFAULT_EXAM_TYPE
        stream = self._get_selected_report_stream()
        logs = db.get_email_logs(cls, term, exam_type, stream, "failed", academic_year)
        self._failed_email_logs_rows = list(logs)

        stream_suffix = f" / {stream}" if stream else ""
        title.config(
            text=f"Failed Email Logs - {cls}{stream_suffix} / {academic_year} / Term {term} / {exam_type}"
        )

        for item in tree.get_children():
            tree.delete(item)

        for log in logs:
            tree.insert(
                "",
                "end",
                iid=str(log["id"]),
                values=(
                    log.get("student_name", ""),
                    log.get("recipient_email", ""),
                    log.get("error_message", ""),
                    log.get("sent_at", ""),
                ),
            )

        if show_panel and not panel.winfo_manager():
            panel.pack(fill="x")

        if logs:
            self._set_failed_email_logs_status(
                f"{len(logs)} failed email record(s) loaded for the current report-card selection.",
                tone="info",
            )
        else:
            self._set_failed_email_logs_status(
                "No failed email logs found for the current class, stream, term, and exam.",
                tone="info",
            )

    def _build_failed_email_logs_panel(self):
        host = getattr(self, "_failed_email_logs_host", None)
        if not host or not host.winfo_exists():
            return

        for child in host.winfo_children():
            child.destroy()

        fl_bo, fl_bi = _card_colors("blossom")
        outer = tk.Frame(host, bg=fl_bo)
        card = tk.Frame(outer, bg=fl_bi, padx=18, pady=16)
        card.pack(fill="both", expand=True, padx=1, pady=1)

        top = tk.Frame(card, bg=fl_bi)
        top.pack(fill="x", pady=(0, 8))

        title = tk.Label(
            top,
            text="Failed Email Logs",
            bg=fl_bi,
            fg=TEXT_PRIMARY,
            font=(FF, 13, "bold"),
        )
        title.pack(side="left")
        self._failed_email_logs_title = title

        tk.Button(
            top,
            text="Hide",
            bg=LEMON_SOFT,
            fg=TEXT_PRIMARY,
            font=(FF, 9, "bold"),
            padx=12,
            pady=6,
            relief="flat",
            cursor="hand2",
            command=self._hide_failed_email_logs_panel,
        ).pack(side="right")

        tk.Label(
            card,
            text="Review failed result emails here, then retry selected rows or export the log as CSV.",
            bg=fl_bi,
            fg=TEXT_SECONDARY,
            font=(FF, 9),
            justify="left",
            wraplength=860,
        ).pack(anchor="w", pady=(0, 10))

        status = tk.Label(
            card,
            text="",
            bg=fl_bi,
            fg=TEXT_SECONDARY,
            font=(FF, 9),
            anchor="w",
            justify="left",
            padx=10,
            pady=8,
        )
        self._failed_email_logs_status = status

        table_wrap = tk.Frame(card, bg=fl_bi)
        table_wrap.pack(fill="both", expand=True)

        cols = ("student", "email", "error", "sent_at")
        tree = ttk.Treeview(
            table_wrap,
            columns=cols,
            show="headings",
            style="App.Treeview",
            height=9,
        )
        tree.heading("student", text="Student")
        tree.heading("email", text="Recipient Email")
        tree.heading("error", text="Error")
        tree.heading("sent_at", text="Time")
        tree.column("student", width=180, anchor="w")
        tree.column("email", width=220, anchor="w")
        tree.column("error", width=360, anchor="w")
        tree.column("sent_at", width=170, anchor="center")

        ysb = ttk.Scrollbar(table_wrap, orient="vertical", command=tree.yview)
        xsb = ttk.Scrollbar(table_wrap, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)

        tree.grid(row=0, column=0, sticky="nsew")
        ysb.grid(row=0, column=1, sticky="ns")
        xsb.grid(row=1, column=0, sticky="ew")
        table_wrap.grid_rowconfigure(0, weight=1)
        table_wrap.grid_columnconfigure(0, weight=1)

        self._failed_email_logs_tree = tree
        self._failed_email_logs_rows = []

        btn_row = tk.Frame(card, bg=fl_bi)
        btn_row.pack(fill="x", pady=(12, 0))

        tk.Button(
            btn_row,
            text="Retry Selected",
            bg=BLUE,
            fg="white",
            font=(FF, 10, "bold"),
            padx=16,
            pady=8,
            command=self._retry_selected_failed_email_logs_inline,
            cursor="hand2",
        ).pack(side="left", padx=(0, 8))
        tk.Button(
            btn_row,
            text="Retry All",
            bg=GREEN,
            fg="white",
            font=(FF, 10, "bold"),
            padx=16,
            pady=8,
            command=self._retry_all_failed_email_logs_inline,
            cursor="hand2",
        ).pack(side="left", padx=(0, 8))
        tk.Button(
            btn_row,
            text="Export Failed CSV",
            bg=PURPLE,
            fg="white",
            font=(FF, 10, "bold"),
            padx=16,
            pady=8,
            command=self._export_failed_email_logs,
            cursor="hand2",
        ).pack(side="left", padx=(0, 8))
        tk.Button(
            btn_row,
            text="Refresh",
            bg="#475569",
            fg="white",
            font=(FF, 10, "bold"),
            padx=16,
            pady=8,
            command=self._refresh_failed_email_logs_panel,
            cursor="hand2",
        ).pack(side="left")

        self._failed_email_logs_panel = outer
        outer.pack_forget()

    def _open_email_settings_dialog(self):
        if self._show_report_email_settings_panel():
            return
        self.show_report_cards()
        self._show_report_email_settings_panel()

    def _validate_email_setup(self):
        settings = self._get_email_settings()
        missing = [
            key
            for key in ("smtp_host", "smtp_port", "smtp_username", "smtp_password")
            if not settings.get(key)
        ]
        if missing:
            self._show_report_email_settings_panel(
                "Please complete the email settings below before sending result emails.",
                tone="warning",
            ) or self._open_email_settings_dialog()
            return None
        return settings

    def _is_valid_email(self, email):
        return bool(re.match(r"[^@]+@[^@]+\.[^@]+", str(email or "").strip()))

    def _create_result_email_html(
        self, student, term, exam_type, settings, academic_year=None
    ):
        academic_year = str(academic_year or datetime.now().year)
        profile = get_school_profile()
        school_name = profile.get("school_app_title", profile.get("school_name", ""))
        school_location = profile.get(
            "school_location", DEFAULT_SCHOOL_PROFILE["school_location"]
        )
        guardian_name = student.get("guardian_name", "").strip() or "Parent/Guardian"
        sender_name = settings.get("smtp_sender_name", "").strip() or school_name
        stream_html = ""
        if student.get("stream", "").strip():
            stream_html = f'<tr><td style="padding: 6px 12px 6px 0;"><b>Stream</b></td><td>{student.get("stream", "")}</td></tr>'
        return f"""
        <html>
        <body style="font-family: Arial, sans-serif; color: #333333; line-height: 1.5; background: #f7f8ee; padding: 20px;">
            <div style="max-width: 640px; margin: 0 auto; background: #ffffff; border: 1px solid #d9e2c3; border-radius: 10px; overflow: hidden;">
                <div style="background: linear-gradient(135deg, #6B764B 0%, #889660 100%); color: #ffffff; padding: 20px 24px;">
                    <div style="font-size: 22px; font-weight: 700;">{school_name}</div>
                    <div style="font-size: 13px; opacity: 0.9; margin-top: 4px;">Official Student Result Delivery</div>
                </div>
                <div style="padding: 24px;">
                <h2 style="margin: 0 0 12px 0; color: #55603A;">Student Report Card</h2>
                <p>Dear {guardian_name},</p>
                <p>Please find attached the academic report for <b>{student.get("name", "")}</b>.</p>
                <table style="border-collapse: collapse; margin: 16px 0;">
                    <tr><td style="padding: 6px 12px 6px 0;"><b>Class</b></td><td>{student.get("class", "")}</td></tr>
                    {stream_html}
                    <tr><td style="padding: 6px 12px 6px 0;"><b>Year</b></td><td>{academic_year}</td></tr>
                    <tr><td style="padding: 6px 12px 6px 0;"><b>Term</b></td><td>{term}</td></tr>
                    <tr><td style="padding: 6px 12px 6px 0;"><b>Exam</b></td><td>{exam_type}</td></tr>
                    <tr><td style="padding: 6px 12px 6px 0;"><b>Admission No</b></td><td>{student.get("admission_no", "")}</td></tr>
                </table>
                <p>We appreciate your continued support in your child's education.</p>
                <br>
                <p>
                    Regards,<br>
                    <b>{sender_name}</b><br>
                    {school_name}<br>
                    {school_location}
                </p>
                </div>
            </div>
        </body>
        </html>
        """

    def _send_email_with_attachment(
        self, to_email, subject, body, attachment_path, settings, html_body=""
    ):
        msg = EmailMessage()
        sender_name = settings.get("smtp_sender_name", "").strip() or get_school_profile().get(
            "school_name", DEFAULT_SCHOOL_PROFILE["school_name"]
        )
        username = settings.get("smtp_username", "").strip()
        msg["Subject"] = subject
        msg["From"] = f"{sender_name} <{username}>"
        msg["To"] = to_email.strip()
        msg.set_content(body)
        if html_body:
            msg.add_alternative(html_body, subtype="html")

        with open(attachment_path, "rb") as handle:
            msg.add_attachment(
                handle.read(),
                maintype="application",
                subtype="pdf",
                filename=os.path.basename(attachment_path),
            )

        host = settings.get("smtp_host", "").strip()
        port = int(settings.get("smtp_port", "465") or 465)
        password = settings.get("smtp_password", "")
        use_tls = settings.get("smtp_use_tls", "0") == "1"

        if use_tls:
            with smtplib.SMTP(host, port, timeout=30) as server:
                server.starttls()
                server.login(username, password)
                server.send_message(msg)
        else:
            with smtplib.SMTP_SSL(host, port, timeout=30) as server:
                server.login(username, password)
                server.send_message(msg)

    def _build_temp_result_pdf(self, result, total_students, term, exam_type):
        temp_dir = tempfile.mkdtemp(prefix="result_mail_")
        student = result["student"]
        year_text = self._get_report_card_context(result, term, exam_type).get(
            "year", str(datetime.now().year)
        )
        pdf_path = os.path.join(
            temp_dir,
            f"{self._get_report_card_file_basename(student, term, year_text)}.pdf",
        )
        ok = self._build_report_card_pdf(
            result, total_students, term, exam_type, pdf_path
        )
        return pdf_path if ok else None

    def _get_result_for_student(
        self, student_id, class_name, term, exam_type, academic_year=None
    ):
        results = self._get_ranked_results(class_name, term, exam_type, academic_year)
        return next(
            (r for r in results if r["student"]["id"] == student_id), None
        ), results

    def _retry_failed_email_logs(self, logs):
        settings = self._validate_email_setup()
        if not settings:
            return
        if not logs:
            messagebox.showinfo(
                "No Failed Emails", "There are no failed email records to retry."
            )
            return

        progress_dialog, status_label, percent_label, progress = (
            self._open_progress_dialog(
                "Retrying Failed Emails", f"Preparing {len(logs)} failed email(s)..."
            )
        )

        def worker():
            sent = 0
            failed = 0
            for index, log in enumerate(logs, start=1):
                pdf_path = None
                try:
                    student = db.get_student(log["student_id"])
                    if not student:
                        failed += 1
                        continue
                    result, results = self._get_result_for_student(
                        log["student_id"],
                        log.get("class_name", ""),
                        log["term"],
                        log.get("exam_type", DEFAULT_EXAM_TYPE),
                        log.get("academic_year"),
                    )
                    if not result:
                        failed += 1
                        continue
                    self.root.after(
                        0,
                        lambda i=index, n=student["name"]: self._update_progress_dialog(
                            progress_dialog,
                            status_label,
                            percent_label,
                            progress,
                            i - 1,
                            len(logs),
                            f"{i}/{len(logs)} Retrying {n}...",
                        ),
                    )
                    pdf_path = self._build_temp_result_pdf(
                        result,
                        len(results),
                        log["term"],
                        log.get("exam_type", DEFAULT_EXAM_TYPE),
                    )
                    if not pdf_path:
                        failed += 1
                        continue
                    subject = f"{student['name']} Report Card - Term {log['term']} {log.get('exam_type', DEFAULT_EXAM_TYPE)}"
                    body = (
                        f"Dear {student.get('guardian_name', 'Parent/Guardian')},\n\n"
                        f"Please find attached the report card for {student['name']}.\n\n"
                        f"Class: {student['class']}\n"
                        f"Year: {log.get('academic_year', datetime.now().year)}\n"
                        f"Term: {log['term']}\n"
                        f"Exam: {log.get('exam_type', DEFAULT_EXAM_TYPE)}\n\n"
                        f"Regards,\n{settings.get('smtp_sender_name', get_school_profile().get('school_name', DEFAULT_SCHOOL_PROFILE['school_name']))}"
                    )
                    html_body = self._create_result_email_html(
                        student,
                        log["term"],
                        log.get("exam_type", DEFAULT_EXAM_TYPE),
                        settings,
                        log.get("academic_year"),
                    )
                    self._send_email_with_attachment(
                        log["recipient_email"],
                        subject,
                        body,
                        pdf_path,
                        settings,
                        html_body,
                    )
                    db.log_email_delivery(
                        student["id"],
                        log["term"],
                        log.get("exam_type", DEFAULT_EXAM_TYPE),
                        log["recipient_email"],
                        "sent",
                        "",
                        log.get("academic_year"),
                    )
                    sent += 1
                except Exception as exc:
                    db.log_email_delivery(
                        log["student_id"],
                        log["term"],
                        log.get("exam_type", DEFAULT_EXAM_TYPE),
                        log["recipient_email"],
                        "failed",
                        str(exc),
                        log.get("academic_year"),
                    )
                    failed += 1
                finally:
                    if pdf_path and os.path.exists(pdf_path):
                        try:
                            shutil.rmtree(os.path.dirname(pdf_path), ignore_errors=True)
                        except Exception:
                            pass

            self.root.after(0, progress_dialog.destroy)
            self.root.after(
                0,
                lambda: (
                    self._refresh_failed_email_logs_panel(show_panel=False),
                    self._set_failed_email_logs_status(
                        f"Retry complete. Sent: {sent} | Failed: {failed}",
                        tone="success" if failed == 0 else "warning",
                    ),
                    messagebox.showinfo(
                        "Retry Complete", f"Sent: {sent}\nFailed: {failed}"
                    ),
                ),
            )

        threading.Thread(target=worker, daemon=True).start()

    def _export_failed_email_logs(self):
        cls = self.rc_cls_cb.get()
        academic_year = self.rc_year_cb.get() or str(datetime.now().year)
        term = self.rc_term_cb.get()
        exam_type = self.rc_exam_cb.get() or DEFAULT_EXAM_TYPE
        stream = self._get_selected_report_stream()
        logs = db.get_email_logs(cls, term, exam_type, stream, "failed", academic_year)
        if not logs:
            messagebox.showinfo(
                "No Failed Emails", "No failed email logs found for this selection."
            )
            return
        file_path = filedialog.asksaveasfilename(
            title="Export Failed Emails",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            initialfile=f"failed_emails_{cls.replace(' ', '_')}_{(stream or 'all_streams').replace(' ', '_')}_{academic_year}_{term}_{exam_type.replace('-', '_')}.csv",
        )
        if not file_path:
            return
        with open(file_path, "w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(
                [
                    "Student",
                    "Admission No",
                    "Class",
                    "Academic Year",
                    "Term",
                    "Exam Type",
                    "Recipient",
                    "Status",
                    "Error",
                    "Sent At",
                ]
            )
            for log in logs:
                writer.writerow(
                    [
                        log.get("student_name", ""),
                        log.get("admission_no", ""),
                        log.get("class_name", ""),
                        log.get("academic_year", ""),
                        log.get("term", ""),
                        log.get("exam_type", ""),
                        log.get("recipient_email", ""),
                        log.get("status", ""),
                        log.get("error_message", ""),
                        log.get("sent_at", ""),
                    ]
                )
        messagebox.showinfo("Exported", f"Failed email log exported to {file_path}")

    def _show_failed_email_logs(self):
        if self._show_failed_email_logs_panel():
            return
        self.show_report_cards()
        self._show_failed_email_logs_panel()

    def _send_result_email(self):
        name = self.rc_stu_cb.get()
        if not name:
            return
        settings = self._validate_email_setup()
        if not settings:
            return
        exam_type = self.rc_exam_cb.get() or DEFAULT_EXAM_TYPE
        term = self.rc_term_cb.get()
        academic_year = self.rc_year_cb.get() or str(datetime.now().year)
        results = self._get_report_card_results()
        result = next((r for r in results if r["student"]["name"] == name), None)
        if not result:
            return
        student = result["student"]
        recipient = student.get("parent_email", "").strip()
        if not recipient:
            messagebox.showwarning(
                "Missing Email", "This student has no parent email saved."
            )
            return
        if not self._is_valid_email(recipient):
            messagebox.showwarning(
                "Invalid Email",
                f"Invalid parent email for {student['name']}: {recipient}",
            )
            return

        progress_dialog, status_label, percent_label, progress = (
            self._open_progress_dialog(
                "Sending Result", f"Preparing result for {student['name']}..."
            )
        )

        def worker():
            pdf_path = None
            try:
                self.root.after(
                    0,
                    lambda: self._update_progress_dialog(
                        progress_dialog,
                        status_label,
                        percent_label,
                        progress,
                        1,
                        3,
                        f"Generating PDF for {student['name']}...",
                    ),
                )
                pdf_path = self._build_temp_result_pdf(
                    result, len(results), term, exam_type
                )
                if not pdf_path:
                    self.root.after(0, progress_dialog.destroy)
                    return
                subject = f"{student['name']} Report Card - Term {term} {exam_type}"
                body = (
                    f"Dear {student.get('guardian_name', 'Parent/Guardian')},\n\n"
                    f"Please find attached the report card for {student['name']}.\n\n"
                    f"Class: {student['class']}\n"
                    f"{'Stream: ' + student.get('stream', '').strip() + chr(10) if student.get('stream', '').strip() else ''}"
                    f"Year: {academic_year}\n"
                    f"Term: {term}\n"
                    f"Exam: {exam_type}\n\n"
                    f"Regards,\n{settings.get('smtp_sender_name', get_school_profile().get('school_name', DEFAULT_SCHOOL_PROFILE['school_name']))}"
                )
                html_body = self._create_result_email_html(
                    student, term, exam_type, settings, academic_year
                )
                self.root.after(
                    0,
                    lambda: self._update_progress_dialog(
                        progress_dialog,
                        status_label,
                        percent_label,
                        progress,
                        2,
                        3,
                        f"Sending email to {recipient}...",
                    ),
                )
                self._send_email_with_attachment(
                    recipient, subject, body, pdf_path, settings, html_body
                )
                db.log_email_delivery(
                    student["id"], term, exam_type, recipient, "sent", "", academic_year
                )
                self.root.after(
                    0,
                    lambda: self._update_progress_dialog(
                        progress_dialog,
                        status_label,
                        percent_label,
                        progress,
                        3,
                        3,
                        f"Email sent to {recipient}.",
                    ),
                )
                self.root.after(0, progress_dialog.destroy)
                self.root.after(
                    0,
                    lambda: messagebox.showinfo(
                        "Email Sent", f"Result sent to {recipient}"
                    ),
                )
            except Exception as exc:
                db.log_email_delivery(
                    student["id"],
                    term,
                    exam_type,
                    recipient,
                    "failed",
                    str(exc),
                    academic_year,
                )
                self.root.after(0, progress_dialog.destroy)
                self.root.after(
                    0, lambda: messagebox.showerror("Email Failed", str(exc))
                )
            finally:
                if pdf_path and os.path.exists(pdf_path):
                    try:
                        shutil.rmtree(os.path.dirname(pdf_path), ignore_errors=True)
                    except Exception:
                        pass

        threading.Thread(target=worker, daemon=True).start()

    def _send_all_results_email(self):
        settings = self._validate_email_setup()
        if not settings:
            return
        cls = self.rc_cls_cb.get()
        term = self.rc_term_cb.get()
        exam_type = self.rc_exam_cb.get() or DEFAULT_EXAM_TYPE
        academic_year = self.rc_year_cb.get() or str(datetime.now().year)
        selected_stream = self._get_selected_report_stream()
        results = self._get_report_card_results()
        if not results:
            scope = f"{cls} - {selected_stream}" if selected_stream else cls
            messagebox.showwarning("No Data", f"No students found for {scope}.")
            return

        progress_dialog, status_label, percent_label, progress = (
            self._open_progress_dialog(
                "Sending Bulk Results", f"Preparing to send {len(results)} result(s)..."
            )
        )

        def worker():
            sent = 0
            failed = 0
            missing = 0
            total = len(results)
            for index, result in enumerate(results, start=1):
                student = result["student"]
                recipient = student.get("parent_email", "").strip()
                if not recipient:
                    missing += 1
                    self.root.after(
                        0,
                        lambda i=index, n=student["name"]: self._update_progress_dialog(
                            progress_dialog,
                            status_label,
                            percent_label,
                            progress,
                            i,
                            total,
                            f"{i}/{total} Skipping {n} - missing parent email.",
                        ),
                    )
                    continue
                if not self._is_valid_email(recipient):
                    failed += 1
                    db.log_email_delivery(
                        student["id"],
                        term,
                        exam_type,
                        recipient,
                        "failed",
                        "Invalid email format",
                        academic_year,
                    )
                    self.root.after(
                        0,
                        lambda i=index, n=student["name"]: self._update_progress_dialog(
                            progress_dialog,
                            status_label,
                            percent_label,
                            progress,
                            i,
                            total,
                            f"{i}/{total} Skipping {n} - invalid email address.",
                        ),
                    )
                    continue
                pdf_path = None
                try:
                    self.root.after(
                        0,
                        lambda i=index, n=student["name"]: self._update_progress_dialog(
                            progress_dialog,
                            status_label,
                            percent_label,
                            progress,
                            i - 1,
                            total,
                            f"{i}/{total} Generating PDF for {n}...",
                        ),
                    )
                    pdf_path = self._build_temp_result_pdf(
                        result, len(results), term, exam_type
                    )
                    if not pdf_path:
                        failed += 1
                        continue
                    subject = f"{student['name']} Report Card - Term {term} {exam_type}"
                    body = (
                        f"Dear {student.get('guardian_name', 'Parent/Guardian')},\n\n"
                        f"Please find attached the report card for {student['name']}.\n\n"
                        f"Class: {student['class']}\n"
                        f"{'Stream: ' + student.get('stream', '').strip() + chr(10) if student.get('stream', '').strip() else ''}"
                        f"Year: {academic_year}\n"
                        f"Term: {term}\n"
                        f"Exam: {exam_type}\n\n"
                        f"Regards,\n{settings.get('smtp_sender_name', get_school_profile().get('school_name', DEFAULT_SCHOOL_PROFILE['school_name']))}"
                    )
                    html_body = self._create_result_email_html(
                        student, term, exam_type, settings, academic_year
                    )
                    self.root.after(
                        0,
                        lambda i=index, n=student["name"]: self._update_progress_dialog(
                            progress_dialog,
                            status_label,
                            percent_label,
                            progress,
                            i,
                            total,
                            f"{i}/{total} Sending to {n}...",
                        ),
                    )
                    self._send_email_with_attachment(
                        recipient, subject, body, pdf_path, settings, html_body
                    )
                    db.log_email_delivery(
                        student["id"], term, exam_type, recipient, "sent", "", academic_year
                    )
                    sent += 1
                except Exception as exc:
                    db.log_email_delivery(
                        student["id"],
                        term,
                        exam_type,
                        recipient,
                        "failed",
                        str(exc),
                        academic_year,
                    )
                    failed += 1
                finally:
                    if pdf_path and os.path.exists(pdf_path):
                        try:
                            shutil.rmtree(os.path.dirname(pdf_path), ignore_errors=True)
                        except Exception:
                            pass

            self.root.after(0, progress_dialog.destroy)
            self.root.after(
                0,
                lambda: messagebox.showinfo(
                    "Bulk Email Complete",
                    f"Sent: {sent}\nFailed: {failed}\nMissing Email: {missing}",
                ),
            )

        threading.Thread(target=worker, daemon=True).start()

    # ==================== CBC INFO ====================
    def show_cbc_info(self):
        """Legacy entry point kept for compatibility; CBC info now lives on the dashboard."""
        self.show_dashboard()

    # ==================== CHARTS ====================
    def show_charts(self):
        self.clear_frame()
        self._set_nav("Charts")
        self._page_header(
            "Performance Charts", "Visual analysis of class and subject performance"
        )

        ctrl = tk.Frame(self.content_frame, bg=CONTENT_BG)
        ctrl.pack(fill="x", pady=(0, 12))

        def lbl(t):
            tk.Label(
                ctrl, text=t, bg=CONTENT_BG, fg=TEXT_SECONDARY, font=(FF, 10)
            ).pack(side="left", padx=(10, 4))

        lbl("Class:")
        self.ch_cls_cb = ttk.Combobox(
            ctrl,
            values=["All"] + self.get_current_classes(),
            state="readonly",
            style="App.TCombobox",
            width=12,
        )
        self.ch_cls_cb.set("All")
        self.ch_cls_cb.pack(side="left", ipady=4)
        lbl("Stream:")
        self.ch_stream_cb = ttk.Combobox(
            ctrl,
            state="readonly",
            style="App.TCombobox",
            width=14,
        )
        self.ch_stream_cb.pack(side="left", ipady=4)
        self._refresh_chart_streams(reload_results=False)
        lbl("Year:")
        self.ch_year_cb = ttk.Combobox(
            ctrl,
            values=self._get_year_options(),
            state="readonly",
            style="App.TCombobox",
            width=10,
        )
        self.ch_year_cb.set(str(datetime.now().year))
        self.ch_year_cb.pack(side="left", ipady=4)
        lbl("Term:")
        self.ch_term_cb = ttk.Combobox(
            ctrl, values=TERMS, state="readonly", style="App.TCombobox", width=10
        )
        self.ch_term_cb.set(TERMS[0])
        self.ch_term_cb.pack(side="left", ipady=4)
        lbl("Exam:")
        self.ch_exam_cb = ttk.Combobox(
            ctrl, values=EXAM_TYPES, state="readonly", style="App.TCombobox", width=12
        )
        self.ch_exam_cb.set(DEFAULT_EXAM_TYPE)
        self.ch_exam_cb.pack(side="left", ipady=4)
        self.ch_cls_cb.bind(
            "<<ComboboxSelected>>", lambda e: self._refresh_chart_streams()
        )
        self.ch_stream_cb.bind("<<ComboboxSelected>>", lambda e: self.load_charts())
        self.ch_year_cb.bind("<<ComboboxSelected>>", lambda e: self.load_charts())
        self.ch_term_cb.bind("<<ComboboxSelected>>", lambda e: self.load_charts())
        self.ch_exam_cb.bind("<<ComboboxSelected>>", lambda e: self.load_charts())

        ch_bo, ch_bi = _card_colors("sky")
        chart_outer = tk.Frame(self.content_frame, bg=ch_bo)
        chart_outer.pack(fill="both", expand=True, pady=4)
        chart_card = tk.Frame(chart_outer, bg=ch_bi, padx=10, pady=10)
        chart_card.pack(fill="both", expand=True, padx=1, pady=1)
        self._chart_card_bg = ch_bi

        self.fig, self.axes = plt.subplots(
            2, 2, figsize=(12, 8), constrained_layout=True
        )
        self.fig.set_facecolor(ch_bi)
        self.canvas = FigureCanvasTkAgg(self.fig, master=chart_card)
        self.canvas.get_tk_widget().pack(fill="both", expand=True)
        self.load_charts()

    def load_charts(self):
        cls = self.ch_cls_cb.get()
        academic_year = self.ch_year_cb.get() or str(datetime.now().year)
        term = self.ch_term_cb.get()
        exam_type = self.ch_exam_cb.get() or DEFAULT_EXAM_TYPE
        selected_stream = self._get_selected_chart_stream()
        results = self._get_ranked_results(cls, term, exam_type, academic_year)
        if selected_stream:
            results = [
                r
                for r in results
                if r.get("student", {}).get("stream", "").strip() == selected_stream
            ]
        subjects = self._get_subjects_for_scope(
            cls, term, exam_type, results, academic_year
        )
        plot_panel = _mix_hex(getattr(self, "_chart_card_bg", CARD_BG), "#ffffff", 0.5)
        for ax in self.axes.flat:
            ax.clear()
            ax.set_facecolor(plot_panel)

        subj_totals = {s: [] for s in subjects}
        for r in results:
            for s in subjects:
                if r["marks"].get(s):
                    subj_totals[s].append(r["marks"][s])

        avgs = [
            round(sum(subj_totals[s]) / len(subj_totals[s]), 1) if subj_totals[s] else 0
            for s in subjects
        ]
        colors = [self._get_subject_color(subject, cls) for subject in subjects]

        subject_labels = [
            self._get_subject_label(subject, cls if cls != "All" else "")
            for subject in subjects
        ]
        bars0 = self.axes[0, 0].bar(
            subject_labels, avgs, color=colors, edgecolor="none", width=0.6
        )
        self.axes[0, 0].set_title(
            "Subject Averages",
            fontweight="bold",
            color=TEXT_PRIMARY,
            pad=10,
            fontsize=11,
        )
        self.axes[0, 0].set_ylim(0, 110)
        self.axes[0, 0].set_ylabel("Average Marks", color=TEXT_SECONDARY, fontsize=9)
        self.axes[0, 0].tick_params(
            axis="x", labelrotation=40, labelcolor=TEXT_SECONDARY, labelsize=8
        )
        self.axes[0, 0].tick_params(axis="y", labelcolor=TEXT_SECONDARY, labelsize=8)
        self.axes[0, 0].spines["top"].set_visible(False)
        self.axes[0, 0].spines["right"].set_visible(False)
        # Align rotated labels to right so they sit under their bar
        for lbl in self.axes[0, 0].get_xticklabels():
            lbl.set_ha("right")
        # Value labels on top of each bar
        for bar, val in zip(bars0, avgs):
            if val > 0:
                self.axes[0, 0].text(
                    bar.get_x() + bar.get_width() / 2,
                    bar.get_height() + 1,
                    str(val),
                    ha="center",
                    va="bottom",
                    fontsize=7,
                    color=TEXT_SECONDARY,
                )

        grade_counts = {}
        for r in results:
            grade_counts[r["grade"]] = grade_counts.get(r["grade"], 0) + 1
        grades = [g for g, v in grade_counts.items() if v > 0]
        counts = [grade_counts[g] for g in grades]
        if counts:
            pie_colors = [self._get_grade_color(g) for g in grades]
            wedges, texts, autotexts = self.axes[0, 1].pie(
                counts,
                labels=grades,
                autopct="%1.1f%%",
                colors=pie_colors,
                startangle=90,
                pctdistance=0.78,
                labeldistance=1.12,
            )
            for t in texts:
                t.set_fontsize(9)
                t.set_color(TEXT_SECONDARY)
            for t in autotexts:
                t.set_color("white")
                t.set_fontsize(8)
            self.axes[0, 1].set_title(
                "Grade Distribution",
                fontweight="bold",
                color=TEXT_PRIMARY,
                pad=10,
                fontsize=11,
            )

        top5 = results[:5]
        if top5:
            names = [r["student"]["name"].split()[0] for r in top5]
            avgs5 = [r["average"] for r in top5]
            top_colors = [BLUE, GREEN, ORANGE, PURPLE, GRADE_COLORS["IE"]][: len(top5)]
            bars = self.axes[1, 0].barh(
                names, avgs5, color=top_colors, edgecolor="none", height=0.5
            )
            self.axes[1, 0].set_title(
                "Top Students",
                fontweight="bold",
                color=TEXT_PRIMARY,
                pad=10,
                fontsize=11,
            )
            self.axes[1, 0].set_xlim(0, 110)
            self.axes[1, 0].set_xlabel(
                "Average Marks", color=TEXT_SECONDARY, fontsize=9
            )
            self.axes[1, 0].tick_params(labelcolor=TEXT_SECONDARY, labelsize=8)
            self.axes[1, 0].spines["top"].set_visible(False)
            self.axes[1, 0].spines["right"].set_visible(False)
            # Value labels at end of each bar
            for bar, val in zip(bars, avgs5):
                self.axes[1, 0].text(
                    val + 1,
                    bar.get_y() + bar.get_height() / 2,
                    f"{val:.1f}",
                    va="center",
                    fontsize=8,
                    color=TEXT_SECONDARY,
                )

        if cls == "All":
            cls_perf = {}
            for c in self.get_current_classes():
                cr = self._get_ranked_results(c, term, exam_type, academic_year)
                if selected_stream:
                    cr = [
                        r
                        for r in cr
                        if r.get("student", {}).get("stream", "").strip()
                        == selected_stream
                    ]
                cls_perf[c] = (
                    round(sum(r["average"] for r in cr) / len(cr), 1) if cr else 0
                )
            cls_bars = self.axes[1, 1].bar(
                list(cls_perf.keys()),
                list(cls_perf.values()),
                color=BLUE,
                edgecolor="none",
                width=0.5,
            )
            self.axes[1, 1].set_title(
                "Class Performance",
                fontweight="bold",
                color=TEXT_PRIMARY,
                pad=10,
                fontsize=11,
            )
            self.axes[1, 1].set_ylim(0, 110)
            self.axes[1, 1].set_ylabel(
                "Average Marks", color=TEXT_SECONDARY, fontsize=9
            )
            self.axes[1, 1].tick_params(labelcolor=TEXT_SECONDARY, labelsize=8)
            self.axes[1, 1].spines["top"].set_visible(False)
            self.axes[1, 1].spines["right"].set_visible(False)
            for bar, val in zip(cls_bars, cls_perf.values()):
                if val > 0:
                    self.axes[1, 1].text(
                        bar.get_x() + bar.get_width() / 2,
                        val + 1,
                        str(val),
                        ha="center",
                        va="bottom",
                        fontsize=8,
                        color=TEXT_SECONDARY,
                    )
        else:
            self.axes[1, 1].set_title(
                'Select "All" for class comparison', color=TEXT_SECONDARY, fontsize=10
            )
            self.axes[1, 1].axis("off")

        self.canvas.draw()

    # ==================== REPORT CARDS ====================
    def show_report_cards(self):
        self.clear_frame()
        self._set_nav("Report Cards")
        self._page_header(
            "Report Cards", "Generate and print learner assessment report cards"
        )

        ctrl = tk.Frame(self.content_frame, bg=CONTENT_BG)
        ctrl.pack(fill="x", pady=(0, 12))

        def lbl(t):
            tk.Label(
                ctrl, text=t, bg=CONTENT_BG, fg=TEXT_SECONDARY, font=(FF, 10)
            ).pack(side="left", padx=(10, 4))

        lbl("Class:")
        self.rc_cls_cb = ttk.Combobox(
            ctrl,
            values=self.get_current_classes(),
            state="readonly",
            style="App.TCombobox",
            width=12,
        )
        rc_classes = self.get_current_classes()
        self.rc_cls_cb.set(rc_classes[0] if rc_classes else "")
        self.rc_cls_cb.pack(side="left", ipady=4)
        lbl("Stream:")
        self.rc_stream_cb = ttk.Combobox(
            ctrl, state="readonly", style="App.TCombobox", width=14
        )
        self.rc_stream_cb.pack(side="left", ipady=4)
        lbl("Year:")
        self.rc_year_cb = ttk.Combobox(
            ctrl,
            values=self._get_year_options(),
            state="readonly",
            style="App.TCombobox",
            width=10,
        )
        self.rc_year_cb.set(str(datetime.now().year))
        self.rc_year_cb.pack(side="left", ipady=4)
        lbl("Term:")
        self.rc_term_cb = ttk.Combobox(
            ctrl, values=TERMS, state="readonly", style="App.TCombobox", width=10
        )
        self.rc_term_cb.set(TERMS[0])
        self.rc_term_cb.pack(side="left", ipady=4)
        lbl("Exam:")
        self.rc_exam_cb = ttk.Combobox(
            ctrl, values=EXAM_TYPES, state="readonly", style="App.TCombobox", width=12
        )
        self.rc_exam_cb.set(DEFAULT_EXAM_TYPE)
        self.rc_exam_cb.pack(side="left", ipady=4)
        lbl("Student:")
        self.rc_stu_cb = ttk.Combobox(
            ctrl, state="readonly", style="App.TCombobox", width=22
        )
        self.rc_stu_cb.pack(side="left", ipady=4)

        self.rc_cls_cb.bind(
            "<<ComboboxSelected>>", lambda e: self._refresh_report_card_streams()
        )
        self.rc_stream_cb.bind("<<ComboboxSelected>>", lambda e: self._load_rc())
        self.rc_year_cb.bind("<<ComboboxSelected>>", lambda e: self._load_rc())
        self.rc_term_cb.bind("<<ComboboxSelected>>", lambda e: self._load_rc())
        self.rc_exam_cb.bind("<<ComboboxSelected>>", lambda e: self._load_rc())
        self.rc_stu_cb.bind("<<ComboboxSelected>>", lambda e: self._display_rc())

        self._toolbar_btn(ctrl, "\U0001f5a8  Print", self._print_rc).pack(
            side="left", padx=14
        )
        self._toolbar_btn(
            ctrl, "Email Result", self._send_result_email, bg=ORANGE
        ).pack(side="left", padx=4)
        self._toolbar_btn(
            ctrl, "Email All", self._send_all_results_email, bg=PURPLE
        ).pack(side="left", padx=4)
        self._toolbar_btn(
            ctrl, "Failed Emails", self._show_failed_email_logs, bg="#8B5CF6"
        ).pack(side="left", padx=4)
        self._toolbar_btn(
            ctrl, "Email Settings", self._open_email_settings_dialog, bg="#475569"
        ).pack(side="left", padx=4)
        self._toolbar_btn(ctrl, "Print All", self._print_all_rc, bg="#475569").pack(
            side="left", padx=4
        )

        self._report_email_settings_host = tk.Frame(self.content_frame, bg=CONTENT_BG)
        self._report_email_settings_host.pack(fill="x", pady=(0, 10))
        self._build_report_email_settings_panel()

        self._failed_email_logs_host = tk.Frame(self.content_frame, bg=CONTENT_BG)
        self._failed_email_logs_host.pack(fill="x", pady=(0, 10))
        self._build_failed_email_logs_panel()

        # Scrollable paper-like preview
        paper_bg = tk.Frame(self.content_frame, bg="#d8dce5")
        paper_bg.pack(fill="both", expand=True, pady=4)

        _cv, _sb, scroll_inner = scrollable_frame(paper_bg, bg="#d8dce5")

        # White "paper" frame with drop shadow
        shadow = tk.Frame(scroll_inner, bg="#aab0bf")
        shadow.pack(pady=(18, 22), padx=40)
        self._rc_paper = tk.Frame(shadow, bg="white", padx=38, pady=26)
        self._rc_paper.pack(padx=3, pady=3)

        self._refresh_report_card_streams(reload_results=False)
        self._load_rc()

    def _load_rc(self):
        results = self._get_report_card_results()
        names = [r["student"]["name"] for r in results]
        self.rc_stu_cb["values"] = names
        if names:
            self.rc_stu_cb.current(0)
            self._display_rc()
        else:
            self.rc_stu_cb.set("")
            for w in self._rc_paper.winfo_children():
                w.destroy()
        panel = getattr(self, "_failed_email_logs_panel", None)
        if panel and panel.winfo_exists() and panel.winfo_manager():
            self._refresh_failed_email_logs_panel(show_panel=False)

    def _display_rc(self):
        name = self.rc_stu_cb.get()
        if not name:
            return
        exam_type = self.rc_exam_cb.get() or DEFAULT_EXAM_TYPE
        results = self._get_report_card_results()
        result = next((r for r in results if r["student"]["name"] == name), None)
        if not result:
            return
        for w in self._rc_paper.winfo_children():
            w.destroy()
        self._render_report_card(
            self._rc_paper, result, len(results), self.rc_term_cb.get(), exam_type
        )

    def _get_student_term_marks(self, student_id, term, academic_year=None):
        """Fetch marks for all exam types for a student in a specific term."""
        academic_year = str(academic_year or datetime.now().year)
        term_marks = {}
        for et in EXAM_TYPES:
            m = db.get_student_marks(student_id, term, et, academic_year)
            if m:
                term_marks[et] = m
        return term_marks

    def _is_pre_primary_level(self, level_name):
        level_key = self._normalize_text(level_name)
        return level_key.startswith("pre-primary")

    def _get_catalog_subject_names_for_level(self, level_name):
        catalog = SUBJECT_CATALOG.get(level_name, [])
        if isinstance(catalog, dict):
            entries = list(catalog.get("core", [])) + list(catalog.get("optional", []))
        else:
            entries = list(catalog)
        return [name for _, name, _, _ in entries]

    def _get_whole_school_template_subject_headers(self, class_name):
        resolved_class = (
            self._match_known_class_name(class_name) or str(class_name or "").strip()
        )
        class_level = self._get_level_for_class(resolved_class)
        subject_names = self._get_catalog_subject_names_for_level(class_level)
        headers = []
        seen = set()

        for subject_name in subject_names:
            meta = self._get_subject_meta(subject_name, resolved_class)
            header_value = (
                str(meta.get("code", "") or "").strip()
                or str(meta.get("abbreviation", "") or "").strip()
                or str(self._get_subject_label(subject_name, resolved_class) or "").strip()
                or str(subject_name or "").strip()
            )
            header_value = header_value.upper()
            if header_value and header_value not in seen:
                headers.append(header_value)
                seen.add(header_value)

        if headers:
            return headers

        fallback_subjects = self._get_subjects_for_level(class_level)
        return [
            str(self._get_subject_label(subject_name, resolved_class) or subject_name)
            .strip()
            .upper()
            for subject_name in fallback_subjects
            if str(subject_name or "").strip()
        ]

    def _get_class_teacher_name(self, class_name, stream_name=""):
        def alias_keys(value):
            raw_value = str(value or "").strip()
            matched_value = self._match_known_class_name(raw_value)
            keys = set()
            for candidate in [raw_value, matched_value]:
                candidate = str(candidate or "").strip()
                if not candidate:
                    continue
                keys.add(self._normalize_key(candidate))
                keys.add(self._normalize_key(self._get_class_label(candidate)))
            return {key for key in keys if key}

        target_keys = alias_keys(class_name)
        if not target_keys:
            return ""

        target_stream = (
            self._match_known_stream_name(stream_name, class_name)
            or str(stream_name or "").strip()
        )
        stream_matches = []
        class_matches = []
        for row in db.get_class_teacher_assignments():
            if not (target_keys & alias_keys(row.get("class_name", ""))):
                continue
            assigned_stream = (
                self._match_known_stream_name(
                    row.get("stream_name", ""), row.get("class_name", "")
                )
                or str(row.get("stream_name", "") or "").strip()
            )
            if target_stream and assigned_stream == target_stream:
                stream_matches.append(row)
            elif not assigned_stream:
                class_matches.append(row)
        if stream_matches:
            return str(stream_matches[0].get("full_name", "") or "").strip()
        if class_matches:
            return str(class_matches[0].get("full_name", "") or "").strip()
        return ""

    def _get_report_comment_text(self, student_id, class_name, term, academic_year=None):
        academic_year = str(academic_year or datetime.now().year)
        comment_row = db.get_student_comment(student_id, term, academic_year)
        if comment_row and str(comment_row.get("comment_text", "") or "").strip():
            return str(comment_row.get("comment_text", "") or "").strip()
        class_comments = db.get_class_comments(class_name, term, academic_year)
        return str(class_comments.get(student_id, "") or "").strip()

    def _get_report_card_context(self, result, term, exam_type=DEFAULT_EXAM_TYPE):
        student = result["student"]
        academic_year = str(result.get("academic_year") or datetime.now().year)
        class_name = student.get("class", "")
        class_level = self._get_level_for_class(class_name)
        is_pp = self._is_pre_primary_level(class_level)

        subjects = list(
            result.get("subjects")
            or self._get_subjects_for_class(class_name, term, exam_type)
        )
        catalog_subjects = self._get_catalog_subject_names_for_level(class_level)
        if catalog_subjects:
            ordered_subjects = [
                subject for subject in catalog_subjects if subject in subjects
            ]
            for subject in subjects:
                if subject not in ordered_subjects:
                    ordered_subjects.append(subject)
            subjects = ordered_subjects

        grade_scales = list(self._get_class_grading_scale(class_name))
        level_meta = GRADING_BY_LEVEL.get(class_level, {}).get("levels", {})
        display_grade_scales = []
        for scale in grade_scales:
            code = str(scale.get("grade_code", "") or "").strip()
            if not code:
                continue
            if is_pp and code == "IE":
                continue
            points = level_meta.get(code, {}).get("points")
            display_grade_scales.append(
                {
                    "code": code,
                    "label": str(scale.get("grade_name", "") or code).strip(),
                    "points": points,
                    "display": f"{code}({points})" if points is not None else code,
                    "min_mark": scale.get("min_mark"),
                    "max_mark": scale.get("max_mark"),
                }
            )
        if not display_grade_scales:
            fallback_codes = (
                ["EE", "ME", "AE", "BE"] if is_pp else ["EE", "ME", "AE", "BE", "IE"]
            )
            display_grade_scales = [
                {
                    "code": code,
                    "label": str(
                        level_meta.get(code, {}).get("label", "") or code
                    ).strip(),
                    "points": level_meta.get(code, {}).get("points"),
                    "display": f"{code}({level_meta.get(code, {}).get('points')})"
                    if level_meta.get(code, {}).get("points") is not None
                    else code,
                    "min_mark": "",
                    "max_mark": "",
                }
                for code in fallback_codes
            ]

        return {
            "class_level": class_level,
            "is_pp": is_pp,
            "subjects": subjects,
            "grade_scales": display_grade_scales,
            "grade_codes": [scale["code"] for scale in display_grade_scales],
            "class_teacher_name": self._get_class_teacher_name(
                class_name, student.get("stream", "")
            ),
            "comment_text": self._get_report_comment_text(
                student.get("id"), class_name, term, academic_year
            ),
            "year": academic_year,
        }

    def _slugify_report_part(self, value):
        text = str(value or "").strip()
        if not text:
            return "unknown"
        text = re.sub(r"\s+", "_", text)
        text = re.sub(r"[^A-Za-z0-9_]+", "", text)
        text = re.sub(r"_+", "_", text).strip("_")
        return text or "unknown"

    def _format_report_card_term(self, term):
        term_map = {
            "One": "TermOne",
            "Two": "TermTwo",
            "Three": "TermThree",
            "1": "TermOne",
            "2": "TermTwo",
            "3": "TermThree",
        }
        normalized = str(term or "").strip()
        return term_map.get(normalized, normalized)

    def _format_report_card_term_display(self, term):
        display_map = {
            "One": "Term One",
            "Two": "Term Two",
            "Three": "Term Three",
            "1": "Term One",
            "2": "Term Two",
            "3": "Term Three",
            "TermOne": "Term One",
            "TermTwo": "Term Two",
            "TermThree": "Term Three",
        }
        normalized = str(term or "").strip()
        return display_map.get(normalized, normalized)

    def _format_term_number(self, term):
        term_map = {
            "One": "1",
            "Two": "2",
            "Three": "3",
            "1": "1",
            "2": "2",
            "3": "3",
            "TermOne": "1",
            "TermTwo": "2",
            "TermThree": "3",
        }
        normalized = str(term or "").strip()
        return term_map.get(normalized, normalized)

    def _format_results_heading(
        self, class_name, term, exam_type, year=None, stream_name=""
    ):
        """Build a report-style heading for results preview and PDF output."""
        year_text = str(year or datetime.now().year)
        matched_class = (
            self._match_known_class_name(class_name) or str(class_name or "").strip()
        )
        stream_text = str(stream_name or "").strip().upper()

        grade_words = {
            "1": "ONE",
            "2": "TWO",
            "3": "THREE",
            "4": "FOUR",
            "5": "FIVE",
            "6": "SIX",
            "7": "SEVEN",
            "8": "EIGHT",
            "9": "NINE",
        }

        if matched_class == "All":
            class_text = "ALL CLASSES"
        else:
            grade_match = re.fullmatch(
                r"Grade\s+(\d+)", matched_class, flags=re.IGNORECASE
            )
            if grade_match:
                grade_num = grade_match.group(1)
                class_text = (
                    f"GRADE {grade_words.get(grade_num, grade_num)} ({grade_num})"
                )
            else:
                class_text = str(matched_class).upper()

        parts = [class_text]
        if stream_text:
            parts.append(stream_text)
        parts.extend(
            [
                f"TERM {self._format_term_number(term)}",
                f"{str(exam_type or '').strip().upper()} ASSESSMENT REPORT",
                year_text,
            ]
        )
        return " ".join(part for part in parts if part).strip()

    def _get_report_card_file_basename(self, student, term, year=None):
        student_name = self._slugify_report_part(student.get("name", "student"))
        grade_name = self._slugify_report_part(student.get("class", "grade"))
        term_name = self._slugify_report_part(self._format_report_card_term(term))
        year_text = self._slugify_report_part(year or datetime.now().year)
        return f"{student_name}_{grade_name}_{term_name}_{year_text}"

    def _get_report_title_for_level(self, class_level, is_pp):
        if is_pp:
            return "ASSESSMENT SUMMARY REPORT\nPRE-PRIMARY"
        level_short = {
            "Lower Primary (Grade 1-3)": "LOWER PRIMARY",
            "Upper Primary (Grade 4-6)": "UPPER PRIMARY",
            "Junior School (Grade 7-9)": "JUNIOR SCHOOL",
        }.get(class_level, "LEARNER")
        return f"ASSESSMENT SUMMARY REPORT\n{level_short}"

    def _canonical_exam_type(self, raw_exam_type):
        raw = str(raw_exam_type or "").strip()
        exam_value = raw.lower()
        if exam_value in ("opener", "opening"):
            return "Opener"
        if exam_value in ("mid-term", "midterm", "mid term"):
            return "Mid-Term"
        if exam_value in ("end-term", "endterm", "end term"):
            return "End-Term"
        return raw

    def _format_assessment_ordinal(self, number):
        number = max(1, int(number or 1))
        if 10 <= (number % 100) <= 20:
            suffix = "TH"
        else:
            suffix = {1: "ST", 2: "ND", 3: "RD"}.get(number % 10, "TH")
        return f"{number}{suffix}"

    def _get_report_assessment_specs(
        self, assessment_types=None, include_exam_type=None
    ):
        ordered = []
        seen = set()

        def add(raw_value):
            value = self._canonical_exam_type(raw_value)
            if not value or value in seen:
                return
            seen.add(value)
            ordered.append(value)

        for exam_type in EXAM_TYPES:
            add(exam_type)
        for exam_type in assessment_types or []:
            add(exam_type)
        add(include_exam_type)

        if not ordered:
            ordered = [self._canonical_exam_type(DEFAULT_EXAM_TYPE)]

        specs = []
        for index, exam_type in enumerate(ordered, start=1):
            ordinal = self._format_assessment_ordinal(index)
            specs.append(
                {
                    "key": exam_type,
                    "title": f"{ordinal} ASSESSMENT",
                    "label": exam_type,
                    "matrix_title": f"{ordinal} {str(exam_type).upper()}",
                }
            )
        return specs

    def _get_exam_type_sort_key(self, raw_exam_type):
        raw_value = str(raw_exam_type or "").strip()
        canonical_value = self._canonical_exam_type(raw_value)
        configured_map = {
            self._canonical_exam_type(exam_type): index
            for index, exam_type in enumerate(EXAM_TYPES)
            if str(exam_type or "").strip()
        }
        return (
            configured_map.get(canonical_value, 999),
            canonical_value.lower(),
            raw_value.lower(),
        )

    def _get_ordered_exam_type_options(
        self, include_available_sessions=False, canonicalize_output=False
    ):
        configured = []
        seen = set()
        for exam_type in EXAM_TYPES:
            raw_value = str(exam_type or "").strip()
            canonical_value = self._canonical_exam_type(raw_value)
            if not canonical_value or canonical_value in seen:
                continue
            configured.append(canonical_value if canonicalize_output else raw_value)
            seen.add(canonical_value)

        available = []
        if include_available_sessions:
            for row in db.get_available_exam_sessions():
                raw_value = str(row.get("exam_type", "") or "").strip()
                canonical_value = self._canonical_exam_type(raw_value)
                if not canonical_value or canonical_value in seen:
                    continue
                available.append(
                    canonical_value if canonicalize_output else raw_value
                )
                seen.add(canonical_value)

        available.sort(key=self._get_exam_type_sort_key)
        return configured + available

    def _get_report_assessment_matrix_spec(
        self, context, assessment_types=None, include_exam_type=None
    ):
        default_codes = (
            ["EE", "ME", "AE", "BE"]
            if context.get("is_pp")
            else ["EE", "ME", "AE", "BE", "IE"]
        )
        scale_codes = []
        for scale in context.get("grade_scales", []):
            code = str(scale.get("code", scale.get("grade_code", "")) or "").strip()
            if code and code not in scale_codes:
                scale_codes.append(code)
        for code in context.get("grade_codes", []) or default_codes:
            code = str(code or "").strip()
            if code and code not in scale_codes:
                scale_codes.append(code)
        if not scale_codes:
            scale_codes = list(default_codes)
        assessment_specs = self._get_report_assessment_specs(
            assessment_types=assessment_types, include_exam_type=include_exam_type
        )
        return {
            "assessment_specs": assessment_specs,
            "assessment_order": [spec["key"] for spec in assessment_specs],
            "assessment_titles": [spec["title"] for spec in assessment_specs],
            "assessment_labels": [spec["label"] for spec in assessment_specs],
            "assessment_matrix_titles": [
                spec["matrix_title"] for spec in assessment_specs
            ],
            "scale_codes": scale_codes,
        }

    def _format_report_mark_value(self, raw_value):
        try:
            value = float(raw_value)
            return str(int(value)) if value.is_integer() else f"{value:.1f}"
        except (TypeError, ValueError):
            return str(raw_value or "")

    def _get_level_theme(self, class_level):
        palette = {
            "Pre-Primary (PP1-PP2)": {
                "title": "#7a7f3f",
                "grid": "#ddd4b6",
                "line": "#ba9a52",
                "muted": "#6a653d",
                "header_bg": "#f8f3e7",
                "accent_soft": "#fdf9ef",
            },
            "Lower Primary (Grade 1-3)": {
                "title": OLIVE_PRIMARY,
                "grid": "#d7cfad",
                "line": "#b38b41",
                "muted": OLIVE_DARK,
                "header_bg": "#f7f3e6",
                "accent_soft": "#fcf8ee",
            },
            "Upper Primary (Grade 4-6)": {
                "title": "#5f6b3e",
                "grid": "#decfa5",
                "line": "#b78f43",
                "muted": "#545937",
                "header_bg": "#faf3e2",
                "accent_soft": "#fdf8ec",
            },
            "Junior School (Grade 7-9)": {
                "title": "#8a6822",
                "grid": "#e5d09b",
                "line": "#bf8f31",
                "muted": "#6d571f",
                "header_bg": "#fbf0da",
                "accent_soft": "#fff8e8",
            },
        }
        return palette.get(class_level, palette["Upper Primary (Grade 4-6)"])

    def _get_report_layout_profile(self, context):
        subject_count = max(1, len(context.get("subjects", [])))
        is_pp = bool(context.get("is_pp"))

        scale = 1.0
        if subject_count >= 11:
            scale = 0.84
        elif subject_count >= 9:
            scale = 0.9
        elif subject_count >= 7:
            scale = 0.95

        if is_pp:
            scale = max(scale, 0.92)

        title_font = 15 if scale >= 0.95 else 14
        body_font = 9 if scale >= 0.95 else 8
        hand_font = 11 if scale >= 0.95 else 10
        comment_font = 11 if scale >= 0.95 else 10
        row_pad = 6 if scale >= 0.95 else 4
        compact_gap = 10 if scale >= 0.95 else 8

        return {
            "subject_count": subject_count,
            "scale": scale,
            "title_font": title_font,
            "meta_font": body_font,
            "table_header_font": body_font,
            "table_sub_font": 7,
            "table_row_font": body_font,
            "hand_font": hand_font,
            "comment_font": comment_font,
            "row_pad": row_pad,
            "section_gap": compact_gap,
            "line_field_width": 250 if scale >= 0.95 else 220,
            "preview_header_width": 620 if scale >= 0.95 else 580,
            "preview_footer_width": 620 if scale >= 0.95 else 580,
            "preview_title_padx": 88 if scale >= 0.95 else 72,
            "preview_comment_wrap": 620 if scale >= 0.95 else 590,
            "pdf_title_font": 16 if scale >= 0.95 else 14,
            "pdf_normal_font": 9 if scale >= 0.95 else 8,
            "pdf_legend_font": 8 if scale >= 0.95 else 7,
            "pdf_row_pad": 8 if scale >= 0.95 else 5,
            "pdf_section_gap": 12 if scale >= 0.95 else 8,
            "pdf_header_min": 70 if scale >= 0.95 else 58,
            "pdf_header_max": 110 if scale >= 0.95 else 86,
            "pdf_footer_min": 40 if scale >= 0.95 else 28,
            "pdf_footer_max": 72 if scale >= 0.95 else 52,
            "pdf_title_width": 380,
            "pdf_marks_widths_standard": [150, 58, 58, 58, 60, 136]
            if scale >= 0.95
            else [142, 54, 54, 54, 54, 142],
            "pdf_marks_left_width_pp": 160 if scale >= 0.95 else 146,
            "pdf_marks_code_width_pp": 28 if scale >= 0.95 else 24,
        }

    def _render_report_header_for_level(self, parent, class_level, layout, theme):
        """Render level-specific report header."""
        # Render letterhead
        self._render_report_letterhead(parent, layout)

        # Determine title and styling based on class level
        title_color = theme["title"]
        title_text = ""
        subtitle_text = ""

        if "Pre-Primary" in class_level:
            title_text = "ASSESSMENT SUMMARY REPORT"
            subtitle_text = "PRE-PRIMARY"
        elif "Lower Primary" in class_level:
            title_text = "ASSESSMENT SUMMARY REPORT"
            subtitle_text = "LOWER PRIMARY"
        elif "Upper Primary" in class_level:
            title_text = "ASSESSMENT SUMMARY REPORT"
            subtitle_text = "UPPER PRIMARY"
        elif "Junior School" in class_level:
            title_text = "ASSESSMENT SUMMARY REPORT"
            subtitle_text = "JUNIOR SCHOOL"
        else:
            title_text = "ASSESSMENT SUMMARY REPORT"
            subtitle_text = class_level

        # Main title
        title_frame = tk.Frame(parent, bg="white")
        title_frame.pack(fill="x", pady=(6, 4))

        tk.Label(
            title_frame,
            text=title_text,
            bg="white",
            fg=title_color,
            font=(FF, layout["title_font"], "bold"),
            justify="center",
        ).pack()

        # Subtitle
        tk.Label(
            title_frame,
            text=subtitle_text,
            bg="white",
            fg=title_color,
            font=(FF, layout["title_font"] - 1, "bold"),
            justify="center",
        ).pack()

        # Title underline
        tk.Frame(parent, bg=title_color, height=3).pack(
            fill="x", padx=layout["preview_title_padx"], pady=(4, 14)
        )

    def _render_report_letterhead(self, parent, layout=None):
        assets = get_letterhead_assets()
        header_path = assets.get("header_path")
        using_header = False
        if header_path and os.path.exists(header_path):
            try:
                header_img = get_processed_letterhead_image(header_path, "header")
                if header_img is not None:
                    header_width = (layout or {}).get("preview_header_width", 620)
                    header_img = header_img.resize(
                        (
                            header_width,
                            int(header_img.height * header_width / header_img.width),
                        ),
                        Image.LANCZOS,
                    )
                    header_photo = ImageTk.PhotoImage(header_img)
                    header_label = tk.Label(parent, image=header_photo, bg="white")
                    header_label.image = header_photo
                    header_label.pack(pady=(0, 8))
                    using_header = True
            except Exception as exc:
                print(f"Failed to load letterhead header: {exc}")

        if not using_header:
            profile = get_school_profile()
            tk.Label(
                parent,
                text=profile.get("school_name", DEFAULT_SCHOOL_PROFILE["school_name"]),
                bg="white",
                fg="#1b5e20",
                font=(FF, 15, "bold"),
            ).pack()
            tk.Label(
                parent,
                text=profile.get("school_motto", DEFAULT_SCHOOL_PROFILE["school_motto"]),
                bg="white",
                fg="#7b8794",
                font=(FF, 9, "italic"),
            ).pack(pady=(0, 8))

    def _render_report_footer_image(self, parent, layout=None):
        assets = get_letterhead_assets()
        footer_path = assets.get("footer_path")
        if footer_path and os.path.exists(footer_path):
            try:
                footer_img = Image.open(footer_path)
                footer_width = (layout or {}).get("preview_footer_width", 620)
                footer_img = footer_img.resize(
                    (
                        footer_width,
                        int(footer_img.height * footer_width / footer_img.width),
                    ),
                    Image.LANCZOS,
                )
                footer_photo = ImageTk.PhotoImage(footer_img)
                footer_label = tk.Label(parent, image=footer_photo, bg="white")
                footer_label.image = footer_photo
                footer_label.pack(pady=(10, 0))
            except Exception as exc:
                print(f"Failed to load letterhead footer: {exc}")

    def _render_unified_report_card(
        self, parent, result, total_students, term, exam_type, context
    ):
        """
        Render a unified, professional report card template for all grades.
        Uses: Learning Areas | Assessment 1 (EE|ME|AE|BE) | Assessment 2 (EE|ME|AE|BE) | ...
        """
        s = result["student"]
        class_name = s.get("class", "")
        grade_text = class_name
        teacher_name = context.get("class_teacher_name") or ""
        comment_text = context.get("comment_text") or " "
        term_marks = self._get_student_term_marks(s["id"], term, context.get("year"))
        subjects = context.get("subjects", [])
        scales = context.get("grade_scales", [])
        scale_codes = context.get(
            "grade_codes", ["EE", "ME", "AE", "BE"]
        )  # Will be sorted by value
        year_text = context.get("year", str(datetime.now().year))

        layout = self._get_report_layout_profile(context)
        theme = self._get_level_theme(context.get("class_level", ""))
        title_color = theme["title"]
        grid_border = theme["grid"]
        line_color = theme["line"]
        text_muted = theme["muted"]
        header_bg = theme["header_bg"]
        accent_soft = theme.get("accent_soft", "#f7f7f7")
        accent_med = theme.get("accent_med", "#e8e8e8")

        matrix_spec = self._get_report_assessment_matrix_spec(
            context,
            assessment_types=term_marks.keys(),
            include_exam_type=exam_type,
        )
        assessments = matrix_spec["assessment_order"]
        assessment_titles = {
            spec["key"]: spec["matrix_title"]
            for spec in matrix_spec.get("assessment_specs", [])
        }

        # Build sorted scale codes with labels and values
        scale_info = {}
        for scale in scales:
            code = scale.get("code", scale.get("grade_code", ""))
            if code:
                scale_info[code] = {
                    "label": scale.get("label", scale.get("grade_name", code)),
                    "value": scale.get("value", scale.get("points", "")),
                    "display": scale.get(
                        "display",
                        f"{code}({scale.get('value', scale.get('points', ''))})",
                    ),
                }

        self._render_report_letterhead(parent, layout)

        tk.Label(
            parent,
            text="PERFORMANCE SUMMARY REPORT",
            bg="white",
            fg=title_color,
            font=(FF, layout["title_font"], "bold"),
            justify="center",
        ).pack(pady=(6, 16))
        tk.Frame(parent, bg=title_color, height=3).pack(
            fill="x", padx=layout["preview_title_padx"], pady=(0, 14)
        )

        # Meta information
        meta = tk.Frame(parent, bg="white")
        meta.pack(fill="x", pady=(0, 4))
        meta.columnconfigure(0, weight=1)
        meta.columnconfigure(1, weight=1)
        meta.columnconfigure(2, weight=1)
        for col, (label, value) in enumerate(
            [
                ("GRADE:", grade_text),
                ("TERM:", self._format_report_card_term_display(term).upper()),
                ("YEAR:", year_text),
            ]
        ):
            cell = tk.Frame(meta, bg="white")
            cell.grid(row=0, column=col, sticky="ew", padx=(0, 10 if col < 2 else 0))
            tk.Label(
                cell,
                text=label,
                bg="white",
                fg=text_muted,
                font=(FF, layout["meta_font"], "bold"),
            ).pack(side="left")
            inner = tk.Frame(cell, bg="white")
            inner.pack(side="left", fill="x", expand=True, padx=(6, 0))
            tk.Label(
                inner,
                text=value,
                bg="white",
                fg="#2b2b2b",
                font=("Comic Sans MS", max(9, layout["hand_font"] - 1)),
            ).pack(anchor="w")
            tk.Frame(inner, bg=line_color, height=1).pack(fill="x", pady=(2, 0))

        # Student info
        info = tk.Frame(parent, bg="white")
        info.pack(fill="x", pady=(6, 10))
        for label, value in [
            ("Student Name:", s.get("name", "")),
            ("Grade Facilitator:", teacher_name),
        ]:
            wrap = tk.Frame(info, bg="white")
            wrap.pack(fill="x", pady=2)
            tk.Label(
                wrap,
                text=label,
                bg="white",
                fg=text_muted,
                font=(FF, layout["meta_font"], "bold"),
                anchor="w",
            ).pack(side="left")
            value_wrap = tk.Frame(wrap, bg="white")
            value_wrap.pack(side="left", fill="x", expand=True, padx=(8, 0))
            tk.Label(
                value_wrap,
                text=value,
                bg="white",
                fg="#2b2b2b",
                font=("Comic Sans MS", layout["hand_font"]),
                anchor="w",
            ).pack(anchor="w")
            tk.Frame(value_wrap, bg=line_color, height=1).pack(fill="x", pady=(2, 0))

        # Main performance table
        tbl = tk.Frame(parent, bg=grid_border, padx=1, pady=1)
        tbl.pack(fill="both", pady=(0, 12))
        tbl_inner = tk.Frame(tbl, bg="white")
        tbl_inner.pack(fill="both")

        # Calculate column structure
        num_assessments = len(assessments)
        num_grades = len(scale_codes)
        scale_codes_sorted = sorted(
            scale_codes,
            key=lambda x: scale_info.get(x, {}).get("value", 0),
            reverse=True,
        )

        # Header row 1: Learning Areas + Assessment titles
        tk.Label(
            tbl_inner,
            text="LEARNING AREAS",
            bg=header_bg,
            fg=text_muted,
            font=(FF, layout["table_header_font"], "bold"),
            anchor="w",
            padx=8,
            pady=layout["row_pad"] + 1,
        ).grid(row=0, column=0, rowspan=2, sticky="nsew", padx=0, pady=0)

        tbl_inner.grid_columnconfigure(0, weight=2, minsize=150)

        current_col = 1
        assessment_col_start = {}
        for idx, assessment in enumerate(assessments):
            assessment_col_start[assessment] = current_col
            title = assessment_titles.get(assessment, assessment)
            tk.Label(
                tbl_inner,
                text=title,
                bg=header_bg,
                fg=text_muted,
                font=(FF, layout["table_header_font"], "bold"),
                pady=layout["row_pad"],
            ).grid(
                row=0,
                column=current_col,
                columnspan=num_grades,
                sticky="nsew",
                padx=0,
                pady=0,
            )
            current_col += num_grades

        # Header row 2: Grade codes for each assessment
        current_col = 1
        for idx, assessment in enumerate(assessments):
            for grade_idx, code in enumerate(scale_codes_sorted):
                grade_display = scale_info.get(code, {}).get("display", code)
                bg_color = accent_soft if grade_idx % 2 == 0 else accent_med
                tk.Label(
                    tbl_inner,
                    text=grade_display,
                    bg=bg_color,
                    fg=text_muted,
                    font=(FF, layout["table_sub_font"], "bold"),
                    pady=max(3, layout["row_pad"] - 1),
                ).grid(row=1, column=current_col, sticky="nsew", padx=0, pady=0)
                tbl_inner.grid_columnconfigure(current_col, weight=1, minsize=50)
                current_col += 1

        # Data rows: Learning areas
        for row_index, subject in enumerate(subjects, start=2):
            row_bg = "#ffffff" if row_index % 2 == 0 else accent_soft

            # Subject name cell
            tk.Label(
                tbl_inner,
                text=subject,
                bg=row_bg,
                fg="#425a70",
                font=(FF, layout["table_row_font"], "bold"),
                anchor="w",
                justify="left",
                padx=8,
                pady=layout["row_pad"],
            ).grid(row=row_index, column=0, sticky="nsew", padx=0, pady=0)

            # Grade cells for each assessment
            current_col = 1
            for assessment in assessments:
                marks_data = term_marks.get(assessment, {})
                mark = marks_data.get(subject, "")

                # Get the grade code for this mark
                if str(mark).strip() not in ("", "None", "0"):
                    try:
                        mark_value = float(mark)
                        grade = self._get_grade_code_for_class(mark_value, class_name)
                    except:
                        grade = None
                else:
                    grade = None

                for grade_idx, code in enumerate(scale_codes_sorted):
                    cell_bg = row_bg
                    cell_text = str(mark) if grade == code and mark else ""
                    cell_fg = "#2f2f2f"
                    cell_font = (FF, layout["table_row_font"])

                    # Highlight cell if this grade matches
                    if grade == code and mark:
                        cell_bg = "#fff3cd"  # Light yellow highlight
                        cell_font = (FF, layout["table_row_font"], "bold")

                    tk.Label(
                        tbl_inner,
                        text=cell_text,
                        bg=cell_bg,
                        fg=cell_fg,
                        font=cell_font,
                        pady=layout["row_pad"],
                    ).grid(
                        row=row_index, column=current_col, sticky="nsew", padx=0, pady=0
                    )
                    current_col += 1

        # Comments section
        comment_shell = tk.Frame(parent, bg=grid_border, padx=1, pady=1)
        comment_shell.pack(fill="x", pady=(4, 10))
        comment_card = tk.Frame(comment_shell, bg=accent_soft, padx=10, pady=8)
        comment_card.pack(fill="x")
        tk.Label(
            comment_card,
            text="General Comments",
            bg=accent_soft,
            fg=title_color,
            font=(FF, layout["meta_font"], "bold"),
            anchor="w",
        ).pack(anchor="w")
        tk.Label(
            comment_card,
            text=comment_text,
            bg=accent_soft,
            fg="#2f2f2f",
            font=("Comic Sans MS", layout["comment_font"]),
            justify="left",
            anchor="w",
            wraplength=layout["preview_comment_wrap"],
        ).pack(anchor="w", pady=(4, 6))
        tk.Frame(comment_card, bg=line_color, height=1).pack(fill="x")

        # Signature section
        footer_shell = tk.Frame(parent, bg=grid_border, padx=1, pady=1)
        footer_shell.pack(fill="x", pady=(10, 0))
        footer_card = tk.Frame(footer_shell, bg="white", padx=10, pady=10)
        footer_card.pack(fill="x")

        sign_row = tk.Frame(footer_card, bg="white")
        sign_row.pack(fill="x", pady=(0, 10))
        for idx, label in enumerate(
            ["Grade Facilitator Signature", "Head Teacher Signature"]
        ):
            cell = tk.Frame(sign_row, bg=accent_soft, padx=10, pady=8)
            cell.grid(row=0, column=idx, sticky="ew", padx=(0, 16 if idx == 0 else 0))
            sign_row.grid_columnconfigure(idx, weight=1)
            tk.Label(
                cell,
                text=label,
                bg=accent_soft,
                fg=title_color,
                font=(FF, layout["meta_font"], "bold"),
            ).pack(anchor="w")
            spacer = tk.Frame(cell, bg=accent_soft, height=14)
            spacer.pack(fill="x")
            tk.Frame(cell, bg=line_color, height=1).pack(fill="x", pady=(2, 2))
            if idx == 0 and teacher_name.strip("_ "):
                tk.Label(
                    cell,
                    text=teacher_name,
                    bg=accent_soft,
                    fg="#2f2f2f",
                    font=(FF, layout["table_row_font"]),
                ).pack(anchor="w")

        self._render_report_footer_image(parent, layout)

    def _render_standard_report_card_aligned(
        self, parent, result, total_students, term, exam_type, context
    ):
        """
        Enhanced standard report card with aligned grading scale columns for all assessments.
        Layout: Learning Areas | 1st Assessment(EE|ME|AE|BE) | 2nd Assessment(EE|ME|AE|BE) | 3rd Assessment(EE|ME|AE|BE)
        """
        s = result["student"]
        class_name = s.get("class", "")
        grade_text = class_name
        teacher_name = context.get("class_teacher_name") or ""
        comment_text = context.get("comment_text") or " "
        term_marks = self._get_student_term_marks(s["id"], term, context.get("year"))
        subjects = context.get("subjects", [])
        year_text = context.get("year", str(datetime.now().year))

        layout = self._get_report_layout_profile(context)
        theme = self._get_level_theme(context.get("class_level", ""))
        title_color = theme["title"]
        grid_border = theme["grid"]
        line_color = theme["line"]
        text_muted = theme["muted"]
        header_bg = theme["header_bg"]
        accent_soft = theme.get("accent_soft", "#f7f7f7")

        # Assessment info
        matrix_spec = self._get_report_assessment_matrix_spec(
            context,
            assessment_types=term_marks.keys(),
            include_exam_type=exam_type,
        )
        assessments = matrix_spec["assessment_order"]
        assessment_titles = matrix_spec["assessment_matrix_titles"]
        scale_codes = matrix_spec["scale_codes"]
        class_level = context.get("class_level", "")

        # Render level-specific header
        self._render_report_header_for_level(parent, class_level, layout, theme)

        # Meta info row
        meta = tk.Frame(parent, bg="white")
        meta.pack(fill="x", pady=(0, 4))
        meta.columnconfigure(0, weight=1)
        meta.columnconfigure(1, weight=1)
        meta.columnconfigure(2, weight=1)
        for col, (label, value) in enumerate(
            [
                ("GRADE:", grade_text),
                ("TERM:", self._format_report_card_term_display(term).upper()),
                ("YEAR:", year_text),
            ]
        ):
            cell = tk.Frame(meta, bg="white")
            cell.grid(row=0, column=col, sticky="ew", padx=(0, 10 if col < 2 else 0))
            tk.Label(
                cell,
                text=label,
                bg="white",
                fg=text_muted,
                font=(FF, layout["meta_font"], "bold"),
            ).pack(side="left")
            inner = tk.Frame(cell, bg="white")
            inner.pack(side="left", fill="x", expand=True, padx=(6, 0))
            tk.Label(
                inner,
                text=value,
                bg="white",
                fg="#2b2b2b",
                font=("Comic Sans MS", max(9, layout["hand_font"] - 1)),
            ).pack(anchor="w")
            tk.Frame(inner, bg=line_color, height=1).pack(fill="x", pady=(2, 0))

        # Student info
        info = tk.Frame(parent, bg="white")
        info.pack(fill="x", pady=(6, 10))
        for label, value in [
            ("Student Name:", s.get("name", "")),
            ("Grade Facilitator:", teacher_name),
        ]:
            wrap = tk.Frame(info, bg="white")
            wrap.pack(fill="x", pady=2)
            tk.Label(
                wrap,
                text=label,
                bg="white",
                fg=text_muted,
                font=(FF, layout["meta_font"], "bold"),
                anchor="w",
            ).pack(side="left")
            value_wrap = tk.Frame(wrap, bg="white")
            value_wrap.pack(side="left", fill="x", expand=True, padx=(8, 0))
            tk.Label(
                value_wrap,
                text=value,
                bg="white",
                fg="#2b2b2b",
                font=("Comic Sans MS", layout["hand_font"]),
                anchor="w",
            ).pack(anchor="w")
            tk.Frame(value_wrap, bg=line_color, height=1).pack(fill="x", pady=(2, 0))

        # Main table with aligned grades
        tbl = tk.Frame(parent, bg=grid_border, padx=1, pady=1)
        tbl.pack(fill="both", expand=True, pady=(0, 12))
        tbl_inner = tk.Frame(tbl, bg="white")
        tbl_inner.pack(fill="both", expand=True)

        num_grades = len(scale_codes)
        separator_columns = []
        assessment_start_cols = []

        # Header row 1: Learning Areas + Assessment titles
        tk.Label(
            tbl_inner,
            text="LEARNING AREAS",
            bg=header_bg,
            fg=text_muted,
            font=(FF, layout["table_header_font"], "bold"),
            anchor="w",
            padx=8,
            pady=layout["row_pad"] + 1,
        ).grid(
            row=0,
            column=0,
            rowspan=2,
            sticky="nsew",
            padx=0,
            pady=0,
        )
        tbl_inner.grid_columnconfigure(0, weight=2, minsize=140)

        current_col = 1
        for assessment_idx, title in enumerate(assessment_titles):
            assessment_start_cols.append(current_col)
            tk.Label(
                tbl_inner,
                text=title,
                bg=header_bg,
                fg=text_muted,
                font=(FF, layout["table_header_font"], "bold"),
                pady=layout["row_pad"],
            ).grid(
                row=0,
                column=current_col,
                columnspan=num_grades,
                sticky="nsew",
                padx=0,
                pady=0,
            )
            for grade_idx, code in enumerate(scale_codes):
                col_index = current_col + grade_idx
                bg = "#f5f5f5" if grade_idx % 2 == 0 else accent_soft
                tk.Label(
                    tbl_inner,
                    text=code,
                    bg=bg,
                    fg=text_muted,
                    font=(FF, layout["table_sub_font"], "bold"),
                    pady=max(3, layout["row_pad"] - 1),
                ).grid(row=1, column=col_index, sticky="nsew", padx=0, pady=0)
                tbl_inner.grid_columnconfigure(col_index, weight=1, minsize=45)
            current_col += num_grades
            if assessment_idx < len(assessment_titles) - 1:
                separator_columns.append(current_col)
                tk.Label(tbl_inner, text="", bg=header_bg).grid(
                    row=0, column=current_col, sticky="nsew", padx=0, pady=0
                )
                tk.Label(tbl_inner, text="", bg=accent_soft).grid(
                    row=1, column=current_col, sticky="nsew", padx=0, pady=0
                )
                tbl_inner.grid_columnconfigure(current_col, minsize=2)
                current_col += 1

        # Data rows
        for row_index, subject in enumerate(subjects, start=2):
            row_bg = "#ffffff" if row_index % 2 == 0 else accent_soft

            # Subject name – use abbreviation so nothing overflows
            tk.Label(
                tbl_inner,
                text=self._get_subject_label(subject, class_name),
                bg=row_bg,
                fg="#425a70",
                font=(FF, layout["table_row_font"], "bold"),
                anchor="w",
                justify="left",
                padx=8,
                pady=layout["row_pad"],
            ).grid(row=row_index, column=0, sticky="nsew", padx=0, pady=0)

            # Grade cells for each assessment
            for assessment_idx, assessment in enumerate(assessments):
                current_col = assessment_start_cols[assessment_idx]
                marks_data = term_marks.get(assessment, {})
                mark = marks_data.get(subject, "")
                if str(mark).strip() == "" and assessment == exam_type:
                    mark = result.get("marks", {}).get(subject, "")

                # Get grade for this mark
                grade = None
                if str(mark).strip() not in ("", "None"):
                    try:
                        mark_value = float(mark)
                        grade = self._get_grade_code_for_class(mark_value, class_name)
                    except:
                        pass

                for code in scale_codes:
                    # Display mark if grade matches
                    cell_bg = row_bg
                    cell_text = (
                        self._format_report_mark_value(mark)
                        if grade == code and str(mark).strip() != ""
                        else ""
                    )
                    cell_font = (FF, layout["table_row_font"])

                    tk.Label(
                        tbl_inner,
                        text=cell_text,
                        bg=cell_bg,
                        fg="#2f2f2f",
                        font=cell_font,
                        pady=layout["row_pad"],
                        anchor="center",
                    ).grid(
                        row=row_index, column=current_col, sticky="nsew", padx=0, pady=0
                    )
                    current_col += 1

        if subjects:
            for sep_col in separator_columns:
                sep = tk.Frame(tbl_inner, bg=grid_border, width=2)
                sep.grid(
                    row=2,
                    column=sep_col,
                    rowspan=len(subjects),
                    sticky="nsew",
                    padx=0,
                    pady=0,
                )

        # Comments
        comment_shell = tk.Frame(parent, bg=grid_border, padx=1, pady=1)
        comment_shell.pack(fill="x", pady=(4, 10))
        comment_card = tk.Frame(comment_shell, bg=accent_soft, padx=10, pady=8)
        comment_card.pack(fill="x")
        tk.Label(
            comment_card,
            text="General Comments",
            bg=accent_soft,
            fg=title_color,
            font=(FF, layout["meta_font"], "bold"),
            anchor="w",
        ).pack(anchor="w")
        tk.Label(
            comment_card,
            text=comment_text,
            bg=accent_soft,
            fg="#2f2f2f",
            font=("Comic Sans MS", layout["comment_font"]),
            justify="left",
            anchor="w",
            wraplength=layout["preview_comment_wrap"],
        ).pack(anchor="w", pady=(4, 6))
        tk.Frame(comment_card, bg=line_color, height=1).pack(fill="x")

        # Signatures
        footer_shell = tk.Frame(parent, bg=grid_border, padx=1, pady=1)
        footer_shell.pack(fill="x", pady=(10, 0))
        footer_card = tk.Frame(footer_shell, bg="white", padx=10, pady=10)
        footer_card.pack(fill="x")

        sign_row = tk.Frame(footer_card, bg="white")
        sign_row.pack(fill="x", pady=(0, 10))
        for idx, label in enumerate(
            ["Grade Facilitator Signature", "Head Teacher Signature"]
        ):
            cell = tk.Frame(sign_row, bg=accent_soft, padx=10, pady=8)
            cell.grid(row=0, column=idx, sticky="ew", padx=(0, 16 if idx == 0 else 0))
            sign_row.grid_columnconfigure(idx, weight=1)
            tk.Label(
                cell,
                text=label,
                bg=accent_soft,
                fg=title_color,
                font=(FF, layout["meta_font"], "bold"),
            ).pack(anchor="w")
            spacer = tk.Frame(cell, bg=accent_soft, height=14)
            spacer.pack(fill="x")
            tk.Frame(cell, bg=line_color, height=1).pack(fill="x", pady=(2, 2))

        self._render_report_footer_image(parent, layout)

    def _render_preprimary_report_card(
        self, parent, result, total_students, term, exam_type, context
    ):
        s = result["student"]
        class_name = s.get("class", "")
        grade_text = class_name
        teacher_name = context.get("class_teacher_name") or ""
        comment_text = context.get("comment_text") or " "
        term_marks = self._get_student_term_marks(s["id"], term, context.get("year"))
        subjects = context.get("subjects", [])
        scales = context.get("grade_scales", [])
        scale_codes = context.get("grade_codes", ["EE", "ME", "AE", "BE"])
        year_text = context.get("year", str(datetime.now().year))
        open_date = ""
        close_date = ""
        layout = self._get_report_layout_profile(context)

        theme = self._get_level_theme(context.get("class_level", ""))
        title_color = theme["title"]
        grid_border = theme["grid"]
        line_color = theme["line"]
        text_muted = theme["muted"]
        header_bg = theme["header_bg"]
        accent_soft = theme.get("accent_soft", "#f7f7f7")

        def line_field(parent_widget, label, value, width=None, show_line=True):
            wrap = tk.Frame(parent_widget, bg="white")
            wrap.pack(fill="x", pady=2)
            tk.Label(
                wrap,
                text=label,
                bg="white",
                fg=text_muted,
                font=(FF, layout["meta_font"], "bold"),
                anchor="w",
            ).pack(side="left")
            value_wrap = tk.Frame(wrap, bg="white")
            value_wrap.pack(side="left", fill="x", expand=True, padx=(8, 0))
            tk.Label(
                value_wrap,
                text=value,
                bg="white",
                fg="#2b2b2b",
                font=("Comic Sans MS", layout["hand_font"]),
                anchor="w",
            ).pack(anchor="w")
            if show_line:
                line = tk.Frame(
                    value_wrap,
                    bg=line_color,
                    height=1,
                    width=width or layout["line_field_width"],
                )
                line.pack(fill="x", pady=(2, 0))

        # Use level-specific header rendering (includes letterhead)
        class_level = context.get("class_level", "Pre-Primary")
        self._render_report_header_for_level(parent, class_level, layout, theme)

        meta = tk.Frame(parent, bg="white")
        meta.pack(fill="x", pady=(0, 4))
        meta.columnconfigure(0, weight=1)
        meta.columnconfigure(1, weight=1)
        meta.columnconfigure(2, weight=1)
        for col, (label, value) in enumerate(
            [
                ("GRADE:", grade_text),
                ("TERM:", self._format_report_card_term_display(term).upper()),
                ("YEAR:", year_text),
            ]
        ):
            cell = tk.Frame(meta, bg="white")
            cell.grid(row=0, column=col, sticky="ew", padx=(0, 10 if col < 2 else 0))
            tk.Label(
                cell,
                text=label,
                bg="white",
                fg=text_muted,
                font=(FF, layout["meta_font"], "bold"),
            ).pack(side="left")
            inner = tk.Frame(cell, bg="white")
            inner.pack(side="left", fill="x", expand=True, padx=(6, 0))
            tk.Label(
                inner,
                text=value,
                bg="white",
                fg="#2b2b2b",
                font=("Comic Sans MS", max(9, layout["hand_font"] - 1)),
            ).pack(anchor="w")
            tk.Frame(inner, bg=line_color, height=1).pack(fill="x", pady=(2, 0))

        info = tk.Frame(parent, bg="white")
        info.pack(fill="x", pady=(6, 10))
        line_field(info, "Student Name:", s.get("name", ""))
        line_field(info, "Grade Facilitator:", teacher_name)

        legend = tk.Frame(parent, bg=grid_border, padx=1, pady=1)
        legend.pack(fill="x", pady=(0, 10))
        legend_inner = tk.Frame(legend, bg="white")
        legend_inner.pack(fill="x")
        for idx, scale in enumerate(scales):
            cell = tk.Frame(legend_inner, bg="white", padx=8, pady=layout["row_pad"])
            cell.grid(row=0, column=idx, sticky="nsew")
            legend_inner.grid_columnconfigure(idx, weight=1)
            tk.Label(
                cell,
                text=scale["display"],
                bg="white",
                fg=title_color,
                font=(FF, layout["table_header_font"], "bold"),
            ).pack()
            tk.Label(
                cell,
                text=scale["label"],
                bg="white",
                fg=text_muted,
                font=(FF, layout["table_sub_font"] + 1, "bold"),
            ).pack()

        tbl = tk.Frame(parent, bg=grid_border, padx=1, pady=1)
        tbl.pack(fill="x", pady=(0, 12))
        tbl_inner = tk.Frame(tbl, bg="white")
        tbl_inner.pack(fill="x")
        table_cell_pad = (0, 0)

        matrix_spec = self._get_report_assessment_matrix_spec(
            context,
            assessment_types=term_marks.keys(),
            include_exam_type=exam_type,
        )
        assessments = matrix_spec["assessment_order"]
        assessment_titles = matrix_spec["assessment_matrix_titles"]

        tk.Label(
            tbl_inner,
            text="LEARNING AREAS",
            bg=header_bg,
            fg=text_muted,
            font=(FF, layout["table_header_font"], "bold"),
            anchor="w",
            padx=8,
            pady=layout["row_pad"] + 1,
        ).grid(
            row=0,
            column=0,
            rowspan=2,
            sticky="nsew",
            padx=table_cell_pad[0],
            pady=table_cell_pad[1],
        )

        current_col = 1
        for idx, title in enumerate(assessment_titles):
            tk.Label(
                tbl_inner,
                text=title,
                bg=header_bg,
                fg=text_muted,
                font=(FF, layout["table_header_font"], "bold"),
                pady=layout["row_pad"],
            ).grid(
                row=0,
                column=current_col,
                columnspan=len(scale_codes),
                sticky="nsew",
                padx=table_cell_pad[0],
                pady=table_cell_pad[1],
            )
            for code_idx, scale in enumerate(scales):
                tk.Label(
                    tbl_inner,
                    text=scale["display"],
                    bg=accent_soft,
                    fg=text_muted,
                    font=(FF, layout["table_sub_font"], "bold"),
                    pady=max(3, layout["row_pad"] - 1),
                ).grid(
                    row=1,
                    column=current_col + code_idx,
                    sticky="nsew",
                    padx=table_cell_pad[0],
                    pady=table_cell_pad[1],
                )
            current_col += len(scale_codes)

        tbl_inner.grid_columnconfigure(0, weight=3)
        for col in range(1, 1 + len(scale_codes) * len(assessments)):
            tbl_inner.grid_columnconfigure(col, weight=1)

        for row_index, subject in enumerate(subjects, start=2):
            row_bg = "#ffffff" if row_index % 2 == 0 else accent_soft
            tk.Label(
                tbl_inner,
                text=self._get_subject_label(subject, class_name),
                bg=row_bg,
                fg="#425a70",
                font=(FF, layout["table_row_font"], "bold"),
                anchor="w",
                justify="left",
                padx=8,
                pady=layout["row_pad"],
            ).grid(
                row=row_index,
                column=0,
                sticky="nsew",
                padx=table_cell_pad[0],
                pady=table_cell_pad[1],
            )
            current_col = 1
            for assessment in assessments:
                marks_map = term_marks.get(assessment, {})
                mark = marks_map.get(subject, "")
                grade = (
                    self._get_grade_code_for_class(mark, class_name)
                    if str(mark).strip() not in ("", "None")
                    else ""
                )
                for code in scale_codes:
                    cell_text = str(mark) if grade == code else ""
                    tk.Label(
                        tbl_inner,
                        text=cell_text,
                        bg=row_bg,
                        fg="#2f2f2f",
                        font=(FF, layout["table_row_font"]),
                        pady=layout["row_pad"],
                    ).grid(
                        row=row_index,
                        column=current_col,
                        sticky="nsew",
                        padx=table_cell_pad[0],
                        pady=table_cell_pad[1],
                    )
                    current_col += 1

        comment_shell = tk.Frame(parent, bg=grid_border, padx=1, pady=1)
        comment_shell.pack(fill="x", pady=(4, 10))
        comment_card = tk.Frame(comment_shell, bg=accent_soft, padx=10, pady=8)
        comment_card.pack(fill="x")
        tk.Label(
            comment_card,
            text="General Comments",
            bg=accent_soft,
            fg=title_color,
            font=(FF, layout["meta_font"], "bold"),
            anchor="w",
        ).pack(anchor="w")
        tk.Label(
            comment_card,
            text=comment_text,
            bg=accent_soft,
            fg="#2f2f2f",
            font=("Comic Sans MS", layout["comment_font"]),
            justify="left",
            anchor="w",
            wraplength=layout["preview_comment_wrap"],
        ).pack(anchor="w", pady=(4, 6))
        tk.Frame(comment_card, bg=line_color, height=1).pack(fill="x")

        footer_shell = tk.Frame(parent, bg=grid_border, padx=1, pady=1)
        footer_shell.pack(fill="x", pady=(10, 0))
        footer_card = tk.Frame(footer_shell, bg="white", padx=10, pady=10)
        footer_card.pack(fill="x")

        sign_row = tk.Frame(footer_card, bg="white")
        sign_row.pack(fill="x", pady=(0, 10))
        for idx, label in enumerate(
            ["Grade Facilitator Signature", "Head Teacher Signature"]
        ):
            cell = tk.Frame(sign_row, bg=accent_soft, padx=10, pady=8)
            cell.grid(row=0, column=idx, sticky="ew", padx=(0, 16 if idx == 0 else 0))
            sign_row.grid_columnconfigure(idx, weight=1)
            tk.Label(
                cell,
                text=label,
                bg=accent_soft,
                fg=title_color,
                font=(FF, layout["meta_font"], "bold"),
            ).pack(anchor="w")
            spacer = tk.Frame(cell, bg=accent_soft, height=14)
            spacer.pack(fill="x")
            tk.Frame(cell, bg=line_color, height=1).pack(fill="x", pady=(2, 2))
            if idx == 0 and teacher_name.strip("_ "):
                tk.Label(
                    cell,
                    text=teacher_name,
                    bg=accent_soft,
                    fg="#2f2f2f",
                    font=(FF, layout["table_row_font"]),
                ).pack(anchor="w")

        date_row = tk.Frame(footer_card, bg="white")
        date_row.pack(fill="x")
        for idx, (label, value) in enumerate(
            [("Opening Date", open_date), ("Closing Date", close_date)]
        ):
            cell = tk.Frame(date_row, bg=accent_soft, padx=10, pady=8)
            cell.grid(row=0, column=idx, sticky="ew", padx=(0, 16 if idx == 0 else 0))
            date_row.grid_columnconfigure(idx, weight=1)
            tk.Label(
                cell,
                text=label,
                bg=accent_soft,
                fg=title_color,
                font=(FF, layout["meta_font"], "bold"),
            ).pack(anchor="w")
            spacer = tk.Frame(cell, bg=accent_soft, height=14)
            spacer.pack(fill="x")
            tk.Frame(cell, bg=line_color, height=1).pack(fill="x", pady=(2, 0))

        self._render_report_footer_image(parent, layout)

    def _render_standard_report_card(
        self, parent, result, total_students, term, exam_type, context
    ):
        s = result["student"]
        marks = result["marks"]
        class_name = s.get("class", "")
        stream_text = s.get("stream", "").strip() or s.get("admission_no", "") or "N/A"
        teacher_name = context.get("class_teacher_name") or ""
        comment_text = context.get("comment_text") or " "
        subjects = context.get("subjects", [])
        scales = context.get("grade_scales", [])
        year_text = context.get("year", str(datetime.now().year))
        theme = self._get_level_theme(context.get("class_level", ""))
        layout = self._get_report_layout_profile(context)
        title_color = theme["title"]
        grid_border = theme["grid"]
        line_color = theme["line"]
        text_muted = theme["muted"]
        header_bg = theme["header_bg"]
        accent_soft = theme.get("accent_soft", "#f7f7f7")
        open_date = ""
        close_date = ""
        term_marks = self._get_student_term_marks(s["id"], term, context.get("year"))

        def line_field(parent_widget, label, value, width=None, show_line=True):
            wrap = tk.Frame(parent_widget, bg="white")
            wrap.pack(fill="x", pady=2)
            tk.Label(
                wrap,
                text=label,
                bg="white",
                fg=text_muted,
                font=(FF, layout["meta_font"], "bold"),
                anchor="w",
            ).pack(side="left")
            value_wrap = tk.Frame(wrap, bg="white")
            value_wrap.pack(side="left", fill="x", expand=True, padx=(8, 0))
            tk.Label(
                value_wrap,
                text=value,
                bg="white",
                fg="#2b2b2b",
                font=("Comic Sans MS", max(9, layout["hand_font"] - 1)),
                anchor="w",
            ).pack(anchor="w")
            if show_line:
                tk.Frame(
                    value_wrap,
                    bg=line_color,
                    height=1,
                    width=width or layout["line_field_width"],
                ).pack(fill="x", pady=(2, 0))

        self._render_report_letterhead(parent, layout)

        tk.Label(
            parent,
            text=self._get_report_title_for_level(
                context.get("class_level", ""), False
            ),
            bg="white",
            fg=title_color,
            font=(FF, layout["title_font"], "bold"),
            justify="center",
        ).pack(pady=(6, 16))
        tk.Frame(parent, bg=title_color, height=3).pack(
            fill="x", padx=layout["preview_title_padx"], pady=(0, 14)
        )

        meta = tk.Frame(parent, bg="white")
        meta.pack(fill="x", pady=(0, 4))
        meta.columnconfigure(0, weight=1)
        meta.columnconfigure(1, weight=1)
        meta.columnconfigure(2, weight=1)
        for col, (label, value) in enumerate(
            [
                ("GRADE:", class_name),
                ("TERM:", self._format_report_card_term_display(term).upper()),
                ("YEAR:", year_text),
            ]
        ):
            cell = tk.Frame(meta, bg="white")
            cell.grid(row=0, column=col, sticky="ew", padx=(0, 10 if col < 2 else 0))
            tk.Label(
                cell,
                text=label,
                bg="white",
                fg=text_muted,
                font=(FF, layout["meta_font"], "bold"),
            ).pack(side="left")
            inner = tk.Frame(cell, bg="white")
            inner.pack(side="left", fill="x", expand=True, padx=(6, 0))
            tk.Label(
                inner,
                text=value,
                bg="white",
                fg="#2b2b2b",
                font=("Comic Sans MS", max(9, layout["hand_font"] - 1)),
            ).pack(anchor="w")
            tk.Frame(inner, bg=line_color, height=1).pack(fill="x", pady=(2, 0))

        info = tk.Frame(parent, bg="white")
        info.pack(fill="x", pady=(6, 10))
        line_field(info, "Name of the Learner:", s.get("name", ""))
        line_field(info, "Name of the Facilitator:", teacher_name)

        info_row = tk.Frame(parent, bg="white")
        info_row.pack(fill="x", pady=(0, 10))
        info_row.columnconfigure(0, weight=1)
        info_row.columnconfigure(1, weight=1)
        left_meta = tk.Frame(info_row, bg="white")
        left_meta.grid(row=0, column=0, sticky="ew", padx=(0, 10))
        right_meta = tk.Frame(info_row, bg="white")
        right_meta.grid(row=0, column=1, sticky="ew")
        line_field(left_meta, "Stream / Adm No:", stream_text)
        line_field(right_meta, "Gender:", s.get("gender", ""))

        if scales:
            legend = tk.Frame(parent, bg=grid_border, padx=1, pady=1)
            legend.pack(fill="x", pady=(0, 10))
            legend_inner = tk.Frame(legend, bg="white")
            legend_inner.pack(fill="x")
            for idx, scale in enumerate(scales):
                cell = tk.Frame(
                    legend_inner, bg="white", padx=6, pady=layout["row_pad"]
                )
                cell.grid(row=0, column=idx, sticky="nsew")
                legend_inner.grid_columnconfigure(idx, weight=1)
                range_txt = ""
                if scale.get("min_mark") not in ("", None) and scale.get(
                    "max_mark"
                ) not in ("", None):
                    range_txt = f"{int(float(scale['min_mark']))}-{int(float(scale['max_mark']))}"
                tk.Label(
                    cell,
                    text=scale["display"],
                    bg="white",
                    fg=title_color,
                    font=(FF, layout["table_header_font"], "bold"),
                ).pack()
                if range_txt:
                    tk.Label(
                        cell,
                        text=range_txt,
                        bg="white",
                        fg="#2b2b2b",
                        font=(FF, layout["table_sub_font"] + 1, "bold"),
                    ).pack()
                tk.Label(
                    cell,
                    text=scale["label"],
                    bg="white",
                    fg=text_muted,
                    font=(FF, layout["table_sub_font"], "bold"),
                    wraplength=110,
                    justify="center",
                ).pack()

        tbl = tk.Frame(parent, bg=grid_border, padx=1, pady=1)
        tbl.pack(fill="x", pady=(0, 12))
        tbl_inner = tk.Frame(tbl, bg="white")
        tbl_inner.pack(fill="x")
        table_cell_pad = (0, 0)
        tk.Label(
            tbl_inner,
            text="LEARNING AREAS",
            bg=header_bg,
            fg=text_muted,
            font=(FF, layout["table_header_font"], "bold"),
            anchor="w",
            padx=8,
            pady=layout["row_pad"] + 1,
        ).grid(
            row=0,
            column=0,
            rowspan=2,
            sticky="nsew",
            padx=table_cell_pad[0],
            pady=table_cell_pad[1],
        )

        matrix_spec = self._get_report_assessment_matrix_spec(
            context,
            assessment_types=term_marks.keys(),
            include_exam_type=exam_type,
        )
        assessment_order = matrix_spec["assessment_order"]
        grouped_headers = [
            (title, label)
            for title, label in zip(
                matrix_spec["assessment_titles"], matrix_spec["assessment_labels"]
            )
        ] + [("AVERAGE", "Score"), ("GRADE / REMARK", "Result")]
        for idx, (header, sublabel) in enumerate(grouped_headers, start=1):
            tbl_inner.grid_columnconfigure(
                idx, weight=2 if idx == len(grouped_headers) else 1
            )
            tk.Label(
                tbl_inner,
                text=header,
                bg=header_bg,
                fg=text_muted,
                font=(FF, layout["table_header_font"], "bold"),
                padx=8,
                pady=layout["row_pad"] + 1,
                anchor="center",
            ).grid(
                row=0,
                column=idx,
                sticky="nsew",
                padx=table_cell_pad[0],
                pady=table_cell_pad[1],
            )
            tk.Label(
                tbl_inner,
                text=sublabel,
                bg=accent_soft,
                fg=text_muted,
                font=(FF, layout["table_sub_font"], "bold"),
                padx=6,
                pady=max(3, layout["row_pad"] - 1),
            ).grid(
                row=1,
                column=idx,
                sticky="nsew",
                padx=table_cell_pad[0],
                pady=table_cell_pad[1],
        )
        tbl_inner.grid_columnconfigure(0, weight=3)

        for row_index, subject in enumerate(subjects, start=2):
            assessment_values = []
            for assessment in assessment_order:
                value = term_marks.get(assessment, {}).get(subject, "")
                if str(value).strip() == "" and assessment == exam_type:
                    value = marks.get(subject, "")
                assessment_values.append(value)
            numeric_marks = []
            for raw in assessment_values:
                try:
                    if str(raw).strip() != "":
                        numeric_marks.append(float(raw))
                except (TypeError, ValueError):
                    continue
            avg_mark = (
                round(sum(numeric_marks) / len(numeric_marks), 1)
                if numeric_marks
                else 0
            )
            grade = self._get_grade_code_for_class(avg_mark, class_name)
            remark = self._get_grade_name_for_class(grade, class_name)
            row_bg = "#ffffff" if row_index % 2 else accent_soft
            tk.Label(
                tbl_inner,
                text=subject,
                bg=row_bg,
                fg="#425a70",
                font=(FF, layout["table_row_font"], "bold"),
                anchor="w",
                justify="left",
                padx=8,
                pady=layout["row_pad"],
            ).grid(
                row=row_index,
                column=0,
                sticky="nsew",
                padx=table_cell_pad[0],
                pady=table_cell_pad[1],
            )
            for col, value in enumerate(assessment_values, start=1):
                tk.Label(
                    tbl_inner,
                    text="" if str(value).strip() == "" else str(value),
                    bg=row_bg,
                    fg="#2f2f2f",
                    font=(FF, layout["table_row_font"]),
                    pady=layout["row_pad"],
                ).grid(
                    row=row_index,
                    column=col,
                    sticky="nsew",
                    padx=table_cell_pad[0],
                    pady=table_cell_pad[1],
                )
            tk.Label(
                tbl_inner,
                text=str(avg_mark),
                bg=row_bg,
                fg=title_color,
                font=(FF, layout["table_row_font"], "bold"),
                pady=layout["row_pad"],
            ).grid(
                row=row_index,
                column=len(assessment_order) + 1,
                sticky="nsew",
                padx=table_cell_pad[0],
                pady=table_cell_pad[1],
            )
            tk.Label(
                tbl_inner,
                text=f"{grade}  {remark}",
                bg=row_bg,
                fg="#2f2f2f",
                font=(FF, layout["table_row_font"]),
                pady=layout["row_pad"],
                anchor="w",
            ).grid(
                row=row_index,
                column=len(assessment_order) + 2,
                sticky="nsew",
                padx=table_cell_pad[0],
                pady=table_cell_pad[1],
            )

        summary_row = len(subjects) + 2
        summary_bg = accent_soft
        summary_values = [
            f"Subjects Taken: {len(subjects)}",
            f"Total: {result['total']}/{result.get('possible_total', len(subjects) * 100)}",
            f"Avg: {result['average']}",
            f"Position: {result['position']} / {total_students}",
        ]
        while len(summary_values) < len(assessment_order) + 3:
            summary_values.append("")
        for col, value in enumerate(summary_values):
            tk.Label(
                tbl_inner,
                text=value,
                bg=summary_bg,
                fg=title_color,
                font=(FF, layout["table_row_font"], "bold"),
                padx=8,
                pady=layout["row_pad"] + 1,
                anchor="w" if col == 0 else "center",
            ).grid(
                row=summary_row,
                column=col,
                sticky="nsew",
                padx=table_cell_pad[0],
                pady=table_cell_pad[1],
            )

        comment_shell = tk.Frame(parent, bg=grid_border, padx=1, pady=1)
        comment_shell.pack(fill="x", pady=(4, 10))
        comment_card = tk.Frame(comment_shell, bg=accent_soft, padx=10, pady=8)
        comment_card.pack(fill="x")
        tk.Label(
            comment_card,
            text="General Comments",
            bg=accent_soft,
            fg=title_color,
            font=(FF, layout["meta_font"], "bold"),
            anchor="w",
        ).pack(anchor="w")
        tk.Label(
            comment_card,
            text=comment_text,
            bg=accent_soft,
            fg="#2f2f2f",
            font=("Comic Sans MS", layout["comment_font"]),
            justify="left",
            anchor="w",
            wraplength=layout["preview_comment_wrap"],
        ).pack(anchor="w", pady=(4, 6))
        tk.Frame(comment_card, bg=line_color, height=1).pack(fill="x")

        footer_shell = tk.Frame(parent, bg=grid_border, padx=1, pady=1)
        footer_shell.pack(fill="x", pady=(10, 0))
        footer_card = tk.Frame(footer_shell, bg="white", padx=10, pady=10)
        footer_card.pack(fill="x")

        sign_row = tk.Frame(footer_card, bg="white")
        sign_row.pack(fill="x", pady=(0, 10))
        for idx, label in enumerate(
            ["Grade Facilitator Signature", "Head Teacher Signature"]
        ):
            cell = tk.Frame(sign_row, bg=accent_soft, padx=10, pady=8)
            cell.grid(row=0, column=idx, sticky="ew", padx=(0, 16 if idx == 0 else 0))
            sign_row.grid_columnconfigure(idx, weight=1)
            tk.Label(
                cell,
                text=label,
                bg=accent_soft,
                fg=title_color,
                font=(FF, layout["meta_font"], "bold"),
            ).pack(anchor="w")
            spacer = tk.Frame(cell, bg=accent_soft, height=14)
            spacer.pack(fill="x")
            tk.Frame(cell, bg=line_color, height=1).pack(fill="x", pady=(2, 2))
            if idx == 0 and teacher_name.strip("_ "):
                tk.Label(
                    cell,
                    text=teacher_name,
                    bg=accent_soft,
                    fg="#2f2f2f",
                    font=(FF, layout["table_row_font"]),
                ).pack(anchor="w")

        date_row = tk.Frame(footer_card, bg="white")
        date_row.pack(fill="x")
        for idx, (label, value) in enumerate(
            [("Opening Date", open_date), ("Closing Date", close_date)]
        ):
            cell = tk.Frame(date_row, bg=accent_soft, padx=10, pady=8)
            cell.grid(row=0, column=idx, sticky="ew", padx=(0, 16 if idx == 0 else 0))
            date_row.grid_columnconfigure(idx, weight=1)
            tk.Label(
                cell,
                text=label,
                bg=accent_soft,
                fg=title_color,
                font=(FF, layout["meta_font"], "bold"),
            ).pack(anchor="w")
            spacer = tk.Frame(cell, bg=accent_soft, height=14)
            spacer.pack(fill="x")
            tk.Frame(cell, bg=line_color, height=1).pack(fill="x", pady=(2, 0))

        self._render_report_footer_image(parent, layout)

    def _render_report_card(
        self, parent, result, total_students, term, exam_type=DEFAULT_EXAM_TYPE
    ):
        """Render the full styled visual report card into parent."""
        s = result["student"]
        marks = result["marks"]
        context = self._get_report_card_context(result, term, exam_type)
        cls_level = context["class_level"]
        is_pp = context["is_pp"]

        def get_grade(m):
            return self._get_grade_code_for_class(m, s.get("class", ""))

        subjects = context["subjects"]

        # Use aligned template with grade codes for all assessments (including pre-primary)
        self._render_standard_report_card_aligned(
            parent, result, total_students, term, exam_type, context
        )
        return
        letterhead_assets = get_letterhead_assets()
        header_path = letterhead_assets.get("header_path")
        using_letterhead_header = False

        if header_path and os.path.exists(header_path):
            try:
                header_img = get_processed_letterhead_image(header_path, "header")
                if header_img is None:
                    raise ValueError("Processed header image unavailable")
                header_width = 620
                header_img = header_img.resize(
                    (
                        header_width,
                        int(header_img.height * header_width / header_img.width),
                    ),
                    Image.LANCZOS,
                )
                header_photo = ImageTk.PhotoImage(header_img)
                header_label = tk.Label(parent, image=header_photo, bg="white")
                header_label.image = header_photo  # Keep reference
                header_label.pack(pady=(0, 8))
                using_letterhead_header = True
            except Exception as e:
                print(f"Failed to load letterhead: {e}")
                # Fallback to text header
                tk.Label(
                    parent,
                    text="MT OLIVES ADVENTIST SCHOOL",
                    bg="white",
                    fg=SCH_BLUE,
                    font=(FF, 15, "bold"),
                ).pack()
        else:
            # Fallback to text header
            tk.Label(
                parent,
                text="MT OLIVES ADVENTIST SCHOOL",
                bg="white",
                fg=SCH_BLUE,
                font=(FF, 15, "bold"),
            ).pack()
            tk.Label(
                parent,
                text="Sajin Close, Along Ngong-Matasia Road,\nNext to Oryx Petrol Station, Ngong",
                bg="white",
                fg="#666",
                font=(FF, 9),
            ).pack()
            tk.Label(
                parent,
                text="https://mountolivessda.org/",
                bg="white",
                fg=SCH_BLUE,
                font=(FF, 9),
            ).pack()
            tk.Label(
                parent,
                text="school@mountolivessda.org",
                bg="white",
                fg=SCH_BLUE,
                font=(FF, 9),
            ).pack()
            tk.Label(
                parent, text="+254 788 700073", bg="white", fg="#666", font=(FF, 9)
            ).pack()

        if not using_letterhead_header:
            tk.Label(
                parent,
                text="In God We Excel",
                bg="white",
                fg="#999",
                font=(FF, 9, "italic"),
            ).pack(pady=(0, 8))
        tk.Frame(parent, bg=SCH_BLUE, height=2).pack(fill="x", pady=(0, 12))

        title_border = tk.Frame(parent, bg=RED_ACC)
        title_border.pack(pady=(0, 14))
        title_inner = tk.Frame(title_border, bg="white", padx=30, pady=7)
        title_inner.pack(padx=2, pady=2)
        rpt_title = (
            "ASSESSMENT SUMMARY REPORT PRE-PRIMARY"
            if is_pp
            else "LEARNER ASSESSMENT REPORT CARD"
        )
        tk.Label(
            title_inner, text=rpt_title, bg="white", fg=TTL_ORG, font=(FF, 12, "bold")
        ).pack()

        # Student Info grid
        grade_num = s.get("class", "Grade 7").replace("Grade ", "")
        stream_display = s.get("stream", "").strip() or s.get("admission_no", "")
        facilitator = s.get(
            "facilitator", "Oma Moriasi"
        )  # In a real app this would come from db or teacher profile
        if is_pp:
            info_rows = [
                ("GRADE", grade_num, "TERM", term.upper()),
                ("YEAR", "2026", "NAME", s["name"]),
                ("FACILITATOR", facilitator, "GENDER", s["gender"]),
            ]
        else:
            info_rows = [
                ("NAME", s["name"], "GRADE", grade_num),
                ("STREAM", stream_display, "YEAR", "2026"),
                ("TERM", f"{term} / {exam_type}", "GENDER", s["gender"]),
            ]
        info_f = tk.Frame(parent, bg="white")
        info_f.pack(fill="x", pady=(0, 14))
        info_f.columnconfigure(0, weight=1)
        info_f.columnconfigure(1, weight=1)
        for ri, (l1, v1, l2, v2) in enumerate(info_rows):
            for ci, (label, val) in enumerate([(l1, v1), (l2, str(v2))]):
                cell = tk.Frame(info_f, bg="white")
                cell.grid(
                    row=ri,
                    column=ci,
                    sticky="ew",
                    padx=(0, 10 if ci == 0 else 0),
                    pady=2,
                )
                row_w = tk.Frame(cell, bg="white")
                row_w.pack(fill="x")
                tk.Label(
                    row_w,
                    text=label,
                    bg="white",
                    fg="#222",
                    font=(FF, 10, "bold"),
                    width=9,
                    anchor="w",
                ).pack(side="left")
                tk.Label(
                    row_w, text=val, bg="white", fg="#333", font=(FF, 10), anchor="w"
                ).pack(side="left")
                dot_sep = tk.Canvas(cell, height=2, bg="white", highlightthickness=0)
                dot_sep.pack(fill="x", pady=(2, 0))

                def _draw_dots(e, c=dot_sep):
                    c.delete("all")
                    for x in range(0, e.width, 5):
                        c.create_oval(x, 0, x + 2, 2, fill="#bbb", outline="")

                dot_sep.bind("<Configure>", _draw_dots)

        if is_pp:
            legend_f = tk.Frame(parent, bg="white")
            legend_f.pack(fill="x", pady=(0, 10))
            for i, (code, full) in enumerate(
                [
                    ("EE(4)", "Exceed Expectation"),
                    ("ME(3)", "Meet Expectation"),
                    ("AE(2)", "Approach Expectation"),
                    ("BE(1)", "Below Expectation"),
                ]
            ):
                tk.Label(
                    legend_f,
                    text=f"{code} {full}",
                    bg="white",
                    fg=PERF_CLR.get(code[:2], "#222"),
                    font=(FF, 8, "bold"),
                    padx=10,
                ).pack(side="left", expand=True)

        # Marks Table
        tbl = tk.Frame(parent, bg="#cccccc")
        tbl.pack(fill="x", pady=(0, 12))

        if is_pp:
            term_marks = self._get_student_term_marks(
                s["id"], term, context.get("year")
            )
            assess_types = [
                spec["key"]
                for spec in self._get_report_assessment_specs(
                    assessment_types=term_marks.keys(), include_exam_type=exam_type
                )
            ]

            # Header Row 1
            tk.Label(
                tbl,
                text="LEARNING AREAS",
                bg=HDR_BG,
                fg=SCH_BLUE,
                font=(FF, 8, "bold"),
                padx=5,
                pady=5,
                anchor="w",
            ).grid(row=0, column=0, rowspan=2, sticky="nsew", padx=1, pady=1)

            for i, at in enumerate(assess_types):
                tk.Label(
                    tbl,
                    text=f"{self._format_assessment_ordinal(i + 1)} ASSESSMENT",
                    bg=HDR_BG,
                    fg=SCH_BLUE,
                    font=(FF, 8, "bold"),
                    padx=2,
                    pady=2,
                ).grid(
                    row=0, column=1 + i * 4, columnspan=4, sticky="nsew", padx=1, pady=1
                )

                for j, subcol in enumerate(["EE(4)", "ME(3)", "AE(2)", "BE(1)"]):
                    tk.Label(
                        tbl,
                        text=subcol,
                        bg=HDR_BG,
                        fg=SCH_BLUE,
                        font=(FF, 7, "bold"),
                        padx=1,
                        pady=1,
                    ).grid(row=1, column=1 + i * 4 + j, sticky="nsew", padx=1, pady=1)

            # Data Rows
            for i, subj in enumerate(subjects):
                subject_style = self._get_subject_colors(subj, s.get("class", ""))
                row_bg = subject_style["soft"] if i % 2 == 0 else subject_style["mid"]
                subj_label = self._get_subject_label(subj, s.get("class", ""))

                tk.Label(
                    tbl,
                    text=subj_label,
                    bg=row_bg,
                    fg=subject_style["dark_text"],
                    font=(FF, 8, "bold"),
                    padx=5,
                    pady=4,
                    anchor="w",
                ).grid(row=i + 2, column=0, sticky="nsew", padx=1, pady=1)

                for k, et in enumerate(assess_types):
                    m_dict = term_marks.get(et, {})
                    mk = m_dict.get(subj, 0)
                    grade = get_grade(mk) if mk else ""

                    for j, subcol in enumerate(["EE", "ME", "AE", "BE"]):
                        val = str(mk) if grade == subcol else ""
                        tk.Label(
                            tbl,
                            text=val,
                            bg=row_bg,
                            fg="#333",
                            font=(FF, 8),
                            padx=1,
                            pady=4,
                        ).grid(
                            row=i + 2,
                            column=1 + k * 4 + j,
                            sticky="nsew",
                            padx=1,
                            pady=1,
                        )

            # Summary - Pre-Primary usually doesn't show total points like high school,
            # but let's show General Comments instead or keep it simple.
            base_row = len(subjects) + 2
        else:
            for col_i, weight in enumerate([3, 1, 1, 4]):
                tbl.columnconfigure(col_i, weight=weight)
            for text, col, anchor in [
                ("LEARNING AREA", 0, "w"),
                ("MARKS", 1, "center"),
                ("AVG", 2, "center"),
                ("PERFORMANCE LEVEL", 3, "w"),
            ]:
                tk.Label(
                    tbl,
                    text=text,
                    bg=HDR_BG,
                    fg=SCH_BLUE,
                    font=(FF, 10, "bold"),
                    padx=10,
                    pady=8,
                    anchor=anchor,
                ).grid(row=0, column=col, sticky="nsew", padx=1, pady=1)
            for i, subj in enumerate(subjects):
                mk = marks.get(subj, 0)
                grade = get_grade(mk)
                subject_style = self._get_subject_colors(subj, s.get("class", ""))
                row_bg = subject_style["soft"] if i % 2 == 0 else subject_style["mid"]
                subj_label = self._get_subject_label(subj, s.get("class", ""))
                tk.Label(
                    tbl,
                    text=subj_label,
                    bg=row_bg,
                    fg=subject_style["dark_text"],
                    font=(FF, 10, "bold"),
                    padx=10,
                    pady=6,
                    anchor="w",
                ).grid(row=i + 1, column=0, sticky="nsew", padx=1, pady=1)
                for col in (1, 2):
                    tk.Label(
                        tbl,
                        text=str(mk),
                        bg=row_bg,
                        fg="#333",
                        font=(FF, 10),
                        padx=10,
                        pady=6,
                    ).grid(row=i + 1, column=col, sticky="nsew", padx=1, pady=1)
                tk.Label(
                    tbl,
                    text=self._get_grade_name_for_class(grade, s.get("class", "")),
                    bg=row_bg,
                    fg=PERF_CLR.get(grade_base_code(grade), PERF_CLR["IE"]),
                    font=(FF, 10, "italic"),
                    padx=10,
                    pady=6,
                    anchor="w",
                ).grid(row=i + 1, column=3, sticky="nsew", padx=1, pady=1)

            total_marks = result["total"]
            avg_score = result["average"]
            grade_ov = result["grade"]
            pos = result["position"]
            possible = result.get("possible_total", len(subjects) * 100)
            base = len(subjects) + 1
            for j, (lab, c1, c2, c3) in enumerate(
                [
                    (
                        "Total Scores",
                        f"{total_marks}/{possible}",
                        f"{avg_score}/100",
                        f"Termly Performance Level: {grade_ov}",
                    ),
                    (
                        "Average Scores",
                        f"{avg_score}/100",
                        "",
                        f"Position: {pos} of {total_students}",
                    ),
                ]
            ):
                tk.Label(
                    tbl,
                    text=lab,
                    bg=MINT_BG,
                    fg="#222",
                    font=(FF, 10, "bold"),
                    padx=10,
                    pady=7,
                    anchor="w",
                ).grid(row=base + j, column=0, sticky="nsew", padx=1, pady=1)
                tk.Label(
                    tbl,
                    text=c1,
                    bg=MINT_BG,
                    fg="#222",
                    font=(FF, 10, "bold"),
                    padx=10,
                    pady=7,
                ).grid(row=base + j, column=1, sticky="nsew", padx=1, pady=1)
                tk.Label(
                    tbl,
                    text=c2,
                    bg=MINT_BG,
                    fg="#222",
                    font=(FF, 10, "bold"),
                    padx=10,
                    pady=7,
                ).grid(row=base + j, column=2, sticky="nsew", padx=1, pady=1)
                tk.Label(
                    tbl,
                    text=c3,
                    bg=MINT_BG,
                    fg="#1b5e20",
                    font=(FF, 10, "bold"),
                    padx=10,
                    pady=7,
                    anchor="w",
                ).grid(row=base + j, column=3, sticky="nsew", padx=1, pady=1)
            base_row = base + 2

        # Performance Trend chart
        if not is_pp:
            tk.Label(
                parent,
                text="Performance Trend",
                bg="white",
                fg=SCH_BLUE,
                font=(FF, 12, "bold"),
            ).pack(pady=(10, 4))
            fig, ax = plt.subplots(figsize=(6.2, 2.8))
            fig.patch.set_facecolor("white")
            ax.set_facecolor("#fafafa")
            chart_subject_labels = [
                self._get_subject_label(sub, s.get("class", "")) for sub in subjects
            ]
            mark_vals = [marks.get(sub, 0) for sub in subjects]
            bar_colors = [
                self._get_subject_color(sub, s.get("class", "")) for sub in subjects
            ]
            ax.bar(
                chart_subject_labels,
                mark_vals,
                color=bar_colors,
                edgecolor="none",
                width=0.55,
            )
            ax.set_ylim(0, 105)
            ax.set_yticks([0, 25, 50, 75, 100])
            ax.tick_params(axis="x", rotation=45, labelsize=8, labelcolor="#444")
            ax.tick_params(axis="y", labelsize=8, labelcolor="#444")
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
            ax.spines["left"].set_color("#ddd")
            ax.spines["bottom"].set_color("#ddd")
            ax.grid(axis="y", color="#eeeeee", linewidth=0.7, zorder=0)
            fig.tight_layout(pad=1.0)
            chart_widget = FigureCanvasTkAgg(fig, master=parent)
            chart_widget.draw()
            chart_widget.get_tk_widget().pack(fill="x", pady=(0, 12))
            plt.close(fig)

        # Comments
        report_comment = context.get("comment_text") or ""
        comment_labels = (
            [("General Comments", "#fffff0", "#d4b800")]
            if is_pp
            else [
                ("Grade Facilitator's Comment", "#f0f8ff", "#90caf9"),
                ("Head Teacher's Comment", "#fffff0", "#d4b800"),
            ]
        )

        for clabel, bg_c, brd_c in comment_labels:
            outer = tk.Frame(parent, bg=brd_c)
            outer.pack(fill="x", pady=4)
            inner = tk.Frame(outer, bg=bg_c, padx=14, pady=10)
            inner.pack(fill="both", expand=True, padx=1, pady=1)
            row_f = tk.Frame(inner, bg=bg_c)
            row_f.pack(fill="x")
            tk.Label(
                row_f, text=f"{clabel}:", bg=bg_c, fg="#222", font=(FF, 10, "bold")
            ).pack(side="left")
            cmnt = report_comment if ("Class" in clabel or is_pp) else ""
            tk.Label(row_f, text=cmnt, bg=bg_c, fg="#555", font=(FF, 10)).pack(
                side="left"
            )

        if is_pp:
            # Signatures and Dates
            sig_f = tk.Frame(parent, bg="white")
            sig_f.pack(fill="x", pady=10)
            tk.Label(
                sig_f,
                text="Grade Facilitator Signature: ...................................",
                bg="white",
                font=(FF, 9),
            ).pack(side="left")
            tk.Label(
                sig_f,
                text="Head Teacher Signature: ...................................",
                bg="white",
                font=(FF, 9),
            ).pack(side="right")

            date_f = tk.Frame(parent, bg="white")
            date_f.pack(fill="x", pady=5)
            tk.Label(
                date_f, text="Opening Date: 28/04/2026", bg="white", font=(FF, 9)
            ).pack(side="left")
            tk.Label(
                date_f, text="Closing Date: 31/03/2026", bg="white", font=(FF, 9)
            ).pack(side="right")

        tk.Frame(parent, bg="white", height=8).pack()
        if not is_pp:
            today = datetime.now().strftime("%m/%d/%Y")
            tk.Label(
                parent,
                text=f"This term closed on: {today}    |    Next term opens on: ___________",
                bg="white",
                fg="#666",
                font=(FF, 9),
            ).pack()
            tk.Label(
                parent,
                text=(
                    "This Exam Report Card has been Issued Without Any Alterations "
                    "Whatsoever. Any Alterations Will Invalidate Its Authenticity."
                ),
                bg="white",
                fg="red",
                font=(FF, 8, "italic"),
                wraplength=480,
                justify="center",
            ).pack(pady=(4, 0))

        # Footer contacts should sit at the very bottom and span wider.
        footer_path = letterhead_assets.get("footer_path")
        if footer_path and os.path.exists(footer_path):
            try:
                footer_img = Image.open(footer_path)
                footer_width = 620
                footer_img = footer_img.resize(
                    (
                        footer_width,
                        int(footer_img.height * footer_width / footer_img.width),
                    ),
                    Image.LANCZOS,
                )
                footer_photo = ImageTk.PhotoImage(footer_img)
                footer_label = tk.Label(parent, image=footer_photo, bg="white")
                footer_label.image = footer_photo  # Keep reference
                footer_label.pack(pady=(10, 0))
            except Exception as e:
                print(f"Failed to load letterhead footer: {e}")

    def _gen_rc_text(self, result, total, term, exam_type=DEFAULT_EXAM_TYPE):
        """Plain-text fallback used for printing."""
        s, m = result["student"], result["marks"]
        context = self._get_report_card_context(result, term, exam_type)
        is_pp = context["is_pp"]
        subjects = context["subjects"]
        comment_text = context.get("comment_text") or ""
        year_text = context.get("year", str(datetime.now().year))

        possible = result.get("possible_total", len(subjects) * 100)
        title = (
            "ASSESSMENT SUMMARY REPORT PRE-PRIMARY"
            if is_pp
            else "LEARNER ASSESSMENT REPORT CARD"
        )
        assessment_specs = self._get_report_assessment_specs(
            assessment_types=self._get_student_term_marks(
                s["id"], term, context.get("year")
            ).keys(),
            include_exam_type=exam_type,
        )
        assessment_order = [spec["key"] for spec in assessment_specs]
        school_name = get_school_profile().get(
            "school_name", DEFAULT_SCHOOL_PROFILE["school_name"]
        )
        lines = [
            "=" * 62,
            f"      {school_name}",
            f"          {title}",
            "=" * 62,
            f"  Name    : {s['name']:<20}  Grade  : {s.get('class', '').replace('Grade ', '')}",
        ]
        if is_pp:
            lines += [
                f"  Year    : {year_text:<19} Term   : {term.upper()}",
                f"  Gender  : {s['gender']:<20}",
            ]
        else:
            lines += [
                f"  Stream  : {s['admission_no']:<20}  Year   : {year_text}",
                f"  Term    : {f'{term} / {exam_type}':<20}  Gender : {s['gender']}",
            ]

        lines.append("-" * 62)
        if is_pp:
            assessment_headers = " ".join(
                f"{self._format_assessment_ordinal(idx):<6}"
                for idx in range(1, len(assessment_order) + 1)
            )
            lines.append(f"  {'Learning Area':<30} {assessment_headers}".rstrip())
        else:
            lines.append(
                f"  {'Subject':<16} {'Marks':>6}  {'Avg':>6}  Performance Level"
            )
        lines.append("-" * 62)

        if is_pp:
            term_marks = self._get_student_term_marks(
                s["id"], term, context.get("year")
            )
            for sub in subjects:
                row = f"  {sub[:30]:<30} "
                for et in assessment_order:
                    mk = term_marks.get(et, {}).get(sub, "")
                    g = (
                        self._get_grade_code_for_class(mk, s.get("class", ""))
                        if mk
                        else ""
                    )
                    row += f"{g:<6}"
                lines.append(row)
        else:
            for sub in subjects:
                mk = m.get(sub, 0)
                g = self._get_grade_code_for_class(mk, s.get("class", ""))
                sub_label = self._get_subject_label(sub, s.get("class", ""))
                lines.append(
                    f"  {sub_label:<16} {mk:>6}  {mk:>6}  {self._get_grade_name_for_class(g, s.get('class', ''))}"
                )

        lines += ["-" * 62]
        if not is_pp:
            lines += [
                f"  Total   : {result['total']}/{possible}   Average: {result['average']}/100",
                f"  Grade   : {result['grade']}   Position: {result['position']} of {total}",
            ]
        else:
            lines += [f"  Comments: {comment_text}"]
        lines += ["=" * 62]
        return "\n".join(lines)

    def _print_rc(self):
        name = self.rc_stu_cb.get()
        if not name:
            return
        exam_type = self.rc_exam_cb.get() or DEFAULT_EXAM_TYPE
        cls = self.rc_cls_cb.get()
        term = self.rc_term_cb.get()
        results = self._get_report_card_results()
        result = next((r for r in results if r["student"]["name"] == name), None)
        if not result:
            return
        s = result["student"]
        stream = self._get_selected_report_stream() or s.get("stream", "").strip()
        year_text = self._get_report_card_context(result, term, exam_type).get(
            "year", str(datetime.now().year)
        )
        file_path = filedialog.asksaveasfilename(
            title="Save Report Card PDF",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=f"{self._get_report_card_file_basename(s, term, year_text)}.pdf",
        )
        if not file_path:
            return
        if self._build_report_card_pdf(
            result, len(results), term, exam_type, file_path
        ):
            messagebox.showinfo("Done", f"Report card PDF saved to {file_path}")

    def _print_all_rc(self):
        cls = self.rc_cls_cb.get()
        term = self.rc_term_cb.get()
        exam_type = self.rc_exam_cb.get() or DEFAULT_EXAM_TYPE
        selected_stream = self._get_selected_report_stream()
        results = self._get_report_card_results()
        if not results:
            messagebox.showwarning("No Data", "No students found")
            return
        output_dir = filedialog.askdirectory(title="Select Folder for Report Card PDFs")
        if not output_dir:
            return

        saved = 0
        for r in results:
            student = r["student"]
            year_text = self._get_report_card_context(r, term, exam_type).get(
                "year", str(datetime.now().year)
            )
            file_path = os.path.join(
                output_dir,
                f"{self._get_report_card_file_basename(student, term, year_text)}.pdf",
            )
            if self._build_report_card_pdf(r, len(results), term, exam_type, file_path):
                saved += 1

        if saved:
            messagebox.showinfo(
                "Done", f"{saved} report card PDF(s) saved to {output_dir}"
            )


# ====================== ENTRY POINT ========================
def main():
    # Initialize school configuration from database (or seed defaults if empty)
    _seed_school_config_to_db()
    refresh_dynamic_school_config()
    
    root = tk.Tk()

    # Get the directory where the script is located
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # Set the app logo
    try:
        logo_path = os.path.join(script_dir, "moas.jpg")
        logo_img = Image.open(logo_path)
        # Resize for icon
        logo_img = logo_img.resize((64, 64), Image.Resampling.LANCZOS)
        logo_photo = ImageTk.PhotoImage(logo_img)
        root.iconphoto(True, logo_photo)
    except Exception as e:
        print(f"Could not load logo: {e}")

    app = SchoolReportApp(root)

    try:
        root.mainloop()
    except KeyboardInterrupt:
        print("\nKeyboardInterrupt received. Shutting down gracefully...")
    finally:
        # Cleanup code - ensure proper shutdown
        print("Cleanup complete. Goodbye!")
        try:
            app.shutdown()
        except Exception:
            pass
        sys.exit(0)


if __name__ == "__main__":
    main()
